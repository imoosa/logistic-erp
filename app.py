from flask import Flask, render_template, render_template_string, request, redirect, url_for, session, flash, jsonify, send_file
from flask import abort
from datetime import date, datetime, timedelta
import random
import hashlib
import secrets
from functools import wraps
from itsdangerous import URLSafeTimedSerializer, BadSignature, SignatureExpired
import os
import json
import re
import math
import pandas as pd
from werkzeug.utils import secure_filename
import io
import base64
from flask import url_for
import base64
from sqlalchemy import text, func, and_, or_
from platform_models import db, SubscriptionPlan, RegisteredUser, Company
import time
import threading
from customer_models import (
    CompanyUser, Client, StockItem,
    Invoice, InvoiceItem,
    Estimate, EstimateItem,
    PurchaseInvoice, PurchaseInvoiceItem, StockPurchaseHistory,
    CashTransaction, Loan, LoanRepayment,
    BankAccount, BankTransaction, CompanyManifest, ManifestEntry, Expense, Supplier, SupplierBrand,
    PriceList, RateLookup, Cheque, CompanyRolePermission, PurchasePayment, WhatsAppLog, StatementClosing,
    DeletedInvoiceLog, CustomerInvoice, CustomerInvoiceItem )
from db_router import get_customer_session, get_customer_session_with_retry, init_customer_db_for_company
from backup_utils import BACKUP_DESTINATIONS
import permissions as perms_module
import permissions as perms_module
from permissions import (
    get_field_permissions,
    can_edit_field,
    can_view_field,
    INVOICE_FIELDS,
    DEFAULT_FIELD_PERMISSIONS
)
from flask_mail import Mail, Message
from utils.ai_assistant import LogisticsAIAssistant
from utils.intent_router import *
from utils.self_learning_ai import SelfLearningAssetAI
from sqlalchemy.exc import OperationalError
from dotenv import load_dotenv
load_dotenv()

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY")
if not app.secret_key:
    raise RuntimeError("FLASK_SECRET_KEY not set in environment")

# --- timezone helpers (server runs in UTC; business timezone is IST) ---
from datetime import timezone as _timezone

IST = _timezone(timedelta(hours=5, minutes=30))

def to_ist(dt):
    """Convert a datetime to IST for display. Assumes naive datetimes are UTC."""
    if dt is None:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=_timezone.utc)
    return dt.astimezone(IST)

def today_ist():
    """Current calendar date in IST. Use this instead of date.today() everywhere."""
    return datetime.now(_timezone.utc).astimezone(IST).date()

app.jinja_env.filters['ist'] = to_ist
# --- end timezone helpers ---

app.config['MAIL_SERVER'] = 'smtp.hostinger.com'
app.config['MAIL_PORT'] = 465
app.config['MAIL_USE_TLS'] = False
app.config['MAIL_USE_SSL'] = True
app.config['MAIL_USERNAME'] = os.getenv("MAIL_USERNAME")
app.config['MAIL_PASSWORD'] = os.getenv("MAIL_PASSWORD")
app.config['MAIL_DEFAULT_SENDER'] = 'support@magnustic.com'
app.config['SESSION_COOKIE_MAX_SIZE'] = 4000


mail = Mail(app)
_ai_assistant = LogisticsAIAssistant(model_name="llama3.2")


_submission_lock = threading.Lock()
_SUBMISSION_TTL_SECONDS = 120 

def send_otp_email(to_email, otp_code):
    # Get company logo path (for the current company)
    company_id = get_current_company()
    company = Company.query.filter_by(company_id=company_id).first()
    
    # Try to get logo from company settings first
    logo_base64 = None
    if company and company.logo_filename:
        logo_path = os.path.join('static', 'company_logos', company.logo_filename)
        if os.path.exists(logo_path):
            with open(logo_path, 'rb') as f:
                logo_base64 = base64.b64encode(f.read()).decode('utf-8')
    
    # Fallback to default Magnustic logo
    if not logo_base64:
        default_logo_path = os.path.join('static', 'logo.png')
        if os.path.exists(default_logo_path):
            with open(default_logo_path, 'rb') as f:
                logo_base64 = base64.b64encode(f.read()).decode('utf-8')
    
    # If no logo found, use a text placeholder
    logo_html = f'''
    <div style="background: #1a237e; color: white; padding: 12px 24px; border-radius: 6px; display: inline-block; font-weight: bold; font-size: 20px; letter-spacing: 1px;">
        MAGNUSTIC ERP
    </div>
    ''' if not logo_base64 else f'''
    <img src="data:image/png;base64,{logo_base64}" alt="Magnustic Logo" style="max-width: 180px; height: auto;">
    '''
    
    # Get company name for personalization
    company_name = company.company_name if company else "Magnustic ERP"
    
    msg = Message(
        subject=f"Verify your email — {company_name}",
        recipients=[to_email],
        sender=app.config['MAIL_DEFAULT_SENDER']
    )
    
    msg.html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
    </head>
    <body style="font-family: Arial, Helvetica, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px; background: #f9fafb; color: #333;">
        <table width="100%" cellpadding="0" cellspacing="0" style="background: #ffffff; border-radius: 12px; box-shadow: 0 2px 8px rgba(0,0,0,0.08); padding: 40px 30px;">
            <tr>
                <td style="text-align: center; padding-bottom: 20px;">
                    {logo_html}
                </td>
            </tr>
            <tr>
                <td style="padding: 10px 0;">
                    <h2 style="color: #1a237e; font-size: 22px; margin: 0 0 8px 0;">Verification Code</h2>
                    <p style="color: #666; font-size: 14px; margin: 0;">Use this code to complete your login</p>
                </td>
            </tr>
            <tr>
                <td style="padding: 20px 0;">
                    <p style="font-size: 15px; color: #444; margin: 0 0 10px 0;">Dear Valued User,</p>
                    <p style="font-size: 15px; color: #444; margin: 0 0 20px 0;">Your verification code for {company_name} is:</p>
                    
                    <div style="background: #f5f7fa; padding: 18px; text-align: center; border-radius: 8px; border: 2px dashed #1a237e; margin: 10px 0 20px 0;">
                        <span style="font-size: 32px; font-weight: 700; color: #1a237e; letter-spacing: 6px; font-family: 'Courier New', monospace;">
                            {otp_code}
                        </span>
                    </div>
                    
                    <p style="font-size: 14px; color: #666; margin: 0 0 5px 0;">
                        <strong>⏱️ This code expires in 10 minutes</strong>
                    </p>
                    <p style="font-size: 13px; color: #888; margin: 0 0 20px 0;">
                        For security, please do not share this code with anyone.
                    </p>
                </td>
            </tr>
            <tr>
                <td style="padding: 15px 0; border-top: 1px solid #e5e7eb;">
                    <p style="font-size: 14px; color: #555; margin: 0 0 4px 0;">
                        Thank you for choosing <strong style="color: #1a237e;">{company_name}</strong>.
                    </p>
                    <p style="font-size: 14px; color: #555; margin: 0 0 4px 0;">
                        We appreciate your trust in us.
                    </p>
                </td>
            </tr>
            <tr>
                <td style="padding: 10px 0 0 0; border-top: 1px solid #e5e7eb;">
                    <p style="font-size: 14px; color: #555; margin: 10px 0 0 0;">
                        Best regards,<br>
                        <strong style="color: #1a237e; font-size: 15px;">Team Magnustic</strong>
                    </p>
                </td>
            </tr>
            <tr>
                <td style="padding: 20px 0 0 0; text-align: center; border-top: 1px solid #e5e7eb;">
                    <p style="font-size: 12px; color: #999; margin: 0;">
                        © 2026 {company_name}. All rights reserved.
                    </p>
                    <p style="font-size: 12px; color: #999; margin: 4px 0 0 0;">
                        <a href="https://www.magnustic.com" style="color: #1a237e; text-decoration: none;">www.magnustic.com</a>
                    </p>
                    <p style="font-size: 11px; color: #bbb; margin: 8px 0 0 0;">
                        This is an automated message, please do not reply to this email.
                    </p>
                </td>
            </tr>
        </table>
    </body>
    </html>
    """
    
    # Plain text fallback
    msg.body = f"""
    {company_name} - Verification Code
    
    Dear Valued User,
    
    Your verification code is: {otp_code}
    This code expires in 10 minutes.
    
    For security, please do not share this code with anyone.
    
    Thank you for choosing {company_name}.
    
    Best regards,
    Team Magnustic
    www.magnustic.com
    """
    
    mail.send(msg)
   

def _generate_and_send_otp(email):
    try:
        otp = f"{secrets.randbelow(1000000):06d}"
        session["otp_email"]   = email
        session["otp_hash"]    = hashlib.sha256(otp.encode()).hexdigest()
        session["otp_expires"] = (datetime.utcnow() + timedelta(minutes=10)).isoformat()
        
        # Try to send email with error handling
        try:
            send_otp_email(email, otp)
            print(f"✅ OTP sent to {email}: {otp}")  # Debug - remove in production
        except Exception as e:
            print(f"❌ Failed to send OTP email: {e}")
            # For testing, we can still allow login with a fallback
            # Store the OTP in session anyway so user can see it in logs
            flash(f"Email sending failed. For testing, your OTP is: {otp}", "warning")
    except Exception as e:
        print(f"❌ Error in _generate_and_send_otp: {e}")
        flash("Error sending verification code. Please try again.", "error")

def _finish_owner_login(reg_user):
    companies = get_owner_companies(reg_user.email)
    if len(companies) == 0:
        session["user"] = {"user_id": reg_user.user_id, "email": reg_user.email,
                            "full_name": reg_user.full_name, "role": reg_user.role,
                            "company_id": None}
        return redirect(url_for("onboard_company"))
    elif len(companies) == 1:
        c = companies[0]
        session["user"] = {"user_id": reg_user.user_id, "email": reg_user.email,
                            "full_name": reg_user.full_name, "role": reg_user.role,
                            "company_id": c.company_id}
        session["active_company_id"] = c.company_id
        session.pop("pending_login_type", None)
        return redirect(url_for("dashboard"))
    else:
        session["pending_login_email"] = reg_user.email
        return redirect(url_for("select_company"))

@app.template_filter('from_json')
def from_json_filter(value):
    """Parse JSON string to Python object in templates"""
    if not value:
        return {}
    try:
        return json.loads(value)
    except (ValueError, TypeError, json.JSONDecodeError):
        return {}

# Also add a filter for JSON parsing with default
@app.template_filter('json_loads')
def json_loads_filter(value, default=None):
    """Parse JSON string to Python object in templates"""
    if not value:
        return default or {}
    try:
        return json.loads(value)
    except (ValueError, TypeError, json.JSONDecodeError):
        return default or {}

# ── Database Configuration ────────────────────────────────────────────────────
PLATFORM_DB_URI = os.environ.get(
    "PLATFORM_DB_URI",
    "sqlite:///platform.db"   # ← change this default
)
app.config["SQLALCHEMY_DATABASE_URI"] = PLATFORM_DB_URI
app.config["SQLALCHEMY_BINDS"] = {}           
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {
    "pool_pre_ping": True,  
    "pool_recycle": 3600,   
    "pool_size": 10,
    "max_overflow": 20,
}
db.init_app(app)

@app.before_request
def _fk_on():
    pass  # MySQL enforces FK by default; no PRAGMA needed

with app.app_context():
    db.create_all()

    # ── Patch existing companies table for new columns ─────────────────────
    # db.create_all() only creates tables that don't exist yet — it never
    # alters a `companies` table that's already there, so a new column added
    # to the Company model (like credit_limit_action) silently does nothing
    # on every database except a brand new one. This app has no migration
    # framework for the platform DB (schema_migrations further down is
    # per-tenant customer DBs only), so patch it in here directly, checking
    # first so this is a no-op once the column actually exists.
    try:
        from sqlalchemy import inspect as _sa_inspect
        _existing_cols = {c["name"] for c in _sa_inspect(db.engine).get_columns("companies")}
        if "credit_limit_action" not in _existing_cols:
            db.session.execute(text(
                "ALTER TABLE companies ADD COLUMN credit_limit_action "
                "VARCHAR(10) NOT NULL DEFAULT 'warn'"
            ))
            db.session.commit()
            print("✅ Added missing companies.credit_limit_action column")
    except Exception as e:
        db.session.rollback()
        print(f"⚠  Could not verify/add companies.credit_limit_action column: {e}")


# ── Create tables and seed on first startup ────────────────────────────────────
with app.app_context():
    # Only create platform tables - customer DBs are created per-company
    db.create_all()

UPLOAD_FOLDER = 'uploads/purchase_invoices'
ALLOWED_EXTENSIONS = {
    'png',
    'jpg',
    'jpeg',
    'pdf',
    'tiff',
    'bmp',
    'xlsx',
    'xls'
}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 10 * 1024 * 1024  # 10MB

os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# ── Company logo uploads (served from /static so no extra route is needed) ──
LOGO_UPLOAD_FOLDER = os.path.join('static', 'company_logos')
ALLOWED_LOGO_EXTENSIONS = {'png', 'jpg', 'jpeg', 'webp'}
os.makedirs(LOGO_UPLOAD_FOLDER, exist_ok=True)

ID_DOCS_UPLOAD_FOLDER = os.path.join('static', 'invoice_docs')
ALLOWED_ID_DOC_EXTENSIONS = {'png', 'jpg', 'jpeg'}
os.makedirs(ID_DOCS_UPLOAD_FOLDER, exist_ok=True)
CLIENT_DOCS_UPLOAD_FOLDER = os.path.join('static', 'client_docs')
os.makedirs(CLIENT_DOCS_UPLOAD_FOLDER, exist_ok=True)

def allowed_logo_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_LOGO_EXTENSIONS

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def allowed_id_doc_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_ID_DOC_EXTENSIONS
# ── Helper / Auth ─────────────────────────────────────────────────────────────
def hash_password(password: str) -> str:
    return hashlib.sha256(password.encode()).hexdigest()

def verify_password(password: str, hashed: str) -> bool:
    return hash_password(password) == hashed


# Add this after the existing template filters in app.py

def wordize_number(n):
    """Convert a number to words (Indian English)."""
    if n is None:
        return "Zero"
    n = int(abs(n))
    if n == 0:
        return "Zero"
    
    ones = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine",
            "Ten", "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen",
            "Seventeen", "Eighteen", "Nineteen"]
    tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]
    
    def words_under_hundred(n):
        if n < 20:
            return ones[n]
        return tens[n // 10] + (" " + ones[n % 10] if n % 10 else "")
    
    def words_under_thousand(n):
        if n < 100:
            return words_under_hundred(n)
        return ones[n // 100] + " Hundred" + (" " + words_under_hundred(n % 100) if n % 100 else "")
    
    def words_under_lakh(n):
        if n < 1000:
            return words_under_thousand(n)
        return words_under_hundred(n // 1000) + " Thousand" + (" " + words_under_thousand(n % 1000) if n % 1000 else "")
    
    def words_under_crore(n):
        if n < 100000:
            return words_under_lakh(n)
        return words_under_hundred(n // 100000) + " Lakh" + (" " + words_under_lakh(n % 100000) if n % 100000 else "")
    
    def words_under_arab(n):
        if n < 10000000:
            return words_under_crore(n)
        return words_under_hundred(n // 10000000) + " Crore" + (" " + words_under_crore(n % 10000000) if n % 10000000 else "")
    
    return words_under_arab(n)

@app.template_filter('wordize')
def wordize_filter(value):
    """Jinja2 filter to convert numbers to words."""
    return wordize_number(value)

def generate_next_user_id():
    """Next USRxxx id, derived from the highest existing numeric suffix —
    NOT from RegisteredUser.query.count(). Count-based generation breaks the
    moment any user is deleted: count() drops, but the highest issued id
    doesn't, so count()+1 collides with an id that still exists."""
    max_num = 0
    for (uid,) in RegisteredUser.query.with_entities(RegisteredUser.user_id).all():
        if uid and uid.startswith("USR"):
            try:
                max_num = max(max_num, int(uid[3:]))
            except ValueError:
                continue
    return f"USR{max_num + 1:03d}"


def save_shipper_id_doc(file_storage, invoice_id, doc_label, old_filename=None):
    """Saves an Aadhaar/PAN upload as '<invoice_id>_<doc_label>.<ext>'.
    Returns the new filename, or old_filename unchanged if nothing was uploaded."""
    if not file_storage or not file_storage.filename:
        return old_filename
    if not allowed_id_doc_file(file_storage.filename):
        flash(f"{doc_label.upper()} file must be a PNG or JPG image.")
        return old_filename
    ext = file_storage.filename.rsplit('.', 1)[1].lower()
    new_filename = secure_filename(f"{invoice_id}_{doc_label}.{ext}")
    if old_filename and old_filename != new_filename:
        old_path = os.path.join(ID_DOCS_UPLOAD_FOLDER, old_filename)
        if os.path.exists(old_path):
            os.remove(old_path)
    file_storage.save(os.path.join(ID_DOCS_UPLOAD_FOLDER, new_filename))
    return new_filename

# ADD (whole function, new)
def save_client_id_doc(file_storage, client_id, doc_label, old_filename=None):
    """Same contract as save_shipper_id_doc(), but for the permanent copy that
    lives on a credit client's own record (static/client_docs/) instead of a
    single booking's (static/invoice_docs/). client_id here is Client.client_id
    (e.g. 'ACM001'), not the numeric PK, so filenames stay readable."""
    if not file_storage or not file_storage.filename:
        return old_filename
    if not allowed_id_doc_file(file_storage.filename):
        flash(f"{doc_label.upper()} file must be a PNG or JPG image.")
        return old_filename
    ext = file_storage.filename.rsplit('.', 1)[1].lower()
    new_filename = secure_filename(f"{client_id}_{doc_label}.{ext}")
    if old_filename and old_filename != new_filename:
        old_path = os.path.join(CLIENT_DOCS_UPLOAD_FOLDER, old_filename)
        if os.path.exists(old_path):
            os.remove(old_path)
    file_storage.save(os.path.join(CLIENT_DOCS_UPLOAD_FOLDER, new_filename))
    return new_filename


def _save_client_id_docs(cdb, client_row, form_files):
    """Handles all 4 client ID-doc uploads for a Client row and commits.
    Shared by client_new() and client_edit() so the save logic can't drift."""
    client_row.aadhar_front_file = save_client_id_doc(
        form_files.get("aadhar_front_file"), client_row.client_id, "aadhar_front", client_row.aadhar_front_file)
    client_row.aadhar_back_file = save_client_id_doc(
        form_files.get("aadhar_back_file"), client_row.client_id, "aadhar_back", client_row.aadhar_back_file)
    client_row.pan_front_file = save_client_id_doc(
        form_files.get("pan_front_file"), client_row.client_id, "pan_front", client_row.pan_front_file)
    client_row.pan_back_file = save_client_id_doc(
        form_files.get("pan_back_file"), client_row.client_id, "pan_back", client_row.pan_back_file)
    cdb.commit()

def _next_numbered_id(session, column, prefix, pad=3, extra_filters=None):
    """
    Safe replacement for count()-based ID generation (e.g. `f"CUST-{count+1:03d}"`).
    That pattern breaks two ways: (1) after any row for this prefix is deleted,
    count() drops but the highest issued number doesn't, so count()+1 collides
    with an id that still exists; (2) under two near-simultaneous requests
    (double-submit, two tabs), both read the same count before either commits,
    so both compute the same next id and the second INSERT fails with a
    duplicate-key IntegrityError.

    This finds the highest existing numeric suffix among ids starting with
    `prefix` and returns prefix + (max + 1). It fixes the delete problem
    outright. It reduces but does not eliminate the race — two truly
    simultaneous requests can still read the same max before either commits.
    For routes with meaningful concurrent-write risk, pair this with a
    retry-on-IntegrityError loop at the call site (catch, session.rollback(),
    regenerate the id, retry once or twice).
    """
    q = session.query(column).filter(column.like(f"{prefix}%"))
    if extra_filters:
        q = q.filter(*extra_filters)
    max_num = 0
    for (val,) in q.all():
        if not val:
            continue
        tail = val[len(prefix):]
        try:
            max_num = max(max_num, int(tail))
        except ValueError:
            continue
    return f"{prefix}{max_num + 1:0{pad}d}"


def _company_name_prefix(company_name, chars=3, from_end=False):
    """
    Derive a 3-letter id prefix from the company name, letters only
    (spaces/punctuation stripped), lowercased.
    from_end=False -> first N letters  (used for supplier ids)
    from_end=True  -> last N letters   (used for client ids)
    e.g. "Demo"      -> "dem" (supplier) / "emo" (client)
         "Magnustic" -> "mag" (supplier) / "tic" (client)
    Falls back to "co" + letters if the name has fewer than `chars` letters.
    """
    letters = "".join(ch for ch in (company_name or "") if ch.isalpha())
    if not letters:
        letters = "co"
    if len(letters) < chars:
        letters = letters.ljust(chars, "x")
    seg = letters[-chars:] if from_end else letters[:chars]
    return seg.lower()


def get_current_user():
    return session.get("user", {})

@app.context_processor
def inject_user():
    return {
        "user": session.get("user", {}),
    }

@app.context_processor
def inject_today():
    """Inject today's date for the subscription banner and other uses"""
    # Named distinctly from the many routes that pass their own
    # today=str(...)/isoformat() kwarg into render_template() for
    # date-input defaults. Those route-level kwargs are for form
    # fields; this one is exclusively for the subscription banner
    # in base.html and must never be shadowed by a string.
    return {"subscription_today": today_ist()}

@app.context_processor
def inject_company_settings():
    company_id = get_current_company()
    is_gst = True  # default safe
    co = None
    if company_id:
        co = Company.query.filter_by(company_id=company_id).first()
        if co and hasattr(co, 'is_gst_registered'):
            is_gst = bool(co.is_gst_registered)
    logo_url = None
    if co and getattr(co, 'logo_filename', None):
        logo_url = url_for('static', filename=f'company_logos/{co.logo_filename}')
    return {'is_gst_registered': is_gst, 'company': co, 'company_logo_url': logo_url}

def get_current_company():
    return session.get("active_company_id") or session.get("user", {}).get("company_id")

@app.context_processor
def inject_field_permissions():
    """Inject field-level permission helpers for templates"""
    user = get_current_user()
    role = user.get("role", "employee")
    user_id = user.get("user_id")
    company_id = get_current_company()
    
    # Get CDB for permission lookups
    cdb = None
    if company_id:
        try:
            cdb = get_customer_session(company_id)
        except:
            pass
    
    def _can_edit_field(field_group):
        return can_edit_field(role, field_group, user_id, company_id, cdb)
    
    def _can_view_field(field_group):
        return can_view_field(role, field_group, user_id, company_id, cdb)
    
    return {
        "can_edit_field": _can_edit_field,
        "can_view_field": _can_view_field,
        "field_permissions": get_field_permissions(role, user_id, company_id, cdb)
    }

@app.errorhandler(OperationalError)
def handle_db_operational_error(e):
    """Handle stale database connections by retrying once"""
    # Check if it's a connection-related error
    if "2006" in str(e) or "2013" in str(e) or "MySQL server has gone away" in str(e):
        # Clear any problematic sessions
        from db_router import _engine_cache, _session_cache
        for company_id in list(_session_cache.keys()):
            try:
                _session_cache[company_id].remove()
            except:
                pass
            # Recreate the engine
            if company_id in _engine_cache:
                _engine_cache[company_id].dispose()
                del _engine_cache[company_id]
        
        flash("Database connection was re-established. Please try again.", "info")
        return redirect(request.url)
    raise e

@app.before_request
def _clear_stale_customer_session():
    """
    Defense against the one case teardown_request can't cover: if the
    gunicorn worker handling a previous request was hard-killed mid-request
    (a --timeout hit on a slow query, or an OOM kill), teardown_request
    never runs, and that company's cached session is left mid-transaction
    for whatever request happens to land on this worker next.
    """
    company_id = get_current_company()
    if not company_id:
        return
    
    try:
        from db_router import _session_cache
        factory = _session_cache.get(company_id)
        if factory is None:
            return
        factory().rollback()
    except Exception:
        pass

@app.teardown_request
def _rollback_customer_session_on_error(exc):
    """
    db_router caches one SQLAlchemy session per company_id and reuses it
    across requests. If any commit on that session fails — an IntegrityError,
    a stale overnight connection, a bad form value, anything — and nothing
    calls rollback() on it, the session is left mid-transaction.

    teardown_request runs after every request — success or failure — and
    unconditionally rolls back the current company's SESSION (not creating
    a new one, but using the cached one). rollback() on a session with no
    open transaction is a harmless no-op.
    """
    company_id = get_current_company()
    if not company_id:
        return
    
    try:
        from db_router import _session_cache
        factory = _session_cache.get(company_id)
        if factory is None:
            return  # nothing cached for this company — nothing to roll back
        factory().rollback()
    except Exception:
        # Don't let cleanup itself take down the error response.
        pass

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user" not in session:
            flash("Please login to continue")
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

def generate_pdf_token(company_id, invoice_id):
    """Signed, time-limited token so WhatsApp's servers can fetch an invoice
    PDF without a login session/cookie."""
    s = URLSafeTimedSerializer(app.secret_key, salt="invoice-pdf")
    return s.dumps({"company_id": company_id, "invoice_id": invoice_id})

def verify_pdf_token(token, max_age=7 * 24 * 3600):
    s = URLSafeTimedSerializer(app.secret_key, salt="invoice-pdf")
    try:
        data = s.loads(token, max_age=max_age)
        return data.get("company_id"), data.get("invoice_id")
    except (BadSignature, SignatureExpired):
        return None, None

def owner_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        user = get_current_user()
        if user.get("role") not in ["owner", "super_admin"]:
            flash("Only company owner can access this page")
            return safe_redirect_after_denial()
        return f(*args, **kwargs)
    return decorated

def super_admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        user = get_current_user()
        if user.get("role") != "super_admin":
            flash("Super admin access required")
            return safe_redirect_after_denial()
        return f(*args, **kwargs)
    return decorated


MODULE_LANDING_ENDPOINT = {
    "dashboard":         "dashboard",
    "analytics":         "reports_dashboard",
    "clients":           "client_list",
    "suppliers":         "supplier_list",
    "estimates":         "estimate_list",
    "stock":             "inventory_list",
    "pricelist":         "price_lists",
    "manifest":          "manifest_list",
    "invoices":          "invoice_list",
    "purchase":          "purchase_invoice_list",
    "creditors":         "creditors_list",
    "debtors":           "debtors_list",
    "expenses":          "expenses",
    "cash":              "cash_in_hand",
    "bank":              "bank_accounts",
    "cheques":           "cheques",
    "loans":             "loan_accounts",
    "receipts_payments": "receipt_new",
    "backup":            "backup",
}


# Preferred "home" module per role — tried before the generic scan below.
ROLE_HOME_MODULE = {
    "employee": "invoices",   # sales lands on their booking invoices, not clients
}


def safe_redirect_after_denial():
    """Where to send someone after an access check fails. NEVER redirects
    back to 'dashboard' blindly — if the person can't see the dashboard
    either, that would just loop forever. Tries the role's preferred home
    module first, then the first module they can view at all, then a
    permission-free 'no access' page."""
    role = get_current_user().get("role")
    home_module = ROLE_HOME_MODULE.get(role)
    if home_module and has_permission(home_module, "view"):
        return redirect(url_for(MODULE_LANDING_ENDPOINT[home_module]))
    for module, endpoint in MODULE_LANDING_ENDPOINT.items():
        if has_permission(module, "view"):
            return redirect(url_for(endpoint))
    return redirect(url_for("no_access"))


@app.route("/no-access")
@login_required
def no_access():
    return render_template("no_access.html")


# ── Module permissions (view/create/edit only — no delete) ───────────────────
def get_effective_permissions(user=None):
    """Full view/create/edit matrix for the current (or given) user.
    Owners and super_admin get None back, which callers should treat as
    'everything allowed' — they never go through the matrix."""
    user = user or get_current_user()
    role = user.get("role")
    if role in ("owner", "super_admin"):
        return None
    company_id = get_current_company()
    if not company_id:
        return perms_module.default_permissions_for(role)
    cdb = get_customer_session(company_id)
    return perms_module.get_effective_permissions(
        role, company_id, user.get("user_id"), cdb,
        CompanyRolePermission, CompanyUser,
    )

def _calculate_client_pending(cdb, company_id, client_id):
    """Calculate the true pending balance from non-void invoices"""
    total = cdb.query(func.sum(Invoice.balance)).filter(
        Invoice.company_id == company_id,
        Invoice.client_id == client_id,
        Invoice.status.notin_(['Void', 'Cancelled'])
    ).scalar() or 0
    return total

def has_permission(module, action="view"):
    user = get_current_user()
    if user.get("role") in ("owner", "super_admin"):
        return True
    perms = get_effective_permissions(user)
    return bool(perms.get(module, {}).get(action, False))


def require_permission(module, action="view", method_actions=None):
    """Gate a route on a (module, action) pair. Pass method_actions={'POST': 'create'}
    etc. when a single route handles both showing a form (view) and submitting it
    (create/edit) so each HTTP method is checked against the right action."""
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            act = action
            if method_actions and request.method in method_actions:
                act = method_actions[request.method]
            if not has_permission(module, act):
                flash("You don't have permission to access this.")
                return safe_redirect_after_denial()
            return f(*args, **kwargs)
        return decorated
    return decorator


@app.context_processor
def inject_permission_helper():
    return {"can": has_permission}

@app.context_processor
def inject_today_global():
    """Inject today's date in IST for all templates"""
    t = today_ist()
    return {
        'today_ist': today_ist,
        'today_str': t.isoformat(),
        'today_display': t.strftime('%d %b %Y'),
        'today_year': t.year,
        'today_month': t.month,
        'today_day': t.day,
        'timedelta': timedelta,
    }

def _get_awb(invoice):
    """Extract docket_no from invoice.terms JSON. Returns '' if absent."""
    try:
        if invoice.terms:
            meta = json.loads(invoice.terms)
            return meta.get("docket_no", "")
    except Exception:
        pass
    return ""

def _get_shipment_meta(invoice):
    """
    Pull the AWB/consignee/destination/carrier-ref block out of invoice.terms
    JSON in one shot, for statement rows that need all four instead of just
    the AWB. "Consignee" here is meta['receiver_name'] — the actual receiving
    party captured on the booking form — not to be confused with
    meta['shipper_name'], which is the sender and has no place on a debtor
    statement (the debtor is the party being billed/received from, not shipped
    to, but receiver_name is what the statement is meant to surface here).
    Returns a dict of empty strings if terms is missing or unparseable.
    """
    empty = {
        "awb": "", "consignee": "", "destination": "", "carrier_ref": "",
        "carrier": "", "chrg_wt": 0, "act_wt": 0, "vol_wt": 0, "other_charges": 0,
        "per_kg": 0,
    }
    try:
        if invoice.terms:
            meta = json.loads(invoice.terms)
            packages = meta.get("packages") or []
            chrg_wt = sum(p.get("chg_weight") or 0 for p in packages)
            act_wt  = sum(p.get("weight") or 0 for p in packages)
            vol_wt  = sum(p.get("vol_weight") or 0 for p in packages)
            # Calculate per_kg rate from freight_amount and freight_weight
            freight_amount = meta.get("freight", 0) or 0
            freight_weight = meta.get("freight_weight", 0) or 0
            per_kg = round(freight_amount / freight_weight, 2) if freight_weight > 0 else 0

            # DEBUG: Print what's being calculated
            print(f"Invoice {invoice.invoice_id}: freight={freight_amount}, weight={freight_weight}, per_kg={per_kg}")

            return {
                "awb":           meta.get("docket_no", ""),
                "consignee":     meta.get("receiver_name", "") or invoice.contact_person or "",
                "destination":   meta.get("destination", ""),
                "carrier_ref":   meta.get("carrier_ref", ""),   # -> statement "Reference No."
                "carrier":       meta.get("carrier", ""),        # -> statement "Service"
                "chrg_wt":       chrg_wt,
                "act_wt":        act_wt,
                "vol_wt":        vol_wt,
                "other_charges": float(meta.get("other", 0) or 0),
                "per_kg":        per_kg,
            }
    except Exception:
        pass
    return empty

def _purchase_shipment_summary(items):
    """
    A PurchaseInvoice can carry several line items, each its own AWB/dest/
    carrier-ref (docket_no, destination, carrier_ref, party_name on
    PurchaseInvoiceItem) — unlike a sales Invoice, there's no single
    shipment per row here. Rather than silently picking the first item and
    hiding the rest, this joins the distinct values with ", " so a
    multi-AWB purchase invoice shows all of them; single-AWB invoices (the
    common case) render exactly as if there were one field.
    party_name is used as the consignor stand-in — it's the customer tied
    to that AWB, not a formal "consignor" field, since no such column
    exists on the purchase side.
    """
    def _joined(attr):
        vals = []
        for it in items:
            v = (getattr(it, attr, None) or "").strip()
            if v and v not in vals:
                vals.append(v)
        return ", ".join(vals)

    return {
        "awb":         _joined("docket_no"),
        "consignor":   _joined("party_name"),
        "consignee":   _joined("consignee_name"),
        "destination": _joined("destination"),
        "carrier_ref": _joined("carrier_ref"),
    }

def _purchase_shipment_rows(items):
    """
    One dict per distinct shipment on a purchase invoice, instead of
    _purchase_shipment_summary's comma-joined single string — lets the
    statement show each AWB/consignee on its own row.
    """
    rows, seen = [], set()
    for it in items:
        awb = (getattr(it, "docket_no", None) or "").strip()
        consignee = (getattr(it, "consignee_name", None) or "").strip()
        destination = (getattr(it, "destination", None) or "").strip()
        carrier_ref = (getattr(it, "carrier_ref", None) or "").strip()   # -> statement "Reference No."
        carrier = (getattr(it, "courier_name", None) or "").strip()      # -> statement "Service"
        weight = getattr(it, "weight_kg", None) or 0
        other_charges = getattr(it, "other_charges", None) or 0
        taxable = getattr(it, "taxable_value", None) or 0
        per_kg = round(taxable / weight, 2) if weight > 0 else 0
        key = (awb, consignee, destination, carrier_ref)
        if key == ("", "", "", "") or key in seen:
            continue
        seen.add(key)
        rows.append({
            "awb": awb, "consignee": consignee, "destination": destination, "carrier_ref": carrier_ref,
            "carrier": carrier,
            # PurchaseInvoiceItem only stores one weight_kg (no chrg/act/vol split like the
            # client-side package JSON does), so charge and actual weight both read from it
            # and volumetric weight has no source -> always blank on this side.
            "chrg_wt": weight, "act_wt": weight, "vol_wt": 0,
            "other_charges": other_charges, "per_kg": per_kg,
        })
    return rows or [{"awb": "", "consignee": "", "destination": "", "carrier_ref": "",
                      "carrier": "", "chrg_wt": 0, "act_wt": 0, "vol_wt": 0, "other_charges": 0,
                      "per_kg": 0}]

def _manifest_entry_shipment_data(cdb, company_id, docket_no):
    """
    Manifest entries only store courier_name/boxes/docket_no — the actual
    weight, dimensions, destination and receiver live on the customer
    invoice that was booked for that docket/AWB (invoice.terms JSON ->
    "packages" list + destination/receiver_name), same place _get_awb()
    reads from. This looks that invoice up by docket_no and aggregates its
    package data for display on the printable manifest.

    Returns None if there's no docket_no, no matching invoice, or the
    invoice has no package data — callers should render "—" in that case.
    """
    if not docket_no:
        return None

    invoice = cdb.query(Invoice).filter_by(company_id=company_id).filter(
        Invoice.terms.like(f'%"docket_no": "{docket_no}"%')
    ).first()
    if not invoice or not invoice.terms:
        return None

    try:
        meta = json.loads(invoice.terms)
    except (ValueError, TypeError):
        return None

    packages = meta.get("packages") or []
    if not packages:
        return None

    total_actual = sum(p.get("weight") or 0 for p in packages)
    total_vol    = sum(p.get("vol_weight") or 0 for p in packages)
    total_chg    = sum(p.get("chg_weight") or 0 for p in packages)

    # L/B/H only render cleanly when every box in the shipment is the same
    # size. Mixed sizes fall back to the first package's dims rather than
    # silently averaging or picking one at random.
    dims = {(p.get("length") or 0, p.get("width") or 0, p.get("height") or 0) for p in packages}
    length, width, height = next(iter(dims)) if len(dims) == 1 else (
        packages[0].get("length") or 0, packages[0].get("width") or 0, packages[0].get("height") or 0
    )

    return {
        "charge_weight": total_chg,
        "actual_weight": total_actual,
        "length": length,
        "width": width,
        "height": height,
        "vol_weight": total_vol,
        "destination": meta.get("destination") or None,
        "receiver": meta.get("receiver_name") or None,
        "carrier_ref": meta.get("carrier_ref") or None,
    }

# ── Seed Data ─────────────────────────────────────────────────────────────────
SUBSCRIPTION_PLANS_DATA = {
    "starter": {
        "name": "Starter Plan",
        "price": "7999",
        "billing_period": "yearly",
        "max_companies": "2",          
        "max_users": "5",             
        "ai_queries_monthly": "100",
        "features": (
            "Android Driver App (ePOD),"
            "Live Logistics Dashboard,"
            "Client & Vendor Profiles,"
            "Docket & Consignment Booking,"
            "Freight Quotations,"
            "Freight Invoicing & Bills,"
            "Expense Tracking,"
            "Basic Hub Inventory,"
            "AI Address Parser,"
            "Email Support"
        ),
    },

    "business": {
        "name": "Business Plan",
        "price": "14999",
        "billing_period": "yearly",
        "max_companies": "4",          
        "max_users": "15",            
        "ai_queries_monthly": "500",
        "features": (
            "All Starter Features,"
            "Multi-Hub Inventory,"
            "Purchase & Fleet Maintenance,"
            "Supplier & Fuel Vendor Management,"
            "Dynamic Freight Price Lists,"
            "Bank & Cash Flow Management,"
            "Cheque & Outstanding Management,"
            "Driver/Staff Loan Management,"
            "WhatsApp Tracking & Invoice Sharing,"
            "Automated Cloud Backups,"
            "Advanced Manifest Reports,"
            "Priority Support"
        ),
    },

    "professional": {
        "name": "Professional Plan",
        "price": "29999",
        "billing_period": "yearly",
        "max_companies": "7",          
        "max_users": "35",
        "ai_queries_monthly": "2000",
        "features": (
            "All Business Features,"
            "Advanced Route Analytics,"
            "AI Fleet/Profit Insights,"
            "Custom Manifest Layouts,"
            "Webhook & API Access,"
            "Advanced Branch Permissions,"
            "Priority Support"
        ),
    },

    "enterprise": {
        "name": "Enterprise Plan",
        "price": "59999",
        "billing_period": "yearly",
        "max_companies": "15",
        "max_users": "100",
        "ai_queries_monthly": "10000",
        "features": (
            "All Professional Features,"
            "High-Volume API Limits,"
            "Advanced AI Manifest Assistant,"
            "E-commerce API Integrations (Shopify, etc.),"
            "Onboarding & Implementation Support,"
            "Dedicated Account Manager,"
            "SLA Options"
        ),
    },

    "custom": {
        "name": "Custom Enterprise",
        "price": "Contact Sales",
        "billing_period": "custom",
        "max_companies": "Unlimited",
        "max_users": "Unlimited",
        "ai_queries_monthly": "Custom",
        "features": (
            "Private Cloud Deployment,"
            "On-premise Deployment,"
            "White-label Android Driver App (Your Branding),"
            "Custom AI Document Parsing Setup,"
            "Tally / Legacy ERP Integration,"
            "Bulk Data Migration,"
            "Custom Feature Development,"
            "On-site Staff Training"
        ),
    },
}

def seed_database():
    """Insert initial plans, users and sample data if the DB is empty."""
    
    # ── Subscription Plans (Platform DB) ─────────────────────────────────────
    if SubscriptionPlan.query.count() == 0:
        for plan_id, data in SUBSCRIPTION_PLANS_DATA.items():
            db.session.add(SubscriptionPlan(
                id=plan_id,
                name=data["name"],
                price=data["price"],
                max_companies=data["max_companies"],
                max_users=data["max_users"],
                features=data["features"],
            ))
        db.session.commit()
        print("✔  Subscription plans seeded.")

    # ── Registered Users (Platform DB) ──────────────────────────────────────
    if RegisteredUser.query.count() == 0:
        demo = RegisteredUser(
            user_id="USR001",
            email="demo@demo.com",
            password_hash=hash_password("Demo@123"),
            full_name="Demo User",
            phone="9999999999",
            role="owner",
            subscription_plan="business",
            created_at=date(2024, 1, 1),
            is_active=True,
            email_verified=True,
            must_change_password=False,
        )
        db.session.add(demo)
        db.session.commit()
        print("✔  Demo user seeded.")

     # ── Super Admin (Platform DB) ────────────────────────────────────────────
    if RegisteredUser.query.filter_by(role="super_admin").count() == 0:
        admin = RegisteredUser(
            user_id="ADMIN001",
            email="admin",
            password_hash=hash_password("Ibrahim@moosa53"),
            full_name="Super Admin",
            role="super_admin",
            is_active=True,
            email_verified=True,
            must_change_password=False,
        )
        db.session.add(admin)
        db.session.commit()
        print("✔  Super admin seeded.")

    # ── Companies (Platform DB) ─────────────────────────────────────────────
    if Company.query.count() == 0:
        comp1 = Company(
            company_id="DEMO001",
            company_name="Demo Company",
            owner_email="demo@demo.com",
            subscription_plan="business",
            subscription_start=date(2024, 1, 1),
            subscription_end=date(2030, 1, 1),
            max_companies_allowed="5",
            max_users_per_company="15",
            gst_number="27AAABC1234F1Z",
            address="Mumbai, Maharashtra",
            phone="9876543210",
            created_at=date(2024, 1, 1),
            is_active=True,
        )
        db.session.add(comp1)
        db.session.commit()
        print("✔  Demo company seeded.")
    
    print("✅ Platform database seeding complete.")

def _ensure_payment_ledger_columns(cdb):
    """One-time, idempotent schema patch: adds the applied_ref_type /
    applied_ref_id / applied_ci_id columns to cash_transactions and
    bank_transactions if they don't exist yet (db.create_all() only
    creates brand-new tables — it never ALTERs existing ones). Safe to
    call on every startup: each ALTER is wrapped so an existing column
    just no-ops instead of crashing the app.
    """
    from sqlalchemy import text
    statements = [
        "ALTER TABLE cash_transactions ADD COLUMN applied_ref_type VARCHAR(20)",
        "ALTER TABLE cash_transactions ADD COLUMN applied_ref_id INTEGER",
        "ALTER TABLE cash_transactions ADD COLUMN applied_ci_id INTEGER",
        "ALTER TABLE bank_transactions ADD COLUMN applied_ref_type VARCHAR(20)",
        "ALTER TABLE bank_transactions ADD COLUMN applied_ref_id INTEGER",
        "ALTER TABLE bank_transactions ADD COLUMN applied_ci_id INTEGER",
    ]
    for stmt in statements:
        try:
            cdb.execute(text(stmt))
            cdb.commit()
        except Exception:
            # Column already exists (or table doesn't exist yet on a brand
            # new company DB, where create_all() already created it with
            # the new columns) — either way, nothing to do.
            cdb.rollback()


def seed_customer_database(company_id):
    """Seed customer data for a specific company in its own database."""
    from db_router import get_customer_session
    
    cdb = get_customer_session(company_id, db_session=db.session)

    _ensure_payment_ledger_columns(cdb)
    
    # ── Company Users ───────────────────────────────────────────────────────
    if cdb.query(CompanyUser).count() == 0:
        # Get company info to know the owner
        company = Company.query.filter_by(company_id=company_id).first()
        owner_reg = RegisteredUser.query.filter_by(email=company.owner_email).first()
        
        users = [
            CompanyUser(
                user_id="EMP001",
                company_id=company_id,
                email=company.owner_email,
                password_hash=hash_password("Demo@123"),
                full_name=owner_reg.full_name if owner_reg else "Demo User",
                role="owner",
                department="Management",
                phone=company.phone,
                is_active=True,
                created_at=today_ist()
            ),
        ]

        cdb.add_all(users)
        cdb.commit()
        print(f"✔  Company users seeded for {company_id}")

    # ── Clients ─────────────────────────────────────────────────────────────
    if cdb.query(Client).count() == 0 and company_id == "DEMO001":
        clients = [
            Client(company_id=company_id, name="ABC Electronics", client_type="Customer",
                   phone="9876543220", status="Active", created_at=today_ist()),
            Client(company_id=company_id, name="XYZ Traders", client_type="Customer",
                   phone="9876543221", status="Active", created_at=today_ist()),
            Client(company_id=company_id, name="PQR Solutions", client_type="Business",
                   phone="9876543222", status="Active", created_at=today_ist()),
            Client(company_id=company_id, name="Reliance Industries", phone="9876543210",
                   pending=0, last_payment=date(2024, 1, 22), status="Paid"),
            Client(company_id=company_id, name="Tata Consultancy", phone="9876543211",
                   pending=89500, last_payment=date(2024, 1, 5), status="Pending"),
            Client(company_id=company_id, name="Infosys Ltd", phone="9876543212",
                   pending=86000, last_payment=date(2024, 1, 18), status="Active"),
        ]
        cdb.add_all(clients)
        cdb.commit()
        print(f"✔  Clients seeded for {company_id}")

    # ── Stock Items ─────────────────────────────────────────────────────────
    if cdb.query(StockItem).count() == 0 and company_id == "DEMO001":
        items = [
            StockItem(company_id=company_id, code="PROD001", name="LED TV 43 inch",
                      category="Electronics", quantity=25, unit="pcs", unit_price=35000,
                      reorder_level=10, last_updated=date(2024, 1, 20)),
            StockItem(company_id=company_id, code="PROD002", name="Smartphone X",
                      category="Electronics", quantity=50, unit="pcs", unit_price=25000,
                      reorder_level=20, last_updated=date(2024, 1, 20)),
        ]
        cdb.add_all(items)
        cdb.commit()
        print(f"✔  Stock items seeded for {company_id}")

    # Close the session
    from db_router import close_customer_session
    close_customer_session(company_id)


# ── Plan helper ───────────────────────────────────────────────────────────────
def get_plan(plan_id):
    p = SubscriptionPlan.query.get(plan_id)
    if not p:
        return {}
    return {
        "name": p.name,
        "price": p.price,
        "max_companies": p.max_companies,
        "max_users_per_company": p.max_users,
        "features": p.features.split(",") if p.features else [],
    }

def get_all_plans():
    return {p.id: get_plan(p.id) for p in SubscriptionPlan.query.all()}


# ── Company helpers ───────────────────────────────────────────────────────────
def get_company_by_id(company_id):
    return Company.query.filter_by(company_id=company_id).first()

def get_owner_companies(owner_email):
    return Company.query.filter_by(owner_email=owner_email, is_active=True).all()

def get_owner_user_stats(owner_email):
    """
    Distinct active users across ALL companies owned by this owner.
    CompanyUser rows live in each company's own separate database, so this
    opens every one of the owner's company DBs and dedupes by email.
    Returns (current_count, max_users, existing_emails_set).
    """
    companies = get_owner_companies(owner_email)
    emails = set()
    for c in companies:
        try:
            _cdb = get_customer_session(c.company_id)
            rows = _cdb.query(CompanyUser).filter_by(is_active=True).all()
            for r in rows:
                if r.email:
                    emails.add(r.email.strip().lower())
        except Exception as e:
            print(f"⚠  Could not read users for {c.company_id}: {e}")

    plan = get_plan(companies[0].subscription_plan) if companies else {}
    max_u = plan.get("max_users_per_company", "Unlimited")
    return len(emails), max_u, emails

def check_company_limit(company_id, user_type="user"):
    company = get_company_by_id(company_id)
    if not company:
        return False, "Company not found"
    plan = get_plan(company.subscription_plan)
    if user_type == "user":
        # Seat cap is owner-wide (across all of the owner's companies),
        # not per company.
        current, max_u, _ = get_owner_user_stats(company.owner_email)
        try:
            max_u = int(max_u)
            if current >= max_u:
                return False, f"Maximum {max_u} users allowed across all your companies under your {plan['name']}. Please upgrade."
        except (ValueError, TypeError):
            pass  # "Unlimited"
    return True, "OK"

def check_new_company_limit(owner_email):
    comps = get_owner_companies(owner_email)
    if not comps:
        return True, "OK"
    plan = get_plan(comps[0].subscription_plan)
    max_c = plan.get("max_companies", 2)
    try:
        max_c = int(max_c)
        if len(comps) >= max_c:
            return False, f"Your {plan['name']} allows up to {max_c} companies. Please upgrade."
    except (ValueError, TypeError):
        pass  # "Unlimited"
    return True, "OK"


def get_cdb():
    """
    Return a customer-database session for the currently active company.
    Use this everywhere you previously used db.session for customer tables.

    Goes through get_customer_session_with_retry(), which self-heals a
    session left broken by a previous request that never got torn down
    (e.g. a killed gunicorn worker) — a PendingRollbackError or dead
    connection here gets rolled back / rebuilt and retried once, instead
    of surfacing as a 500 on this request.

    Example:
        cdb = get_cdb()
        clients = cdb.query(Client).filter_by(company_id=company_id).all()
    """
    company_id = get_current_company()
    if not company_id:
        return None
    return get_customer_session_with_retry(company_id)

def _first_or_404(obj):
    """Replacement for Flask-SQLAlchemy's first_or_404() for plain SQLAlchemy queries."""
    if obj is None:
        from flask import abort
        abort(404)
    return obj

_MULTI_WORD_COUNTRIES = [
    'NEW ZELAND', 'NEW ZEALAND', 'SOUTH AFRICA', 'SOUTH KOREA', 'NORTH KOREA',
    'SAUDI ARABIA', 'UNITED KINGDOM', 'UNITED STATES', 'UNITED ARAB EMIRATES',
    'CZECH REPUBLIC', 'COSTA RICA', 'SRI LANKA', 'HONG KONG', 'PUERTO RICO',
    'EL SALVADOR', 'DOMINICAN REPUBLIC',
]


def _split_country_label(col_label):
    """
    Splits a column header that may contain multiple country names jammed
    together with single spaces (e.g. 'Germany Belgium Netherlands' or
    'Australia New Zealand') into individual country names, without breaking
    apart known multi-word country names like 'New Zealand' or 'Czech Republic'.
    """
    label_upper = col_label.upper()
    found = []
    remaining = label_upper

    # Pull out known multi-word names first so they aren't split on whitespace.
    for name in _MULTI_WORD_COUNTRIES:
        if name in remaining:
            found.append(name)
            remaining = remaining.replace(name, ' ')

    # Whatever's left, split on whitespace as single-word country names.
    found.extend(remaining.split())
    return found


def _find_header_row(filepath, max_scan_rows=20):
    """
    Real rate-card files in this system have several blank/title rows before
    the actual header row (e.g. row 0-7 blank, row 8 = title, row 9 = header).
    pd.read_excel() with no skiprows treats row 0 as the header, which produces
    'Unnamed: N' columns and silently breaks parsing. This scans raw cells to
    find the row that actually looks like a header (contains COUNTRY, WEIGHT/KG,
    or a country/weight-shaped row) and returns its 0-indexed row number for
    pandas' `header=` argument.
    """
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]

    HEADER_HINTS = ('COUNTRY', 'WEIGHT', 'KG', 'DESTINATION')

    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True)):
        cells = [str(c).strip().upper() for c in row if c is not None and str(c).strip() != '']
        if len(cells) < 2:
            continue
        if any(hint in c for c in cells for hint in HEADER_HINTS):
            return i  # 0-indexed -> matches pandas header=i

    return 0  # fallback: assume no blank rows


def _extract_weight_from_label(label):
    import re
    s = str(label).strip().upper()

    gms_match = re.search(r'(\d+(?:\.\d+)?)\s*GMS?\b', s)
    if gms_match:
        return float(gms_match.group(1)) / 1000.0

    num_match = re.search(r'(\d+(?:\.\d+)?)', s)
    if not num_match:
        return None

    return float(num_match.group(1))

def get_employee_companies(email):
    results = []
    active_company_id = session.get("active_company_id")
    for comp in Company.query.filter_by(is_active=True).all():
        if getattr(comp, "hidden_on_mobile", False):
            continue
        try:
            _cdb = get_customer_session(comp.company_id)
            emp = _cdb.query(CompanyUser).filter_by(email=email, is_active=True).first()
            if emp:
                comp.emp_role = emp.role
                comp.emp_user_id = emp.user_id
                comp.is_active_selection = (comp.company_id == active_company_id)
                results.append(comp)
        except Exception:
            continue
    return results

def _is_weight_label(label):
    """True if a column header looks like a weight tier (not TIME/DAY/COUNTRY/etc)."""
    s = str(label).strip().upper()
    if not s or s in ('NAN', 'NONE'):
        return False
    if 'TIME' in s or 'DAY' in s or 'COUNTRY' in s:
        return False
    return ('KG' in s) or ('GMS' in s) or s.replace('.', '', 1).isdigit()

def update_customer_invoice_from_booking(cdb, company_id, booking_invoice_id):
    """
    When a booking (Invoice) is edited/updated, find all CustomerInvoices
    that contain this booking and update their totals.
    """
    # Find all customer invoices that contain this booking
    customer_invoices = cdb.query(CustomerInvoice).filter(
        CustomerInvoice.company_id == company_id,
        CustomerInvoice.booking_ids_json.isnot(None),
        CustomerInvoice.status != 'Void'
    ).all()
    
    updated_invoices = []
    for ci in customer_invoices:
        try:
            booking_ids = json.loads(ci.booking_ids_json)
            if booking_invoice_id in booking_ids:
                # This customer invoice contains the booking being edited
                # Recalculate totals from all bookings in this invoice
                updated_invoices.append(ci.id)
                _recalculate_customer_invoice_totals(cdb, company_id, ci)
        except (ValueError, TypeError):
            continue
    
    return updated_invoices


def _recalculate_customer_invoice_totals(cdb, company_id, cust_inv):
    """
    Recalculate all totals for a customer invoice from its constituent bookings.
    Called when a booking in the invoice is edited.
    """
    try:
        booking_ids = json.loads(cust_inv.booking_ids_json) if cust_inv.booking_ids_json else []
    except (ValueError, TypeError):
        booking_ids = []
    
    if not booking_ids:
        return
    
    # Get all bookings
    bookings = cdb.query(Invoice).filter(
        Invoice.id.in_(booking_ids),
        Invoice.company_id == company_id,
        Invoice.status.notin_(['Void', 'Draft'])
    ).all()
    
    # Recalculate totals
    subtotal = 0.0
    tax_total = 0.0
    cgst_total = 0.0
    sgst_total = 0.0
    igst_total = 0.0
    grand_total = 0.0
    
    for inv in bookings:
        meta = {}
        if inv.terms:
            try:
                meta = json.loads(inv.terms)
            except:
                pass
        
        freight = meta.get("freight", inv.subtotal or 0)
        gst = meta.get("gst", inv.tax_amount or 0)
        total = inv.grand_total or 0
        
        cgst = meta.get("cgst", 0)
        sgst = meta.get("sgst", 0)
        igst = meta.get("igst", 0)
        
        subtotal += freight
        tax_total += gst
        cgst_total += cgst
        sgst_total += sgst
        igst_total += igst
        grand_total += total
    
    # Update the customer invoice
    cust_inv.subtotal = subtotal
    cust_inv.tax_amount = tax_total
    cust_inv.cgst_total = cgst_total
    cust_inv.sgst_total = sgst_total
    cust_inv.igst_total = igst_total
    cust_inv.grand_total = grand_total
    cust_inv.balance = max(0, grand_total - (cust_inv.paid_amount or 0))
    
    # Update status based on balance
    if cust_inv.balance <= 0:
        cust_inv.status = "Paid"
    elif (cust_inv.paid_amount or 0) > 0:
        cust_inv.status = "Partial"
    else:
        cust_inv.status = "Pending"
    
    # Update the items for each booking
    for inv in bookings:
        # Update or create the corresponding CustomerInvoiceItem
        item = cdb.query(CustomerInvoiceItem).filter_by(
            customer_invoice_id=cust_inv.id,
            booking_invoice_id=inv.id
        ).first()
        
        if item:
            meta = {}
            if inv.terms:
                try:
                    meta = json.loads(inv.terms)
                except:
                    pass
            
            packages = meta.get("packages", [])
            total_weight = sum(p.get("weight", 0) * p.get("qty", 1) for p in packages) or meta.get("freight_weight", 0)
            freight = meta.get("freight", inv.subtotal or 0)
            
            cgst = meta.get("cgst", 0)
            sgst = meta.get("sgst", 0)
            igst = meta.get("igst", 0)
            
            item.docket_no = meta.get("docket_no", "")
            item.receiver_name = meta.get("receiver_name", "")
            item.destination = meta.get("destination", "")
            item.carrier = meta.get("carrier", "")
            item.carrier_ref = meta.get("carrier_ref", "")
            item.weight_kg = total_weight
            item.rate_per_kg = freight / total_weight if total_weight > 0 else 0
            item.taxable_amount = freight
            item.cgst_amount = cgst
            item.sgst_amount = sgst
            item.igst_amount = igst
            item.total_amount = inv.grand_total or 0
            item.gst_type = 'interstate' if meta.get('is_interstate', False) else 'intrastate'
    
    cdb.commit()

def _build_tiers_and_bands(weight_value_pairs):
    weight_value_pairs = sorted(weight_value_pairs, key=lambda x: x[0])

    band_rows = []
    tier_candidates = []
    prev_price = None
    band_started = False  # once we cross into bands, everything after stays banded

    for weight, is_band_flag, price in weight_value_pairs:
        if is_band_flag:
            band_started = True
            band_rows.append((weight, price))
            continue

        if band_started:
            # Already past the breakpoint (explicitly marked earlier) -> stay banded.
            band_rows.append((weight, price))
            continue

        if prev_price is not None and prev_price > 0 and price < prev_price * 0.5:
            # Sharp drop (price roughly halved or more) -> real band transition,
            # not tier-to-tier rounding noise.
            band_started = True
            band_rows.append((weight, price))
        else:
            tier_candidates.append((weight, price))
            if price > 0:
                prev_price = max(prev_price or 0, price)  # ignore small dips/noise

    tiers = [{'weight': w, 'price': p} for w, p in tier_candidates]

    bands = []
    band_rows.sort(key=lambda x: x[0])
    for weight, price in band_rows:
        if bands and abs(bands[-1]['rate_per_kg'] - price) < max(0.01, bands[-1]['rate_per_kg'] * 0.005):
            bands[-1]['max_kg'] = weight  # extend current band (within 0.5% = same rate, just rounding)
        else:
            bands.append({'min_kg': weight, 'max_kg': None, 'rate_per_kg': price})

    for i in range(len(bands) - 1):
        bands[i]['max_kg'] = bands[i + 1]['min_kg']
    if bands:
        bands[-1]['max_kg'] = None  # last band is open-ended

    return tiers, bands

def get_party_name(client_id=None, supplier_id=None, form=None, fallback_name=None):
    """
    Single source of truth for getting the correct party name for transactions.
    Ensures consistency across receipts, payments, and booking edits.
    
    Args:
        client_id: The client ID (for receipts/debtors)
        supplier_id: The supplier ID (for payments/creditors)
        form: The request form object (for cash/walk-in bookings)
        fallback_name: A fallback name if all else fails
    
    Returns:
        The correct party name string
    """
    cdb = get_cdb()
    company_id = get_current_company()
    
    # Priority 1: If we have a client_id, get the exact name from Client table
    if client_id:
        client = cdb.query(Client).filter_by(id=client_id, company_id=company_id).first()
        if client:
            return client.name
    
    # Priority 2: If we have a supplier_id, get the exact name from Supplier table
    if supplier_id:
        supplier = cdb.query(Supplier).filter_by(id=supplier_id, company_id=company_id).first()
        if supplier:
            return supplier.name
    
    # Priority 3: For cash/walk-in bookings, get from form
    if form:
        shipper_name = form.get("shipper_name", "").strip()
        if shipper_name:
            return shipper_name
        # Alternative: customer_name field
        customer_name = form.get("customer_name", "").strip()
        if customer_name:
            return customer_name
        # Alternative: client_name field
        client_name = form.get("client_name", "").strip()
        if client_name:
            return client_name
        # Alternative: party_name field
        party_name = form.get("party_name", "").strip()
        if party_name:
            return party_name
    
    # Priority 4: Use fallback
    if fallback_name:
        return fallback_name
    
    # Last resort: return a generic name
    return "Unknown Party"

def _supplier_closing_balance(cdb, company_id, s):
    """Live running balance exactly as the supplier statement page computes
    it: opening balance + Σ(grand_total − paid_amount) across purchase invoices."""
    total = s.opening_balance or 0
    for inv in cdb.query(PurchaseInvoice).filter_by(company_id=company_id, supplier_id=s.id).all():
        total += (inv.grand_total or 0) - (inv.paid_amount or 0)
    return total


def _supplier_close_statement(cdb, company_id, s, action, scope="till_yesterday", as_of_date=None):
    """Archives the supplier's current live ledger into StatementClosing,
    then marks the closed PurchaseInvoice rows as fully paid (so the
    payable, which is driven directly off PurchaseInvoice.paid_amount/
    balance, nets to zero for the closed period) and moves the statement
    cutoff forward so the next statement load starts blank (action=
    'cleared') or with just the carried-forward balance (action=
    'carried_forward'). Invoices are kept for GST/audit — only
    paid_amount/balance/status change, same fields a normal payment
    update would touch.

    `as_of_date`: the LAST date to include in the archived/closed
    statement — same semantics as _client_close_statement's as_of_date.
    If not given, derived from `scope` (only relevant for action=
    'cleared'): 'till_yesterday' (default) -> yesterday; 'complete' ->
    today. Clamped so it can't precede the day before the existing
    statement_cutoff, or fall after today.
    Returns the amount that was payable at closing time."""
    today = today_ist()

    if as_of_date is None:
        if action == "cleared" and scope == "complete":
            as_of_date = today
        else:
            as_of_date = today - timedelta(days=1)

    if as_of_date > today:
        as_of_date = today
    if s.statement_cutoff:
        floor_date = s.statement_cutoff.date() - timedelta(days=1)
        if as_of_date < floor_date:
            as_of_date = floor_date

    archive_until = as_of_date + timedelta(days=1)  # exclusive upper bound

    ledger, total_debit, total_credit, closing = _build_supplier_ledger(
        cdb, company_id, s, since=s.statement_cutoff, until=archive_until)

    cdb.add(StatementClosing(
        company_id=company_id,
        entity_type="supplier",
        entity_id=s.id,
        entity_name=s.name,
        action=action,
        closing_balance=closing,
        total_debit=total_debit,
        total_credit=total_credit,
        ledger_snapshot=json.dumps(ledger, default=str),
        closed_by=session.get("username", "unknown"),
        closed_at=datetime.utcnow(),
    ))

    closed_invoices_q = cdb.query(PurchaseInvoice).filter_by(
        company_id=company_id, supplier_id=s.id
    ).filter(PurchaseInvoice.date < archive_until)
    if s.statement_cutoff:
        closed_invoices_q = closed_invoices_q.filter(
            PurchaseInvoice.date >= s.statement_cutoff.date())
    for inv in closed_invoices_q.all():
        if (inv.balance or 0) != 0 or (inv.paid_amount or 0) != (inv.grand_total or 0):
            inv.paid_amount = inv.grand_total or 0
            inv.balance = 0
            inv.status = "Paid"

    s.statement_cutoff = datetime.combine(archive_until, datetime.min.time())
    s.opening_balance = closing if action == "carried_forward" else 0
    s.payable = s.opening_balance
    return closing

def round_billable_weight(w):
    """Server-side mirror of booking.html's roundBillableWeight(): 0 < w <= 10kg
    rounds UP to the next 0.5kg; above 10kg rounds UP to the next whole 1kg.
    Any code path that turns a chargeable weight into a rate-card lookup (or a
    billing amount) must round through this first — the DISPLAYED/stored
    chargeable weight on the invoice stays the exact actual/volumetric figure
    (1.75kg, 10.1kg); only the figure used to pick a price is rounded up to
    the slab."""
    if not w or w <= 0:
        return 0
    if w <= 10:
        return math.ceil(w / 0.5) * 0.5
    return math.ceil(w)


def calculate_rate(rate_data, country_key, weight):
    """
    Single source of truth for turning (rate_data, country, weight) into a price.
    Used by both the sales and purchase rate-lookup endpoints so the logic only
    lives in one place.

    rate_data['countries'][country] is either:
      - new format: {'tiers': [...], 'bands': [...]}
      - legacy format: {weight_str: price} (old flat dict from price lists
        uploaded before this fix; kept working for backward compatibility)

    Returns (rate, weight_used, pricing_type) or (None, None, None) if no match.
    pricing_type is 'tier' or 'per_kg' so the UI can show how the figure was derived.
    """
    entry = rate_data['countries'].get(country_key)
    if not entry:
        return None, None, None

    if 'tiers' not in entry and 'bands' not in entry:
        # Legacy flat dict
        rate_keys = sorted(float(k) for k in entry.keys())
        if not rate_keys:
            return None, None, None
        closest = rate_keys[-1]
        for w in rate_keys:
            if w >= weight:
                closest = w
                break
        rate = entry.get(closest)
        if rate is None:
            rate = entry.get(str(closest))
        return rate, closest, 'tier'

    tiers = entry.get('tiers', [])
    bands = entry.get('bands', [])

    for band in bands:
        min_kg, max_kg = band['min_kg'], band['max_kg']
        if weight >= min_kg and (max_kg is None or weight < max_kg):
            return round(band['rate_per_kg'] * weight, 2), weight, 'per_kg'

    if bands and weight >= bands[-1]['min_kg']:
        # Heavier than the last defined band's min -> still use it (open-ended)
        return round(bands[-1]['rate_per_kg'] * weight, 2), weight, 'per_kg'

    if tiers:
        tiers_sorted = sorted(tiers, key=lambda t: t['weight'])

        # Exact match (or effectively exact, floating point) - use the tier price as-is.
        for t in tiers_sorted:
            if abs(t['weight'] - weight) < 1e-9:
                return t['price'], t['weight'], 'tier'

        if weight <= tiers_sorted[0]['weight']:
            return tiers_sorted[0]['price'], tiers_sorted[0]['weight'], 'tier'

        if weight >= tiers_sorted[-1]['weight']:
            # Heavier than the last defined tier: extend using the per-kg rate implied
            # by the last two tiers instead of just returning the last tier's flat price.
            last = tiers_sorted[-1]
            if len(tiers_sorted) >= 2:
                prev = tiers_sorted[-2]
                per_kg = (last['price'] - prev['price']) / (last['weight'] - prev['weight']) if last['weight'] != prev['weight'] else 0
                price = round(last['price'] + per_kg * (weight - last['weight']), 2)
                return price, weight, 'per_kg'
            return last['price'], last['weight'], 'tier'

        # Weight falls between two tiers - interpolate linearly instead of rounding up
        # to the next whole-kg tier, so fractional (decimal) weights are billed fairly.
        lower = tiers_sorted[0]
        upper = tiers_sorted[-1]
        for i in range(len(tiers_sorted) - 1):
            if tiers_sorted[i]['weight'] <= weight <= tiers_sorted[i + 1]['weight']:
                lower = tiers_sorted[i]
                upper = tiers_sorted[i + 1]
                break
        if upper['weight'] == lower['weight']:
            return lower['price'], lower['weight'], 'tier'
        fraction = (weight - lower['weight']) / (upper['weight'] - lower['weight'])
        price = round(lower['price'] + fraction * (upper['price'] - lower['price']), 2)
        return price, weight, 'per_kg'

    return None, None, None


def compute_invoice_gst(taxable_amount, apply_gst, shipper_state, receiver_state):
    """
    Single source of truth for customer-invoice GST, mirroring how the purchase
    invoice flow already splits CGST/SGST vs IGST (see purchase_invoice_new).

    - CGST+SGST (9%+9%) when shipper and receiver are in the same state (intra-state).
    - IGST (18%) when they're in different states (inter-state) - this is how GST
      actually works in India; a flat "18% GST" line was never technically correct.
    - Grand total is rounded to the nearest whole rupee, with the rounding
      difference broken out as a separate "Round Off" line, matching standard
      Indian tax-invoice practice (Rule 3, GST invoicing rounding conventions).

    Returns a dict: taxable, cgst, sgst, igst, gst_total, pre_round_total,
    round_off, grand_total, is_interstate.
    """
    taxable = round(taxable_amount, 2)
    s1 = (shipper_state or "").strip().lower()
    s2 = (receiver_state or "").strip().lower()
    # If either state is missing we can't determine interstate vs intrastate, so
    # fall back to intrastate (CGST+SGST) rather than guessing - this matches the
    # system's prior default behavior for incomplete address data.
    is_interstate = bool(s1 and s2 and s1 != s2)

    gst_total = round(taxable * 0.18, 2) if apply_gst else 0.0
    if apply_gst and is_interstate:
        cgst, sgst, igst = 0.0, 0.0, gst_total
    elif apply_gst:
        cgst = round(gst_total / 2, 2)
        sgst = round(gst_total - cgst, 2)
        igst = 0.0
    else:
        cgst = sgst = igst = 0.0

    pre_round_total = round(taxable + gst_total, 2)
    grand_total = round(pre_round_total)  # nearest whole rupee, per standard invoice rounding
    round_off = round(grand_total - pre_round_total, 2)

    return {
        "taxable": taxable,
        "cgst": cgst,
        "sgst": sgst,
        "igst": igst,
        "gst_total": gst_total,
        "pre_round_total": pre_round_total,
        "round_off": round_off,
        "grand_total": grand_total,
        "is_interstate": is_interstate,
    }

# Add this helper function near other company helpers (around line 200)
# ── Add this helper function near other company helpers ──
def is_gst_number_taken(gst_number, exclude_company_id=None):
    """
    Check if a GST number is already used by ANY company (global uniqueness).
    Only checks ACTIVE companies.
    """
    if not gst_number or not gst_number.strip():
        return False
    
    query = Company.query.filter(
        func.lower(Company.gst_number) == func.lower(gst_number.strip()),
        Company.is_active == True
    )
    if exclude_company_id:
        query = query.filter(Company.company_id != exclude_company_id)
    return query.first() is not None

def is_company_name_taken(owner_email, company_name, exclude_company_id=None):
    """
    Check if a company name is already taken by the SAME owner.
    Only checks ACTIVE companies.
    """
    query = Company.query.filter(
        Company.owner_email == owner_email,
        func.lower(Company.company_name) == func.lower(company_name.strip()),
        Company.is_active == True
    )
    if exclude_company_id:
        query = query.filter(Company.company_id != exclude_company_id)
    return query.first() is not None

def parse_price_list(filepath, courier):
    """
    Parse a courier Excel rate card into structured JSON.

    Handles the real-world rate card shapes used in this system:
      - DPD: COUNTRY column + per-weight columns, some explicitly marked as
        bands ("8 KG +", "11 KG +", "21 KG +") whose values are per-kg rates.
      - FEDEX / DHL: WEIGHT column + one column per destination/country-group,
        where rows below ~10.5kg are full tier prices and rows from ~11kg
        onward are flat per-kg rates (no explicit '+' marker - detected by
        the price no longer increasing with weight).

    Output shape (per country):
      data['countries'][COUNTRY] = {
          'tiers': [{'weight': 1.0, 'price': 3381.44}, ...],
          'bands': [{'min_kg': 11.0, 'max_kg': 21.0, 'rate_per_kg': 980}, ...]
      }

    NOTE: signature changed from parse_price_list(df, courier) to
    parse_price_list(filepath, courier) because the old caller's
    pd.read_excel(filepath) with no header row offset produced 'Unnamed: N'
    columns on every real file in this system (header is row 8/9, not row 0),
    so format detection always failed silently. This function now reads the
    file itself after locating the real header row.
    """
    import re

    data = {
        'courier': courier,
        'format': 'unknown',
        'countries': {},
        'weights': []
    }

    header_row = _find_header_row(filepath)
    df = pd.read_excel(filepath, engine='openpyxl', header=header_row)
    df = df.dropna(how='all')

    headers = df.columns.tolist()
    print(f"📊 Parsing {courier} - header_row={header_row} - Columns found: {headers}")

    # ── DPD-style format: has a COUNTRY column ──────────────────────────
    country_col = None
    for h in headers:
        if 'COUNTRY' in str(h).upper():
            country_col = h
            break

    if country_col:
        print(f"✅ Found country column: {country_col}")
        data['format'] = 'dpd'

        weight_cols = []  # (col_name, weight_val)
        for h in headers:
            if h == country_col or not _is_weight_label(h):
                continue
            weight_val = _extract_weight_from_label(h)
            if weight_val is not None:
                weight_cols.append((h, weight_val))
                print(f"   Weight column: {h} -> {weight_val}kg")

        weight_cols.sort(key=lambda x: x[1])
        data['weights'] = [w[1] for w in weight_cols]

        for idx, row in df.iterrows():
            country = str(row[country_col]).strip().upper()
            if not country or country in ('NAN', 'NONE', '') or len(country) > 60:
                continue  # skip blanks and footer/notes rows

            pairs = []
            for col_name, weight_val in weight_cols:
                try:
                    val = row[col_name]
                    if pd.notna(val) and val != '':
                        pairs.append((float(weight_val), False, float(val)))
                except Exception as e:
                    print(f"   Error parsing {col_name} for {country}: {e}")

            if pairs:
                tiers, bands = _build_tiers_and_bands(pairs)
                data['countries'][country] = {'tiers': tiers, 'bands': bands}

        print(f"✅ Parsed {len(data['countries'])} countries for {courier}")
        return data

    # ── FEDEX/DHL-style format: WEIGHT/KG column + one column per destination ──
    weight_col = None
    for h in headers:
        if str(h).strip().upper() in ('WEIGHT', 'KG') or 'WEIGHT' in str(h).upper():
            weight_col = h
            break

    if weight_col:
        print(f"✅ Found weight column: {weight_col}")
        data['format'] = 'fedex'

        country_cols = [h for h in headers if h != weight_col]
        raw_by_country = {}  # destination -> [(weight, is_band, price)]

        for idx, row in df.iterrows():
            wv = row[weight_col]
            if pd.isna(wv):
                continue

            # WEIGHT column can be a pure number (FEDEX: 0.5, 1, 1.5 ...) or a
            # text label (DHL: '1ST 500gms', '1 kg', '21 KG'). Handle both.
            if isinstance(wv, (int, float)):
                weight_val = float(wv)
            else:
                weight_val = _extract_weight_from_label(wv)
                if weight_val is None:
                    continue  # footer/notes row, not a weight row

            if weight_val <= 0:
                continue

            data['weights'].append(weight_val)

            for col in country_cols:
                col_label = str(col).replace('/', ' ').replace('_', ' ').replace('&', ' ')
                col_label = re.sub(r'\s+', ' ', col_label).strip()
                country_list = _split_country_label(col_label)
                for country in country_list:
                    country = country.strip().upper()
                    if not country or len(country) <= 1 or 'TIME' in country:
                        continue
                    val = row[col]
                    if pd.isna(val):
                        continue
                    raw_by_country.setdefault(country, []).append(
                        (weight_val, False, float(val))
                    )

        for country, pairs in raw_by_country.items():
            tiers, bands = _build_tiers_and_bands(pairs)
            data['countries'][country] = {'tiers': tiers, 'bands': bands}

        data['weights'] = sorted(set(data['weights']))
        print(f"✅ Parsed {len(data['countries'])} countries for {courier}")
        return data

    print(f"❌ Could not detect format for {courier}. Headers: {headers}")
    return data

# ─────────────────────────────────────────────────────────────────────────────
# ── Auth Routes ───────────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    if "user" in session:
        return redirect(url_for("dashboard"))
    return redirect(url_for("login"))

@app.route("/api/ai-chat", methods=["POST"])
@login_required
def api_ai_chat():
    """
    Chat endpoint for the ERP AI assistant. company_id is ALWAYS taken
    from the server-side session (get_current_company()) — the request
    body is never trusted for it, so there is no way for the client to
    ask about a different company's data by editing the payload.
    """
    payload = request.get_json(silent=True) or {}
    user_message = (payload.get("message") or "").strip()

    if not user_message:
        return jsonify({"response": "Please type a question.", "source": "error"}), 400
    if len(user_message) > 500:
        return jsonify({"response": "That's a bit long — please ask in under 500 characters.",
                         "source": "error"}), 400

    company_id = get_current_company()

    result = _ai_assistant.chat(
        user_message=user_message,
        company_id=company_id,
        has_permission=has_permission,   # your existing permissions.py-backed function
    )

    return jsonify(result)


@app.route("/api/ai-chat/clear", methods=["POST"])
@login_required
def api_ai_chat_clear():
    """Clears the assistant's short conversational memory (not DB data)."""
    _ai_assistant.clear_context()
    return jsonify({"ok": True})

@app.route("/health")
def health_check():
    """Health check endpoint to warm up database connections"""
    try:
        # Try a simple query on platform DB
        db.session.execute(text("SELECT 1")).fetchone()
        
        # Try customer DB if company is active
        company_id = get_current_company()
        if company_id:
            cdb = get_cdb()
            if cdb:
                cdb.execute(text("SELECT 1")).fetchone()
        
        return jsonify({"status": "healthy"})
    except Exception as e:
        return jsonify({"status": "unhealthy", "error": str(e)}), 500

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email    = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")

        # Super-admin / registered-user login
        reg_user = RegisteredUser.query.filter_by(email=email, is_active=True).first()
        if reg_user and verify_password(password, reg_user.password_hash):
            if reg_user.role == "super_admin":
                session["user"] = {
                    "user_id": reg_user.user_id, "email": reg_user.email,
                    "full_name": reg_user.full_name, "role": "super_admin",
                    "company_id": None,
                }
                return redirect(url_for("admin_dashboard"))

            # Owner: may have zero (not yet onboarded), one, or multiple companies
            if not reg_user.email_verified and reg_user.must_change_password:
                _generate_and_send_otp(reg_user.email)
                session["pending_password_change_email"] = reg_user.email
                return redirect(url_for("account_setup"))
            elif not reg_user.email_verified:
                _generate_and_send_otp(reg_user.email)
                return redirect(url_for("verify_otp"))
            elif reg_user.must_change_password:
                session["pending_password_change_email"] = reg_user.email
                return redirect(url_for("force_change_password"))
            return _finish_owner_login(reg_user)

        # Company employee login — search each company's DB
        emp_matches = []  # list of (company_id, CompanyUser row)
        for comp in Company.query.filter_by(is_active=True).all():
            try:
                _cdb = get_customer_session(comp.company_id)
                _emp = _cdb.query(CompanyUser).filter_by(email=email, is_active=True).first()
                if _emp and verify_password(password, _emp.password_hash):
                    emp_matches.append((comp.company_id, _emp))
            except Exception:
                continue

        if emp_matches:
            if len(emp_matches) == 1:
                comp_id, emp = emp_matches[0]
                session["user"] = {
                    "user_id": emp.user_id, "email": emp.email,
                    "full_name": emp.full_name, "role": emp.role,
                    "company_id": comp_id,
                }
                session["active_company_id"] = comp_id
                session.pop("pending_login_type", None)
                return redirect(url_for("dashboard"))
            else:
                # multiple companies — send to the picker instead of guessing
                session["pending_login_email"] = email
                session["pending_login_type"] = "employee"
                return redirect(url_for("select_company"))

        flash("Invalid email or password")
    return render_template("login.html")

@app.route("/company/update-terms", methods=["POST"])
@login_required
@owner_required
def update_company_terms():
    company_id = get_current_company()
    company    = get_company_by_id(company_id)
    if company:
        company.terms_footer   = request.form.get("terms_footer", "").strip() or None
        company.terms_annexure = request.form.get("terms_annexure", "").strip() or None

        # ── Terms Visibility (per print format) ──
        company.show_terms_customer_invoice = "show_terms_customer_invoice" in request.form
        company.show_terms_awb_invoice = "show_terms_awb_invoice" in request.form
        company.show_terms_performa_invoice = "show_terms_performa_invoice" in request.form
        company.show_terms_box_label = "show_terms_box_label" in request.form
        company.show_terms_shipping_label = "show_terms_shipping_label" in request.form

        db.session.commit()
        flash("Invoice terms updated.")
    return redirect(url_for("company_settings"))

@app.route("/verify-otp", methods=["GET", "POST"])
def verify_otp():
    email = session.get("otp_email")
    if not email:
        return redirect(url_for("login"))

    if request.method == "POST":
        entered = request.form.get("otp", "").strip()
        expires = session.get("otp_expires")
        if not expires or datetime.utcnow() > datetime.fromisoformat(expires):
            flash("Code expired. Request a new one.", "error")
            return redirect(url_for("verify_otp"))
        if hashlib.sha256(entered.encode()).hexdigest() != session.get("otp_hash"):
            flash("Incorrect code.", "error")
            return redirect(url_for("verify_otp"))

        reg_user = RegisteredUser.query.filter_by(email=email).first()
        reg_user.email_verified = True
        db.session.commit()
        session.pop("otp_email", None)
        session.pop("otp_hash", None)
        session.pop("otp_expires", None)

        if reg_user.must_change_password:
            session["pending_password_change_email"] = email
            return redirect(url_for("force_change_password"))
        return _finish_owner_login(reg_user)

    return render_template("verify_otp.html", email=email)


@app.route("/verify-otp/resend")
def resend_otp():
    email = session.get("otp_email")
    if email:
        _generate_and_send_otp(email)
        flash("A new code has been sent.", "success")
    return redirect(url_for("verify_otp"))


@app.route("/force-change-password", methods=["GET", "POST"])
def force_change_password():
    email = session.get("pending_password_change_email")
    if not email:
        return redirect(url_for("login"))
    reg_user = RegisteredUser.query.filter_by(email=email).first()
    if not reg_user:
        return redirect(url_for("login"))

    if request.method == "POST":
        password = request.form.get("password", "")
        confirm  = request.form.get("confirm_password", "")
        if len(password) < 8:
            flash("Password must be at least 8 characters", "error")
            return redirect(url_for("force_change_password"))
        if password != confirm:
            flash("Passwords don't match", "error")
            return redirect(url_for("force_change_password"))

        reg_user.password_hash = hash_password(password)
        reg_user.must_change_password = False
        db.session.commit()
        session.pop("pending_password_change_email", None)
        return _finish_owner_login(reg_user)

    return render_template("force_change_password.html", email=email)

@app.route("/account-setup", methods=["GET", "POST"])
def account_setup():
    email = session.get("otp_email") or session.get("pending_password_change_email")
    if not email:
        return redirect(url_for("login"))
    reg_user = RegisteredUser.query.filter_by(email=email).first()
    if not reg_user:
        return redirect(url_for("login"))

    if request.method == "POST":
        entered  = request.form.get("otp", "").strip()
        password = request.form.get("password", "")
        confirm  = request.form.get("confirm_password", "")

        expires = session.get("otp_expires")
        if not expires or datetime.utcnow() > datetime.fromisoformat(expires):
            flash("Code expired. Request a new one.", "error")
            return redirect(url_for("account_setup"))
        if hashlib.sha256(entered.encode()).hexdigest() != session.get("otp_hash"):
            flash("Incorrect code.", "error")
            return redirect(url_for("account_setup"))
        if len(password) < 8:
            flash("Password must be at least 8 characters", "error")
            return redirect(url_for("account_setup"))
        if password != confirm:
            flash("Passwords don't match", "error")
            return redirect(url_for("account_setup"))

        reg_user.email_verified = True
        reg_user.password_hash = hash_password(password)
        reg_user.must_change_password = False
        db.session.commit()

        session.pop("otp_email", None)
        session.pop("otp_hash", None)
        session.pop("otp_expires", None)
        session.pop("pending_password_change_email", None)
        return _finish_owner_login(reg_user)

    return render_template("account_setup.html", email=email)


@app.route("/account-setup/resend")
def resend_account_setup_otp():
    email = session.get("otp_email") or session.get("pending_password_change_email")
    if email:
        _generate_and_send_otp(email)
        flash("A new code has been sent.", "success")
    return redirect(url_for("account_setup"))

@app.route("/company/add", methods=["GET", "POST"])
@login_required
@owner_required
def add_new_company():
    
    company_id = get_current_company()
    user = get_current_user()
    
    if request.method == "POST":
        company_name = request.form.get("company_name", "").strip()
        gst_number = request.form.get("gst_number", "")
        address = request.form.get("address", "")
        phone = request.form.get("phone", "")
        
        if not company_name:
            flash("Company name is required")
            return redirect(url_for("add_new_company"))
        
        if is_company_name_taken(user.get("email"), company_name):
            flash(f"A company named '{company_name}' already exists. Please choose a different name.", "error")
            return redirect(url_for("add_new_company"))
        
        # ── Check if GST number is already used by ANY company ──
        if gst_number and is_gst_number_taken(gst_number):
            flash(f"GST number '{gst_number}' is already registered to another active company. Please check and try again.", "error")
            return redirect(url_for("add_new_company"))

        # Check if user can add more companies based on their plan
        can_add, message = check_new_company_limit(user.get("email"))
        if not can_add:
            flash(message)
            return redirect(url_for("company_settings"))
        
        # Create new company
        new_company_id = _next_numbered_id(db.session, Company.company_id, "COMP")
        
        # Get user's plan
        reg_user = RegisteredUser.query.filter_by(email=user.get("email")).first()
        plan = reg_user.subscription_plan if reg_user else None
        plan_obj = SubscriptionPlan.query.get(plan) or SubscriptionPlan.query.order_by(SubscriptionPlan.id).first()

        is_gst = request.form.get('is_gst_registered', '1') == '1'
        gst_number = request.form.get('gst_number', '').strip() if is_gst else None
        
        new_company = Company(
            company_id=new_company_id,
            company_name=company_name,
            owner_email=user.get("email"),
            subscription_plan=plan,
            subscription_start=today_ist(),
            subscription_end=today_ist() + timedelta(days=365),
            max_companies_allowed=plan_obj.max_companies,
            max_users_per_company=plan_obj.max_users,
            gst_number=gst_number if is_gst else '',
            address=address,
            phone=phone,
            created_at=today_ist(),
            is_active=True,
            is_gst_registered=is_gst,
            awb_prefix=(request.form.get("awb_prefix", "AHL") or "AHL").strip().upper(),
            awb_start=int(request.form.get("awb_start", 81000) or 81000),
        )
        db.session.add(new_company)
        db.session.commit()

        # ── Create dedicated VPS MySQL database and all customer tables ────────
        init_customer_db_for_company(new_company)

        # ── Create the owner as first CompanyUser in the new customer DB ───────
        cdb = get_customer_session(new_company_id)
        emp_id    = _next_numbered_id(cdb, CompanyUser.user_id, "EMP")
        new_emp   = CompanyUser(
            user_id=emp_id,
            company_id=new_company_id,
            email=user.get("email"),
            password_hash=hash_password(request.form.get("password", "Temp@123")),
            full_name=user.get("full_name", ""),
            role="owner",
            department="Management",
            phone=phone,
            is_active=True,
            created_at=today_ist(),
        )
        cdb.add(new_emp)
        cdb.commit()

        flash(f"Company '{company_name}' created successfully! A dedicated database has been provisioned.")
        return redirect(url_for("dashboard"))
    
    # GET request - show the form with user's plan information
    user = get_current_user()
    reg_user = RegisteredUser.query.filter_by(email=user.get("email")).first()
    plan_key = reg_user.subscription_plan if reg_user else None
    plan_obj = SubscriptionPlan.query.get(plan_key) or SubscriptionPlan.query.order_by(SubscriptionPlan.id).first()
    
    # Get current companies count for this owner
    companies_count = Company.query.filter_by(owner_email=user.get("email")).count()
    max_companies_allowed = plan_obj.max_companies
    
    # Parse max companies (handle "Unlimited" string)
    if max_companies_allowed == "Unlimited":
        max_companies = None
        remaining_companies = "Unlimited"
        can_add_more = True
    else:
        max_companies = int(max_companies_allowed)
        remaining_companies = max(0, max_companies - companies_count)
        can_add_more = remaining_companies > 0
    
    plan_config = {
        "name": plan_obj.name,
        "price": plan_obj.price,
        "max_companies": plan_obj.max_companies,
        "max_users": plan_obj.max_users,
        "features": plan_obj.features.split(",") if plan_obj.features else [],
        "companies_used": companies_count,
        "remaining_companies": remaining_companies,
        "max_companies_int": max_companies,
        "can_add_more": can_add_more,
    }
    
    return render_template(
                "add_company.html",
                plan_config=plan_config,
                current_count=companies_count,
                max_companies=plan_obj.max_companies if plan_obj.max_companies == "Unlimited" else int(plan_obj.max_companies),
                can_add=can_add_more,
                awb_prefix="AHL",
                awb_start=81000,
            )

@app.route("/select-company", methods=["GET", "POST"])
def select_company():
    if "user" in session:
        current_role = session["user"].get("role")
        login_type = "owner" if current_role in ("owner", "super_admin") else "employee"
        pending_email = session["user"].get("email")
    else:
        login_type = session.get("pending_login_type", "owner")
        pending_email = session.get("pending_login_email")
    
    if not pending_email:
        return redirect(url_for("login"))

    if request.method == "POST":
        company_id = request.form.get("company_id")

        if login_type == "employee":
            comp = get_company_by_id(company_id)
            emp = None
            if comp:
                _cdb = get_customer_session(company_id)
                emp = _cdb.query(CompanyUser).filter_by(
                    email=pending_email, company_id=company_id, is_active=True
                ).first()
            if emp:
                session["user"] = {
                    "user_id": emp.user_id, "email": emp.email,
                    "full_name": emp.full_name, "role": emp.role,
                    "company_id": company_id,
                }
                session["active_company_id"] = company_id
                session.pop("pending_login_email", None)
                session.pop("pending_login_type", None)
                return redirect(url_for("dashboard"))
            flash("Invalid company selection.")

        else:  # owner (existing logic, unchanged)
            company = get_company_by_id(company_id)
            if company and company.owner_email == pending_email:
                reg_user = RegisteredUser.query.filter_by(email=pending_email).first()
                session["user"] = {
                    "email": reg_user.email, "full_name": reg_user.full_name,
                    "role": reg_user.role, "user_id": reg_user.user_id,
                }
                session["active_company_id"] = company_id
                session.pop("pending_login_email", None)
                session.pop("pending_login_type", None)
                return redirect(url_for("dashboard"))
            flash("Invalid company selection.")

    # GET — build the list to render
    if login_type == "employee":
        companies = get_employee_companies(pending_email)
        first_emp = None
        if companies:
            _cdb = get_customer_session(companies[0].company_id)
            first_emp = _cdb.query(CompanyUser).filter_by(email=pending_email, is_active=True).first()
        user = {
            "full_name": first_emp.full_name if first_emp else pending_email,
            "email": pending_email,
            "role": first_emp.role if first_emp else "employee",
        }
        owner_user_count = owner_max_users = None
    else:
        companies = get_owner_companies(pending_email)
        user = get_current_user() or {"full_name": pending_email, "email": pending_email, "role": "owner"}
        owner_user_count, owner_max_users, _ = get_owner_user_stats(pending_email)

    return render_template("select_company.html", companies=companies, user=user,
                           owner_user_count=owner_user_count, owner_max_users=owner_max_users)


@app.route("/company/toggle-mobile-visibility/<company_id>", methods=["POST"])
def toggle_company_mobile_visibility(company_id):
    """Owner-only: show/hide one of their companies everywhere (desktop AND
    mobile) on the Select Company screen.

    No @login_required/@owner_required here on purpose: an owner picking
    between multiple companies reaches this page via session["pending_login_email"]
    BEFORE session["user"] is ever set, so those decorators would 404/redirect
    every single time this button is clicked. Auth is handled manually below,
    mirroring select_company()'s own two-path logic.
    """
    if "user" in session:
        user = get_current_user()
        if user.get("role") not in ("owner", "super_admin"):
            flash("Only company owner can access this page")
            return safe_redirect_after_denial()
        owner_email = user.get("email", "")
    else:
        if session.get("pending_login_type", "owner") != "owner":
            flash("Please login to continue")
            return redirect(url_for("login"))
        owner_email = session.get("pending_login_email") or ""
        if not owner_email:
            flash("Please login to continue")
            return redirect(url_for("login"))

    company = Company.query.filter_by(company_id=company_id, owner_email=owner_email).first()
    if not company:
        flash("Company not found.")
        return redirect(url_for("select_company"))
    company.hidden_on_mobile = not company.hidden_on_mobile
    db.session.commit()
    flash(f"{company.company_name} is now {'hidden' if company.hidden_on_mobile else 'visible'}.")
    return redirect(url_for("select_company"))


@app.route("/switch-company/<company_id>")
@login_required
def switch_company(company_id):
    user = get_current_user()
    company = get_company_by_id(company_id)
    if not company:
        flash("Company not found.")
        return redirect(url_for("dashboard"))

    if user.get("role") in ("owner", "super_admin") and company.owner_email == user.get("email"):
        session["active_company_id"] = company_id
        session["user"]["company_id"] = company_id
        flash(f"Switched to {company.company_name}")
        return redirect(url_for("dashboard"))

    # Non-owner: must have an active CompanyUser row in the target company
    cdb = get_customer_session(company_id)
    emp = cdb.query(CompanyUser).filter_by(
        email=user.get("email"), company_id=company_id, is_active=True
    ).first()
    if emp:
        session["active_company_id"] = company_id
        session["user"]["company_id"] = company_id
        session["user"]["role"] = emp.role
        session["user"]["user_id"] = emp.user_id
        flash(f"Switched to {company.company_name}")
    else:
        flash("You don't have access to that company.")
    return redirect(url_for("dashboard"))

@app.route("/onboarding/create-company", methods=["GET", "POST"])
@login_required
def onboard_company():
    """
    First-login step for accounts created by the super admin (register_client).
    The RegisteredUser already exists with no Company yet — this is where the
    client sets up their own company profile (GST, AWB numbering, etc).
    """
    user = get_current_user()
    email = user.get("email")

    reg_user = RegisteredUser.query.filter_by(email=email).first()
    if not reg_user:
        return redirect(url_for("login"))

    # If they already have a company, this step is done — don't let them repeat it
    existing = get_owner_companies(email)
    if existing:
        session["active_company_id"] = existing[0].company_id
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        company_name = request.form.get("company_name", "").strip()
        if not company_name:
            flash("Company name is required", "error")
            return redirect(url_for("onboard_company"))
        
        is_gst     = request.form.get("is_gst_registered", "1") == "1"
        gst_number = request.form.get('gst_number', '').strip() if is_gst else None
        
        if is_company_name_taken(email, company_name):
            flash(f"A company named '{company_name}' already exists. Please choose a different name.", "error")
            return redirect(url_for("onboard_company"))

        if gst_number and is_gst_number_taken(gst_number):
            flash(f"GST number '{gst_number}' is already registered to another active company. Please check and try again.", "error")
            return redirect(url_for("onboard_company"))
        
        is_gst     = request.form.get("is_gst_registered", "1") == "1"
        gst_number = request.form.get("gst_number", "").strip() if is_gst else ""
        address    = request.form.get("address", "").strip()
        phone      = request.form.get("phone", reg_user.phone or "").strip()
        awb_prefix = (request.form.get("awb_prefix", "AHL") or "AHL").strip().upper()
        awb_start  = int(request.form.get("awb_start", 81000) or 81000)

        plan_obj = SubscriptionPlan.query.get(reg_user.subscription_plan) or SubscriptionPlan.query.order_by(SubscriptionPlan.id).first()
        end_days = 730 if plan_obj.id == "custom" else 365

        new_company_id = _next_numbered_id(db.session, Company.company_id, "COMP")

        # ── NEW: Use custom plan values if available ───────────────────────
        max_companies = plan_obj.max_companies
        max_users = plan_obj.max_users
        
        # Override with custom values if this user has them
        if plan_obj.id == "custom":
            if reg_user.custom_max_companies:
                max_companies = str(reg_user.custom_max_companies)
            if reg_user.custom_max_users:
                max_users = str(reg_user.custom_max_users)

        new_company = Company(
            company_id=new_company_id,
            company_name=company_name,
            owner_email=email,
            subscription_plan=plan_obj.id,
            subscription_start=today_ist(),
            subscription_end=today_ist() + timedelta(days=end_days),
            max_companies_allowed=max_companies,
            max_users_per_company=max_users,
            gst_number=gst_number,
            address=address,
            phone=phone,
            created_at=today_ist(),
            is_active=True,
            is_gst_registered=is_gst,
            storage_type="cloud",
            awb_prefix=awb_prefix,
            awb_start=awb_start,
        )
        db.session.add(new_company)
        db.session.commit()

        init_customer_db_for_company(new_company)

        # Owner becomes the first CompanyUser — reuse their existing password,
        # don't ask for a second one.
        try:
            cdb = get_customer_session(new_company_id)
            new_emp = CompanyUser(
                user_id="EMP001",
                company_id=new_company_id,
                email=email,
                password_hash=reg_user.password_hash,
                full_name=reg_user.full_name,
                role="owner",
                department="Management",
                phone=phone,
                is_active=True,
                created_at=today_ist(),
            )
            cdb.add(new_emp)
            cdb.commit()
        except Exception as e:
            cdb.rollback()
            print(f"⚠  Could not create CompanyUser for {new_company_id}: {e}")

        session["active_company_id"] = new_company_id
        session["user"]["company_id"] = new_company_id
        flash(f"Company '{company_name}' created. Welcome to Nexa ERP!", "success")
        return redirect(url_for("dashboard"))

    return render_template("onboard_company.html", user=reg_user, awb_prefix="AHL", awb_start=81000)


@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        # ── Pull owner / account fields ───────────────────────────────────────
        email            = request.form.get("email", "").strip().lower()
        password         = request.form.get("password", "")
        confirm_password = request.form.get("confirm_password", "")
        full_name        = request.form.get("full_name", "").strip()
        phone            = request.form.get("phone", "").strip()
        plan_key         = request.form.get("subscription_plan", "starter")
        awb_prefix = request.form.get("awb_prefix", "AHL").strip().upper() or "AHL"
        awb_start  = int(request.form.get("awb_start", 81000) or 81000)
        # ── Pull primary company fields ───────────────────────────────────────
        company_name     = request.form.get("company_name", "").strip()
        address          = request.form.get("address", request.form.get("company_address_1", "")).strip()
        company_phone    = request.form.get("company_phone_1", phone).strip()
        is_gst           = request.form.get("is_gst_registered", "1") == "1"
        gst_number = request.form.get('gst_number', '').strip() if is_gst else None

        # ── Extra companies (from hidden JSON field) ──────────────────────────
        extra_companies_raw = request.form.get("extra_companies", "[]")
        try:
            extra_companies = json.loads(extra_companies_raw)
            if not isinstance(extra_companies, list):
                extra_companies = []
        except (ValueError, TypeError):
            extra_companies = []

        # ── Validations ───────────────────────────────────────────────────────
        if not email:
            flash("Email is required", "error")
            return redirect(url_for("register"))

        if RegisteredUser.query.filter_by(email=email).first():
            flash("An account with this email already exists", "error")
            return redirect(url_for("register"))

        if password != confirm_password:
            flash("Passwords do not match", "error")
            return redirect(url_for("register"))

        if len(password) < 8:
            flash("Password must be at least 8 characters", "error")
            return redirect(url_for("register"))

        if not company_name:
            flash("Company name is required", "error")
            return redirect(url_for("register"))

        # ── Plan lookup ───────────────────────────────────────────────────────
        plan_obj = SubscriptionPlan.query.get(plan_key) or SubscriptionPlan.query.order_by(SubscriptionPlan.id).first()
        if not plan_obj:
            flash("No subscription plans are configured. Contact support.", "error")
            return redirect(url_for("register"))
        end_days = 730 if plan_obj.id == "custom" else 365

        # ── Check extra companies don't exceed plan limit ─────────────────────
        max_c = plan_obj.max_companies
        try:
            max_c_int = int(max_c)
            total_requested = 1 + len(extra_companies)
            if total_requested > max_c_int:
                flash(
                    f"Your {plan_obj.name} allows up to {max_c_int} "
                    f"{'company' if max_c_int == 1 else 'companies'}. "
                    f"You requested {total_requested}.",
                    "error"
                )
                return redirect(url_for("register"))
        except (ValueError, TypeError):
            pass  # "Unlimited" — no cap

        # ── Create RegisteredUser (platform DB) ───────────────────────────────
        user_id = generate_next_user_id()

        new_user = RegisteredUser(
            user_id=user_id,
            email=email,
            password_hash=hash_password(password),
            full_name=full_name,
            phone=phone,
            role="owner",
            subscription_plan=plan_obj.id,
            created_at=today_ist(),
            is_active=True,
        )
        db.session.add(new_user)
        db.session.flush()  # get id without committing

        # ── Helper: create one Company record + its customer DB ───────────────
        def _create_company(c_name, c_address, c_phone, c_gst_registered, c_gst_number, c_awb_prefix="AHL", c_awb_start=81000):
            if is_company_name_taken(email, c_name):
                # Can't flash inside nested function, so we raise an exception
                raise ValueError(f"Company name '{c_name}' is already taken. Please choose a different name.")
            
            # ── Check if GST number is already used by ANY company ──
            if c_gst_number and is_gst_number_taken(c_gst_number):
                raise ValueError(f"GST number '{c_gst_number}' is already registered to another active company. Please check and try again.")
            c_id       = _next_numbered_id(db.session, Company.company_id, "COMP")

            company = Company(
                company_id=c_id,
                company_name=c_name,
                owner_email=email,
                subscription_plan=plan_obj.id,
                subscription_start=today_ist(),
                subscription_end=today_ist() + timedelta(days=end_days),
                max_companies_allowed=plan_obj.max_companies,
                max_users_per_company=plan_obj.max_users,
                address=c_address,
                phone=c_phone or phone,
                gst_number = c_gst_number if c_gst_registered else None,
                is_gst_registered=c_gst_registered,
                created_at=today_ist(),
                is_active=True,
                storage_type="cloud",
                awb_prefix=c_awb_prefix,
                awb_start=c_awb_start,
            )
            db.session.add(company)
            db.session.flush()  # make company_id available before commit

            return c_id

        # ── Create primary company ─────────────────────────────────────────────
        try:
            primary_company_id = _create_company(
                company_name, address, company_phone, is_gst, gst_number,
                awb_prefix, awb_start
            )
        except ValueError as e:
            flash(str(e), "error")
            return redirect(url_for("register"))

        # ── Create extra companies ────────────────────────────────────────────
        extra_company_ids = []
        for ec in extra_companies:
            ec_name = ec.get("name", "").strip()
            if not ec_name:
                continue
            try:
                ec_id = _create_company(
                    ec_name,
                    ec.get("address", ""),
                    ec.get("phone", ""),
                    bool(ec.get("is_gst_registered", True)),
                    ec.get("gst_number", "").strip() if ec.get("is_gst_registered", True) else None,
                    (ec.get("awb_prefix", "") or "AHL").strip().upper() or "AHL",
                    int(ec.get("awb_start", 81000) or 81000),
                )
                extra_company_ids.append(ec_id)
            except ValueError as e:
                flash(str(e), "error")
                return redirect(url_for("register"))

        # ── Commit all platform records at once ───────────────────────────────
        db.session.commit()

        # ── Bootstrap customer databases ──────────────────────────────────────
        all_company_ids = [primary_company_id] + extra_company_ids

        for c_id in all_company_ids:
            try:
                company_obj = Company.query.filter_by(company_id=c_id).first()
                init_customer_db_for_company(company_obj)
            except Exception as e:
                print(f"⚠  Could not init customer DB for {c_id}: {e}")

        # ── Create owner as CompanyUser in primary company's DB ───────────────
        try:
            cdb       = get_customer_session(primary_company_id)
            emp_id    = _next_numbered_id(cdb, CompanyUser.user_id, "EMP")

            new_emp = CompanyUser(
                user_id=emp_id,
                company_id=primary_company_id,
                email=email,
                password_hash=hash_password(password),
                full_name=full_name,
                role="owner",
                department="Management",
                phone=phone,
                is_active=True,
                created_at=today_ist(),
            )
            cdb.add(new_emp)
            cdb.commit()
        except Exception as e:
            cdb.rollback()
            print(f"⚠  Could not create CompanyUser for {primary_company_id}: {e}")

        # ── Also add owner as CompanyUser in any extra company DBs ────────────
        for c_id in extra_company_ids:
            try:
                cdb       = get_customer_session(c_id)
                emp_id    = _next_numbered_id(cdb, CompanyUser.user_id, "EMP")
                extra_emp = CompanyUser(
                    user_id=emp_id,
                    company_id=c_id,
                    email=email,
                    password_hash=hash_password(password),
                    full_name=full_name,
                    role="owner",
                    department="Management",
                    phone=phone,
                    is_active=True,
                    created_at=today_ist(),
                )
                cdb.add(extra_emp)
                cdb.commit()
            except Exception as e:
                cdb.rollback()
                print(f"⚠  Could not create CompanyUser for {c_id}: {e}")

        total = len(all_company_ids)
        flash(
            f"Welcome to Nexa ERP! Your account and "
            f"{total} {'company has' if total == 1 else 'companies have'} been set up. Please login.",
            "success"
        )
        return redirect(url_for("login"))

    return render_template("register.html", plans=get_all_plans())

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ─────────────────────────────────────────────────────────────────────────────
# ── Dashboard ─────────────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────


def _ensure_export_history_table(cdb):
    """Create the export_history table in the company DB if it doesn't exist yet."""
    cdb.execute(text("""
        CREATE TABLE IF NOT EXISTS export_history (
            id INT AUTO_INCREMENT PRIMARY KEY,
            company_id VARCHAR(64) NOT NULL,
            export_type VARCHAR(20) NOT NULL,
            sales_from DATE NULL,
            sales_to DATE NULL,
            purchase_from DATE NULL,
            purchase_to DATE NULL,
            filename VARCHAR(255) NOT NULL,
            exported_by VARCHAR(255) NULL,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    """))
    cdb.commit()


def _log_export_history(cdb, company_id, export_type, sales_from_date, sales_to_date,
                         purchase_from_date, purchase_to_date, filename, exported_by):
    try:
        _ensure_export_history_table(cdb)
        cdb.execute(
            text("""
                INSERT INTO export_history
                    (company_id, export_type, sales_from, sales_to, purchase_from, purchase_to, filename, exported_by)
                VALUES
                    (:company_id, :export_type, :sales_from, :sales_to, :purchase_from, :purchase_to, :filename, :exported_by)
            """),
            {
                "company_id": company_id,
                "export_type": export_type,
                "sales_from": sales_from_date,
                "sales_to": sales_to_date,
                "purchase_from": purchase_from_date,
                "purchase_to": purchase_to_date,
                "filename": filename,
                "exported_by": exported_by,
            },
        )
        cdb.commit()
    except Exception as e:
        # Never let history logging break the actual export/download.
        print(f"Could not log export history: {e}")
        cdb.rollback()


def _get_export_history(cdb, company_id, limit=15):
    try:
        _ensure_export_history_table(cdb)
        rows = cdb.execute(
            text("""
                SELECT export_type, sales_from, sales_to, purchase_from, purchase_to,
                       filename, exported_by, created_at
                FROM export_history
                WHERE company_id = :company_id
                ORDER BY created_at DESC
                LIMIT :limit
            """),
            {"company_id": company_id, "limit": limit},
        ).mappings().all()
        return [dict(r) for r in rows]
    except Exception as e:
        print(f"Could not load export history: {e}")
        return []


@app.route("/reports/export-selector")
@login_required
@require_permission("analytics", "view")
def export_selector():
    """
    Standalone page for picking a date range and metric (sales / purchase /
    both) before exporting. Independent of reports_dashboard() / 
    report_dashboard.html — this only feeds query params into the existing
    export_reports_excel() route below via a plain GET form.
    """
    company_id = get_current_company()
    company = get_company_by_id(company_id)
    if not company:
        flash("Company not found")
        return redirect(url_for("logout"))

    cdb = get_cdb()
    history = _get_export_history(cdb, company_id) if cdb else []

    default_from = today_ist().replace(day=1).strftime("%Y-%m-%d")
    default_to = today_ist().strftime("%Y-%m-%d")

    return render_template(
        "export_selector.html",
        active="export_selector",
        company=company,
        default_from=default_from,
        default_to=default_to,
        history=history,
    )


@app.route("/reports/export-excel")
@login_required
@require_permission("analytics", "view")
def export_reports_excel():
    """
    Export Sales, Purchase, and Pending (receivables + payables) to a single
    multi-sheet Excel file. Pending is computed from Invoice.balance /
    PurchaseInvoice.balance directly rather than the cached Client.pending /
    Supplier.payable fields, so the export always matches what the individual
    invoices actually say — no risk of it drifting from a stale cache field.
    Pending Receivable/Payable are filtered by invoice date using the same
    Sales/Purchase date range as the rest of the export (by choice — an
    unpaid invoice outside the selected range won't appear, even though it
    may still be owed today).

    Query params:
      - sales_from / sales_to: date range for Sales
      - purchase_from / purchase_to: date range for Purchase
      - type: 'sales', 'purchase', or 'both' (default: 'both')
      - from / to: legacy fallback (applies to both)
    """
    company_id = get_current_company()
    company = get_company_by_id(company_id)
    if not company:
        flash("Company not found")
        return redirect(url_for("logout"))

    cdb = get_cdb()
    if not cdb:
        flash("Could not connect to company database")
        return redirect(url_for("logout"))

    # ── Parse export type ────────────────────────────────────────────────────────
    export_type = request.args.get("type", "both").strip().lower()
    if export_type not in ("sales", "purchase", "both"):
        export_type = "both"

    # ── Parse Sales date range ──────────────────────────────────────────────────
    sales_from_str = request.args.get("sales_from", request.args.get("from", "")).strip()
    sales_to_str = request.args.get("sales_to", request.args.get("to", "")).strip()

    date_parse_warnings = []

    try:
        sales_from_date = datetime.strptime(sales_from_str, "%Y-%m-%d").date() if sales_from_str else None
    except ValueError:
        sales_from_date = None
        date_parse_warnings.append(f"Sales 'from' date '{sales_from_str}' was not understood and was ignored.")
    try:
        sales_to_date = datetime.strptime(sales_to_str, "%Y-%m-%d").date() if sales_to_str else None
    except ValueError:
        sales_to_date = None
        date_parse_warnings.append(f"Sales 'to' date '{sales_to_str}' was not understood and was ignored.")

    # ── Parse Purchase date range ────────────────────────────────────────────────
    purchase_from_str = request.args.get("purchase_from", request.args.get("from", "")).strip()
    purchase_to_str = request.args.get("purchase_to", request.args.get("to", "")).strip()

    try:
        purchase_from_date = datetime.strptime(purchase_from_str, "%Y-%m-%d").date() if purchase_from_str else None
    except ValueError:
        purchase_from_date = None
        date_parse_warnings.append(f"Purchase 'from' date '{purchase_from_str}' was not understood and was ignored.")
    try:
        purchase_to_date = datetime.strptime(purchase_to_str, "%Y-%m-%d").date() if purchase_to_str else None
    except ValueError:
        purchase_to_date = None
        date_parse_warnings.append(f"Purchase 'to' date '{purchase_to_str}' was not understood and was ignored.")

    if date_parse_warnings:
        flash(
            "Some date filters couldn't be applied, so those sheets were exported unfiltered: "
            + " ".join(date_parse_warnings),
            "error",
        )

    # ── Legacy fallback: if no sales-specific dates, use from/to ───────────────
    legacy_from_str = request.args.get("from", "").strip()
    legacy_to_str = request.args.get("to", "").strip()
    if sales_from_date is None and sales_to_date is None and (legacy_from_str or legacy_to_str):
        try:
            sales_from_date = datetime.strptime(legacy_from_str, "%Y-%m-%d").date() if legacy_from_str else None
        except ValueError:
            pass
        try:
            sales_to_date = datetime.strptime(legacy_to_str, "%Y-%m-%d").date() if legacy_to_str else None
        except ValueError:
            pass
    if purchase_from_date is None and purchase_to_date is None and (legacy_from_str or legacy_to_str):
        try:
            purchase_from_date = datetime.strptime(legacy_from_str, "%Y-%m-%d").date() if legacy_from_str else None
        except ValueError:
            pass
        try:
            purchase_to_date = datetime.strptime(legacy_to_str, "%Y-%m-%d").date() if legacy_to_str else None
        except ValueError:
            pass

    # ── Helper: apply date filters to a query ──────────────────────────────────
    def apply_date_filters(query, date_column, from_date, to_date):
        if from_date:
            query = query.filter(date_column >= from_date)
        if to_date:
            query = query.filter(date_column <= to_date)
        return query

    # ── Sales ────────────────────────────────────────────────────────────────────
    sales_rows = []
    sales_df = None
    if export_type in ("sales", "both"):
        sales_q = cdb.query(Invoice).filter(Invoice.company_id == company_id)
        sales_q = apply_date_filters(sales_q, Invoice.date, sales_from_date, sales_to_date)
        sales_invoices = sales_q.order_by(Invoice.date.asc()).all()

        clients_by_id = {c.id: c for c in cdb.query(Client).filter_by(company_id=company_id).all()}

        for inv in sales_invoices:
            client = clients_by_id.get(inv.client_id)
            sales_rows.append({
                "Invoice No":   inv.invoice_id,
                "Date":         inv.date.strftime("%Y-%m-%d") if inv.date else "",
                "Due Date":     inv.due_date.strftime("%Y-%m-%d") if inv.due_date else "",
                "Client":       client.name if client else (inv.contact_person or "—"),
                "Phone":        client.phone if client else (inv.phone or ""),
                "Subtotal":     round(float(inv.subtotal or 0), 2),
                "Tax":          round(float(inv.tax_amount or 0), 2),
                "Grand Total":  round(float(inv.grand_total or 0), 2),
                "Paid":         round(float(inv.paid_amount or 0), 2),
                "Balance":      round(float(inv.balance or 0), 2),
                "Status":       inv.status,
            })
        sales_df = pd.DataFrame(sales_rows, columns=[
            "Invoice No", "Date", "Due Date", "Client", "Phone",
            "Subtotal", "Tax", "Grand Total", "Paid", "Balance", "Status",
        ])

    # ── Purchase ─────────────────────────────────────────────────────────────────
    purchase_rows = []
    purchase_df = None
    if export_type in ("purchase", "both"):
        purchase_q = cdb.query(PurchaseInvoice).filter(PurchaseInvoice.company_id == company_id)
        purchase_q = apply_date_filters(purchase_q, PurchaseInvoice.date, purchase_from_date, purchase_to_date)
        purchase_invoices = purchase_q.order_by(PurchaseInvoice.date.asc()).all()

        suppliers_by_id = {s.id: s for s in cdb.query(Supplier).filter_by(company_id=company_id).all()}

        for pur in purchase_invoices:
            supplier = suppliers_by_id.get(pur.supplier_id)
            purchase_rows.append({
                "Invoice No":       pur.invoice_id,
                "Supplier Invoice #": pur.invoice_number or "",
                "Date":             pur.date.strftime("%Y-%m-%d") if pur.date else "",
                "Due Date":         pur.due_date.strftime("%Y-%m-%d") if pur.due_date else "",
                "Supplier":         supplier.name if supplier else (pur.supplier_name or "—"),
                "Phone":            supplier.phone if supplier else "",
                "Subtotal":         round(float(pur.subtotal or 0), 2),
                "Tax":              round(float(pur.tax_amount or 0), 2),
                "Grand Total":      round(float(pur.grand_total or 0), 2),
                "Paid":             round(float(pur.paid_amount or 0), 2),
                "Balance":          round(float(pur.balance or 0), 2),
                "Status":           pur.status,
            })
        purchase_df = pd.DataFrame(purchase_rows, columns=[
            "Invoice No", "Supplier Invoice #", "Date", "Due Date", "Supplier", "Phone",
            "Subtotal", "Tax", "Grand Total", "Paid", "Balance", "Status",
        ])

    # ── Pending: Receivables (unpaid/partial sales) ─────────────────────────
    # Filtered by invoice date, same range as the Sales sheet (sales_from_date/
    # sales_to_date) — this is a deliberate choice: it means an unpaid invoice
    # from outside the selected range will NOT show up here even though it's
    # still owed today. If you want total outstanding exposure regardless of
    # period, remove this date filter and query all unpaid invoices instead.
    receivable_rows = []
    receivable_df = None
    if export_type in ("sales", "both"):
        receivable_q = cdb.query(Invoice).filter(Invoice.company_id == company_id)
        receivable_q = apply_date_filters(receivable_q, Invoice.date, sales_from_date, sales_to_date)
        all_sales = receivable_q.all()
        clients_by_id = {c.id: c for c in cdb.query(Client).filter_by(company_id=company_id).all()}
        for inv in all_sales:
            bal = round(float(inv.balance or 0), 2)
            if bal > 0:
                client = clients_by_id.get(inv.client_id)
                receivable_rows.append({
                    "Invoice No":  inv.invoice_id,
                    "Date":        inv.date.strftime("%Y-%m-%d") if inv.date else "",
                    "Client":      client.name if client else (inv.contact_person or "—"),
                    "Phone":       client.phone if client else (inv.phone or ""),
                    "Grand Total": round(float(inv.grand_total or 0), 2),
                    "Paid":        round(float(inv.paid_amount or 0), 2),
                    "Balance Due": bal,
                })
        receivable_df = pd.DataFrame(receivable_rows, columns=[
            "Invoice No", "Date", "Client", "Phone", "Grand Total", "Paid", "Balance Due",
        ])
        total_receivable = round(sum(r["Balance Due"] for r in receivable_rows), 2)
    else:
        total_receivable = 0

    # ── Pending: Payables (unpaid/partial purchases) ────────────────────────
    # Filtered by invoice date, same range as the Purchase sheet — see the
    # note above Pending: Receivables for the tradeoff this implies.
    payable_rows = []
    payable_df = None
    if export_type in ("purchase", "both"):
        payable_q = cdb.query(PurchaseInvoice).filter(PurchaseInvoice.company_id == company_id)
        payable_q = apply_date_filters(payable_q, PurchaseInvoice.date, purchase_from_date, purchase_to_date)
        all_purchases = payable_q.all()
        suppliers_by_id = {s.id: s for s in cdb.query(Supplier).filter_by(company_id=company_id).all()}
        for pur in all_purchases:
            bal = round(float(pur.balance or 0), 2)
            if bal > 0:
                supplier = suppliers_by_id.get(pur.supplier_id)
                payable_rows.append({
                    "Invoice No":  pur.invoice_id,
                    "Date":        pur.date.strftime("%Y-%m-%d") if pur.date else "",
                    "Supplier":    supplier.name if supplier else (pur.supplier_name or "—"),
                    "Phone":       supplier.phone if supplier else "",
                    "Grand Total": round(float(pur.grand_total or 0), 2),
                    "Paid":        round(float(pur.paid_amount or 0), 2),
                    "Balance Due": bal,
                })
        payable_df = pd.DataFrame(payable_rows, columns=[
            "Invoice No", "Date", "Supplier", "Phone", "Grand Total", "Paid", "Balance Due",
        ])
        total_payable = round(sum(p["Balance Due"] for p in payable_rows), 2)
    else:
        total_payable = 0

    # ── Write workbook ───────────────────────────────────────────────────────
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        # Sales sheet
        if sales_df is not None and not sales_df.empty:
            sales_df.to_excel(writer, sheet_name="Sales", index=False)
        elif sales_df is not None:
            # Empty sheet with headers
            sales_df.to_excel(writer, sheet_name="Sales", index=False)

        # Purchase sheet
        if purchase_df is not None and not purchase_df.empty:
            purchase_df.to_excel(writer, sheet_name="Purchase", index=False)
        elif purchase_df is not None:
            purchase_df.to_excel(writer, sheet_name="Purchase", index=False)

        # Pending sheets
        if receivable_df is not None and not receivable_df.empty:
            receivable_df.to_excel(writer, sheet_name="Pending - Receivable", index=False)
        elif receivable_df is not None:
            receivable_df.to_excel(writer, sheet_name="Pending - Receivable", index=False)

        if payable_df is not None and not payable_df.empty:
            payable_df.to_excel(writer, sheet_name="Pending - Payable", index=False)
        elif payable_df is not None:
            payable_df.to_excel(writer, sheet_name="Pending - Payable", index=False)

        from openpyxl.styles import Font, PatternFill

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")

        # Apply formatting to all sheets
        sheet_df_map = {
            "Sales": sales_df,
            "Purchase": purchase_df,
            "Pending - Receivable": receivable_df,
            "Pending - Payable": payable_df,
        }

        for sheet_name, df in sheet_df_map.items():
            if df is None or sheet_name not in writer.sheets:
                continue
            ws = writer.sheets[sheet_name]
            # Only apply header formatting if there's at least one row
            if not df.empty:
                for col_idx, col_name in enumerate(df.columns, start=1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.font = header_font
                    cell.fill = header_fill
                    max_len = max([len(str(col_name))] + [len(str(v)) for v in df[col_name].astype(str)]) if len(df) else len(str(col_name))
                    ws.column_dimensions[cell.column_letter].width = min(max(max_len + 3, 12), 40)
                ws.freeze_panes = "A2"

        # Totals row under each pending sheet
        if "Pending - Receivable" in writer.sheets and receivable_df is not None:
            rec_ws = writer.sheets["Pending - Receivable"]
            rec_total_row = len(receivable_df) + 3
            rec_ws.cell(row=rec_total_row, column=6, value="Total Receivable:").font = Font(bold=True)
            rec_ws.cell(row=rec_total_row, column=7, value=total_receivable).font = Font(bold=True)

        if "Pending - Payable" in writer.sheets and payable_df is not None:
            pay_ws = writer.sheets["Pending - Payable"]
            pay_total_row = len(payable_df) + 3
            pay_ws.cell(row=pay_total_row, column=6, value="Total Payable:").font = Font(bold=True)
            pay_ws.cell(row=pay_total_row, column=7, value=total_payable).font = Font(bold=True)

    buf.seek(0)

    # ── Build filename ──────────────────────────────────────────────────────
    # Show date ranges in filename
    parts = []
    if export_type in ("sales", "both") and (sales_from_date or sales_to_date):
        from_str = sales_from_date.strftime("%Y%m%d") if sales_from_date else "start"
        to_str = sales_to_date.strftime("%Y%m%d") if sales_to_date else "end"
        parts.append(f"Sales_{from_str}-{to_str}")
    if export_type in ("purchase", "both") and (purchase_from_date or purchase_to_date):
        from_str = purchase_from_date.strftime("%Y%m%d") if purchase_from_date else "start"
        to_str = purchase_to_date.strftime("%Y%m%d") if purchase_to_date else "end"
        parts.append(f"Purchase_{from_str}-{to_str}")

    suffix = "_".join(parts) if parts else "all_time"
    filename = f"{company.company_name.replace(' ', '_')}_{export_type.title()}_{suffix}.xlsx"

    _log_export_history(
        cdb, company_id, export_type,
        sales_from_date, sales_to_date,
        purchase_from_date, purchase_to_date,
        filename, get_current_user().get("email"),
    )

    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/reports-dashboard")
@login_required
@require_permission("analytics", "view")
def reports_dashboard():
    
    company_id = get_current_company()
    company = get_company_by_id(company_id)
    
    if not company:
        flash("Company not found")
        return redirect(url_for("logout"))
    
    cdb = get_cdb()
    if not cdb:
        flash("Could not connect to company database")
        return redirect(url_for("logout"))
    
    # Set default dates (current month)
    from_date = today_ist().replace(day=1)
    to_date = today_ist()
    
    # Get initial data for the template
    # Cash in Hand
    cash_transactions = cdb.query(CashTransaction).filter_by(company_id=company_id).all()
    cash_balance = sum(t.amount for t in cash_transactions if t.type == 'income') - \
                   sum(t.amount for t in cash_transactions if t.type == 'expense')
    
    # Bank Balance
    bank_accounts = cdb.query(BankAccount).filter_by(company_id=company_id, status='Active').all()
    bank_balance = sum(acc.balance for acc in bank_accounts)
    
    # Total Revenue (current month)
    sales_invoices = cdb.query(Invoice).filter(
        Invoice.company_id == company_id,
        Invoice.date >= from_date,
        Invoice.date <= to_date
    ).all()
    total_revenue = sum(float(inv.grand_total or 0) for inv in sales_invoices)
    
    # Total Purchases (current month)
    purchase_invoices = cdb.query(PurchaseInvoice).filter(
        PurchaseInvoice.company_id == company_id,
        PurchaseInvoice.date >= from_date,
        PurchaseInvoice.date <= to_date
    ).all()
    total_purchases = sum(float(pur.grand_total or 0) for pur in purchase_invoices)
    
    # Profit
    profit = total_revenue - total_purchases
    
    # Pending Amount
    all_invoices = cdb.query(Invoice).filter_by(company_id=company_id).all()
    pending_amount = sum(float(getattr(inv, 'balance', 0) or 0) for inv in all_invoices)
    
    # Cash flow for period
    period_cash_income = cdb.query(CashTransaction).filter(
        CashTransaction.company_id == company_id,
        CashTransaction.type == 'income',
        CashTransaction.date >= from_date,
        CashTransaction.date <= to_date
    ).all()
    period_cash_expense = cdb.query(CashTransaction).filter(
        CashTransaction.company_id == company_id,
        CashTransaction.type == 'expense',
        CashTransaction.date >= from_date,
        CashTransaction.date <= to_date
    ).all()
    cash_inflow_period = sum(t.amount for t in period_cash_income)
    cash_outflow_period = sum(t.amount for t in period_cash_expense)
    cash_net_period = cash_inflow_period - cash_outflow_period
    
    # Chart Data (Last 6 months)
    chart_labels = []
    revenue_data = []
    purchase_data = []
    profit_trend = []
    profit_labels = []
    
    for i in range(5, -1, -1):
        month_date = today_ist().replace(day=1) - timedelta(days=30 * i)
        month_start = month_date.replace(day=1)
        if month_date.month == 12:
            month_end = month_date.replace(day=31)
        else:
            month_end = month_date.replace(month=month_date.month + 1, day=1) - timedelta(days=1)
        
        month_label = month_date.strftime('%b %Y')
        chart_labels.append(month_label)
        
        month_revenue = sum(
            float(inv.grand_total or 0) for inv in cdb.query(Invoice).filter(
                Invoice.company_id == company_id,
                Invoice.date >= month_start,
                Invoice.date <= month_end
            ).all()
        )
        revenue_data.append(month_revenue / 100000)

        month_purchases = sum(
            float(pur.grand_total or 0) for pur in cdb.query(PurchaseInvoice).filter(
                PurchaseInvoice.company_id == company_id,
                PurchaseInvoice.date >= month_start,
                PurchaseInvoice.date <= month_end
            ).all()
        )
        purchase_data.append(month_purchases / 100000)

        month_profit = month_revenue - month_purchases
        profit_trend.append(month_profit / 1000)
        profit_labels.append(month_label)
    
    # Status counts for all shipments
    all_customer_invoices = cdb.query(Invoice).filter(
        Invoice.company_id == company_id,
        Invoice.invoice_id.like("CUST-%")
    ).all()
    
    status_counts = {
        "delivered": sum(1 for i in all_customer_invoices if i.status == "Paid"),
        "in_transit": sum(1 for i in all_customer_invoices if i.status == "Partial"),
        "pending": sum(1 for i in all_customer_invoices if i.status not in ["Paid", "Partial", "Draft"]),
        "draft": sum(1 for i in all_customer_invoices if i.status == "Draft"),
        "total": len(all_customer_invoices)
    }
    
    # Payment methods breakdown
    cash_txns = cdb.query(CashTransaction).filter_by(company_id=company_id, type='income').all()
    bank_txns = cdb.query(BankTransaction).filter_by(company_id=company_id, type='credit').all()
    
    payment_methods = {
        "Cash": sum(t.amount for t in cash_txns),
        "Online/UPI": sum(t.amount for t in bank_txns if t.transaction_mode == "Online"),
        "Cheque": sum(t.amount for t in bank_txns if t.transaction_mode == "Cheque"),
    }
    
    # Top clients
    clients = cdb.query(Client).filter_by(company_id=company_id).all()
    top_clients_data = []
    for client in clients[:10]:
        client_invoices = cdb.query(Invoice).filter_by(company_id=company_id, client_id=client.id).all()
        client_shipments = [i for i in client_invoices if i.invoice_id.startswith("CUST-")]
        total_billed = sum(float(inv.grand_total or 0) for inv in client_invoices)
        pending = sum(float(getattr(inv, 'balance', 0) or 0) for inv in client_invoices)
        top_clients_data.append({
            "name": client.name,
            "total_billed": total_billed,
            "pending": pending,
            "shipment_count": len(client_shipments)
        })
    top_clients_data.sort(key=lambda x: x["total_billed"], reverse=True)
    top_clients_data = top_clients_data[:5]
    
    # Recent shipments (last 10)
    recent_shipments = []
    for inv in all_customer_invoices[:10]:
        meta = {}
        if inv.terms:
            try:
                meta = json.loads(inv.terms)
            except:
                pass
        status_label = "Delivered" if inv.status == "Paid" else "In Transit" if inv.status == "Partial" else "Pending" if inv.status != "Draft" else "Draft"
        status_class = "delivered" if inv.status == "Paid" else "transit" if inv.status == "Partial" else "pending"
        recent_shipments.append({
            "docket_no": meta.get("docket_no", inv.invoice_id),
            "customer_name": inv.client_obj.name if inv.client_obj else (inv.contact_person or "—"),
            "destination": meta.get("destination", ""),
            "total": float(inv.grand_total or 0),
            "status": inv.status,
            "status_label": status_label,
            "status_class": status_class
        })
    
    # Recent payments
    recent_payments = []
    for txn in cash_txns[:10]:
        recent_payments.append({
            "date": txn.date.strftime("%d %b %Y"),
            "customer": txn.description[:30],
            "invoice_id": txn.reference or "—",
            "amount": txn.amount,
            "mode": "Cash"
        })
    for txn in bank_txns[:5]:
        recent_payments.append({
            "date": txn.date.strftime("%d %b %Y"),
            "customer": txn.description[:30],
            "invoice_id": txn.reference or "—",
            "amount": txn.amount,
            "mode": txn.transaction_mode or "Bank"
        })
    recent_payments.sort(key=lambda x: x['date'], reverse=True)
    recent_payments = recent_payments[:10]
    
    # Pending invoices
    pending_invoices = []
    for inv in all_invoices:
        balance = float(getattr(inv, 'balance', 0) or 0)
        if balance > 0:
            pending_invoices.append({
                "invoice_id": inv.invoice_id,
                "customer": inv.client_obj.name if inv.client_obj else (inv.contact_person or "—"),
                "date": inv.date.strftime("%d %b %Y") if inv.date else "—",
                "due_date": inv.due_date.strftime("%d %b %Y") if inv.due_date else "—",
                "balance": balance
            })
    pending_invoices = pending_invoices[:10]
    
    kpi = {
        "cash_balance": cash_balance,
        "bank_balance": bank_balance,
        "total_revenue": total_revenue,
        "total_purchases": total_purchases,
        "profit": profit,
        "pending_amount": pending_amount,
        "cash_inflow_period": cash_inflow_period,
        "cash_outflow_period": cash_outflow_period,
        "cash_net_period": cash_net_period,
    }
    
    return render_template("report_dashboard.html",
                         company=company,
                         kpi=kpi,
                         from_date=from_date.strftime('%Y-%m-%d'),
                         to_date=to_date.strftime('%Y-%m-%d'),
                         chart_labels=chart_labels,
                         revenue_data=revenue_data,
                         purchase_data=purchase_data,
                         profit_labels=profit_labels,
                         profit_trend=profit_trend,
                         cash_inflow_period=cash_inflow_period,
                         cash_outflow_period=cash_outflow_period,
                         cash_net_period=cash_net_period,
                         top_clients_data=top_clients_data,
                         status_counts=status_counts,
                         payment_methods=payment_methods,
                         recent_shipments=recent_shipments,
                         recent_payments=recent_payments,
                         pending_invoices=pending_invoices,
                         total_shipments=status_counts["total"])

@app.route("/dashboard")
@login_required
@require_permission("dashboard", "view")
def dashboard():
    """Main business dashboard"""
    company_id = get_current_company()
    company = get_company_by_id(company_id)
    
    if not company:
        flash("Company not found")
        return redirect(url_for("logout"))
    
    cdb = get_cdb()
    if not cdb:
        flash("Could not connect to company database")
        return redirect(url_for("logout"))

    # Get date filters (default to all-time to match AJAX endpoint)
    from_date_str = request.args.get('from_date', '')
    to_date_str = request.args.get('to_date', '')
    
    if not from_date_str:
        from_date = date(2000, 1, 1)
    else:
        from_date = date.fromisoformat(from_date_str)
    
    if not to_date_str:
        to_date = today_ist()
    else:
        to_date = date.fromisoformat(to_date_str)

    # Cash in Hand
    cash_transactions = cdb.query(CashTransaction).filter_by(company_id=company_id).all()
    cash_balance = sum(t.amount for t in cash_transactions if t.type == 'income') - \
                   sum(t.amount for t in cash_transactions if t.type == 'expense')
    
    # Bank Balance
    bank_accounts = cdb.query(BankAccount).filter_by(company_id=company_id, status='Active').all()
    bank_balance = sum(acc.balance for acc in bank_accounts)
    
    # Sales Invoices (Revenue) - EXCLUDE Void and Draft
    sales_invoices = cdb.query(Invoice).filter(
        Invoice.company_id == company_id,
        Invoice.date >= from_date,
        Invoice.date <= to_date,
        Invoice.status.notin_(['Void', 'Draft'])  # ← EXCLUDE Void and Draft
    ).all()
    total_revenue = sum(float(inv.grand_total or 0) for inv in sales_invoices)
    
    # Purchase Invoices - EXCLUDE Void and Draft
    purchase_invoices = cdb.query(PurchaseInvoice).filter(
        PurchaseInvoice.company_id == company_id,
        PurchaseInvoice.date >= from_date,
        PurchaseInvoice.date <= to_date,
        PurchaseInvoice.status.notin_(['Void', 'Draft'])  # ← EXCLUDE Void and Draft
    ).all()
    total_purchases = sum(float(pur.grand_total or 0) for pur in purchase_invoices)
    
    # Expenses for the period
    period_expenses = cdb.query(Expense).filter(
        Expense.company_id == company_id,
        Expense.date >= from_date,
        Expense.date <= to_date
    ).all()
    total_expenses = sum(float(exp.amount or 0) for exp in period_expenses)
    
    gross_profit = total_revenue - total_purchases
    net_profit = gross_profit - total_expenses
    
    # Pending Amount - EXCLUDE Void and Draft
    all_active_invoices = cdb.query(Invoice).filter(
        Invoice.company_id == company_id,
        Invoice.status.notin_(['Void', 'Draft'])  # ← EXCLUDE Void and Draft
    ).all()
    pending_amount = sum(float(getattr(inv, 'balance', 0) or 0) for inv in all_active_invoices)
    
    # Cash flow for period
    period_cash_income = cdb.query(CashTransaction).filter(
        CashTransaction.company_id == company_id,
        CashTransaction.type == 'income',
        CashTransaction.date >= from_date,
        CashTransaction.date <= to_date
    ).all()
    period_cash_expense = cdb.query(CashTransaction).filter(
        CashTransaction.company_id == company_id,
        CashTransaction.type == 'expense',
        CashTransaction.date >= from_date,
        CashTransaction.date <= to_date
    ).all()
    cash_inflow_period = sum(t.amount for t in period_cash_income)
    cash_outflow_period = sum(t.amount for t in period_cash_expense)
    cash_net_period = cash_inflow_period - cash_outflow_period
    
    # Chart Data (Last 6 months) - EXCLUDE Void and Draft
    chart_labels = []
    revenue_data = []
    purchase_data = []
    expense_data = []
    profit_trend = []
    profit_labels = []
    
    for i in range(5, -1, -1):
        month_date = today_ist().replace(day=1) - timedelta(days=30 * i)
        month_start = month_date.replace(day=1)
        if month_date.month == 12:
            month_end = month_date.replace(day=31)
        else:
            month_end = month_date.replace(month=month_date.month + 1, day=1) - timedelta(days=1)
        
        month_label = month_date.strftime('%b %Y')
        chart_labels.append(month_label)
        
        # Monthly Revenue - EXCLUDE Void and Draft
        month_revenue = sum(
            float(inv.grand_total or 0) for inv in cdb.query(Invoice).filter(
                Invoice.company_id == company_id,
                Invoice.date >= month_start,
                Invoice.date <= month_end,
                Invoice.status.notin_(['Void', 'Draft'])  # ← EXCLUDE Void and Draft
            ).all()
        )
        revenue_data.append(month_revenue / 100000)
        
        # Monthly Purchases - EXCLUDE Void and Draft
        month_purchases = sum(
            float(pur.grand_total or 0) for pur in cdb.query(PurchaseInvoice).filter(
                PurchaseInvoice.company_id == company_id,
                PurchaseInvoice.date >= month_start,
                PurchaseInvoice.date <= month_end,
                PurchaseInvoice.status.notin_(['Void', 'Draft'])  # ← EXCLUDE Void and Draft
            ).all()
        )
        purchase_data.append(month_purchases / 100000)
        
        # Monthly Expenses
        month_expenses = sum(
            float(exp.amount or 0) for exp in cdb.query(Expense).filter(
                Expense.company_id == company_id,
                Expense.date >= month_start,
                Expense.date <= month_end
            ).all()
        )
        expense_data.append(month_expenses / 100000)
        
        month_gross_profit = month_revenue - month_purchases
        month_net_profit = month_gross_profit - month_expenses
        profit_trend.append(month_net_profit / 1000)
        profit_labels.append(month_label)
    
    # Status counts for shipments - EXCLUDE Void and Draft
    all_customer_invoices = cdb.query(Invoice).filter(
        Invoice.company_id == company_id,
        Invoice.invoice_id.like("CUST-%"),
        Invoice.status.notin_(['Void', 'Draft'])  # ← EXCLUDE Void and Draft
    ).all()
    
    status_counts = {
        "delivered": sum(1 for i in all_customer_invoices if i.status == "Paid"),
        "in_transit": sum(1 for i in all_customer_invoices if i.status == "Partial"),
        "pending": sum(1 for i in all_customer_invoices if i.status not in ["Paid", "Partial"]),
        "draft": 0,  # Excluded
        "total": len(all_customer_invoices)
    }
    
    # Payment methods breakdown
    cash_txns = cdb.query(CashTransaction).filter_by(company_id=company_id, type='income').all()
    bank_txns = cdb.query(BankTransaction).filter_by(company_id=company_id, type='credit').all()
    
    payment_methods = {
        "Cash": sum(t.amount for t in cash_txns),
        "Online/UPI": sum(t.amount for t in bank_txns if t.transaction_mode == "Online"),
        "Cheque": sum(t.amount for t in bank_txns if t.transaction_mode == "Cheque"),
    }
    
    # Top clients - EXCLUDE Void and Draft
    clients = cdb.query(Client).filter_by(company_id=company_id).all()
    top_clients_data = []
    for client in clients[:10]:
        client_invoices = cdb.query(Invoice).filter(
            Invoice.company_id == company_id,
            Invoice.client_id == client.id,
            Invoice.status.notin_(['Void', 'Draft'])  # ← EXCLUDE Void and Draft
        ).all()
        client_shipments = [i for i in client_invoices if i.invoice_id.startswith("CUST-")]
        total_billed = sum(float(inv.grand_total or 0) for inv in client_invoices)
        pending = sum(float(getattr(inv, 'balance', 0) or 0) for inv in client_invoices)
        top_clients_data.append({
            "name": client.name,
            "total_billed": total_billed,
            "pending": pending,
            "shipment_count": len(client_shipments)
        })
    top_clients_data.sort(key=lambda x: x["total_billed"], reverse=True)
    top_clients_data = top_clients_data[:5]
    
    # Recent shipments - EXCLUDE Void and Draft
    recent_shipments = []
    for inv in all_customer_invoices[:10]:
        meta = {}
        if inv.terms:
            try:
                meta = json.loads(inv.terms)
            except:
                pass
        status_label = "Delivered" if inv.status == "Paid" else "In Transit" if inv.status == "Partial" else "Pending"
        status_class = "delivered" if inv.status == "Paid" else "transit" if inv.status == "Partial" else "pending"
        recent_shipments.append({
            "docket_no": meta.get("docket_no", inv.invoice_id),
            "customer_name": inv.client_obj.name if inv.client_obj else (inv.contact_person or "—"),
            "destination": meta.get("destination", ""),
            "total": float(inv.grand_total or 0),
            "status": inv.status,
            "status_label": status_label,
            "status_class": status_class
        })
    
    # Recent payments
    recent_payments = []
    for txn in cash_txns[:10]:
        recent_payments.append({
            "date": txn.date.strftime("%d %b %Y"),
            "customer": txn.description[:30],
            "invoice_id": txn.reference or "—",
            "amount": txn.amount,
            "mode": "Cash"
        })
    for txn in bank_txns[:5]:
        recent_payments.append({
            "date": txn.date.strftime("%d %b %Y"),
            "customer": txn.description[:30],
            "invoice_id": txn.reference or "—",
            "amount": txn.amount,
            "mode": txn.transaction_mode or "Bank"
        })
    recent_payments.sort(key=lambda x: x['date'], reverse=True)
    recent_payments = recent_payments[:10]
    
    # Pending invoices - EXCLUDE Void and Draft
    pending_invoices = []
    for inv in all_active_invoices:
        balance = float(getattr(inv, 'balance', 0) or 0)
        if balance > 0:
            pending_invoices.append({
                "invoice_id": inv.invoice_id,
                "customer": inv.client_obj.name if inv.client_obj else (inv.contact_person or "—"),
                "date": inv.date.strftime("%d %b %Y") if inv.date else "—",
                "due_date": inv.due_date.strftime("%d %b %Y") if inv.due_date else "—",
                "balance": balance
            })
    pending_invoices = pending_invoices[:10]

    # Recent purchase invoices
    recent_purchases_raw = cdb.query(PurchaseInvoice).filter(
        PurchaseInvoice.company_id == company_id,
        PurchaseInvoice.status.notin_(['Void', 'Draft'])  # ← EXCLUDE Void and Draft
    ).order_by(PurchaseInvoice.date.desc()).limit(10).all()

    recent_purchases_data = []
    for p in recent_purchases_raw:
        try:
            supplier_name = p.supplier.name if p.supplier else (getattr(p, 'supplier_name', None) or "—")
        except Exception:
            supplier_name = getattr(p, 'supplier_name', None) or "—"
        recent_purchases_data.append({
            "id": p.invoice_id,
            "supplier": supplier_name,
            "date": p.date.strftime("%d %b %Y") if p.date else "—",
            "total": float(p.grand_total or 0),
            "status": p.status or "Unpaid"
        })

    expense_categories = {}
    for exp in period_expenses:
        expense_categories[exp.category] = expense_categories.get(exp.category, 0) + exp.amount

    kpi = {
        "cash_balance": cash_balance,
        "bank_balance": bank_balance,
        "total_revenue": total_revenue,
        "total_purchases": total_purchases,
        "total_expenses": total_expenses,
        "gross_profit": gross_profit,
        "net_profit": net_profit,
        "pending_amount": pending_amount,
        "cash_inflow_period": cash_inflow_period,
        "cash_outflow_period": cash_outflow_period,
        "cash_net_period": cash_net_period,
    }

    return render_template("dashboard.html",
                         company=company,
                         kpi=kpi,
                         from_date=from_date.strftime('%Y-%m-%d'),
                         to_date=to_date.strftime('%Y-%m-%d'),
                         chart_labels=chart_labels,
                         revenue_data=revenue_data,
                         purchase_data=purchase_data,
                         expense_data=expense_data,
                         profit_labels=profit_labels,
                         profit_trend=profit_trend,
                         cash_inflow_period=cash_inflow_period,
                         cash_outflow_period=cash_outflow_period,
                         cash_net_period=cash_net_period,
                         top_clients_data=top_clients_data,
                         status_counts=status_counts,
                         payment_methods=payment_methods,
                         recent_shipments=recent_shipments,
                         recent_payments=recent_payments,
                         pending_invoices=pending_invoices,
                         recent_purchases_data=recent_purchases_data,
                         total_shipments=status_counts["total"])


@app.route("/api/dashboard-data")
@login_required
@require_permission("dashboard", "view")
def api_dashboard_data():
    """API endpoint for dashboard data with date filters"""
    cdb = get_cdb()
    company_id = get_current_company()
    
    from_date_str = request.args.get('from_date', '')
    to_date_str = request.args.get('to_date', '')
    
    if not from_date_str:
        from_date = date(2000, 1, 1)
    else:
        from_date = date.fromisoformat(from_date_str)
    
    if not to_date_str:
        to_date = today_ist()
    else:
        to_date = date.fromisoformat(to_date_str)
    
    # Cash in Hand (all time, not filtered by date)
    cash_transactions = cdb.query(CashTransaction).filter_by(company_id=company_id).all()
    cash_balance = sum(t.amount for t in cash_transactions if t.type == 'income') - \
                   sum(t.amount for t in cash_transactions if t.type == 'expense')
    
    # Bank Balance (all time, not filtered by date)
    bank_accounts = cdb.query(BankAccount).filter_by(company_id=company_id, status='Active').all()
    bank_balance = sum(acc.balance for acc in bank_accounts)
    
    # ── FILTERED: Exclude Void and Draft invoices ──────────────────────────────
    # Filtered Sales Invoices (exclude Void and Draft)
    sales_invoices = cdb.query(Invoice).filter(
        Invoice.company_id == company_id,
        Invoice.date >= from_date,
        Invoice.date <= to_date,
        Invoice.status.notin_(['Void', 'Draft'])  # ← EXCLUDE Void and Draft
    ).all()
    total_revenue = sum(inv.grand_total or 0 for inv in sales_invoices)
    
    # Filtered Purchase Invoices
    purchase_invoices = cdb.query(PurchaseInvoice).filter(
        PurchaseInvoice.company_id == company_id,
        PurchaseInvoice.date >= from_date,
        PurchaseInvoice.date <= to_date,
        PurchaseInvoice.status.notin_(['Void', 'Draft'])  # ← EXCLUDE Void and Draft
    ).all()
    total_purchases = sum(pur.grand_total or 0 for pur in purchase_invoices)
    
    # ── Expenses for period ─────────────────────────────────────────────────────
    period_expenses = cdb.query(Expense).filter(
        Expense.company_id == company_id,
        Expense.date >= from_date,
        Expense.date <= to_date
    ).all()
    total_expenses = sum(exp.amount or 0 for exp in period_expenses)
    
    gross_profit = total_revenue - total_purchases
    net_profit = gross_profit - total_expenses
    
    # ── Pending Amount: Only from non-Void, non-Draft invoices ────────────────
    all_active_invoices = cdb.query(Invoice).filter(
        Invoice.company_id == company_id,
        Invoice.status.notin_(['Void', 'Draft'])  # ← EXCLUDE Void and Draft
    ).all()
    pending_amount = sum(getattr(inv, 'balance', 0) or 0 for inv in all_active_invoices)
    
    # Cash flow for period (exclude Void/Draft invoices from payment calculations)
    period_cash_income = cdb.query(CashTransaction).filter(
        CashTransaction.company_id == company_id,
        CashTransaction.type == 'income',
        CashTransaction.date >= from_date,
        CashTransaction.date <= to_date
    ).all()
    period_cash_expense = cdb.query(CashTransaction).filter(
        CashTransaction.company_id == company_id,
        CashTransaction.type == 'expense',
        CashTransaction.date >= from_date,
        CashTransaction.date <= to_date
    ).all()
    cash_inflow_period = sum(t.amount for t in period_cash_income)
    cash_outflow_period = sum(t.amount for t in period_cash_expense)
    
    # Chart data (last 6 months) - EXCLUDE Void and Draft
    chart_labels = []
    revenue_data = []
    purchase_data = []
    expense_data = []
    profit_trend = []
    profit_labels = []
    
    for i in range(5, -1, -1):
        month_date = today_ist().replace(day=1) - timedelta(days=30 * i)
        month_start = month_date.replace(day=1)
        if month_date.month == 12:
            month_end = month_date.replace(day=31)
        else:
            month_end = month_date.replace(month=month_date.month + 1, day=1) - timedelta(days=1)
        
        month_label = month_date.strftime('%b %Y')
        chart_labels.append(month_label)
        
        # Monthly Revenue - EXCLUDE Void and Draft
        month_revenue = sum(
            inv.grand_total or 0 for inv in cdb.query(Invoice).filter(
                Invoice.company_id == company_id,
                Invoice.date >= month_start,
                Invoice.date <= month_end,
                Invoice.status.notin_(['Void', 'Draft'])  # ← EXCLUDE Void and Draft
            ).all()
        )
        revenue_data.append(month_revenue / 100000)
        
        # Monthly Purchases - EXCLUDE Void and Draft
        month_purchases = sum(
            pur.grand_total or 0 for pur in cdb.query(PurchaseInvoice).filter(
                PurchaseInvoice.company_id == company_id,
                PurchaseInvoice.date >= month_start,
                PurchaseInvoice.date <= month_end,
                PurchaseInvoice.status.notin_(['Void', 'Draft'])  # ← EXCLUDE Void and Draft
            ).all()
        )
        purchase_data.append(month_purchases / 100000)
        
        # Monthly Expenses
        month_expenses = sum(
            exp.amount or 0 for exp in cdb.query(Expense).filter(
                Expense.company_id == company_id,
                Expense.date >= month_start,
                Expense.date <= month_end
            ).all()
        )
        expense_data.append(month_expenses / 100000)
        
        if i <= 5:
            profit_labels.append(month_label)
            month_net_profit = (month_revenue - month_purchases) - month_expenses
            profit_trend.append(month_net_profit / 1000)
    
    # Top clients (exclude Void and Draft invoices)
    clients = cdb.query(Client).filter_by(company_id=company_id).all()
    top_clients_data = []
    for client in clients[:10]:
        client_invoices = cdb.query(Invoice).filter(
            Invoice.company_id == company_id,
            Invoice.client_id == client.id,
            Invoice.status.notin_(['Void', 'Draft'])  # ← EXCLUDE Void and Draft
        ).all()
        total_billed = sum(inv.grand_total or 0 for inv in client_invoices)
        pending = sum(getattr(inv, 'balance', 0) or 0 for inv in client_invoices)
        top_clients_data.append({
            "name": client.name,
            "total_billed": total_billed,
            "pending": pending,
            "status": client.status or "Active"
        })
    top_clients_data.sort(key=lambda x: x["total_billed"], reverse=True)
    top_clients_data = top_clients_data[:5]
    
    # Recent invoices (exclude Void and Draft)
    recent_invoices_raw = cdb.query(Invoice).filter(
        Invoice.company_id == company_id,
        Invoice.status.notin_(['Void', 'Draft'])  # ← EXCLUDE Void and Draft
    ).order_by(Invoice.date.desc()).limit(10).all()
    recent_invoices_data = []
    for inv in recent_invoices_raw:
        client_name = inv.client_obj.name if inv.client_obj else (inv.contact_person or "—")
        total = inv.grand_total or 0
        balance = getattr(inv, 'balance', 0) or 0
        if balance <= 0:
            status = "Paid"
        elif inv.status == "Partial":
            status = "Partial"
        else:
            status = "Pending"
        recent_invoices_data.append({
            "id": inv.invoice_id,
            "customer": client_name,
            "date": inv.date.strftime("%d %b %Y") if inv.date else "—",
            "total": total,
            "status": status
        })
    
    # Low stock
    stock_items = cdb.query(StockItem).filter_by(company_id=company_id).all()
    low_stock_items = []
    for item in stock_items:
        reorder = item.reorder_level or 10
        if item.quantity <= reorder and item.quantity > 0:
            low_stock_items.append({
                "code": item.code,
                "name": item.name,
                "quantity": item.quantity,
                "reorder_level": reorder
            })
    low_stock_items = low_stock_items[:8]
    
    return jsonify({
        "kpi": {
            "cash_balance": cash_balance,
            "bank_balance": bank_balance,
            "total_revenue": total_revenue,
            "total_purchases": total_purchases,
            "total_expenses": total_expenses,
            "gross_profit": gross_profit,
            "net_profit": net_profit,
            "pending_amount": pending_amount,
            "cash_inflow_period": cash_inflow_period,
            "cash_outflow_period": cash_outflow_period,
            "cash_net_period": cash_inflow_period - cash_outflow_period,
        },
        "chart_labels": chart_labels,
        "revenue_data": revenue_data,
        "purchase_data": purchase_data,
        "expense_data": expense_data,
        "profit_labels": profit_labels,
        "profit_trend": profit_trend,
        "top_clients": top_clients_data,
        "recent_invoices": recent_invoices_data,
        "low_stock": low_stock_items,
    })


# ── Price List Routes ─────────────────────────────────────────────────────────

@app.route("/price-lists")
@login_required
@require_permission("pricelist", "view")
def price_lists():
    """Manage price lists"""
    cdb = get_cdb()
    company_id = get_current_company()
    price_lists = cdb.query(PriceList).filter_by(company_id=company_id, is_active=True).all()
    return render_template("price_lists.html", price_lists=price_lists, active='price_lists')

@app.route("/price-lists/view/<int:price_list_id>")
@login_required
@require_permission("pricelist", "view")
def view_price_list(price_list_id):
    """Preview a single price list's parsed rate data and upload date."""
    cdb = get_cdb()
    company_id = get_current_company()
    price_list = cdb.query(PriceList).filter_by(
        id=price_list_id, company_id=company_id
    ).first()
    if not price_list:
        abort(404)

    rate_data = json.loads(price_list.rate_data) if price_list.rate_data else {}
    countries = rate_data.get("countries", {})

    return render_template("view_price_list.html",
                           price_list=price_list,
                           rate_data=rate_data,
                           countries=countries,
                           active='price_lists')


@app.route("/debug/price-lists-data")
@login_required
def debug_price_lists_data():
    """Debug endpoint to check price list data in database"""
    cdb = get_cdb()
    company_id = get_current_company()
    
    price_lists = cdb.query(PriceList).filter_by(company_id=company_id).all()
    
    result = []
    for pl in price_lists:
        try:
            data = json.loads(pl.rate_data) if pl.rate_data else {}
            result.append({
                'id': pl.id,
                'courier': pl.courier,
                'filename': pl.filename,
                'is_active': pl.is_active,
                'countries': list(data.get('countries', {}).keys())[:5] if data.get('countries') else [],
                'weights': data.get('weights', []),
                'has_data': bool(data.get('countries'))
            })
        except Exception as e:
            result.append({
                'id': pl.id,
                'courier': pl.courier,
                'filename': pl.filename,
                'is_active': pl.is_active,
                'error': str(e)
            })
    
    return jsonify({
        'total': len(price_lists),
        'lists': result
    })

@app.route("/price-lists/delete/<int:price_list_id>", methods=["POST"])
@login_required
@owner_required
def delete_price_list(price_list_id):
    """Hard delete a price list"""
    cdb = get_cdb()
    company_id = get_current_company()
    
    if not cdb:
        flash("Could not connect to company database", "error")
        return redirect(url_for("price_lists"))
    
    try:
        # Use a direct DELETE with filter
        result = cdb.query(PriceList).filter_by(
            id=price_list_id, 
            company_id=company_id
        ).delete(synchronize_session='fetch')
        
        cdb.commit()
        
        if result > 0:
            flash(f"Price list deleted successfully!", "success")
        else:
            flash("Price list not found", "error")
            
    except Exception as e:
        cdb.rollback()
        import traceback
        traceback.print_exc()
        flash(f"Error deleting price list: {str(e)}", "error")
    
    return redirect(url_for("price_lists"))

@app.route("/debug/price-list/<int:price_list_id>")
@login_required
@owner_required
def debug_price_list(price_list_id):
    """Debug endpoint to check if a price list really exists"""
    cdb = get_cdb()
    company_id = get_current_company()
    
    price_list = cdb.query(PriceList).filter_by(
        id=price_list_id, 
        company_id=company_id
    ).first()
    
    if price_list:
        return jsonify({
            'exists': True,
            'id': price_list.id,
            'courier': price_list.courier,
            'filename': price_list.filename,
            'is_active': price_list.is_active,
            'file_path': price_list.file_path,
            'list_type': price_list.list_type
        })
    else:
        return jsonify({'exists': False})

@app.route("/debug/excel-columns", methods=["POST"])
@login_required
def debug_excel_columns():
    """Debug endpoint to check Excel file columns"""
    if 'price_file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    
    file = request.files['price_file']
    try:
        df = pd.read_excel(file, engine='openpyxl')
        columns = df.columns.tolist()
        first_row = df.iloc[0].to_dict() if len(df) > 0 else {}
        
        return jsonify({
            'columns': columns,
            'first_row': {str(k): str(v) for k, v in first_row.items()},
            'row_count': len(df)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route("/price-lists/upload", methods=["GET", "POST"])
@login_required
@require_permission("pricelist", "view", method_actions={'POST': 'create'})
def upload_price_list():
    """Upload a price list Excel file"""
    cdb = get_cdb()
    company_id = get_current_company()
    
    if request.method == "POST":
        print("=" * 80)
        print("UPLOAD ROUTE EXECUTED")
        print("=" * 80)
        if 'price_file' not in request.files:
            flash("No file uploaded", "error")
            return redirect(url_for("price_lists"))
        
        file = request.files['price_file']
        courier = request.form.get('courier', '').strip().upper()
        print("File object:", file)
        print("Filename:", repr(file.filename))
        print("Courier:", repr(courier))
        print("Allowed:", allowed_file(file.filename) if file.filename else False)
        
        if not courier:
            flash("Courier name is required", "error")
            return redirect(url_for("price_lists"))
        
        print("Entering upload block...")

        if file is None:
            print("File is None")

        elif file.filename == "":
            print("Filename is empty")

        elif not allowed_file(file.filename):
            print("Extension not allowed:", file.filename)

        else:
            print("Everything OK")
            try:
                # Save the file first
                filename = secure_filename(f"{courier}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}")
                filepath = os.path.join('uploads/price_lists', filename)
                os.makedirs('uploads/price_lists', exist_ok=True)
                file.save(filepath)
                
                print(f"📊 File: {file.filename}")

                # Parse rates based on format. parse_price_list() now finds the
                # real header row itself (rate cards have several blank/title
                # rows before the header, which used to break pd.read_excel()
                # with no header offset and silently produce zero countries).
                rate_data = parse_price_list(filepath, courier)
                
                # Check if we parsed any data
                if not rate_data['countries']:
                    flash(f"No countries parsed from {file.filename}. Detected format: {rate_data.get('format', 'unknown')}. Check that the file has a COUNTRY column (DPD-style) or a WEIGHT/KG column (FEDEX/DHL-style).", "error")
                    return redirect(url_for("price_lists"))
                
                # Deactivate old price lists for this courier — matched on the
                # normalized key so re-uploading under a differently-spelled
                # courier name ("Bluedart" vs "BLUE_DART") still retires the old one.
                upload_list_type = request.form.get('list_type', 'sales')
                target_key = normalize_courier(courier)
                old_lists = cdb.query(PriceList).filter_by(
                    company_id=company_id, list_type=upload_list_type, is_active=True
                ).all()
                for old in old_lists:
                    if normalize_courier(old.courier) == target_key:
                        old.is_active = False
                
                # Save new price list
                price_list = PriceList(
                    company_id=company_id,
                    courier=courier,
                    filename=file.filename,
                    file_path=filepath,
                    rate_data=json.dumps(rate_data),
                    is_active=True,
                    list_type   = request.form.get('list_type', 'sales'),
                    uploaded_by=get_current_user().get('email')
                )
                cdb.add(price_list)
                cdb.commit()
                
                flash(f"✅ Price list for {courier} uploaded! {len(rate_data['countries'])} countries, {len(rate_data['weights'])} weight tiers.", "success")
                
            except Exception as e:
                cdb.rollback()
                flash(f"Error processing file: {str(e)}", "error")
                print(f"Upload error: {e}")
                import traceback
                traceback.print_exc()
        
        return redirect(url_for("price_lists"))
    
    return render_template("upload_price_list.html", active='price_lists')


@app.route("/api/rate-lookup")
@login_required
@require_permission("pricelist", "view")
def api_rate_lookup():
    """API endpoint to lookup shipping rate"""
    cdb = get_cdb()
    company_id = get_current_company()
    
    courier = request.args.get('courier', '').strip().upper()
    destination = request.args.get('destination', '').strip().upper()
    weight = float(request.args.get('weight', 0))
    
    print("=" * 60)
    print(f"🔍 RATE LOOKUP REQUEST:")
    print(f"   Courier: {courier}")
    print(f"   Destination: {destination}")
    print(f"   Weight: {weight}")
    print("=" * 60)
    
    if not courier or not destination or weight <= 0:
        return jsonify({'error': 'Missing parameters'}), 400
    
    # Get active price list for this courier - SALES only
    price_list = find_price_list(cdb, company_id, courier, 'sales')
    
    print(f"📋 Price list found: {price_list is not None}")
    if price_list:
        print(f"   List type: {price_list.list_type}")
    
    if not price_list:
        # Check if there's a purchase list as fallback (for debugging)
        purchase_list = find_price_list(cdb, company_id, courier, 'purchase')
        if purchase_list:
            print(f"⚠️ Found PURCHASE list for {courier}, but sales lookup only uses SALES lists")
        return jsonify({'error': f'No active sales price list found for {courier}'}), 404
    
    try:
        rate_data = json.loads(price_list.rate_data)
        print(f"📊 Rate data loaded: {len(rate_data.get('countries', {}))} countries")
        
        countries = rate_data.get('countries', {})
        weights = sorted(rate_data.get('weights', []))
        
        print(f"📍 Available countries: {list(countries.keys())[:5]}...")
        print(f"⚖️ Available weights: {weights}")
        
        # Find matching country
        matched_country = None
        matched_rates = None
        
        # 1. Try exact match
        if destination in countries:
            matched_country = destination
            matched_rates = countries[destination]
            print(f"✅ Exact match: {matched_country}")
        
        # 2. Try partial match (destination contains country or vice versa)
        if not matched_rates:
            for country, rates in countries.items():
                if destination in country or country in destination:
                    matched_country = country
                    matched_rates = rates
                    print(f"✅ Partial match: {matched_country}")
                    break
        
        # 3. Try word matching (split by spaces)
        if not matched_rates:
            dest_words = destination.split()
            for country, rates in countries.items():
                country_words = country.split()
                for dw in dest_words:
                    if len(dw) > 2:
                        for cw in country_words:
                            if dw in cw or cw in dw:
                                matched_country = country
                                matched_rates = rates
                                print(f"✅ Word match: {matched_country}")
                                break
                    if matched_rates:
                        break
                if matched_rates:
                    break
        
        if not matched_rates:
            print(f"❌ No match found for: {destination}")
            return jsonify({'error': f'No rate found for {destination}. Available countries: {", ".join(list(countries.keys())[:10])}'}), 404
        
        # Use calculate_rate() so band (per-kg) pricing above the tier table
        # is multiplied by weight instead of returned as a raw stored number.
        rate, weight_used, pricing_type = calculate_rate(rate_data, matched_country, weight)

        print(f"💰 Rate: {rate} for {weight_used}kg ({pricing_type})")

        if not rate or rate <= 0:
            return jsonify({'error': f'No rate found for {weight}kg in {matched_country}'}), 404
        
        # Log lookup
        try:
            lookup = RateLookup(
                company_id=company_id,
                courier=courier,
                destination=destination,
                weight=weight,
                rate=rate
            )
            cdb.add(lookup)
            cdb.commit()
        except Exception as e:
            cdb.rollback()
            print(f"⚠️ Could not log lookup: {e}")
        
        # Per-kg rate is what the invoice's Rate per kg field actually needs —
        # computed here (backend) rather than left for the frontend to divide,
        # since weight_used is the rounded slab weight the money was priced at,
        # not necessarily the raw weight passed in.
        rate_per_kg = round(rate / weight_used, 2) if weight_used else 0

        return jsonify({
            'success': True,
            'rate': rate,
            'rate_per_kg': rate_per_kg,
            'weight_used': weight_used,
            'pricing_type': pricing_type,
            'country_matched': matched_country,
            'courier': courier,
            'destination': destination,
            'list_type': 'sales'  # ← Add this for debugging
        })
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

def normalize_courier(name):
    """Canonical courier key for matching — strips spaces/underscores/punctuation
    and uppercases, so 'Blue Dart', 'Bluedart', and 'BLUE_DART' all collapse to
    the same key 'BLUEDART'. The courier name gets typed in at least three
    different places (booking Carrier dropdown, Supplier > Brands, price-list
    upload form) and there's no guarantee they're spelled identically — this is
    the single source of truth for "are these the same courier" everywhere in
    the app. Never compare courier strings with == directly; use this."""
    return re.sub(r'[^A-Z0-9]', '', (name or '').strip().upper())


def find_price_list(cdb, company_id, courier, list_type):
    """Look up an active PriceList by courier name, matching on the normalized
    key instead of an exact string, so upload-time spelling differences don't
    silently produce 'no price list found'. Small tables (price lists per
    company are never more than a handful), so filtering in Python after a
    narrow SQL query is simpler and safer here than a DB-side normalization
    expression that would need to work identically on SQLite and MySQL."""
    target = normalize_courier(courier)
    if not target:
        return None
    candidates = cdb.query(PriceList).filter_by(
        company_id=company_id, is_active=True, list_type=list_type
    ).all()
    for pl in candidates:
        if normalize_courier(pl.courier) == target:
            return pl
    return None


def _auto_fetch_purchase_rate(cdb, company_id, courier, destination, weight):
    """
    Server-side purchase rate lookup, used by the booking auto-generation
    hook (invoice_customer_save). Mirrors the matching logic in
    /api/purchase-rate-lookup below, but returns a dict/None directly instead
    of a Flask response, since this runs mid-request rather than over HTTP.
    """
    destination = (destination or "").strip().upper()

    if not courier:
        return {'ok': False, 'reason': "no carrier / courier company was set on this booking"}
    if not weight or weight <= 0:
        return {'ok': False, 'reason': "chargeable weight is 0 — no Freight Weight or package weight was entered"}
    if not destination:
        return {'ok': False, 'reason': "no destination was entered on this booking"}

    price_list = find_price_list(cdb, company_id, courier, 'purchase')
    if not price_list:
        return {'ok': False, 'reason': f"no active purchase price list found for '{courier}'"}

    try:
        rate_data = json.loads(price_list.rate_data)
        countries = rate_data.get('countries', {})

        matched_country = None
        matched_rates = None

        if destination in countries:
            matched_country = destination
            matched_rates = countries[destination]

        if not matched_rates:
            for country, rates in countries.items():
                if destination in country or country in destination:
                    matched_country = country
                    matched_rates = rates
                    break

        if not matched_rates:
            dest_words = destination.split()
            for country, rates in countries.items():
                country_words = country.split()
                for dw in dest_words:
                    if len(dw) > 2:
                        for cw in country_words:
                            if dw in cw or cw in dw:
                                matched_country = country
                                matched_rates = rates
                                break
                    if matched_rates:
                        break
                if matched_rates:
                    break

        if not matched_rates:
            sample = ', '.join(list(countries.keys())[:8])
            more = '…' if len(countries) > 8 else ''
            return {'ok': False, 'reason': (
                f"purchase price list for '{courier}' has no rate for destination "
                f"'{destination}' — it only covers {len(countries)} countries: {sample}{more}"
            )}

        rate, weight_used, pricing_type = calculate_rate(rate_data, matched_country, weight)
        if not rate or rate <= 0:
            return {'ok': False, 'reason': (
                f"matched destination '{matched_country}' in the purchase price list "
                f"but no rate found for {weight}kg"
            )}

        return {
            'ok': True,
            'rate': rate,
            'weight_used': weight_used,
            'pricing_type': pricing_type,
            'country_matched': matched_country,
        }
    except Exception as e:
        print(f"[auto-purchase-rate] lookup failed for {courier}/{destination}: {e}")
        return {'ok': False, 'reason': f"price list lookup error: {e}"}


def _sync_auto_purchase_invoice_line(cdb, company_id, form, packages_data,
                                      freight_weight, apply_gst, gst_calc,
                                      invoice_date, docket_no, invoice_id,
                                      inv_pk, action):
    """
    Create (or, on re-edit, update in place) the auto-generated purchase
    invoice line for a booking. Shared by invoice_customer_save() (new
    bookings) and invoice_customer_update() (edits) — previously this logic
    only lived in invoice_customer_save(), so any booking whose *first*
    successful save happened to land on the update route (stale
    edit_invoice_id, resubmit, etc.) silently never got a purchase line and
    there was no error/flash to show for it. Bookings saved before this fix
    was deployed can be repaired by simply re-saving them (Edit -> Save) --
    this function will detect the missing line and create it.

    Upsert key is PurchaseInvoiceItem.source_invoice_id == inv_pk, so
    re-saving an already-linked booking updates its existing line instead
    of creating a duplicate.

    Wrapped in try/except on purpose: a failure here must never break the
    invoice save that already happened (and, for invoice_customer_save(),
    already committed) before this runs. Every failure is flashed to the
    user AND printed to the server log so it's never silent again.
    """
    if action == "draft":
        return
    courier_company_id = (form.get("courier_company_id", "") or "").strip()
    carrier_name = (form.get("carrier", "") or "").strip()
    if not (courier_company_id and carrier_name):
        return

    try:
        supplier_for_pi = cdb.query(Supplier).filter_by(
            id=int(courier_company_id), company_id=company_id
        ).first() if courier_company_id.isdigit() else None

        if not supplier_for_pi:
            flash(
                f"Note: courier company on {docket_no or invoice_id} did not match a "
                f"known Supplier record — no purchase line was generated. Fix the "
                f"Courier Company selection and save again.",
                "warning",
            )
            print(f"[purchase-auto-gen] no matching Supplier for courier_company_id="
                  f"{courier_company_id!r} company_id={company_id!r} on {invoice_id}")
            return

        # Purchase weight follows the discounted weight (actual − Disc. Wt) per
        # package, same as booking.html's "Discounted weight" column — a weight
        # discount entered on the booking must reduce what we're billed here.
        actual_weight_pi = sum(
            max((p.get("weight") or 0) - (p.get("discount_wt") or 0), 0) * (p.get("qty") or 1)
            for p in packages_data
        )
        chg_weight = actual_weight_pi if actual_weight_pi > 0 else (freight_weight or 0.0)
        # Rate-card lookups always run on the rounded slab weight (1.75kg -> 2kg,
        # 10.1kg -> 11kg) — chg_weight itself stays the actual weight for display
        # and for weight_kg below.
        rating_weight_pi = round_billable_weight(chg_weight)
        purchase_rate = 0.0
        taxable_pi = 0.0
        rate_result = _auto_fetch_purchase_rate(
            cdb, company_id, carrier_name, form.get("destination", ""), rating_weight_pi
        )
        if rate_result and rate_result.get('ok'):
            # Bill the FULL rate the price list returns for the matched slab.
            # Do NOT re-derive a per-kg rate (rate / weight_used) and multiply
            # it back by the actual weight — that silently discounts the bill
            # any time actual weight is below the slab it was rounded up to
            # (e.g. 1.75kg billed at 500 for the 2kg slab was coming out as
            # 500/2*1.75 = 437.50 before this fix).
            taxable_pi = round(rate_result["rate"], 2)
            purchase_rate = round(taxable_pi / chg_weight, 4) if chg_weight else 0.0
        gst_pct_pi = 18.0 if apply_gst else 0.0
        gst_amt_pi = round(taxable_pi * gst_pct_pi / 100, 2) if apply_gst else 0.0
        if apply_gst and gst_calc.get("is_interstate"):
            cgst_pi, sgst_pi, igst_pi = 0.0, 0.0, gst_amt_pi
        else:
            cgst_pi = round(gst_amt_pi / 2, 2)
            sgst_pi = gst_amt_pi - cgst_pi
            igst_pi = 0.0
        line_total_pi = round(taxable_pi + gst_amt_pi, 2)
        carrier_ref_value = (form.get("carrier_ref") or "").strip()
        total_boxes_for_awb = sum((p.get("qty") or 1) for p in packages_data) or 1

        existing_item = cdb.query(PurchaseInvoiceItem).filter_by(
            source_invoice_id=inv_pk
        ).first()

        if existing_item:
            old_taxable = existing_item.taxable_value or 0.0
            old_line_total = existing_item.total_amount or 0.0
            today_pi = cdb.query(PurchaseInvoice).filter_by(
                id=existing_item.purchase_invoice_id
            ).first()

            existing_item.description   = docket_no or invoice_id
            existing_item.quantity      = total_boxes_for_awb
            existing_item.purchase_rate = purchase_rate
            existing_item.taxable_value = taxable_pi
            existing_item.gst_percent   = gst_pct_pi
            existing_item.cgst_amount   = cgst_pi
            existing_item.sgst_amount   = sgst_pi
            existing_item.igst_amount   = igst_pi
            existing_item.total_amount  = line_total_pi
            existing_item.docket_no     = docket_no or None
            existing_item.carrier_ref   = carrier_ref_value or None
            existing_item.party_name    = form.get("shipper_name", "") or None
            existing_item.consignee_name = form.get("receiver_name", "") or None
            existing_item.destination   = form.get("destination", "") or None
            existing_item.courier_name  = carrier_name
            existing_item.weight_kg     = chg_weight
            existing_item.rate_per_kg   = purchase_rate

            if today_pi:
                delta = line_total_pi - old_line_total
                today_pi.subtotal    = (today_pi.subtotal or 0) + (taxable_pi - old_taxable)
                today_pi.grand_total = (today_pi.grand_total or 0) + delta
                today_pi.balance     = (today_pi.balance or 0) + delta
                if supplier_for_pi:
                    supplier_for_pi.payable = (supplier_for_pi.payable or 0) + delta
            return

        today_pi = cdb.query(PurchaseInvoice).filter_by(
            company_id=company_id,
            supplier_id=supplier_for_pi.id,
            date=date.fromisoformat(invoice_date),
        ).first()

        if not today_pi:
            pi_id = _next_numbered_id(
                cdb, PurchaseInvoice.invoice_id,
                "PURCHASE-INV-" + datetime.now().strftime("%Y%m%d") + "-"
            )
            today_pi = PurchaseInvoice(
                invoice_id=pi_id,
                company_id=company_id,
                supplier_id=supplier_for_pi.id,
                invoice_number=None,
                date=date.fromisoformat(invoice_date),
                subtotal=0, tax_amount=0, grand_total=0,
                paid_amount=0, balance=0, status="Pending",
                created_at=datetime.utcnow(),
            )
            cdb.add(today_pi)
            cdb.flush()

        cdb.add(PurchaseInvoiceItem(
            purchase_invoice_id=today_pi.id,
            source_invoice_id=inv_pk,
            description=docket_no or invoice_id,
            quantity=total_boxes_for_awb,
            unit="pcs",
            purchase_rate=purchase_rate,
            taxable_value=taxable_pi,
            gst_percent=gst_pct_pi,
            cgst_amount=cgst_pi,
            sgst_amount=sgst_pi,
            igst_amount=igst_pi,
            total_amount=line_total_pi,
            docket_no=docket_no or None,
            carrier_ref=carrier_ref_value or None,
            party_name=form.get("shipper_name", "") or None,
            consignee_name=form.get("receiver_name", "") or None,
            destination=form.get("destination", "") or None,
            courier_name=carrier_name,
            weight_kg=chg_weight,
            rate_per_kg=purchase_rate,
        ))
        today_pi.subtotal    = (today_pi.subtotal or 0) + taxable_pi
        today_pi.tax_amount  = (today_pi.tax_amount or 0) + gst_amt_pi
        today_pi.grand_total = (today_pi.grand_total or 0) + line_total_pi
        today_pi.balance     = (today_pi.balance or 0) + line_total_pi
        supplier_for_pi.payable = (supplier_for_pi.payable or 0) + line_total_pi

        if not rate_result or not rate_result.get('ok'):
            reason = rate_result.get('reason') if rate_result else "unknown error"
            flash(
                f"Note: {reason} — the auto-generated purchase line for "
                f"{docket_no or invoice_id} has rate ₹0, fix it manually in Purchases.",
                "warning",
            )
    except Exception as e:
        cdb.rollback()
        print(f"[purchase-auto-gen] FAILED for invoice {invoice_id} (pk={inv_pk}): {e}")
        flash(
            f"Warning: could not auto-generate the purchase line for {docket_no or invoice_id} "
            f"({e}). The invoice itself saved fine — add the purchase entry manually.",
            "warning",
        )


def _repair_manifest_shipper_mismatches(cdb, company_id):
    """
    One-off data repair, NOT part of normal request flow.

    Fixes ManifestEntry rows that were orphaned under the wrong
    CompanyManifest because of the pre-fix shipper_changed bug:
    the booking's terms.shipper_name is correct, but the entry never
    got moved to a manifest matching that shipper. Re-derives the
    correct shipper from each entry's linked booking (matched by
    docket_no) and moves the entry + fixes total_boxes on both the
    old and new manifest. Safe to run more than once — no-ops once
    everything matches.
    """
    fixed = []
    entries = cdb.query(ManifestEntry).join(
        CompanyManifest, ManifestEntry.manifest_id == CompanyManifest.id
    ).filter(
        CompanyManifest.company_id == company_id,
        ManifestEntry.docket_no.isnot(None),
        ManifestEntry.docket_no != "",
    ).all()

    for entry in entries:
        inv = cdb.query(Invoice).filter_by(
            company_id=company_id
        ).filter(
            Invoice.terms.like(f'%"docket_no": "{entry.docket_no}"%')
        ).first()
        if not inv or not inv.terms:
            continue

        try:
            meta = json.loads(inv.terms)
        except (TypeError, ValueError):
            continue

        true_shipper = (meta.get("shipper_name") or "").strip()
        if not true_shipper:
            continue

        current_manifest = cdb.query(CompanyManifest).filter_by(id=entry.manifest_id).first()
        if not current_manifest or current_manifest.shipper_client_name == true_shipper:
            continue  # already correct

        shipper_client = cdb.query(Client).filter_by(
            company_id=company_id, name=true_shipper
        ).first()
        if not shipper_client:
            continue  # can't safely create a new client here — skip, handle manually

        # 1. Get or create a manifest for TODAY with the same shipper
        today = today_ist()
        target_manifest = cdb.query(CompanyManifest).filter_by(
            company_id=company_id,
            shipper_client_id=shipper_client_id,  # Get this from the entry
            date=today
        ).first()

        if not target_manifest:
            # Create new manifest for today
            last_mf = cdb.query(CompanyManifest).filter_by(company_id=company_id).order_by(CompanyManifest.id.desc()).first()
            target_manifest = CompanyManifest(
                manifest_id=f"MFT-{(last_mf.id + 1) if last_mf else 1:04d}",
                company_id=company_id,
                date=today,
                shipper_client_id=shipper_client_id,
                shipper_client_name=shipper_client_name,
                total_boxes=0,
                notes=f"Auto-created for dispatch on {today}"
            )
            cdb.add(target_manifest)
            cdb.flush()

        # 2. Move each generated entry to today's manifest
        for entry in selected_entries:
            entry.manifest_id = target_manifest.id
            entry.status = 'Generated'
            entry.generated_at = datetime.utcnow()
            entry.generated_by = user_email

        # 3. Update box counts
        target_manifest.total_boxes = cdb.query(ManifestEntry).filter_by(manifest_id=target_manifest.id).count()

        old_manifest_id = current_manifest.id
        entry.manifest_id = target_manifest.id
        fixed.append((entry.docket_no, current_manifest.shipper_client_name, true_shipper))

        old_remaining = cdb.query(ManifestEntry).filter_by(manifest_id=old_manifest_id).all()
        current_manifest.total_boxes = len(old_remaining)
        _recompute_manifest_status(current_manifest)

        new_entries = cdb.query(ManifestEntry).filter_by(manifest_id=target_manifest.id).all()
        target_manifest.total_boxes = len(new_entries)
        _recompute_manifest_status(target_manifest)

    cdb.commit()
    return fixed


@app.route("/admin/repair-manifest-shippers")
@login_required
def repair_manifest_shippers():
    cdb = get_cdb()
    company_id = get_current_company()
    fixed = _repair_manifest_shipper_mismatches(cdb, company_id)
    if fixed:
        flash(f"Repaired {len(fixed)} manifest entr{'y' if len(fixed)==1 else 'ies'}: " +
              "; ".join(f"{d} ({old} → {new})" for d, old, new in fixed), "success")
    else:
        flash("No mismatched manifest entries found.", "info")
    return redirect(url_for("manifest_list"))


CASH_CLIENT_ID = "CASH"


# Add this function near the top of app.py (around line 600, after _get_or_create_cash_client)

def _get_or_create_cash_client(cdb, company_id, shipper_name):
    """
    Get or create a cash client record for a walk-in customer.
    Creates ONE client per unique shipper_name so we can track repeat
    cash customers and their stock history.
    """
    if not shipper_name or not shipper_name.strip():
        # Fallback to generic cash client
        return _get_or_create_generic_cash_client(cdb, company_id)
    
    # Normalize and clean the name
    shipper_name = shipper_name.strip()
    
    # Check if this cash client already exists
    cash_client = cdb.query(Client).filter_by(
        company_id=company_id,
        name=shipper_name,
        client_type="Cash-Only"  # Special type to filter out of debtors
    ).first()
    
    if cash_client:
        return cash_client
    
    # Create a new cash client
    company_obj = Company.query.filter_by(company_id=company_id).first()
    client_prefix = _company_name_prefix(company_obj.company_name if company_obj else "", from_end=True)
    
    # Generate a unique client_id with 'CASH' prefix
    cash_client_id = _next_numbered_id(
        cdb, Client.client_id, 
        f"{client_prefix}CASH",  # e.g., "demCASH001"
        extra_filters=[Client.company_id == company_id]
    )
    
    cash_client = Client(
        client_id=cash_client_id,
        company_id=company_id,
        name=shipper_name,
        client_type="Cash-Only",  # This filters them out of debtors list
        status="Active",
        pending=0.0,  # Cash clients don't have pending balances
        opening_balance=0.0,
        credit_limit=0,
        created_at=today_ist(),
        notes=f"Cash/Walk-in customer - created on {today_ist().strftime('%d %b %Y')}"
    )
    cdb.add(cash_client)
    cdb.flush()
    
    return cash_client


def _get_or_create_generic_cash_client(cdb, company_id):
    """
    Fallback: returns the shared CASH placeholder client.
    Used when no shipper_name is provided.
    """
    cash_client = cdb.query(Client).filter_by(
        company_id=company_id, 
        client_type="Cash-Only",
        name="Cash / Walk-in"  # The generic one
    ).first()
    
    if not cash_client:
        cash_client = Client(
            client_id="CASH001",  # Simple fixed ID for the generic one
            company_id=company_id,
            name="Cash / Walk-in",
            client_type="Cash-Only",
            status="Active",
            pending=0.0,
            opening_balance=0.0,
            created_at=today_ist(),
            notes="Generic cash/walk-in customer (no name provided)"
        )
        cdb.add(cash_client)
        cdb.flush()
    
    return cash_client


def _sync_auto_manifest_entry(cdb, company_id, shipper_name, carrier_name, action,
                               invoice_date, docket_no, invoice_id, total_boxes,
                               primary_stock_id=None, primary_stock_name=None,
                               item_type="Box", old_docket_no=None, booking_type="credit"):
    """
    Create (or update in place, on re-save) the ManifestEntry/CompanyManifest
    for a booking.
    """
    if action == "draft":
        return
    shipper_name_mf = (shipper_name or "").strip()
    carrier_name = (carrier_name or "").strip()
    
    if not (shipper_name_mf and carrier_name):
        return

    try:
        # ── Get or create the shipper client ─────────────────────────────
        if booking_type == "cash":
            # For cash bookings: create a dedicated cash client per shipper_name
            if shipper_name_mf:
                shipper_mf = _get_or_create_cash_client(cdb, company_id, shipper_name_mf)
            else:
                shipper_mf = _get_or_create_generic_cash_client(cdb, company_id)
        else:
            # For credit bookings: use the existing client
            shipper_mf = cdb.query(Client).filter_by(
                company_id=company_id, name=shipper_name_mf
            ).first()
            if not shipper_mf:
                # If no client exists for a credit booking, create one
                company_obj_mf = Company.query.filter_by(company_id=company_id).first()
                mf_client_prefix = _company_name_prefix(company_obj_mf.company_name if company_obj_mf else "", from_end=True)
                mf_client_id = _next_numbered_id(cdb, Client.client_id, mf_client_prefix, extra_filters=[Client.company_id == company_id])
                shipper_mf = Client(
                    client_id=mf_client_id,
                    company_id=company_id,
                    name=shipper_name_mf,
                    client_type="Customer",  # Regular customer for credit
                    status="Active",
                    created_at=today_ist()
                )
                cdb.add(shipper_mf)
                cdb.flush()

        # ── Rest of the function continues as before ──────────────────────
        total_boxes_mf = int(total_boxes) or 1
        lookup_docket = old_docket_no or docket_no

        # Find existing manifest entries for this docket
        existing_rows = cdb.query(ManifestEntry).join(
            CompanyManifest, ManifestEntry.manifest_id == CompanyManifest.id
        ).filter(
            ManifestEntry.docket_no == lookup_docket,
            CompanyManifest.company_id == company_id,
        ).all() if lookup_docket else []

        # Get stock type from primary stock
        stock_type_mf = item_type or "Box"
        if primary_stock_id:
            stock_obj = cdb.query(StockItem).filter_by(id=primary_stock_id).first()
            if stock_obj:
                stock_type_mf = stock_obj.item_type or stock_obj.category or "Box"

        if existing_rows:
            # Get the parent manifest
            parent_manifest = cdb.query(CompanyManifest).filter_by(
                id=existing_rows[0].manifest_id
            ).first()

            # ── Shipper mismatch: this manifest may hold OTHER bookings'
            # dockets too (it's grouped by date, not by booking). Renaming
            # parent_manifest.shipper_client_name here would relabel every
            # other docket sharing this manifest. Instead, move ONLY this
            # docket's rows to the manifest matching the current shipper —
            # creating one for that date if it doesn't exist yet.
            if parent_manifest and shipper_mf and parent_manifest.shipper_client_id != shipper_mf.id:
                target_manifest = cdb.query(CompanyManifest).filter_by(
                    company_id=company_id,
                    shipper_client_id=shipper_mf.id,
                    date=parent_manifest.date,
                ).first()
                if not target_manifest:
                    last_mf = cdb.query(CompanyManifest).filter_by(company_id=company_id) \
                                  .order_by(CompanyManifest.id.desc()).first()
                    target_manifest = CompanyManifest(
                        manifest_id=f"MFT-{(last_mf.id + 1) if last_mf else 1:04d}",
                        company_id=company_id,
                        date=parent_manifest.date,
                        shipper_client_id=shipper_mf.id,
                        shipper_client_name=shipper_mf.name,
                        total_boxes=0,
                        notes=f"Auto-created from booking {invoice_id}",
                    )
                    cdb.add(target_manifest)
                    cdb.flush()

                old_manifest_id = parent_manifest.id
                for row in existing_rows:
                    row.manifest_id = target_manifest.id
                parent_manifest = target_manifest

                old_remaining = cdb.query(ManifestEntry).filter_by(manifest_id=old_manifest_id).all()
                stale_manifest = cdb.query(CompanyManifest).filter_by(id=old_manifest_id).first()
                if stale_manifest:
                    if old_remaining:
                        stale_manifest.total_boxes = len(old_remaining)
                        _recompute_manifest_status(stale_manifest)
                    else:
                        cdb.delete(stale_manifest)

            # UPDATE ALL ENTRIES (including Generated ones) with latest metadata
            for row in existing_rows:
                # Update core fields on ALL entries
                row.courier_name = carrier_name
                if docket_no and row.docket_no != docket_no:
                    row.docket_no = docket_no
                # Cash bookings have no dedicated per-shipper Client row (see
                # _get_or_create_cash_client), so the typed customer name only
                # lives here. This used to be set at creation and never
                # touched again — renaming the walk-in customer on a booking
                # edit silently left every existing manifest entry showing
                # the old name.
                if booking_type == "cash":
                    row.notes = shipper_name_mf
                # Update stock metadata on ALL entries
                if primary_stock_name:
                    row.stock_item_name = primary_stock_name
                if primary_stock_id:
                    row.stock_item_id = primary_stock_id
                    stock_obj = cdb.query(StockItem).filter_by(id=primary_stock_id).first()
                    if stock_obj:
                        row.item_type = stock_obj.item_type or stock_obj.category or item_type or "Box"
                elif not primary_stock_id:
                    # Keep existing item_type if no new stock is linked
                    row.item_type = row.item_type or item_type or "Box"

            # Count entries by status
            pending_rows = [r for r in existing_rows if r.status != 'Generated']
            generated_rows = [r for r in existing_rows if r.status == 'Generated']
            pending_count = len(pending_rows)
            generated_count = len(generated_rows)
            
            # Calculate how many Pending entries we need
            # Generated entries are locked - we can only add/remove Pending ones
            target_pending_count = max(0, total_boxes_mf - generated_count)
            delta = target_pending_count - pending_count

            if delta > 0:
                # Need to add more Pending entries
                for _ in range(delta):
                    cdb.add(ManifestEntry(
                        manifest_id=parent_manifest.id,
                        courier_name=carrier_name,
                        boxes=1,
                        docket_no=docket_no or None,
                        stock_item_id=primary_stock_id,
                        stock_item_name=primary_stock_name,
                        # Cash bookings all share the CASH client, so the
                        # real customer name has nowhere else to live on
                        # this entry — stamp it here instead of losing it.
                        notes=shipper_name_mf if booking_type == "cash" else None,
                        item_type=stock_type_mf,
                        status='Pending',
                    ))
            elif delta < 0:
                # Need to remove some Pending entries (remove from the end)
                to_remove = pending_rows[:min(-delta, len(pending_rows))]
                for row in to_remove:
                    cdb.delete(row)

            # Update manifest total boxes and status
            if parent_manifest:
                all_entries = cdb.query(ManifestEntry).filter_by(
                    manifest_id=parent_manifest.id
                ).all()
                parent_manifest.total_boxes = len(all_entries)
                _recompute_manifest_status(parent_manifest)
            return

        # No existing entries - create a new manifest
        today_manifest = cdb.query(CompanyManifest).filter_by(
            company_id=company_id,
            shipper_client_id=shipper_mf.id,
            date=date.fromisoformat(invoice_date),
        ).first()

        if not today_manifest:
            last_mf = cdb.query(CompanyManifest).filter_by(company_id=company_id) \
                          .order_by(CompanyManifest.id.desc()).first()
            next_num_mf = (last_mf.id + 1) if last_mf else 1
            today_manifest = CompanyManifest(
                manifest_id=f"MFT-{next_num_mf:04d}",
                company_id=company_id,
                date=date.fromisoformat(invoice_date),
                shipper_client_id=shipper_mf.id,
                shipper_client_name=shipper_mf.name,
                total_boxes=0,
                notes=f"Auto-created from booking {invoice_id}",
                created_by=session.get("user", {}).get("email", ""),
            )
            cdb.add(today_manifest)
            cdb.flush()

        for _ in range(total_boxes_mf):
            cdb.add(ManifestEntry(
                manifest_id=today_manifest.id,
                courier_name=carrier_name,
                boxes=1,
                docket_no=docket_no or None,
                stock_item_id=primary_stock_id,
                stock_item_name=primary_stock_name,
                # Same reasoning as the existing-manifest branch above.
                notes=shipper_name_mf if booking_type == "cash" else None,
                item_type=stock_type_mf,
                status='Pending',
            ))
        today_manifest.total_boxes = total_boxes_mf
        
    except Exception as e:
        cdb.rollback()
        print(f"[manifest-auto-gen] FAILED for invoice {invoice_id}: {e}")
        flash(
            f"Warning: could not sync manifest for {docket_no or invoice_id} "
            f"({e}). Check the server log.",
            "warning",
        )

@app.route("/company/permissions/fields/<role>", methods=["POST"])
@login_required
@owner_required
def save_field_permissions(role):
    if role not in ("employee", "accountant", "manager"):
        flash("Invalid role")
        return redirect(url_for("company_settings"))
    
    company_id = get_current_company()
    cdb = get_customer_session(company_id)
    
    row = cdb.query(CompanyRolePermission).filter_by(company_id=company_id, role=role).first()
    if not row:
        row = CompanyRolePermission(company_id=company_id, role=role)
        cdb.add(row)
    
    # Build field permissions from form
    field_perms = {}
    for field_key in INVOICE_FIELDS:
        field_perms[field_key] = {
            "view": request.form.get(f"field__{field_key}__view") == "on",
            "edit": request.form.get(f"field__{field_key}__edit") == "on"
        }
    
    if role in perms_module.HARD_LOCKED_EDIT:
        for locked_group in perms_module.HARD_LOCKED_EDIT[role]:
            if locked_group in field_perms:
                field_perms[locked_group]["edit"] = False

    row.field_permissions_json = json.dumps(field_perms)
    row.updated_at = datetime.utcnow()
    cdb.commit()
    
    flash(f"Field permissions for {role.title()} updated")
    return redirect(url_for("company_settings"))

@app.route("/company/permissions/fields/user/<user_id>", methods=["POST"])
@login_required
@owner_required
def save_user_field_permissions(user_id):
    company_id = get_current_company()
    cdb = get_customer_session(company_id)
    cu = cdb.query(CompanyUser).filter_by(user_id=user_id, company_id=company_id).first()
    if not cu:
        flash("User not found")
        return redirect(url_for("company_settings"))
    if cu.role in ("owner", "super_admin"):
        flash("Owner access can't be limited this way")
        return redirect(url_for("company_settings"))

    field_perms = {}
    for field_key in INVOICE_FIELDS:
        field_perms[field_key] = {
            "view": request.form.get(f"field__{field_key}__view") == "on",
            "edit": request.form.get(f"field__{field_key}__edit") == "on",
        }
    for locked_group in perms_module.HARD_LOCKED_EDIT.get(cu.role, set()):
        if locked_group in field_perms:
            field_perms[locked_group]["edit"] = False

    cu.field_permissions = json.dumps(field_perms)
    cdb.commit()
    flash(f"Field access updated for {cu.full_name}")
    return redirect(url_for("company_settings"))

@app.route("/inventory/clear_party_stock", methods=["POST"])
@login_required
@owner_required
def inventory_clear_party_stock():
    company_id = get_current_company()
    if not company_id:
        return redirect(url_for('login'))
    cdb = get_customer_session(company_id)

    party_name    = (request.form.get("party_name") or "").strip()
    is_cash       = request.form.get("is_cash") == "1"
    client_id_raw = request.form.get("client_id")
    client_id     = int(client_id_raw) if client_id_raw and client_id_raw.isdigit() else None

    if not party_name:
        flash("No party specified.", "danger")
        return redirect(url_for('inventory_list'))
    if not is_cash and not client_id:
        flash("Missing party reference — could not clear stock.", "danger")
        return redirect(url_for('inventory_list'))

    user_email = session.get('user', {}).get('email', '')

    if is_cash:
        items = cdb.query(StockItem).filter_by(company_id=company_id, client_id=None).all()
    else:
        items = cdb.query(StockItem).filter_by(company_id=company_id, client_id=client_id).all()
    item_ids = {i.id: i for i in items}

    if not item_ids:
        flash(f"'{party_name}' had no outstanding stock to clear.", "info")
        return redirect(url_for('inventory_list'))

    hist_rows = cdb.query(StockPurchaseHistory).filter(
        StockPurchaseHistory.stock_item_id.in_(item_ids.keys()),
        StockPurchaseHistory.awb_no.isnot(None),
    ).all()

    dispatched = {(h.stock_item_id, h.awb_no) for h in hist_rows if h.movement_type == "OUT"}
    outstanding = [
        h for h in hist_rows
        if h.movement_type != "OUT" and (h.stock_item_id, h.awb_no) not in dispatched
    ]

    # Cash StockItem rows are shared across shippers, so only close out the
    # AWBs that actually belong to THIS party — matched via the booking
    # invoice's shipper_name, same as inventory_list() does.
    if is_cash:
        refs = {h.reference for h in outstanding if h.reference}
        invoices_by_ref = {
            inv.invoice_id: inv
            for inv in cdb.query(Invoice).filter(
                Invoice.invoice_id.in_(refs), Invoice.company_id == company_id
            ).all()
        } if refs else {}

        def matches(h):
            inv = invoices_by_ref.get(h.reference)
            if not inv:
                return False
            try:
                meta = json.loads(inv.terms) if inv.terms else {}
            except Exception:
                meta = {}
            return (meta.get("shipper_name") or "").strip() == party_name

        outstanding = [h for h in outstanding if matches(h)]

    cleared_count = 0
    for h in outstanding:
        cdb.add(StockPurchaseHistory(
            stock_item_id=h.stock_item_id,
            purchase_invoice_id=None,
            quantity=-(h.quantity or 0),
            purchase_rate=0,
            movement_type="OUT",
            purchase_date=today_ist(),
            reference=f"Manual clear by {user_email}",
            awb_no=h.awb_no,   # must match the outstanding row's AWB exactly —
                                # this is what Package Log's exclusion check needs
        ))
        item = item_ids.get(h.stock_item_id)
        if item and (item.quantity or 0) > 0:
            item.quantity = max(0, (item.quantity or 0) - (h.quantity or 0))
            item.last_updated = today_ist()
        cleared_count += 1

    cdb.commit()

    if cleared_count:
        flash(f"Cleared stock for '{party_name}' — {cleared_count} package(s) closed out.", "success")
    else:
        flash(f"'{party_name}' had no outstanding stock to clear.", "info")

    return redirect(url_for('inventory_list'))

@app.route("/api/purchase-rate-lookup")
@login_required
@require_permission("pricelist", "view")
def api_purchase_rate_lookup():
    """Purchase rate lookup — uses purchase price lists only"""
    cdb = get_cdb()
    company_id = get_current_company()

    courier     = request.args.get('courier', '').strip().upper()
    destination = request.args.get('destination', '').strip().upper()
    weight      = float(request.args.get('weight', 0))

    if not courier or not destination or weight <= 0:
        return jsonify({'error': 'Missing parameters'}), 400

    price_list = find_price_list(cdb, company_id, courier, 'purchase')

    if not price_list:
        return jsonify({'error': f'No active purchase price list found for {courier}'}), 404

    try:
        rate_data = json.loads(price_list.rate_data)
        countries = rate_data.get('countries', {})
        weights   = sorted(rate_data.get('weights', []))

        matched_country = None
        matched_rates   = None

        if destination in countries:
            matched_country = destination
            matched_rates   = countries[destination]
        
        if not matched_rates:
            for country, rates in countries.items():
                if destination in country or country in destination:
                    matched_country = country
                    matched_rates   = rates
                    break

        if not matched_rates:
            dest_words = destination.split()
            for country, rates in countries.items():
                country_words = country.split()
                for dw in dest_words:
                    if len(dw) > 2:
                        for cw in country_words:
                            if dw in cw or cw in dw:
                                matched_country = country
                                matched_rates   = rates
                                break
                    if matched_rates:
                        break
                if matched_rates:
                    break

        if not matched_rates:
            return jsonify({'error': f'No rate found for {destination}'}), 404

        rate, weight_used, pricing_type = calculate_rate(rate_data, matched_country, weight)

        if not rate or rate <= 0:
            return jsonify({'error': f'No rate found for {weight}kg in {matched_country}'}), 404

        return jsonify({
            'success': True,
            'rate': rate,
            'weight_used': weight_used,
            'pricing_type': pricing_type,
            'country_matched': matched_country,
            'courier': courier,
            'destination': destination
        })

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route("/api/price-lists/list")
@login_required
@require_permission("pricelist", "view")
def api_price_lists_list():
    """Return list of available couriers with price lists"""
    cdb = get_cdb()
    company_id = get_current_company()
    
    price_lists = cdb.query(PriceList).filter_by(company_id=company_id, is_active=True).all()
    
    return jsonify([{
        'courier': pl.courier,
        'filename': pl.filename,
        'uploaded_at': pl.uploaded_at.strftime('%d %b %Y'),
        'countries': len(json.loads(pl.rate_data).get('countries', {}))
    } for pl in price_lists])

# ─────────────────────────────────────────────────────────────────────────────
# ── Clients ───────────────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────

def _normalize_client(c, outstanding=None):
    """Return a dict whose keys match what clients.html / client_form.html expect.

    `outstanding`: pass a live-computed total (opening_balance + all unpaid
    invoices - receipts) to show the real current dues. If omitted, falls
    back to the cached c.pending field (balance-carry-forward, only reset
    on statement close/shift — NOT the same as total outstanding)."""
    return {
        # identity
        "id":              c.id,
        "client_id":       c.client_id or "—",
        "client_name":     c.name,
        "client_type":     c.client_type     or "Business",
        "contact_person":  c.contact_person  or "",
        # contact
        "phone":           c.phone           or "",
        "alternate_phone": c.alternate_phone or "",
        "email":           c.email           or "",
        "website":         c.website         or "",
        # address
        "address_line1":   c.address_line1   or "",
        "address_line2":   c.address_line2   or "",
        "city":            c.city            or "",
        "state":           c.state           or "",
        "pincode":         c.pincode         or "",
        "country":         c.country         or "India",
        # GST & tax
        "gst_number":      c.gst_number      or "",
        "pan_number":      c.pan_number      or "",
        "aadhar_number":   c.aadhar_number   or "",
        "aadhar_front_file": c.aadhar_front_file or "",
        "aadhar_back_file":  c.aadhar_back_file  or "",
        "pan_front_file":    c.pan_front_file    or "",
        "pan_back_file":     c.pan_back_file     or "",
        "gst_type":        c.gst_type        or "Regular",
        # financial
        "credit_limit":    c.credit_limit    or 0.0,
        "credit_days":     c.credit_days     or 30,
        "outstanding":     outstanding if outstanding is not None else (c.pending or 0.0),
        "opening_balance": c.opening_balance or 0.0,
        "last_payment":    c.last_payment,
        # status
        "status":          c.status          or "Active",
        "notes":           c.notes           or "",
        "created_at":      c.created_at,
    }


@app.route("/clients")
@login_required
@require_permission("clients", "view")
def client_list():
    cdb = get_cdb()
    company_id    = get_current_company()
    filter_status = request.args.get("status", "All")

    query = cdb.query(Client).filter(
        Client.company_id == company_id,
        Client.client_type.in_(["Customer", "Business", "Individual"]),
        Client.status != "Deleted",
    )
    if filter_status != "All":
        query = query.filter_by(status=filter_status)

    client_rows = query.all()
    client_ids  = [c.id for c in client_rows]

    # Live totals — same formula as the debtor statement's closing balance:
    # opening_balance + invoices since cutoff − receipts since cutoff.
    # A statement_cutoff means opening_balance already nets out everything
    # before that date, so pre-cutoff invoices/receipts must NOT be summed
    # again on top of it — that double-counted Infosys/Tata Consultancy
    # here and on /debtors. Grouped (unfiltered) queries stay as the fast
    # path for the common case of no cutoff; clients with a cutoff get a
    # filtered per-client query instead, mirroring debtor_statement().
    invoiced_by_client = dict(
        cdb.query(Invoice.client_id, func.sum(Invoice.grand_total))
           .filter(Invoice.company_id == company_id, Invoice.client_id.in_(client_ids))
           .group_by(Invoice.client_id).all()
    ) if client_ids else {}
    cash_by_name = dict(
        cdb.query(CashTransaction.party_name, func.sum(CashTransaction.amount))
           .filter(CashTransaction.company_id == company_id,
                   CashTransaction.category.in_(["Receipt", "Adjustment"]),
                   CashTransaction.reference != "WRITE-OFF")
           .group_by(CashTransaction.party_name).all()
    )
    bank_by_name = dict(
        cdb.query(BankTransaction.party_name, func.sum(BankTransaction.amount))
           .filter(BankTransaction.company_id == company_id, BankTransaction.type == "credit")
           .group_by(BankTransaction.party_name).all()
    )

    clients = []
    for c in client_rows:
        cutoff_date = c.statement_cutoff.date() if c.statement_cutoff else None

        if cutoff_date:
            inv_q = (cdb.query(func.sum(Invoice.grand_total))
                     .filter(Invoice.company_id == company_id, Invoice.client_id == c.id,
                             Invoice.date >= cutoff_date))
            total_invoiced = float(inv_q.scalar() or 0)

            cash_q = (cdb.query(func.sum(CashTransaction.amount))
                      .filter(CashTransaction.company_id == company_id, CashTransaction.party_name == c.name,
                              CashTransaction.category.in_(["Receipt", "Adjustment"]),
                              CashTransaction.reference != "WRITE-OFF",
                              CashTransaction.date >= cutoff_date))
            cash_received = float(cash_q.scalar() or 0)

            bank_q = (cdb.query(func.sum(BankTransaction.amount))
                      .filter(BankTransaction.company_id == company_id, BankTransaction.party_name == c.name,
                              BankTransaction.type == "credit", BankTransaction.date >= cutoff_date))
            bank_received = float(bank_q.scalar() or 0)
        else:
            total_invoiced = float(invoiced_by_client.get(c.id, 0) or 0)
            cash_received = 0
            for k, v in cash_by_name.items():
                if k and k.lower() == c.name.lower():
                    cash_received = float(v or 0)
                    break
            bank_received  = float(bank_by_name.get(c.name, 0) or 0)

        true_outstanding = (c.opening_balance or 0) + total_invoiced - cash_received - bank_received
        clients.append(_normalize_client(c, outstanding=true_outstanding))

    return render_template("clients.html", clients=clients, current_status=filter_status)

@app.route("/clients/<int:client_pk>/remove", methods=["POST"])
@login_required
@owner_required
def client_remove(client_pk):
    """
    Soft-delete: removes the client from the Clients list only. Every
    invoice, manifest entry, stock item, and ledger row that references
    this client_id is left untouched — they keep resolving to this same
    row, it just no longer shows up in client_list(). Never a hard DELETE:
    that would violate the FK every invoice/purchase/stock row holds.
    """
    cdb = get_cdb()
    company_id = get_current_company()
    c = _first_or_404(cdb.query(Client).filter_by(id=client_pk, company_id=company_id).first())
    c.status = "Deleted"
    cdb.commit()
    flash(f"'{c.name}' removed from the client list. Their bookings and ledger history are unaffected.")
    return redirect(url_for("client_list"))

# /clients/new  ── template links here for new client
@app.route("/clients/new", methods=["GET", "POST"])
@login_required
@require_permission("clients", "view", method_actions={'POST': 'create'})
def client_new():
    cdb = get_cdb()
    company_id = get_current_company()
    if request.method == "POST":
        f = request.form

        # GST uniqueness check (per company)
        gst = f.get("gst_number", "").strip().upper()
        if gst:
            existing_gst = cdb.query(Client).filter_by(
                company_id=company_id, gst_number=gst
            ).first()
            if existing_gst:
                flash(f"GST number {gst} is already registered to client '{existing_gst.name}'. Please check and try again.", "error")
                return render_template("client_form.html", form_data=f)

        company_obj = Company.query.filter_by(company_id=company_id).first()
        client_prefix = _company_name_prefix(company_obj.company_name if company_obj else "", from_end=True)
        new_client_id = _next_numbered_id(cdb, Client.client_id, client_prefix, extra_filters=[Client.company_id == company_id])

        new_client = Client(
            client_id       = new_client_id,
            company_id      = company_id,
            name            = f.get("client_name", "").strip(),
            client_type     = f.get("client_type", "Business"),
            contact_person  = f.get("contact_person", "").strip(),
            phone           = f.get("phone", "").strip(),
            alternate_phone = f.get("alternate_phone", "").strip(),
            email           = f.get("email", "").strip().lower(),
            website         = f.get("website", "").strip(),
            address_line1   = f.get("address_line1", "").strip(),
            address_line2   = f.get("address_line2", "").strip(),
            city            = f.get("city", "").strip(),
            state           = f.get("state", "").strip(),
            pincode         = f.get("pincode", "").strip(),
            country         = f.get("country", "India").strip(),
            gst_number      = gst or None,
            pan_number      = f.get("pan_number", "").strip().upper() or None,
            aadhar_number   = f.get("aadhar_number", "").strip() or None,
            gst_type        = f.get("gst_type", "Regular"),
            credit_limit    = float(f.get("credit_limit", 0) or 0),
            credit_days     = int(f.get("credit_days", 30) or 30),
            pending         = float(f.get("opening_balance", 0) or 0),
            opening_balance = float(f.get("opening_balance", 0) or 0),
            status          = f.get("status", "Active"),
            notes           = f.get("notes", "").strip(),
            created_at      = today_ist(),
        )
        cdb.add(new_client)
        cdb.commit()
        _save_client_id_docs(cdb, new_client, request.files)
        flash(f"Client '{new_client.name}' added successfully!")
        return redirect(url_for("client_list"))
    return render_template("client_form.html", form_data={})


# Keep /clients/add as an alias so old links still work
@app.route("/clients/add", methods=["GET", "POST"])
@login_required
@require_permission("clients", "create")
def client_add():
    return client_new()


# /clients/<id>  ── view detail (template links here with 👁️)
@app.route("/clients/<int:client_pk>")
@login_required
@require_permission("clients", "view")
def client_view(client_pk):
    cdb = get_cdb()
    company_id = get_current_company()
    c = _first_or_404(cdb.query(Client).filter_by(id=client_pk, company_id=company_id).first())
    client = _normalize_client(c)
    invoices = cdb.query(Invoice).filter_by(company_id=company_id, client_id=c.id).order_by(Invoice.date.desc()).all()
    # Orders feature retired — client_detail.html isn't uploaded here, so still
    # passing orders=[] rather than dropping the kwarg, to avoid breaking that
    # template if it references `orders` directly. Safe to remove once that
    # template's Orders section is also cleaned up.
    orders   = []
    return render_template("client_detail.html", client=client, invoices=invoices, orders=orders)

def _build_client_ledger(cdb, company_id, c, since=None, until=None):
    """Builds the debtor ledger for a client. `since` (a datetime) is the
    statement cutoff — only invoices/receipts dated ON or AFTER its date are
    included, and the opening line reflects the carried-forward balance as
    of that cutoff instead of the original account-opening balance. `until`
    (a date, exclusive) is only passed when archiving a statement being
    closed today — it caps the archive at everything dated BEFORE today, so
    today's entries stay live and land in the new statement instead of the
    one being closed."""
    since_date = since.date() if since else None

    # Full-detail statement (client page): every booking shows as its own
    # line, regardless of whether it's since been grouped into a customer
    # invoice — customer invoices are not shown on this statement at all,
    # so there's nothing to dedupe against.
    invoices_q = cdb.query(Invoice).filter_by(company_id=company_id, client_id=c.id)
    invoices_q = invoices_q.filter(Invoice.status.notin_(['Cancelled', 'Void']))
    if since_date:
        invoices_q = invoices_q.filter(Invoice.date >= since_date)
    if until:
        invoices_q = invoices_q.filter(Invoice.date < until)
    invoices = invoices_q.order_by(Invoice.date.asc()).all()

    # Real payment events — same reasoning as debtor_statement(): a
    # transaction's own date/amount, not the invoice's date and a
    # back-computed grand_total-minus-balance figure that hid advance
    # payments and multi-part payments entirely.
    cash_txns_q = cdb.query(CashTransaction).filter(
        CashTransaction.company_id == company_id,
        func.lower(CashTransaction.party_name) == func.lower(c.name)
    ).filter(CashTransaction.category.in_(["Receipt", "Adjustment"]))
    # The write-off/carry-forward adjustment row itself is the mechanism
    # that produces this cutoff — it must never appear as a ledger line,
    # otherwise every "new" statement would open with a phantom credit.
    cash_txns_q = cash_txns_q.filter(CashTransaction.reference != "WRITE-OFF")
    if since_date:
        cash_txns_q = cash_txns_q.filter(CashTransaction.date >= since_date)
    if until:
        cash_txns_q = cash_txns_q.filter(CashTransaction.date < until)
    cash_txns = cash_txns_q.all()

    bank_txns_q = cdb.query(BankTransaction).filter(
        BankTransaction.company_id == company_id,
        func.lower(BankTransaction.party_name) == func.lower(c.name)
    ).filter(BankTransaction.type == "credit")
    if since_date:
        bank_txns_q = bank_txns_q.filter(BankTransaction.date >= since_date)
    if until:
        bank_txns_q = bank_txns_q.filter(BankTransaction.date < until)
    bank_txns = bank_txns_q.all()

    events = []

    ledger = []
    running_balance = c.opening_balance or 0.0

    # Opening balance / balance carried forward. When `since` is set, this
    # is a carried-forward balance and its date must be the first day of
    # THIS statement (the cutoff date) — not the client's original
    # created_at — otherwise the statement's displayed "from" date is wrong.
    if running_balance:
        ledger.append({
            "date": since.date() if since else (c.created_at or today_ist()),
            "type": "Balance Carried Forward" if since else "Opening Balance",
            "ref": "—",
            "awb": "", "consignee": "", "destination": "", "carrier_ref": "", "carrier": "",
            "chrg_wt": 0, "act_wt": 0, "vol_wt": 0,
            "grand_total": 0, "other_charges": 0, "billing_amount": 0,
            "debit": running_balance,
            "credit": 0,
            "balance": running_balance,
            "status": "",
            "id": None,
        })

    for inv in invoices:
        ship = _get_shipment_meta(inv)
        grand_total = inv.grand_total or 0
        events.append({
            "date": inv.date,
            "type": "Invoice",
            "ref": inv.invoice_id,
            "awb": ship["awb"],
            "consignee": ship["consignee"],
            "destination": ship["destination"],
            "carrier_ref": ship["carrier_ref"],
            "carrier": ship["carrier"],
            "chrg_wt": ship["chrg_wt"],
            "act_wt": ship["act_wt"],
            "vol_wt": ship["vol_wt"],
            "grand_total": grand_total,
            "other_charges": ship["other_charges"],
            "billing_amount": ship["other_charges"] + grand_total,
            "per_kg": ship.get("per_kg", 0),
            "debit": grand_total,
            "credit": 0,
            "status": inv.status,
            "id": inv.invoice_id,
            "_sort": 0,
        })

    blank_shipment = {"awb": "", "consignee": "", "destination": "", "carrier_ref": "", "carrier": "",
                       "chrg_wt": 0, "act_wt": 0, "vol_wt": 0,
                       "grand_total": 0, "other_charges": 0, "billing_amount": 0,
                       "per_kg": 0}

    for ct in cash_txns:
        ref = ct.reference or ""
        events.append({
            "date": ct.date,
            "type": "Payment Received",
            "ref": "—" if ref == "ADVANCE" else ref,
            **blank_shipment,
            "debit": 0, "credit": ct.amount or 0, "status": "",
            "id": ref or None, "_sort": 1,
        })

    for bt in bank_txns:
        ref = bt.reference or ""
        events.append({
            "date": bt.date, "type": "Payment Received",
            "ref": "—" if ref == "ADVANCE" else ref,
            **blank_shipment,
            "debit": 0, "credit": bt.amount or 0, "status": "",
            "id": ref or None, "_sort": 1,
        })

    events.sort(key=lambda e: (e["date"] or date.min, e["_sort"]))
    for e in events:
        running_balance += (e["debit"] or 0) - (e["credit"] or 0)
        e["balance"] = running_balance
        del e["_sort"]
        ledger.append(e)

    total_debit = sum(r["debit"] for r in ledger)
    total_credit = sum(r["credit"] for r in ledger)

    return ledger, total_debit, total_credit, running_balance


@app.route("/clients/<int:client_pk>/statement")
@login_required
@require_permission("clients", "view")
def client_statement(client_pk):
    """Statement view for a client (debtor-style ledger)"""
    cdb = get_cdb()
    company_id = get_current_company()
    c = _first_or_404(cdb.query(Client).filter_by(id=client_pk, company_id=company_id).first())

    ledger, total_debit, total_credit, running_balance = _build_client_ledger(
        cdb, company_id, c, since=c.statement_cutoff)

    archives = (cdb.query(StatementClosing)
                .filter_by(company_id=company_id, entity_type="client", entity_id=c.id)
                .order_by(StatementClosing.closed_at.desc())
                .all())

    return render_template("ledger_statement.html",
                           entity=_normalize_client(c),
                           company=get_company_by_id(company_id),
                           ledger=ledger,
                           total_debit=total_debit,
                           total_credit=total_credit,
                           closing_balance=running_balance,
                           mode="debtor",
                           nav_active="clients",
                           back_url=f"/clients/{client_pk}",
                           archive_base_url=f"/clients/{client_pk}",
                           archives=archives,
                           archived=False,
                           today=today_ist().strftime("%d %b %Y"))


@app.route("/clients/<int:client_pk>/statement/archive/<int:archive_id>")
@login_required
@require_permission("clients", "view")
def client_statement_archive(client_pk, archive_id):
    """Prints a frozen old statement exactly as it looked at the moment the
    outstanding was cleared/shifted — recomputing from live data would drift
    if invoices are edited later, so this reads the saved snapshot instead."""
    cdb = get_cdb()
    company_id = get_current_company()
    c = _first_or_404(cdb.query(Client).filter_by(id=client_pk, company_id=company_id).first())
    archive = _first_or_404(cdb.query(StatementClosing).filter_by(
        id=archive_id, company_id=company_id, entity_type="client", entity_id=client_pk).first())

    return render_template("ledger_statement.html",
                           entity=_normalize_client(c),
                           company=get_company_by_id(company_id),
                           ledger=json.loads(archive.ledger_snapshot or "[]"),
                           total_debit=archive.total_debit,
                           total_credit=archive.total_credit,
                           closing_balance=archive.closing_balance,
                           mode="debtor",
                           nav_active="clients",
                           back_url=f"/clients/{client_pk}/statement",
                           archived=True,
                           archived_at=archive.closed_at,
                           today=today_ist().strftime("%d %b %Y"))

# /clients/<id>/edit
@app.route("/clients/<int:client_pk>/edit", methods=["GET", "POST"])
@login_required
@require_permission("clients", "view", method_actions={'POST': 'edit'})
def client_edit(client_pk):
    cdb = get_cdb()
    company_id = get_current_company()
    c          = _first_or_404(cdb.query(Client).filter_by(id=client_pk, company_id=company_id).first())
    if request.method == "POST":
        f   = request.form
        gst = f.get("gst_number", "").strip().upper()

        # GST uniqueness: check no OTHER client has the same GST
        if gst:
            existing_gst = cdb.query(Client).filter(
                Client.company_id == company_id,
                Client.gst_number == gst,
                Client.id != c.id
            ).first()
            if existing_gst:
                flash(f"GST number {gst} is already registered to client '{existing_gst.name}'.", "error")
                return render_template("client_form.html", client=_normalize_client(c), form_data=f)

        c.name            = f.get("client_name", c.name).strip()
        c.client_type     = f.get("client_type",     c.client_type)
        c.contact_person  = f.get("contact_person",  c.contact_person or "").strip()
        c.phone           = f.get("phone",            c.phone or "").strip()
        c.alternate_phone = f.get("alternate_phone",  c.alternate_phone or "").strip()
        c.email           = f.get("email",            c.email or "").strip().lower()
        c.website         = f.get("website",          c.website or "").strip()
        c.address_line1   = f.get("address_line1",    c.address_line1 or "").strip()
        c.address_line2   = f.get("address_line2",    c.address_line2 or "").strip()
        c.city            = f.get("city",             c.city or "").strip()
        c.state           = f.get("state",            c.state or "").strip()
        c.pincode         = f.get("pincode",          c.pincode or "").strip()
        c.country         = f.get("country",          c.country or "India").strip()
        c.gst_number      = gst or None
        c.pan_number      = f.get("pan_number",  c.pan_number or "").strip().upper() or None
        c.aadhar_number   = f.get("aadhar_number", c.aadhar_number or "").strip() or None
        c.gst_type        = f.get("gst_type",    c.gst_type)
        c.credit_limit    = float(f.get("credit_limit",    c.credit_limit    or 0) or 0)
        c.credit_days     = int(f.get("credit_days",       c.credit_days     or 30) or 30)
        c.opening_balance = float(f.get("opening_balance", c.opening_balance or 0) or 0)
        c.status          = f.get("status", c.status)
        c.notes           = f.get("notes",   c.notes or "").strip()
        cdb.commit()
        _save_client_id_docs(cdb, c, request.files)
        flash(f"Client '{c.name}' updated successfully!")
        return redirect(url_for("client_list"))
    return render_template("client_form.html", client=_normalize_client(c), form_data={})


def _client_closing_balance(cdb, company_id, c):
    """Live running balance exactly as the client statement page computes it:
    opening balance + all invoice totals − all recorded receipts (cash + bank)."""
    total_invoiced = sum(
        inv.grand_total or 0
        for inv in cdb.query(Invoice).filter_by(company_id=company_id, client_id=c.id).all()
    )
    cash_received = sum(
        t.amount or 0
        for t in cdb.query(CashTransaction).filter_by(
            company_id=company_id, party_name=c.name, category="Receipt").all()
    )
    bank_received = sum(
        t.amount or 0
        for t in cdb.query(BankTransaction).filter_by(
            company_id=company_id, party_name=c.name
        ).filter(BankTransaction.type == "credit").all()
    )
    return (c.opening_balance or 0) + total_invoiced - cash_received - bank_received


def _client_close_statement(cdb, company_id, c, action, scope="till_yesterday", as_of_date=None):
    """Archives the client's current live ledger into StatementClosing (so
    it can be printed later exactly as it stood), then moves the statement
    cutoff forward so the next statement load starts blank (action=
    'cleared') or with just the carried-forward balance (action=
    'carried_forward').

    `as_of_date`: the LAST date to include in the archived/closed
    statement. Everything dated ON or BEFORE it is archived; everything
    AFTER it stays live and becomes the first entries of the new
    statement (whose "Balance Carried Forward" date is as_of_date + 1
    day). Lets someone carry forward through, say, 30 June even though
    today is 5 July — the 1–5 July bills stay live in the new statement.
    If not given, it's derived from `scope` (only relevant for
    action='cleared'):
      - 'till_yesterday' (default): as_of_date = yesterday — today's
        entries stay live.
      - 'complete': as_of_date = today — nothing stays live.
    Clamped so it can never be before the day preceding the current
    statement_cutoff (which would resurrect already-archived entries)
    or after today (can't close a future date).
    Returns the amount that was outstanding at closing time."""
    today = today_ist()

    if as_of_date is None:
        if action == "cleared" and scope == "complete":
            as_of_date = today
        else:
            as_of_date = today - timedelta(days=1)

    if as_of_date > today:
        as_of_date = today
    if c.statement_cutoff:
        floor_date = c.statement_cutoff.date() - timedelta(days=1)
        if as_of_date < floor_date:
            as_of_date = floor_date

    archive_until = as_of_date + timedelta(days=1)  # exclusive upper bound

    ledger, total_debit, total_credit, closing = _build_client_ledger(
        cdb, company_id, c, since=c.statement_cutoff, until=archive_until)

    cdb.add(StatementClosing(
        company_id=company_id,
        entity_type="client",
        entity_id=c.id,
        entity_name=c.name,
        action=action,
        closing_balance=closing,
        total_debit=total_debit,
        total_credit=total_credit,
        ledger_snapshot=json.dumps(ledger, default=str),
        closed_by=session.get("username", "unknown"),
        closed_at=datetime.utcnow(),
    ))

    c.statement_cutoff = datetime.combine(archive_until, datetime.min.time())
    c.opening_balance = closing if action == "carried_forward" else 0
    c.pending = c.opening_balance
    return closing


# /clients/<id>/delete  ── kept at the old URL/template link so nothing else
# breaks, but this NO LONGER deletes the client row. Deleting Invoice rows
# would violate GST retention and break every FK pointing at invoices.id
# (purchase invoice lines, WhatsApp logs, cheques). This now archives the old
# statement and clears the live statement/outstanding — the client record
# and invoice history stay, viewable via the archived statement link.
@app.route("/clients/<int:client_pk>/delete", methods=["GET", "POST"])
@login_required
@owner_required
def client_delete(client_pk):
    cdb = get_cdb()
    company_id = get_current_company()
    c          = _first_or_404(cdb.query(Client).filter_by(id=client_pk, company_id=company_id).first())
    scope = request.args.get("scope", "till_yesterday")
    if scope not in ("complete", "till_yesterday"):
        scope = "till_yesterday"
    amount = _client_close_statement(cdb, company_id, c, action="cleared", scope=scope)
    cdb.commit()
    if amount:
        if scope == "complete":
            flash(f"Outstanding of ₹{amount:,.2f} cleared for '{c.name}', including today's entries. Old statement archived — client record and invoices were kept.")
        else:
            flash(f"Outstanding of ₹{amount:,.2f} cleared for '{c.name}' up to yesterday. Old statement archived — today's entries remain in the new statement.")
    else:
        flash(f"'{c.name}' had no outstanding to clear.")
    return redirect(url_for("client_list"))


# /clients/<id>/shift-to-opening  ── archives the itemised ledger the same
# way as above, but carries the amount forward as a single opening_balance
# figure instead of writing it off to zero. Defaults to yesterday, but an
# explicit ?as_of=YYYY-MM-DD lets the user pick an earlier cutoff (e.g.
# carry forward through 30 June even though today is 5 July) — anything
# after that date stays live in the new statement regardless.
@app.route("/clients/<int:client_pk>/shift-to-opening", methods=["GET", "POST"])
@login_required
@owner_required
def client_shift_to_opening(client_pk):
    cdb = get_cdb()
    company_id = get_current_company()
    c          = _first_or_404(cdb.query(Client).filter_by(id=client_pk, company_id=company_id).first())
    as_of_date = None
    as_of_raw = request.args.get("as_of")
    if as_of_raw:
        try:
            as_of_date = datetime.strptime(as_of_raw, "%Y-%m-%d").date()
        except ValueError:
            as_of_date = None
    amount = _client_close_statement(cdb, company_id, c, action="carried_forward", as_of_date=as_of_date)
    cdb.commit()
    flash(f"₹{amount:,.2f} carried forward as opening balance for '{c.name}', as of "
          f"{(c.statement_cutoff - timedelta(days=1)).strftime('%d %b %Y')}. New statement starts "
          f"{c.statement_cutoff.strftime('%d %b %Y')}; entries from then on stay live.")
    return redirect(url_for("client_list"))


# ─────────────────────────────────────────────────────────────────────────────
# ── Stock / Inventory ─────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────


@app.route("/inventory")
@login_required
@require_permission("stock", "view")
def inventory_list():
    cdb = get_customer_session(get_current_company())
    company_id = get_current_company()

    stock_items = cdb.query(StockItem).filter_by(company_id=company_id).all()

    # ── Top summary: total + by item_type ────────────────────────────────────
    total_inventory  = 0
    inventory_by_type = {}   # {"Box": 455, "Envelope": 3000, ...}

    for item in stock_items:
        qty = int(item.quantity or 0)
        if qty <= 0:
            continue
        total_inventory += qty
        itype = (item.item_type or item.category or "Other").strip()
        inventory_by_type[itype] = inventory_by_type.get(itype, 0) + qty

    # ── Party-wise breakdown ──────────────────────────────────────────────────
    # Build: {client_id: {"name": str, "items": {item_type: qty}}}
    client_ids = set(i.client_id for i in stock_items if i.client_id)
    clients    = {
        c.id: c.name
        for c in cdb.query(Client).filter(
            Client.id.in_(client_ids),
            Client.company_id == company_id
        ).all()
    } if client_ids else {}

    # All item types that appear anywhere in stock — for table columns.
    # NOTE: this used to be scoped to `if i.client_id`, which meant a type
    # that only ever showed up on cash (client-less) stock had no column at
    # all. Now that cash stock gets its own party_stock row below, it needs
    # a column too.
    all_types = sorted(set(
        (i.item_type or i.category or "Other").strip()
        for i in stock_items if int(i.quantity or 0) > 0
    ))

    party_stock = {}       # {party_label: {item_type: qty}}
    party_client_ids = {}  # {party_label: client_id} — used for click-to-filter in the template
    party_is_cash = {}     # {party_label: True} — cash/anonymous buckets, for the UI badge
    for item in stock_items:
        if not item.client_id:
            continue
        qty = int(item.quantity or 0)
        if qty <= 0:
            continue
        cname = clients.get(item.client_id, f"Party #{item.client_id}")
        itype = (item.item_type or item.category or "Other").strip()
        if cname not in party_stock:
            party_stock[cname] = {}
        party_stock[cname][itype] = party_stock[cname].get(itype, 0) + qty
        party_client_ids[cname] = item.client_id

    # ── Package-level shipment log (AWB / source / destination / dims / weight) ──
    # Pulled from StockPurchaseHistory, which gets ONE row per package added at
    # booking time — unlike StockItem, which merges quantities by name+client_id
    # and can't hold per-shipment detail without overwriting it on every booking.
    stock_item_ids = {i.id for i in stock_items}
    all_history_rows = (
        cdb.query(StockPurchaseHistory)
        .filter(StockPurchaseHistory.stock_item_id.in_(stock_item_ids))
        .filter(StockPurchaseHistory.awb_no.isnot(None))
        .order_by(StockPurchaseHistory.purchase_date.desc())
        .all()
    ) if stock_item_ids else []

    # Figure out which intake quantity is still on hand, per stock_item_id,
    # using running remaining quantity (FIFO by intake date) instead of
    # matching IN/OUT rows by awb_no. The old approach only works when a
    # stock item's intake and dispatch share the same awb_no — true for a
    # per-client item, false for a shared cash bucket bulk-purchased under
    # one intake awb and drained by several different customer dockets.
    in_rows_by_item = {}
    dispatched_qty_by_item = {}
    for h in all_history_rows:
        if h.movement_type == "OUT":
            dispatched_qty_by_item[h.stock_item_id] = (
                dispatched_qty_by_item.get(h.stock_item_id, 0) + abs(h.quantity or 0)
            )
        else:
            in_rows_by_item.setdefault(h.stock_item_id, []).append(h)

    history_rows = []  # list of (row, remaining_qty), not bare rows
    for sid, rows in in_rows_by_item.items():
        rows.sort(key=lambda r: (r.purchase_date or date.min, r.id))
        remaining_to_clear = dispatched_qty_by_item.get(sid, 0)
        for r in rows:
            qty = int(r.quantity or 0)
            if remaining_to_clear >= qty:
                remaining_to_clear -= qty
                continue
            still_here = qty - remaining_to_clear
            remaining_to_clear = 0
            history_rows.append((r, still_here))

    # Pull payment_mode / shipper_name off the invoice each history row is
    # tagged with. `reference` is a plain string column (no FK) storing the
    # invoice's human-readable invoice_id, set at booking time in
    # invoice_customer_save / invoice_customer_update. purchase_invoice_id
    # can't be reused for this — it's a real FK to purchase_invoices.id.
    # Older rows saved before this tagging existed will just fall back to
    # "—" below.
    invoice_refs = {h.reference for h, _ in history_rows if h.reference}
    invoices_by_ref = {
        inv.invoice_id: inv
        for inv in cdb.query(Invoice).filter(
            Invoice.invoice_id.in_(invoice_refs), Invoice.company_id == company_id
        ).all()
    } if invoice_refs else {}

    stock_items_by_id = {i.id: i for i in stock_items}
    package_log = []
    for h, remaining_qty in history_rows:
        if remaining_qty <= 0:
            continue
        si = stock_items_by_id.get(h.stock_item_id)
        if not si:
            continue

        booking_type = "—"
        shipper_name = "—"
        inv = invoices_by_ref.get(h.reference)
        if inv:
            try:
                meta = json.loads(inv.terms) if inv.terms else {}
            except Exception:
                meta = {}
            # booking_type ("cash"/"credit") is the ledger classification —
            # the same field that drives the CASH/CREDIT badge and the
            # All Sales/Cash/Credit filter tabs on the Booking list.
            # payment_mode is a separate instrument field (cash/upi/cheque)
            # that defaults to "cash" even on credit bookings, so it's the
            # wrong field to key this off of.
            booking_type = (meta.get("booking_type") or "—").strip().title() or "—"
            shipper_name = (meta.get("shipper_name") or "—").strip() or "—"

        # Cash / walk-in bookings have no client_id, so there's no real
        # "party". Use the shipper_name off the invoice as the party label
        # instead, so cash stock is attributable to *someone* rather than
        # collapsing into "—" and becoming untrackable. This same label is
        # used to build the cash rows in party_stock below, so
        # click-to-filter (data-party) lines up between the two tables.
        # No "Cash — " prefix on the name itself — is_cash below already
        # carries that signal for the template to render as a badge.
        if si.client_id:
            party_label = clients.get(si.client_id, f"Party #{si.client_id}")
        else:
            party_label = shipper_name if shipper_name != "—" else "Unknown"

        package_log.append({
            "awb_no":       h.awb_no or "—",
            "party_name":   party_label,
            "client_id":    si.client_id,
            "is_cash":      not bool(si.client_id),
            "item_name":    si.name,
            "item_type":    (si.item_type or si.category or "Other"),
            "qty":          remaining_qty,
            "source":       h.source or "—",
            "destination":  h.destination or "—",
            "booking_date": h.purchase_date.strftime("%d %b %Y") if h.purchase_date else "—",
            "length":       h.length or 0,
            "width":        h.width or 0,
            "height":       h.height or 0,
            "weight":       h.weight or 0,
            "payment_mode": booking_type,
            "shipper_name": shipper_name,
        })

        # Fold cash stock into party_stock too, under the same label, so it
        # shows up in the Party-wise Stock Breakdown instead of being
        # invisible (StockItem rows with no client_id are skipped above).
        # We aggregate off history_rows (not StockItem) because a single
        # client-less StockItem can be fed by bookings from several
        # different cash shippers — StockItem only tracks a merged total
        # per item name, not per shipper, so history is the only place
        # that per-shipper quantity actually exists.
        if not si.client_id:
            itype = (si.item_type or si.category or "Other").strip()
            if remaining_qty > 0:
                party_stock.setdefault(party_label, {})
                party_stock[party_label][itype] = party_stock[party_label].get(itype, 0) + remaining_qty
                party_is_cash[party_label] = True

    return render_template("inventory.html",
        total_inventory=total_inventory,
        inventory_by_type=inventory_by_type,
        party_stock=party_stock,
        party_client_ids=party_client_ids,
        party_is_cash=party_is_cash,
        all_types=all_types,
        package_log=package_log,
    )

# ── Stock JSON API (used by inventory.html JS modals) ────────────────────────
@app.route("/stock/item/<code>")
@login_required
@require_permission("stock", "view")
def stock_item_get(code):
    cdb = get_cdb()
    company_id = get_current_company()
    item = _first_or_404(cdb.query(StockItem).filter_by(company_id=company_id, code=code.upper()).first())
    return jsonify({
        "code":          item.code,
        "name":          item.name,
        "category":      item.category or "",
        "quantity":      item.quantity,
        "unit":          item.unit or "pcs",
        "unit_price":    item.unit_price,
        "reorder_level": item.reorder_level or 10,
        "hsn":           item.hsn or "",
    })

@app.route("/api/stock/items")
@login_required
@require_permission("stock", "view")
def api_stock_items():
    cdb = get_cdb()
    
    company_id = get_current_company()
    items = cdb.query(StockItem).filter_by(company_id=company_id).order_by(StockItem.name).all()
    return jsonify([{
        "id":            item.id,
        "code":          item.code or "",
        "name":          item.name,
        "unit":          item.unit or "pcs",
        "quantity":      item.quantity,
        "unit_price":    float(item.unit_price or 0),
        "purchase_rate": float(item.purchase_rate or item.last_purchase_rate or 0),
        "gst_percent":   float(item.gst_percent or 18),
        "hsn":           item.hsn or "",
        "category":      item.category or "",
        "reorder_level": item.reorder_level or 10,
    } for item in items])

@app.route("/stock/save", methods=["POST"])
@login_required
@require_permission("stock", "view", method_actions={'POST': 'create'})
def stock_save():
    """Create or update a stock item via JSON (called from the modal form)."""
    cdb = get_cdb()
    company_id = get_current_company()
    data       = request.get_json(force=True)

    code = data.get("code", "").strip().upper()
    item = cdb.query(StockItem).filter_by(company_id=company_id, code=code).first() if code else None

    if item:
        # update existing
        item.name          = data.get("name", item.name)
        item.category      = data.get("category", item.category)
        item.quantity      = float(data.get("quantity", item.quantity))
        item.unit          = data.get("unit", item.unit)
        item.unit_price    = float(data.get("unit_price", item.unit_price))
        item.reorder_level = float(data.get("reorder_level", item.reorder_level))
        item.last_updated  = today_ist()
    else:
        # auto-generate a code if none provided
        if not code:
            code  = _next_numbered_id(cdb, StockItem.code, "PROD")
        item = StockItem(
            company_id    = company_id,
            code          = code,
            name          = data.get("name", ""),
            category      = data.get("category", "Other"),
            quantity      = float(data.get("quantity", 0)),
            unit          = data.get("unit", "pcs"),
            unit_price    = float(data.get("unit_price", 0)),
            reorder_level = float(data.get("reorder_level", 10)),
            hsn           = data.get("hsn", ""),
            last_updated  = today_ist(),
        )
        cdb.add(item)

    cdb.commit()
    return jsonify({"success": True, "code": item.code})


@app.route("/stock/adjust", methods=["POST"])
@login_required
@require_permission("stock", "edit")
def stock_adjust():
    """Quick quantity adjustment from the Adj button in the table."""
    cdb = get_cdb()
    company_id = get_current_company()
    data       = request.get_json(force=True)
    code       = data.get("code", "").strip().upper()
    item       = _first_or_404(cdb.query(StockItem).filter_by(company_id=company_id, code=code).first())
    item.quantity     = float(data.get("quantity", item.quantity))
    item.last_updated = today_ist()
    cdb.commit()
    return jsonify({"success": True})


@app.route("/stock/movements/<code>")
@login_required
@require_permission("stock", "view")
def stock_movements(code):
    """Return full movement history for a stock item (purchases IN, invoices OUT)."""
    cdb = get_cdb()
    company_id = get_current_company()
    item = cdb.query(StockItem).filter_by(
        company_id=company_id, code=code.upper()
    ).first()

    history = (
        cdb.query(StockPurchaseHistory)
        .filter_by(stock_item_id=item.id)
        .order_by(StockPurchaseHistory.purchase_date.desc())
        .all()
    )

    movements = []
    total_in  = 0
    total_out = 0

    for h in history:
        qty = h.quantity or 0
        is_in = qty > 0

        # Determine movement type and reference
        if h.purchase_invoice_id:
            inv = cdb.get(PurchaseInvoice, h.purchase_invoice_id)
            ref  = inv.invoice_number or inv.invoice_id if inv else f"PUR-{h.purchase_invoice_id}"
            mtype = "Purchase"
        else:
            # Negative qty = dispatched via customer invoice
            mtype = "Dispatched"
            ref   = "Customer Invoice"

        if is_in:
            total_in += abs(qty)
        else:
            total_out += abs(qty)

        movements.append({
            "date":     h.purchase_date.strftime("%d %b %Y") if h.purchase_date else "",
            "type":     mtype,
            "ref":      ref,
            "quantity": qty,
            "rate":     float(h.purchase_rate or 0),
        })

    return jsonify({
        "code":       item.code,
        "name":       item.name,
        "movements":  movements,
        "total_in":   total_in,
        "total_out":  total_out,
    })


@login_required
def inventory_add():
    company_id = get_current_company()
    if request.method == "POST":
        item = StockItem(
            company_id=company_id,
            code=request.form.get("code", "").upper(),
            name=request.form.get("name", ""),
            category=request.form.get("category", ""),
            quantity=float(request.form.get("quantity", 0)),
            unit=request.form.get("unit", "pcs"),
            unit_price=float(request.form.get("unit_price", 0)),
            reorder_level=float(request.form.get("reorder_level", 0)),
            hsn=request.form.get("hsn", ""),
            last_updated=today_ist(),
        )
        cdb.add(item)
        cdb.commit()
        flash("Stock item added!")
        return redirect(url_for("inventory_list"))
    return render_template("inventory_form.html")


@app.route("/inventory/edit/<int:item_pk>", methods=["GET", "POST"])
@login_required
@require_permission("stock", "edit")
def inventory_edit(item_pk):
    cdb = get_cdb()
    company_id = get_current_company()
    item       = _first_or_404(cdb.query(StockItem).filter_by(id=item_pk, company_id=company_id).first())
    if request.method == "POST":
        item.name          = request.form.get("name", item.name)
        item.category      = request.form.get("category", item.category)
        item.quantity      = float(request.form.get("quantity", item.quantity))
        item.unit          = request.form.get("unit", item.unit)
        item.unit_price    = float(request.form.get("unit_price", item.unit_price))
        item.reorder_level = float(request.form.get("reorder_level", item.reorder_level))
        item.hsn           = request.form.get("hsn", item.hsn)
        item.last_updated  = today_ist()
        cdb.commit()
        flash("Stock item updated!")
        return redirect(url_for("inventory_list"))
    return render_template("inventory_form.html", item=item)


@app.route("/inventory/delete/<int:item_pk>", methods=["POST"])
@login_required
@owner_required
def inventory_delete(item_pk):
    cdb = get_cdb()
    company_id = get_current_company()
    item       = _first_or_404(cdb.query(StockItem).filter_by(id=item_pk, company_id=company_id).first())
    cdb.delete(item)
    cdb.commit()
    flash("Stock item deleted.")
    return redirect(url_for("inventory_list"))

# ── Purchase Invoice Routes ─────────────────────────────────────────────────────────

@app.route("/purchase/list")
@login_required
@require_permission("purchase", "view")
def purchase_invoice_list():
    cdb = get_cdb()
    company_id = get_current_company()
    invoices = cdb.query(PurchaseInvoice).filter_by(company_id=company_id).order_by(PurchaseInvoice.date.desc()).all()

    print("=== Purchase Invoice Debug ===")
    for inv in invoices:
        print(f"ID: {inv.id}, invoice_id: {inv.invoice_id}, supplier: {inv.supplier.name if inv.supplier else 'None'}")
    
    total_amount = sum(p.grand_total for p in invoices)
    total_paid = sum(p.paid_amount for p in invoices)
    total_due = sum(p.balance for p in invoices)
    
    return render_template("purchases.html",
        purchases=invoices,
        total_amount=total_amount,
        total_paid=total_paid,
        total_due=total_due
    )


@app.route("/purchase/generate-from-booking", methods=["POST"])
@login_required
@require_permission("purchase", "view", method_actions={'POST': 'create'})
def purchase_generate_from_booking():
    """
    Manual repair button for the "auto-generate purchase line + manifest
    entry" flow — for a booking that never got one or both of those for
    any reason (an old edit that predates the invoice_customer_update()
    fix, a rate-lookup that threw before that was hardened, etc). Takes
    the booking's AWB/docket number or its invoice ID (e.g.
    CUST-20260723-034), rebuilds the same inputs the save routes would
    have had from the booking's stored terms JSON, and syncs the purchase
    line and the manifest entry independently — each is idempotent and
    skipped on its own if it already exists, so re-running this on a
    booking that already got its purchase line (but not its manifest
    entry, or vice versa) still repairs whichever one is missing.
    """
    cdb = get_cdb()
    company_id = get_current_company()
    lookup = (request.form.get("booking_ref") or "").strip()

    if not lookup:
        flash("Enter the booking's AWB/docket number or invoice ID.", "error")
        return redirect(request.referrer or url_for("purchase_invoice_list"))

    inv = cdb.query(Invoice).filter_by(company_id=company_id, docket_no=lookup).first()
    if not inv:
        inv = cdb.query(Invoice).filter_by(company_id=company_id, invoice_id=lookup).first()
    if not inv:
        # The dedicated docket_no column is NULL on a lot of bookings — it's
        # only reliably populated inside terms JSON. Match on that instead.
        inv = cdb.query(Invoice).filter_by(company_id=company_id).filter(
            Invoice.terms.like(f'%"docket_no": "{lookup}"%')
        ).first()
    if not inv:
        flash(f"No booking found matching '{lookup}'.", "error")
        return redirect(request.referrer or url_for("purchase_invoice_list"))

    try:
        meta = json.loads(inv.terms) if inv.terms else {}
    except Exception:
        meta = {}

    company_obj = Company.query.filter_by(company_id=company_id).first()
    apply_gst = company_obj.is_gst_registered if (company_obj and hasattr(company_obj, "is_gst_registered")) else True
    gst_calc = {"is_interstate": bool(meta.get("is_interstate", False))}
    packages_data = meta.get("packages") or []
    freight_weight = float(meta.get("freight_weight") or 0)
    invoice_date = inv.date.isoformat() if inv.date else str(today_ist())
    docket_no = inv.docket_no or meta.get("docket_no", "")

    booking_form = {
        "courier_company_id": meta.get("courier_company_id", ""),
        "carrier":            meta.get("carrier", ""),
        "carrier_ref":        meta.get("carrier_ref", ""),
        "destination":        meta.get("destination", ""),
        "shipper_name":       meta.get("shipper_name", ""),
        "booking_type":       meta.get("booking_type", "credit"),
    }

    if not (booking_form["courier_company_id"] and booking_form["carrier"]):
        flash(f"'{lookup}' has no Courier Company / Carrier saved on it, so a purchase "
              f"line/manifest entry can't be generated — open the booking, set those, "
              f"and save first.", "error")
        return redirect(request.referrer or url_for("purchase_invoice_list"))

    # ── Purchase line — skip if it already exists, don't touch it. ───────────
    existing_pi_item = cdb.query(PurchaseInvoiceItem).filter_by(source_invoice_id=inv.id).first()
    if existing_pi_item:
        flash(f"'{lookup}' already has a purchase line — left it as-is. "
              f"Edit that line from Purchases if its rate/weight needs correcting.", "warning")
    else:
        _sync_auto_purchase_invoice_line(
            cdb, company_id, booking_form, packages_data,
            freight_weight, apply_gst, gst_calc,
            invoice_date, docket_no, inv.invoice_id, inv.id, "final",
        )
        cdb.commit()
        if cdb.query(PurchaseInvoiceItem).filter_by(source_invoice_id=inv.id).first():
            flash(f"Purchase line created for {docket_no or inv.invoice_id}.", "success")
        else:
            flash(f"Could not create a purchase line for {docket_no or inv.invoice_id} — "
                  f"check the flash warnings above for the reason, or the server log.", "error")

    # ── Manifest entry — same idempotent-skip pattern, checked independently
    # of the purchase line above so this button repairs whichever of the two
    # is actually missing. ────────────────────────────────────────────────
    existing_manifest_entry = cdb.query(ManifestEntry).join(
        CompanyManifest, ManifestEntry.manifest_id == CompanyManifest.id
    ).filter(
        ManifestEntry.docket_no == docket_no,
        CompanyManifest.company_id == company_id,
    ).first() if docket_no else None

    if existing_manifest_entry:
        flash(f"'{lookup}' is already on a manifest — left it as-is.", "warning")
    else:
        total_boxes_mf = sum((p.get("qty") or 1) for p in packages_data) or 1
        primary_stock_name = packages_data[0].get("name") if packages_data else None
        _sync_auto_manifest_entry(
            cdb, company_id, booking_form["shipper_name"], booking_form["carrier"], "final",
            invoice_date, docket_no, inv.invoice_id, total_boxes_mf,
            primary_stock_id=None,  # repair path never links/creates stock — see helper docstring
            primary_stock_name=primary_stock_name,
            booking_type=booking_form["booking_type"],
        )
        cdb.commit()
        still_missing = not (cdb.query(ManifestEntry).join(
            CompanyManifest, ManifestEntry.manifest_id == CompanyManifest.id
        ).filter(
            ManifestEntry.docket_no == docket_no,
            CompanyManifest.company_id == company_id,
        ).first() if docket_no else False)
        if still_missing:
            flash(f"Could not add {docket_no or inv.invoice_id} to the manifest — "
                  f"check the flash warnings above for the reason, or the server log.", "error")
        else:
            flash(f"{docket_no or inv.invoice_id} added to the manifest.", "success")

    return redirect(request.referrer or url_for("purchase_invoice_list"))

COURIER_OPTIONS = ["Bluedart", "DHL", "DTDC", "DPD", "FedEx", "Delhivery", "Ecom Express", "India Post", "Other"]
ITEM_TYPE_OPTIONS = ["Box", "Envelope", "Crate", "Pouch", "Carton"]

@app.route("/purchase/delete/<invoice_id>", methods=["POST"])
@login_required
@owner_required
def purchase_invoice_delete(invoice_id):
    cdb        = get_cdb()
    company_id = get_current_company()

    invoice = cdb.query(PurchaseInvoice).filter_by(
        invoice_id=invoice_id, company_id=company_id
    ).first()

    if not invoice:
        abort(404)

    # ── Reverse stock deductions ──────────────────────────────────────────────
    # When a purchase bill was created, stock was DEDUCTED (OUT movement).
    # Deleting the bill reverses that: add stock back.
    for item in invoice.items:
        if item.stock_item_id and item.quantity:
            stock = cdb.query(StockItem).filter_by(
                id=item.stock_item_id, company_id=company_id
            ).first()
            if stock:
                stock.quantity    = (stock.quantity or 0) + item.quantity
                stock.last_updated = today_ist()

    # ── Reverse supplier payable ──────────────────────────────────────────────
    # Only reverse the UNPAID portion (paid_amount was already deducted from
    # supplier.payable when payments were recorded).
    if invoice.supplier_id:
        supplier = cdb.get(Supplier, invoice.supplier_id)
        if supplier:
            unpaid = invoice.balance or 0
            supplier.payable = max(0, (supplier.payable or 0) - unpaid)

    # cascade="all, delete-orphan" on items + purchase_history handles child rows
    cdb.delete(invoice)
    cdb.commit()

    flash(f"Purchase {invoice_id} deleted and stock restored.")
    return redirect(url_for("purchase_invoice_list"))


@app.route("/purchase/new", methods=["GET", "POST"])
@login_required
@require_permission("purchase", "view", method_actions={'POST': 'create'})
def purchase_invoice_new():
    cdb = get_cdb()
    company_id = get_current_company()
    

    if request.method == "POST":
        supplier_id   = request.form.get("supplier_id")
        supplier_name = request.form.get("supplier_name", "").strip()

        # Auto-create supplier if typed manually and not in list
        if not supplier_id and supplier_name:
            existing = cdb.query(Supplier).filter_by(  # ← Client → Supplier
                company_id=company_id, name=supplier_name
            ).first()
            if existing:
                supplier_id = existing.id
            else:
                new_supplier = Supplier(          # ← Client → Supplier
                    company_id=company_id,
                    name=supplier_name,
                    status="Active",
                    created_at=today_ist()
                )
                cdb.add(new_supplier)
                cdb.flush()
                supplier_id = new_supplier.id

        invoice_number = request.form.get("invoice_number", "").strip()
        invoice_date   = request.form.get("invoice_date") or str(today_ist())
        notes          = request.form.get("notes", "").strip()
        is_interstate  = bool(request.form.get("is_interstate"))   # ← NEW

        # ── GST flag from company settings ──────────────────────────────────
        co = Company.query.filter_by(company_id=company_id).first()
        apply_gst = bool(co.is_gst_registered) if (co and hasattr(co, 'is_gst_registered')) else True

        # ── Line items ───────────────────────────────────────────────────────
        docket_nos    = request.form.getlist("docket_no[]")
        carrier_refs  = request.form.getlist("carrier_ref[]")
        party_names   = request.form.getlist("party_name[]")
        destinations  = request.form.getlist("destination[]")
        couriers      = request.form.getlist("courier_name[]")
        stock_item_ids= request.form.getlist("stock_item_id[]")
        item_names    = request.form.getlist("item_name[]")
        item_qtys     = request.form.getlist("item_qty[]")
        weights       = request.form.getlist("weight_kg[]")
        rates         = request.form.getlist("rate_per_kg[]")
        gst_percents  = request.form.getlist("gst_percent[]")
        other_charges = request.form.getlist("other_charges[]")
        resale_charges = request.form.getlist("resale_charges[]")

        line_items = []
        subtotal  = 0.0
        tax_total = 0.0

        for i in range(len(couriers)):
            courier = (couriers[i] if i < len(couriers) else "").strip()
            if not courier:
                continue

            qty    = float(item_qtys[i])    if i < len(item_qtys)    and item_qtys[i]    else 0
            weight = float(weights[i])      if i < len(weights)      and weights[i]      else 0
            rate   = float(rates[i])        if i < len(rates)        and rates[i]        else 0
            oc     = float(other_charges[i]) if (i < len(other_charges) and other_charges[i]) else 0.0
            resale = float(resale_charges[i]) if (i < len(resale_charges) and resale_charges[i]) else 0.0
            gst_pct = float(gst_percents[i]) if (apply_gst and i < len(gst_percents) and gst_percents[i]) else 0.0

            taxable    = round((weight * rate) + oc + resale, 2)
            gst_amount = round(taxable * gst_pct / 100, 2) if apply_gst else 0.0
            line_total = round(taxable + gst_amount, 2)

            # Split GST: IGST for interstate, CGST+SGST for intrastate
            if apply_gst and is_interstate:
                cgst_amt = 0.0
                sgst_amt = 0.0
                igst_amt = gst_amount
            else:
                cgst_amt = round(gst_amount / 2, 2)
                sgst_amt = gst_amount - cgst_amt   # avoids rounding gap
                igst_amt = 0.0

            subtotal  += taxable
            tax_total += gst_amount

            line_items.append({
                "docket_no":     docket_nos[i].strip()    if i < len(docket_nos)    else "",
                "carrier_ref":   carrier_refs[i].strip()  if i < len(carrier_refs)  else "",
                "party_name":    party_names[i].strip()   if i < len(party_names)   else "",
                "destination":   destinations[i].strip()  if i < len(destinations)  else "",
                "courier_name":  courier,
                "stock_item_id": int(stock_item_ids[i])   if i < len(stock_item_ids) and stock_item_ids[i] else None,
                "item_name":     item_names[i].strip()    if i < len(item_names)    else "",
                "qty":           qty,
                "weight_kg":     weight,
                "rate_per_kg":   rate,
                "other_charges": oc,
                "resale_charges": resale,
                "gst_percent":   gst_pct,
                "taxable_value": taxable,
                "cgst_amount":   cgst_amt,
                "sgst_amount":   sgst_amt,
                "igst_amount":   igst_amt,
                "total_amount":  line_total,
            })

        if not line_items:
            flash("Add at least one item row (courier + weight + rate).", "danger")
            return redirect(url_for("purchase_invoice_new"))

        subtotal    = round(subtotal, 2)
        tax_total   = round(tax_total, 2)
        grand_total = round(subtotal + tax_total, 2)

        invoice_id = _next_numbered_id(cdb, PurchaseInvoice.invoice_id, "PURCHASE-INV-" + datetime.now().strftime("%Y%m%d") + "-")

        purchase_inv = PurchaseInvoice(
            invoice_id=invoice_id,
            company_id=company_id,
            supplier_id=int(supplier_id) if supplier_id else None,
            supplier_name=supplier_name if not supplier_id else None,
            invoice_number=invoice_number,
            date=date.fromisoformat(invoice_date),
            subtotal=subtotal,
            tax_amount=tax_total,
            grand_total=grand_total,
            paid_amount=0,
            balance=grand_total,
            status="Pending",
            notes=notes,
            created_at=datetime.utcnow()
        )
        cdb.add(purchase_inv)
        cdb.flush()

        # Create item rows + DEDUCT stock
        stock_deductions = {}
        skipped_dockets = []
        for li in line_items:
            cdb.add(PurchaseInvoiceItem(
                purchase_invoice_id=purchase_inv.id,
                stock_item_id=li["stock_item_id"],
                description=li["item_name"] or li["courier_name"],
                quantity=li["qty"],
                unit="pcs",
                purchase_rate=li["rate_per_kg"],
                other_charges=li["other_charges"],
                resale_charges=li.get("resale_charges", 0),
                taxable_value=li["taxable_value"],
                gst_percent=li["gst_percent"],
                cgst_amount=li["cgst_amount"],
                sgst_amount=li["sgst_amount"],
                igst_amount=li["igst_amount"],
                total_amount=li["total_amount"],
                docket_no=li["docket_no"] or None,
                carrier_ref=li["carrier_ref"] or None,
                party_name=li["party_name"] or None,
                destination=li["destination"] or None,
                courier_name=li["courier_name"],
                weight_kg=li["weight_kg"],
                rate_per_kg=li["rate_per_kg"],
            ))

            # Don't deduct twice: if this docket's manifest row was already
            # marked Generated (stock pulled at manifest-generate time), skip
            # it here. NOTE: since a docket/AWB can now correspond to several
            # box rows (one per box), this still only checks "does at least
            # one row for this docket exist with status Generated" — same
            # coarse granularity as before, just re-pointed at row status
            # instead of the old manifest-wide flag.
            already_deducted_via_manifest = False
            if li["docket_no"]:
                already_deducted_via_manifest = cdb.query(ManifestEntry).join(
                    CompanyManifest, ManifestEntry.manifest_id == CompanyManifest.id
                ).filter(
                    ManifestEntry.docket_no == li["docket_no"],
                    CompanyManifest.company_id == company_id,
                    ManifestEntry.status == 'Generated',
                ).first() is not None

            if already_deducted_via_manifest:
                skipped_dockets.append(li["docket_no"])
            elif li["stock_item_id"] and li["qty"] > 0:
                stock_deductions[li["stock_item_id"]] = stock_deductions.get(li["stock_item_id"], 0) + li["qty"]

        for sid, qty in stock_deductions.items():
            stock = cdb.query(StockItem).filter_by(id=sid, company_id=company_id).first()
            if stock:
                stock.quantity = (stock.quantity or 0) - qty
                stock.last_updated = today_ist()
                cdb.add(StockPurchaseHistory(
                    stock_item_id=sid,
                    purchase_invoice_id=purchase_inv.id,
                    quantity=qty,
                    purchase_rate=0,
                    movement_type="OUT",
                    purchase_date=date.fromisoformat(invoice_date),
                    reference=invoice_id,
                ))

        if skipped_dockets:
            flash(
                f"Note: stock for docket(s) {', '.join(skipped_dockets)} was already deducted when their "
                f"manifest was generated — not deducted again here.",
                "info",
            )

        # Update supplier pending payable
        if supplier_id:
            supplier = cdb.get(Supplier, int(supplier_id))
            if supplier:
                supplier.payable = (supplier.payable or 0) + grand_total

        cdb.commit()

        # NOTE: Manifest auto-creation used to happen here (on Purchase Bill
        # save). It has moved to invoice_customer_save() — a manifest now
        # represents "shipments booked today", not "shipments billed by the
        # courier today". See that function for the new logic.

        # Handle file upload
        if "invoice_file" in request.files:
            file = request.files["invoice_file"]
            if file and file.filename and allowed_file(file.filename):
                filename = secure_filename(f"{invoice_id}_{file.filename}")
                filepath = os.path.join(app.config["UPLOAD_FOLDER"], filename)
                file.save(filepath)
                purchase_inv.file_path = filepath
                cdb.commit()

        flash(f"Purchase invoice {invoice_id} saved successfully! Stock deducted and manifest(s) created.")
        return redirect(url_for("purchase_invoice_list"))

    stock_items = cdb.query(StockItem).filter_by(company_id=company_id).order_by(StockItem.name).all()
    suppliers = cdb.query(Supplier).filter_by(company_id=company_id, status="Active").order_by(Supplier.name).all()
    purchase_price_lists = cdb.query(PriceList).filter_by(
        company_id=company_id,
        is_active=True,
        list_type='purchase'
    ).all()
    return render_template("purchase_new.html",
                           suppliers=suppliers,
                           stock_items=stock_items,
                           courier_options=COURIER_OPTIONS,
                           item_type_options=ITEM_TYPE_OPTIONS,
                           today=str(today_ist()),
                           purchase_price_lists=purchase_price_lists)


# ── Add these two routes to app.py (after purchase_invoice_new, before purchase_invoice_view) ──

@app.route("/purchase/edit/<invoice_id>", methods=["GET", "POST"])
@login_required
@require_permission("purchase", "view", method_actions={'POST': 'edit'})
def purchase_invoice_edit(invoice_id):
    cdb        = get_cdb()
    company_id = get_current_company()

    invoice = cdb.query(PurchaseInvoice).filter_by(
        invoice_id=invoice_id, company_id=company_id
    ).first()
    if not invoice:
        abort(404)

    if request.method == "POST":
        # ── Bill-level fields ────────────────────────────────────────────────
        invoice.invoice_number = request.form.get("invoice_number", "").strip() or invoice.invoice_number
        inv_date = request.form.get("invoice_date")
        if inv_date:
            invoice.date = date.fromisoformat(inv_date)
        invoice.notes = request.form.get("notes", "").strip()

        is_interstate = bool(request.form.get("is_interstate"))

        # ── GST flag ─────────────────────────────────────────────────────────
        co = Company.query.filter_by(company_id=company_id).first()
        apply_gst = bool(co.is_gst_registered) if (co and hasattr(co, 'is_gst_registered')) else True

        # ── Line items ────────────────────────────────────────────────────────
        item_ids    = request.form.getlist("item_id[]")
        weights     = request.form.getlist("weight_kg[]")
        rates       = request.form.getlist("rate_per_kg[]")
        other_charges = request.form.getlist("other_charges[]")
        resale_charges = request.form.getlist("resale_charges[]")
        gst_percents= request.form.getlist("gst_percent[]")

        subtotal  = 0.0
        tax_total = 0.0
        old_grand = invoice.grand_total or 0.0

        for i, item_id in enumerate(item_ids):
            item = cdb.query(PurchaseInvoiceItem).filter_by(
                id=int(item_id), purchase_invoice_id=invoice.id
            ).first()
            if not item:
                continue

            weight  = float(weights[i])      if i < len(weights)      and weights[i]      else item.weight_kg or 0
            rate    = float(rates[i])        if i < len(rates)        and rates[i]        else item.rate_per_kg or 0
            oc      = float(other_charges[i]) if (i < len(other_charges) and other_charges[i]) else (item.other_charges or 0)
            resale  = float(resale_charges[i]) if (i < len(resale_charges) and resale_charges[i]) else (item.resale_charges or 0)
            gst_pct = float(gst_percents[i]) if (apply_gst and i < len(gst_percents) and gst_percents[i]) else (item.gst_percent or 0)

            taxable    = round((weight * rate) + oc + resale, 2)
            gst_amount = round(taxable * gst_pct / 100, 2) if apply_gst else 0.0
            line_total = round(taxable + gst_amount, 2)

            if apply_gst and is_interstate:
                cgst_amt, sgst_amt, igst_amt = 0.0, 0.0, gst_amount
            else:
                cgst_amt = round(gst_amount / 2, 2)
                sgst_amt = gst_amount - cgst_amt
                igst_amt = 0.0

            item.weight_kg      = weight
            item.rate_per_kg    = rate
            item.purchase_rate  = rate
            item.other_charges  = oc
            item.resale_charges = resale
            item.taxable_value  = taxable
            item.gst_percent    = gst_pct
            item.cgst_amount    = cgst_amt
            item.sgst_amount    = sgst_amt
            item.igst_amount    = igst_amt
            item.total_amount   = line_total

            subtotal  += taxable
            tax_total += gst_amount

        new_grand = round(subtotal + tax_total, 2)

        invoice.subtotal   = round(subtotal, 2)
        invoice.tax_amount = round(tax_total, 2)
        invoice.grand_total= new_grand

        # Adjust balance: keep paid_amount fixed, recompute balance
        invoice.balance = round(max(0, new_grand - (invoice.paid_amount or 0)), 2)
        if invoice.balance <= 0:
            invoice.status = "Paid"
        elif (invoice.paid_amount or 0) > 0:
            invoice.status = "Partial"
        else:
            invoice.status = "Pending"

        # Adjust supplier payable for the difference
        diff = new_grand - old_grand
        if diff != 0 and invoice.supplier_id:
            supplier = cdb.get(Supplier, invoice.supplier_id)
            if supplier:
                supplier.payable = max(0, (supplier.payable or 0) + diff)

        cdb.commit()
        flash(f"Purchase bill {invoice_id} updated successfully.", "success")
        return redirect(url_for("purchase_invoice_view", invoice_id=invoice_id))

    return render_template("purchase_edit.html", invoice=invoice, today=str(today_ist()))


@app.route("/purchase/view/<invoice_id>")
@login_required
@require_permission("purchase", "view")
def purchase_invoice_view(invoice_id):
    cdb = get_cdb()
    company_id = get_current_company()
    invoice = cdb.query(PurchaseInvoice).filter_by(invoice_id=invoice_id, company_id=company_id).first()
    if not invoice:
        abort(404)

    bank_accounts = cdb.query(BankAccount).filter_by(company_id=company_id, status='Active').all()
    return render_template("purchase_view.html", invoice=invoice, bank_accounts=bank_accounts, today=str(today_ist()))

@app.route("/purchase/pay/<int:pk>", methods=["POST"])
@login_required
@require_permission("purchase", "edit")
def purchase_make_payment(pk):
    cdb = get_cdb()
    company_id = get_current_company()
    invoice = _first_or_404(cdb.query(PurchaseInvoice).filter_by(id=pk, company_id=company_id).first())

    amount     = float(request.form.get("amount", 0))
    pay_mode   = request.form.get("pay_mode", "Cash")
    narration  = request.form.get("narration", "")
    pay_date_s = request.form.get("pay_date")
    pay_date   = date.fromisoformat(pay_date_s) if pay_date_s else today_ist()
    bank_account_id = request.form.get("bank_account_id", type=int)

    if amount <= 0:
        flash("Invalid payment amount.")
        return redirect(url_for("purchase_invoice_view", invoice_id=invoice.invoice_id))

    if amount > (invoice.balance or 0):
        amount = invoice.balance or 0

    # Non-cash modes need a real bank account to post the debit against —
    # same requirement /payments/save enforces, so this route stays
    # consistent with it instead of silently going nowhere.
    bank_account = None
    if pay_mode.lower() != "cash":
        if not bank_account_id:
            flash("Please select a bank account for non-cash payments.", "error")
            return redirect(url_for("purchase_invoice_view", invoice_id=invoice.invoice_id))
        bank_account = cdb.query(BankAccount).filter_by(
            id=bank_account_id, company_id=company_id, status='Active'
        ).first()
        if not bank_account:
            flash("Selected bank account not found or inactive.", "error")
            return redirect(url_for("purchase_invoice_view", invoice_id=invoice.invoice_id))

    invoice.paid_amount = (invoice.paid_amount or 0) + amount
    invoice.balance     = max(0, (invoice.balance or 0) - amount)

    if invoice.balance <= 0:
        invoice.status = "Paid"
    elif invoice.paid_amount > 0:
        invoice.status = "Partial"

    if invoice.supplier:
        invoice.supplier.payable = max(0, (invoice.supplier.payable or 0) - amount)

    # Kept for backward-compat / audit trail — NOT what creditors, the
    # creditor statement, or the Payments page read from. Those all read
    # CashTransaction/BankTransaction, so the write below is the one that
    # actually makes this payment visible anywhere else in the app.
    pmt = PurchasePayment(
        company_id  = company_id,
        invoice_id  = invoice.id,
        supplier_id = invoice.supplier_id,
        date        = pay_date,
        amount      = amount,
        pay_mode    = pay_mode,
        narration   = narration,
        created_by  = session.get("user", {}).get("user_id")
    )
    cdb.add(pmt)

    supplier_name = invoice.supplier.name if invoice.supplier else ""
    txn_reference = invoice.invoice_number or invoice.invoice_id
    desc = f"Payment made for purchase invoice {txn_reference}"
    if narration:
        desc += f" - {narration}"

    if pay_mode.lower() == "cash":
        cash_txn = CashTransaction(
            company_id=company_id,
            type="expense",
            date=pay_date,
            category="Payment",
            description=desc,
            amount=amount,
            reference=txn_reference,
            notes=f"Payment of ₹{amount:,.2f} to supplier via Cash",
            party_name=supplier_name,
            created_by=get_current_user().get('email'),
            applied_ref_type="purchase_invoice",
            applied_ref_id=invoice.id,
        )
        cdb.add(cash_txn)
    else:
        bank_txn = BankTransaction(
            bank_account_id=bank_account.id,
            company_id=company_id,
            type="debit",
            date=pay_date,
            description=desc,
            amount=amount,
            reference=txn_reference,
            transaction_mode=pay_mode.title(),
            notes=narration,
            party_name=supplier_name,
            created_by=get_current_user().get('email'),
            applied_ref_type="purchase_invoice",
            applied_ref_id=invoice.id,
        )
        cdb.add(bank_txn)
        bank_account.balance -= amount

    cdb.commit()
    flash(f"Payment of ₹{amount:,.2f} via {pay_mode} recorded. {narration}")
    return redirect(url_for("purchase_invoice_view", invoice_id=invoice.invoice_id))

# ─────────────────────────────────────────────────────────────────────────────
# ── Invoices ──────────────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────

@app.route("/booking/list")
@login_required
@require_permission("invoices", "view")
def invoice_list():
    cdb = get_cdb()
    company_id    = get_current_company()
    filter_status = request.args.get("status", "All")
    filter_btype  = request.args.get("btype", "All")   # All | cash | credit

    # Map template tab names -> DB status values
    status_map = {
        "paid":    "Paid",
        "partial": "Partial",
    }

    query = cdb.query(Invoice).filter_by(company_id=company_id)
    if filter_status == "pending":
        query = query.filter(Invoice.status.notin_(["Paid", "Partial", "Draft", "Void"]))
    elif filter_status == "draft":
        query = query.filter_by(status="Draft")
    elif filter_status == "void":
        query = query.filter_by(status="Void")
    elif filter_status != "All":
        db_status = status_map.get(filter_status)
        if db_status:
            query = query.filter_by(status=db_status)

    raw_invoices = query.order_by(Invoice.created_at.desc()).all()

    invoices = []
    for inv in raw_invoices:
        # For cash/walk-in bookings there's no client_obj, so this used to
        # fall back straight to inv.contact_person — but contact_person is
        # the "Name (person at company)" field on the booking form, NOT the
        # "Company / Customer Name" field. That company name only lives in
        # terms.shipper_name (parsed below), so parse it first and prefer
        # it; only fall back to the contact person if shipper_name itself
        # was left blank.
        _meta_for_name = {}
        if inv.terms:
            try:
                _meta_for_name = json.loads(inv.terms)
            except (ValueError, TypeError):
                _meta_for_name = {}

        if inv.client_obj:
            customer_name = inv.client_obj.name
        elif _meta_for_name.get("shipper_name"):
            customer_name = _meta_for_name.get("shipper_name")
        elif inv.contact_person:
            customer_name = inv.contact_person
        else:
            customer_name = "—"

        # ── Get resale charges ──────────────────────────────────────────────
        resale_charges = getattr(inv, 'resale_charges', 0) or 0
        resale_gst = resale_charges * 0.18  # 18% GST
        resale_total = resale_charges + resale_gst
        
        # ── Total includes resale ────────────────────────────────────────────
        total = inv.grand_total or 0.0

        if inv.status == "Paid":
            paid       = total
            balance    = 0.0
            tab_status = "paid"
        elif inv.status == "Partial":
            paid       = inv.paid_amount if hasattr(inv, "paid_amount") and inv.paid_amount else (inv.subtotal or 0.0)
            balance    = inv.balance if hasattr(inv, "balance") and inv.balance is not None else total - paid
            tab_status = "partial"
        elif inv.status == "Draft":
            paid       = 0.0
            balance    = total
            tab_status = "draft"
        elif inv.status == "Void":
            paid       = 0.0
            balance    = 0.0
            tab_status = "void"
        else:
            paid       = 0.0
            balance    = total
            tab_status = "pending"

        is_draft = inv.status == "Draft"
        is_void  = inv.status == "Void"

        # Unpack shipment metadata stored as JSON in inv.terms
        meta = {}
        if inv.terms:
            try:
                meta = json.loads(inv.terms)
            except (ValueError, TypeError):
                meta = {}

        # Determine if this is a customer/shipment invoice (has AWB docket)
        docket_no = meta.get("docket_no", "")
        is_shipment = bool(docket_no) or inv.invoice_id.startswith("CUST-")

        # ── NEW: Check if this invoice has a linked performa invoice ──────────
        # Query for an Estimate linked to this invoice
        linked_est = cdb.query(Estimate).filter(
            Estimate.company_id == company_id,
            Estimate.terms.like(f'%"linked_invoice_id": "{inv.invoice_id}"%')
        ).first()
        has_performa = linked_est is not None

        invoices.append({
            "id":             inv.invoice_id,
            "customer_name":  customer_name,
            "date":           inv.date,
            "bill_type":      "credit",
            "booking_type":   meta.get("booking_type", "credit"),
            "total":          total,  # ← Now includes resale
            "paid":           paid,
            "balance":        balance,
            "status":         tab_status,
            "is_draft":       is_draft,
            "is_void":        is_void,
            "completion_status": "Draft" if is_draft else "Completed",
            # ── NEW: Perfoma invoice status ───────────────────────────────────
            "has_performa":   has_performa,
            # Shipment-specific fields unpacked from JSON terms
            "docket_no":      docket_no,
            "receiver_name":  meta.get("receiver_name", ""),
            "destination":    meta.get("destination", ""),
            "carrier":        meta.get("carrier", ""),
            "carrier_ref": meta.get("carrier_ref", ""),
            "tracking_number": meta.get("tracking_number", ""),
            # Manual courier-tracking status. Not set explicitly on creation —
            # defaults to "Booked" for any finalized (non-draft, non-void)
            # invoice until someone picks a different stage from the dropdown,
            # so every existing booking gets the right default with no backfill.
            "tracking_status": meta.get("tracking_status") or ("Booked" if not is_draft and not is_void else ""),
            "shipment_type":  meta.get("shipment_type", ""),
            "mode":           meta.get("mode", ""),
            "is_shipment":    is_shipment,
            # Resale fields
            "has_resale":     getattr(inv, 'has_resale', False),
            "resale_charges": resale_charges,
            "resale_reason":  getattr(inv, 'resale_reason', ''),
            "resale_date":    getattr(inv, 'resale_date', None),
            "resale_total":   resale_total,  # ← NEW: total including GST
        })

    if filter_btype in ("cash", "credit"):
        invoices = [inv for inv in invoices if inv["booking_type"] == filter_btype]

    search_q = request.args.get("q", "").strip()
    if search_q:
        needle = search_q.lower()
        invoices = [
            inv for inv in invoices
            if needle in (inv["customer_name"] or "").lower()
            or needle in (inv["docket_no"] or "").lower()
            or needle in (inv["carrier_ref"] or "").lower()
            or needle in (inv["tracking_number"] or "").lower()
        ]

    from_date_str = request.args.get("from_date", "").strip()
    to_date_str   = request.args.get("to_date", "").strip()

    def _parse_date(s):
        try:
            return datetime.strptime(s, "%Y-%m-%d").date()
        except (ValueError, TypeError):
            return None

    from_date = _parse_date(from_date_str)
    to_date   = _parse_date(to_date_str)

    if from_date or to_date:
        def _inv_date(inv):
            d = inv["date"]
            if isinstance(d, datetime):
                d = d.date()
            return d

        if from_date:
            invoices = [inv for inv in invoices if _inv_date(inv) and _inv_date(inv) >= from_date]
        if to_date:
            invoices = [inv for inv in invoices if _inv_date(inv) and _inv_date(inv) <= to_date]

    return render_template("booking_list.html",
                           invoices=invoices,
                           current_status=filter_status,
                           current_btype=filter_btype,
                           current_from_date=from_date_str,
                           current_to_date=to_date_str)


@app.route("/booking/list/update-tracking/<invoice_id>", methods=["POST"])
@login_required
@require_permission("invoices", "edit")
def invoice_list_update_tracking(invoice_id):
    """AJAX endpoint used from the Bookings list page: lets a user update just
    Tracking Number / Carrier Ref No. inline, without opening the full booking
    edit form. Both fields live inside the JSON blob in Invoice.terms (same
    place invoice_customer_update reads/writes them), so this is a scoped
    version of that save — it only touches these two keys, leaves everything
    else in terms untouched, and mirrors the same downstream syncs
    (PurchaseInvoiceItem.carrier_ref, WhatsApp tracking-update notification)
    so behaviour matches editing from the booking form."""
    cdb = get_cdb()
    company_id = get_current_company()
    invoice = cdb.query(Invoice).filter_by(invoice_id=invoice_id, company_id=company_id).first()
    if not invoice:
        return jsonify({"ok": False, "error": "Booking not found."}), 404

    try:
        meta = json.loads(invoice.terms) if invoice.terms else {}
    except (ValueError, TypeError):
        meta = {}

    old_carrier_ref = (meta.get("carrier_ref", "") or "").strip()
    old_tracking_number = (meta.get("tracking_number", "") or "").strip()

    new_carrier_ref = (request.form.get("carrier_ref") or "").strip()
    new_tracking_number = (request.form.get("tracking_number") or "").strip()

    meta["carrier_ref"] = new_carrier_ref
    meta["tracking_number"] = new_tracking_number
    invoice.terms = json.dumps(meta)
    cdb.commit()

    docket_no = meta.get("docket_no", "")

    # ── Carrier ref sync onto linked purchase-side records (mirrors the
    # same block in invoice_customer_update) ────────────────────────────────
    if new_carrier_ref and new_carrier_ref != old_carrier_ref:
        try:
            updated_count = cdb.query(PurchaseInvoiceItem).filter_by(
                source_invoice_id=invoice.id
            ).update({"carrier_ref": new_carrier_ref})

            if not updated_count and docket_no:
                fallback_items = cdb.query(PurchaseInvoiceItem).filter_by(
                    docket_no=docket_no
                ).all()
                for pi_item in fallback_items:
                    pi_item.carrier_ref = new_carrier_ref
                    if not pi_item.source_invoice_id:
                        pi_item.source_invoice_id = invoice.id

            cdb.commit()
        except Exception as e:
            cdb.rollback()
            print(f"[purchase-sync] could not update carrier_ref on linked purchase item for {invoice.invoice_id}: {e}")

    # ── WhatsApp tracking-update notification (mirrors invoice_customer_update) ──
    if new_tracking_number and new_tracking_number != old_tracking_number:
        try:
            from tasks import send_tracking_update_notification_async
            send_tracking_update_notification_async(
                company_id=company_id,
                invoice_id=invoice.invoice_id,
                carrier=meta.get("carrier", ""),
                tracking_number=new_tracking_number,
            )
        except Exception as e:
            print(f"[whatsapp] could not queue tracking-update notification for {invoice.invoice_id}: {e}")

    return jsonify({
        "ok": True,
        "carrier_ref": new_carrier_ref,
        "tracking_number": new_tracking_number,
    })


TRACKING_STATUS_STAGES = ["Booked", "In Transit", "Out for Delivery", "Delivered"]


@app.route("/booking/list/update-tracking-status/<invoice_id>", methods=["POST"])
@login_required
@require_permission("invoices", "edit")
def invoice_list_update_tracking_status(invoice_id):
    """AJAX endpoint for the Bookings list page's Tracking Status dropdown.
    Lets a user manually set the courier stage (Booked / In Transit / Out for
    Delivery / Delivered) after checking the courier's own site — there's no
    live courier API integration, so this is a manual log, same pattern as
    update-tracking for carrier_ref/tracking_number. Stored as
    meta['tracking_status'] inside Invoice.terms."""
    cdb = get_cdb()
    company_id = get_current_company()
    invoice = cdb.query(Invoice).filter_by(invoice_id=invoice_id, company_id=company_id).first()
    if not invoice:
        return jsonify({"ok": False, "error": "Booking not found."}), 404

    new_status = (request.form.get("tracking_status") or "").strip()
    if new_status not in TRACKING_STATUS_STAGES:
        return jsonify({"ok": False, "error": "Invalid tracking status."}), 400

    try:
        meta = json.loads(invoice.terms) if invoice.terms else {}
    except (ValueError, TypeError):
        meta = {}

    old_status = meta.get("tracking_status", "")
    meta["tracking_status"] = new_status
    invoice.terms = json.dumps(meta)
    cdb.commit()

    # WhatsApp tracking-stage notification, same fire-and-forget pattern used
    # for tracking_number updates above.
    if new_status != old_status:
        try:
            from tasks import send_tracking_update_notification_async
            send_tracking_update_notification_async(
                company_id=company_id,
                invoice_id=invoice.invoice_id,
                carrier=meta.get("carrier", ""),
                tracking_number=meta.get("tracking_number", ""),
                tracking_status=new_status,
            )
        except TypeError:
            # tasks.send_tracking_update_notification_async may not accept a
            # tracking_status kwarg yet — don't let that break the save.
            pass
        except Exception as e:
            print(f"[whatsapp] could not queue tracking-status notification for {invoice.invoice_id}: {e}")

    return jsonify({"ok": True, "tracking_status": new_status})


@app.route("/booking/hard-delete/<invoice_id>", methods=["POST"])
@login_required
@owner_required
def invoice_hard_delete(invoice_id):
    """
    TEMPORARY cleanup tool — actually deletes the Invoice row and its
    InvoiceItems, unlike invoice_void() which only flags status="Void" and
    keeps everything for GST audit. Use this only for bookings that were
    never real (duplicate/bug-generated) and were never actually reported
    or paid against. Reverses the same downstream effects as void
    (stock/manifest/purchase-line/ledger) before removing the row.
    Remove this route once cleanup is done — it is not meant to stay
    reachable long-term; that's what Void is for.
    """
    company_id = get_current_company()
    if not company_id:
        return redirect(url_for("login"))
    cdb = get_customer_session(company_id)

    inv = cdb.query(Invoice).filter_by(invoice_id=invoice_id, company_id=company_id).first()
    if not inv:
        flash("Invoice not found.", "danger")
        return redirect(url_for("invoice_list"))

    # Refuse if a real payment (cheque) is tied to this invoice — deleting
    # the invoice would orphan that cheque's invoice_id FK. Void it instead
    # if this booking was ever actually paid against.
    cheque_hit = cdb.query(Cheque).filter_by(invoice_id=inv.id).first()
    if cheque_hit:
        flash(
            f"{invoice_id} has a cheque record linked to it — refusing to hard-delete. "
            f"Use Void instead if this booking was real.", "danger"
        )
        return redirect(url_for("invoice_view", invoice_id=invoice_id))

    docket_no = _get_awb(inv)

    try:
        # 1) Reverse client outstanding
        if inv.balance and inv.client_id:
            client = cdb.query(Client).filter_by(id=inv.client_id, company_id=company_id).first()
            if client:
                client.pending = max(0, (client.pending or 0) - inv.balance)

        # 2) Reverse manifest entries + add back stock they'd deducted
        if docket_no:
            entries = cdb.query(ManifestEntry).join(
                CompanyManifest, ManifestEntry.manifest_id == CompanyManifest.id
            ).filter(
                ManifestEntry.docket_no == docket_no,
                CompanyManifest.company_id == company_id,
            ).all()
            touched_manifest_ids = set()
            for entry in entries:
                if entry.status == "Generated" and entry.stock_item_id and entry.boxes:
                    stock = cdb.query(StockItem).filter_by(
                        id=entry.stock_item_id, company_id=company_id
                    ).first()
                    if stock:
                        stock.quantity = (stock.quantity or 0) + entry.boxes
                        stock.last_updated = today_ist()
                        cdb.add(StockPurchaseHistory(
                            stock_item_id=stock.id, purchase_invoice_id=None,
                            quantity=entry.boxes, purchase_rate=0, movement_type="IN",
                            purchase_date=today_ist(),
                            reference=f"HARDDEL-MANIFEST-{invoice_id}", awb_no=docket_no,
                        ))
                if entry.manifest_id:
                    touched_manifest_ids.add(entry.manifest_id)
                cdb.delete(entry)
            cdb.flush()
            for manifest_id in touched_manifest_ids:
                parent_manifest = cdb.query(CompanyManifest).filter_by(id=manifest_id).first()
                if parent_manifest:
                    remaining = cdb.query(ManifestEntry).filter_by(manifest_id=parent_manifest.id).count()
                    if remaining == 0:
                        cdb.delete(parent_manifest)
                    else:
                        _recompute_manifest_status(parent_manifest)

        # 3) Reverse the auto-generated purchase invoice line
        pi_item = cdb.query(PurchaseInvoiceItem).filter_by(source_invoice_id=inv.id).first()
        if pi_item:
            parent_pi = cdb.query(PurchaseInvoice).filter_by(id=pi_item.purchase_invoice_id).first()
            line_total = pi_item.total_amount or 0
            if pi_item.docket_no:
                booking_invoice_items = cdb.query(InvoiceItem).filter(
                    InvoiceItem.invoice_id == inv.id,
                    InvoiceItem.stock_item_id.isnot(None)
                ).all()
                for inv_item in booking_invoice_items:
                    stock = cdb.query(StockItem).filter_by(
                        id=inv_item.stock_item_id, company_id=company_id
                    ).first()
                    if stock:
                        stock.quantity = (stock.quantity or 0) + (inv_item.qty or 0)
                        stock.last_updated = today_ist()
                        cdb.add(StockPurchaseHistory(
                            stock_item_id=stock.id, purchase_invoice_id=None,
                            quantity=inv_item.qty or 0, purchase_rate=0, movement_type="IN",
                            purchase_date=today_ist(),
                            reference=f"HARDDEL-PURCHASE-{invoice_id}", awb_no=docket_no,
                        ))
            if parent_pi:
                supplier = cdb.query(Supplier).filter_by(
                    id=parent_pi.supplier_id, company_id=company_id
                ).first()
                if supplier:
                    supplier.payable = max(0, (supplier.payable or 0) - line_total)
                parent_pi.subtotal = max(0, (parent_pi.subtotal or 0) - (pi_item.taxable_value or 0))
                parent_pi.grand_total = max(0, (parent_pi.grand_total or 0) - line_total)
                parent_pi.balance = max(0, (parent_pi.balance or 0) - line_total)
            cdb.delete(pi_item)
            cdb.flush()
            if parent_pi:
                remaining_items = cdb.query(PurchaseInvoiceItem).filter_by(
                    purchase_invoice_id=parent_pi.id
                ).count()
                if remaining_items == 0:
                    cdb.delete(parent_pi)

        # 4) Remove the stock this booking originally added
        stock_history_entries = cdb.query(StockPurchaseHistory).filter(
            StockPurchaseHistory.awb_no == docket_no,
            StockPurchaseHistory.purchase_invoice_id.is_(None)
        ).all()
        for hist in stock_history_entries:
            stock = cdb.query(StockItem).filter_by(
                id=hist.stock_item_id, company_id=company_id
            ).first()
            if stock:
                stock.quantity = max(0, (stock.quantity or 0) - (hist.quantity or 0))
                stock.last_updated = today_ist()
            cdb.delete(hist)

        # 5) Delete the linked proforma invoice (Estimate), if any
        linked_est = cdb.query(Estimate).filter_by(company_id=company_id).filter(
            Estimate.terms.like(f'%"linked_invoice_id": "{invoice_id}"%')
        ).first()
        if linked_est:
            cdb.query(EstimateItem).filter_by(estimate_id=linked_est.id).delete()
            cdb.delete(linked_est)

        # ── Log the deletion before the row disappears ────────────────────
        try:
            meta = json.loads(inv.terms) if inv.terms else {}
        except Exception:
            meta = {}
        client_name = None
        if inv.client_id:
            c = cdb.query(Client).filter_by(id=inv.client_id, company_id=company_id).first()
            client_name = c.name if c else None
        cdb.add(DeletedInvoiceLog(
            company_id=company_id,
            invoice_id=invoice_id,
            awb_no=docket_no,
            client_name=client_name,
            shipper_name=meta.get("shipper_name"),
            grand_total=inv.grand_total,
            deleted_by=session.get('user', {}).get('email', ''),
            deleted_at=datetime.utcnow(),
            reason=request.form.get("reason", "").strip() or None,
        ))
        
        # 6) Actually delete the invoice — this is the part Void doesn't do.
        cdb.query(InvoiceItem).filter_by(invoice_id=inv.id).delete()
        cdb.delete(inv)

        cdb.commit()
        flash(f"Booking {invoice_id} permanently deleted, stock/manifest/ledger reversed.", "success")
    except Exception as e:
        cdb.rollback()
        flash(f"Delete failed: {e}", "danger")

    return redirect(url_for("invoice_list"))

@app.route("/booking/deleted-log")
@login_required
@owner_required
def invoice_deleted_log():
    cdb = get_customer_session(get_current_company())
    rows = cdb.query(DeletedInvoiceLog).filter_by(
        company_id=get_current_company()
    ).order_by(DeletedInvoiceLog.deleted_at.desc()).all()
    return render_template("booking_deleted_log.html", rows=rows)

@app.route("/booking/new", methods=["GET", "POST"])
@login_required
@require_permission("invoices", "view", method_actions={'POST': 'create'})
def invoice_new():
    cdb = get_cdb()
    company_id = get_current_company()
    clients = cdb.query(Client).filter(
        Client.company_id == company_id,
        ~Client.client_type.in_(["Supplier", "Both", "Cash-Only"])  
    ).all()

    price_lists = cdb.query(PriceList).filter_by(
        company_id=company_id, 
        is_active=True,
        list_type='sales'
    ).all()

    suppliers = cdb.query(Supplier).filter_by(company_id=company_id, status="Active").order_by(Supplier.name).all()

    # Check if we're editing an existing invoice
    edit_id = request.args.get("edit")
    existing_invoice = None
    if edit_id:
        existing_invoice = cdb.query(Invoice).filter_by(invoice_id=edit_id, company_id=company_id).first()
        if not existing_invoice:
            flash("Invoice not found")
            return redirect(url_for("invoice_list"))

    if request.method == "POST":
        # Handle customer invoice POST (save/update)
        # This is for the customer invoice form
        client_id_raw = request.form.get("customer_id")
        client_id = int(client_id_raw) if client_id_raw else None
        invoice_date = request.form.get("invoice_date") or str(today_ist())
        docket_no = request.form.get("docket_no", "")
        action = request.form.get("action", "final")

        # ── AWB/docket uniqueness — checked here at save time, not just
        # when _next_awb_number() suggested a value back at form-render
        # time. On edit, exclude the invoice being edited (it's allowed to
        # keep its own docket_no). ──
        _edit_invoice_id_for_dupe_check = request.form.get("edit_invoice_id")
        dupe_invoice_id = _docket_no_in_use(
            cdb, company_id, docket_no,
            exclude_invoice_id=_edit_invoice_id_for_dupe_check
        )
        if dupe_invoice_id:
            old_docket_no = docket_no
            docket_no = _next_awb_number(company_id)
            flash(f"AWB {old_docket_no} was already used on invoice {dupe_invoice_id} — "
                  f"this invoice was automatically assigned {docket_no} instead.")

        # Charges & totals
        freight = float(request.form.get("freight_amount", 0) or 0)
        freight_weight = float(request.form.get("freight_weight", 0) or 0)
        freight_rate = float(request.form.get("freight_rate_per_kg", 0) or 0)
        # Rounded rate-card slab weight from the rate lookup — persisted so a
        # later edit-load recomputes freight against the right weight instead
        # of falling back to the actual/display weight (see booking.html fix).
        freight_billing_weight = float(request.form.get("freight_billing_weight", 0) or 0) or freight_weight
        fuel = float(request.form.get("fuel_surcharge", 0) or 0)
        other = float(request.form.get("other_charges", 0) or 0)
        base = freight + fuel + other
        gst = round(base * 0.18, 2)
        grand_total = round(base + gst, 2)
        amount_paid = float(request.form.get("amount_paid", 0) or 0)
        balance = round(grand_total - amount_paid, 2)

        # Payment info
        payment_mode = request.form.get("payment_mode", "cash")
        upi_app = request.form.get("upi_app", "")
        upi_ref = request.form.get("upi_ref", "")
        cheque_no = request.form.get("cheque_no", "")
        cheque_date = request.form.get("cheque_date", "")
        cheque_bank = request.form.get("cheque_bank", "")

        # Status
        if action == "draft":
            status = "Draft"
        elif balance <= 0:
            status = "Paid"
        elif amount_paid > 0:
            status = "Partial"
        else:
            status = "Pending"

        notes = request.form.get("notes", "")
        
        # Process Packages
        pkg_names = request.form.getlist("pkg_name[]")
        pkg_types = request.form.getlist("pkg_type[]")
        pkg_units = request.form.getlist("pkg_unit[]")
        pkg_qtys = request.form.getlist("pkg_qty[]")
        pkg_l = request.form.getlist("pkg_l[]")
        pkg_w = request.form.getlist("pkg_w[]")
        pkg_h = request.form.getlist("pkg_h[]")
        pkg_wt = request.form.getlist("pkg_wt[]")
        pkg_division = request.form.getlist("pkg_division[]")
        pkg_discount = request.form.getlist("pkg_discount[]")
        pkg_discwt = request.form.getlist("pkg_discwt[]")
        pkg_volwt = request.form.getlist("pkg_volwt[]")
        pkg_chgwt = request.form.getlist("pkg_chgwt[]")
        pkg_rates = request.form.getlist("pkg_rate[]")
        
        packages_data = []
        for i in range(len(pkg_names)):
            if pkg_names[i] and pkg_names[i].strip():
                packages_data.append({
                    "name": pkg_names[i],
                    "type": pkg_types[i] if i < len(pkg_types) else "",
                    "unit": pkg_units[i] if i < len(pkg_units) else "cm",
                    "qty": float(pkg_qtys[i] or 1) if pkg_qtys[i] else 1,
                    "length": float(pkg_l[i] or 0) if i < len(pkg_l) else 0,
                    "width": float(pkg_w[i] or 0) if i < len(pkg_w) else 0,
                    "height": float(pkg_h[i] or 0) if i < len(pkg_h) else 0,
                    "weight": float(pkg_wt[i] or 0) if i < len(pkg_wt) else 0,
                    "division": float(pkg_division[i] or 5000) if i < len(pkg_division) and pkg_division[i] else 5000,
                    "discount": float(pkg_discount[i] or 0) if i < len(pkg_discount) and pkg_discount[i] else 0,
                    "discount_wt": float(pkg_discwt[i] or 0) if i < len(pkg_discwt) and pkg_discwt[i] else 0,
                    "vol_weight": float(pkg_volwt[i] or 0) if i < len(pkg_volwt) and pkg_volwt[i] else 0,
                    "chg_weight": float(pkg_chgwt[i] or 0) if i < len(pkg_chgwt) and pkg_chgwt[i] else 0,
                    "rate": float(pkg_rates[i] or 0) if i < len(pkg_rates) else 0,
                })

        # Additional (extra) receivers — a booking can have more than one
        # consignee beyond the main Receiver/Consignee fields above.
        add_recv_names    = request.form.getlist("additional_receiver_name[]")
        add_recv_companies = request.form.getlist("additional_receiver_company[]")
        add_recv_phones   = request.form.getlist("additional_receiver_phone[]")
        add_recv_addresses = request.form.getlist("additional_receiver_address[]")
        add_recv_doc_types = request.form.getlist("additional_receiver_doc_type[]")
        add_recv_doc_nos  = request.form.getlist("additional_receiver_doc_no[]")

        additional_receivers_data = []
        for i in range(len(add_recv_names)):
            if add_recv_names[i] and add_recv_names[i].strip():
                additional_receivers_data.append({
                    "name": add_recv_names[i],
                    "company": add_recv_companies[i] if i < len(add_recv_companies) else "",
                    "phone": add_recv_phones[i] if i < len(add_recv_phones) else "",
                    "address": add_recv_addresses[i] if i < len(add_recv_addresses) else "",
                    "doc_type": add_recv_doc_types[i] if i < len(add_recv_doc_types) else "",
                    "doc_no": add_recv_doc_nos[i] if i < len(add_recv_doc_nos) else "",
                })
        
        # Shipment metadata
        shipment_meta = json.dumps({
            "docket_no": docket_no,
            "shipper_name": request.form.get("shipper_name", ""),
            "shipper_address": request.form.get("shipper_address", ""),
            "client_code": request.form.get("client_code", ""),
            "receiver_name": request.form.get("receiver_name", ""),
            "receiver_company": request.form.get("receiver_company", ""),
            "receiver_phone": request.form.get("receiver_phone", ""),
            "receiver_address": request.form.get("receiver_address", ""),
            "destination": request.form.get("destination", ""),
            "shipment_type": request.form.get("shipment_type", ""),
            "mode": request.form.get("mode", ""),
            "carrier": request.form.get("carrier", ""),
            "tracking_number": meta.get("tracking_number", ""),
            "carrier_ref": request.form.get("carrier_ref", ""),
            "origin": request.form.get("origin", "India"),
            "pickup_date": request.form.get("pickup_date", ""),
            "departure_time": request.form.get("departure_time", ""),
            "expected_delivery": request.form.get("expected_delivery", ""),
            "comments": request.form.get("comments", ""),
            "payment_mode": payment_mode,
            "upi_app": upi_app,
            "upi_ref": upi_ref,
            "cheque_no": cheque_no,
            "cheque_date": cheque_date,
            "cheque_bank": cheque_bank,
            "freight": freight,
            "fuel": fuel,
            "other": other,
            "freight_weight": freight_weight,
            "freight_rate_per_kg": freight_rate,
            "freight_billing_weight": freight_billing_weight,
            "other_charges_reason": request.form.get("other_charges_reason", ""),
            "gst": gst,
            "amount_paid": amount_paid,
            "packages": packages_data,
            "additional_receivers": additional_receivers_data,
        })

        # Check if we're updating an existing invoice
        edit_invoice_id = request.form.get("edit_invoice_id")
        if edit_invoice_id:
            # Update existing invoice
            invoice = cdb.query(Invoice).filter_by(invoice_id=edit_invoice_id, company_id=company_id).first()
            if invoice:
                invoice.client_id = client_id
                invoice.date = date.fromisoformat(invoice_date)
                invoice.status = status
                invoice.contact_person = request.form.get("shipper_contact_name", "")
                invoice.phone = request.form.get("customer_phone", "")
                invoice.subtotal = base
                invoice.tax_amount = gst
                invoice.grand_total = grand_total
                invoice.terms = shipment_meta
                invoice.email = notes
                invoice.paid_amount = amount_paid
                invoice.balance = balance
                
                cdb.commit()
                flash(f"Customer invoice {invoice.invoice_id} updated successfully!")
                return redirect(url_for("invoice_list"))
        else:
            # Create new invoice
            invoice_id = _next_numbered_id(cdb, Invoice.invoice_id, "", extra_filters=[Invoice.company_id == company_id])
            
            inv = Invoice(
                invoice_id=invoice_id,
                company_id=company_id,
                client_id=client_id,
                date=date.fromisoformat(invoice_date),
                status=status,
                contact_person=request.form.get("shipper_contact_name", ""),
                phone=request.form.get("customer_phone", ""),
                subtotal=base,
                tax_amount=gst,
                grand_total=grand_total,
                terms=shipment_meta,
                email=notes,
                paid_amount=amount_paid,
                balance=balance,
            )
            cdb.add(inv)
            cdb.commit()

            # Fire-and-forget WhatsApp notification — never blocks the response,
            # never fails the invoice if WhatsApp isn't configured or the API errors.
            try:
                from tasks import send_invoice_generate_notification_async
                send_invoice_generate_notification_async(company_id=company_id, invoice_id=invoice_id)
            except Exception as e:
                print(f"[whatsapp] could not queue generate-notification for {invoice_id}: {e}")

            flash(f"Customer invoice {invoice_id} created successfully!")
            return redirect(url_for("invoice_list"))

    # GET request - prepare form data
    form_data = {}
    packages = []
    invoice_id = None
    invoice_date = str(today_ist())
    docket_no = ""
    is_edit = False
    client_display_id = ""
    
    if existing_invoice:
        is_edit = True
        invoice_id = existing_invoice.invoice_id
        invoice_date = existing_invoice.date.strftime('%Y-%m-%d')
        client_display_id = existing_invoice.client_obj.client_id if existing_invoice.client_obj else ""
        
        # Parse the terms JSON to get all the stored data
        try:
            meta = json.loads(existing_invoice.terms) if existing_invoice.terms else {}
        except:
            meta = {}
        
        # Build form_data with all existing values
        form_data = {
            "status": existing_invoice.status or "",
            "customer_id": existing_invoice.client_id,
            "customer_phone": existing_invoice.phone or "",
            "shipper_name": meta.get("shipper_name", ""),
            "shipper_contact_name": meta.get("shipper_contact_name", existing_invoice.contact_person or ""),
            "courier_company_id": meta.get("courier_company_id", ""),
             "shipper_address1": meta.get("shipper_address1", meta.get("shipper_address", "")),  
            "shipper_address2": meta.get("shipper_address2", ""),  
            "shipper_city": meta.get("shipper_city", ""),  
            "shipper_state": meta.get("shipper_state", ""),  
            "shipper_pincode": meta.get("shipper_pincode", ""),  
            "shipper_country": meta.get("shipper_country", "India"), 
            "shipper_doc_type": meta.get("shipper_doc_type", ""),
            "shipper_doc_no": meta.get("shipper_doc_no", ""),
            "client_code": meta.get("client_code", ""),
            "receiver_name": meta.get("receiver_name", ""),
            "receiver_company": meta.get("receiver_company", ""),
            "receiver_phone": meta.get("receiver_phone", ""),
            "receiver_address1": meta.get("receiver_address1", meta.get("receiver_address", "")),  
            "receiver_address2": meta.get("receiver_address2", ""),  
            "receiver_city": meta.get("receiver_city", ""),  
            "receiver_state": meta.get("receiver_state", ""),  
            "receiver_pincode": meta.get("receiver_pincode", ""),  
            "receiver_country": meta.get("receiver_country", "India"),
            "receiver_doc_type": meta.get("receiver_doc_type", ""),
            "receiver_doc_no": meta.get("receiver_doc_no", ""),
            "destination": meta.get("destination", ""),
            "shipment_type": meta.get("shipment_type", ""),
            "mode": meta.get("mode", ""),
            "carrier": meta.get("carrier", ""),
            "tracking_number": meta.get("tracking_number", ""),
            "carrier_ref": meta.get("carrier_ref", ""),
            "origin": meta.get("origin", "India"),
            "pickup_date": meta.get("pickup_date", ""),
            "departure_time": meta.get("departure_time", ""),
            "expected_delivery": meta.get("expected_delivery", ""),
            "comments": meta.get("comments", ""),
            "freight": meta.get("freight", existing_invoice.subtotal or 0),
            "fuel": meta.get("fuel", 0),
            "other": meta.get("other", 0),
            "freight_weight": meta.get("freight_weight", 0),
            "freight_rate_per_kg": meta.get("freight_rate_per_kg", 0),
            "freight_billing_weight": meta.get("freight_billing_weight", 0),
            "other_charges_reason": meta.get("other_charges_reason", ""),
            "amount_paid": meta.get("amount_paid", existing_invoice.paid_amount or 0),
            "payment_mode": meta.get("payment_mode", "cash"),
            "booking_type": meta.get("booking_type", "credit" if meta.get("payment_mode") != "cash" else "cash"),
            "discount": meta.get("discount", 0),
            "upi_app": meta.get("upi_app", ""),
            "upi_ref": meta.get("upi_ref", ""),
            "cheque_no": meta.get("cheque_no", ""),
            "cheque_date": meta.get("cheque_date", ""),
            "cheque_bank": meta.get("cheque_bank", ""),
            "notes": existing_invoice.email or "",
            "docket_no": meta.get("docket_no", ""),
            "has_resale": getattr(existing_invoice, 'has_resale', False),
            "resale_charges": getattr(existing_invoice, 'resale_charges', 0),
            "resale_reason": getattr(existing_invoice, 'resale_reason', ''),
            "resale_date": getattr(existing_invoice, 'resale_date', ''),
            "resale_notes": getattr(existing_invoice, 'resale_notes', ''),
            "vendor": meta.get("vendor", ""),
            "additional_receivers": meta.get("additional_receivers", []),
        }
        # ── Load linked Performa Invoice items ──────────────────────────────
        linked_est = cdb.query(Estimate).filter_by(company_id=company_id).filter(
            Estimate.terms.like(f'%"linked_invoice_id": "{existing_invoice.invoice_id}"%')
        ).first()
        if linked_est and linked_est.terms:
            try:
                perf_meta = json.loads(linked_est.terms)
                form_data["performa_items"] = perf_meta.get("line_items", [])
                form_data["perf_weight"]    = perf_meta.get("weight", "")
                form_data["perf_reference"] = perf_meta.get("reference", "")
                form_data["performa_invoice_no"]   = perf_meta.get("invoice_no", "")
                form_data["performa_invoice_date"] = perf_meta.get("invoice_date", "")
                form_data["export_reason"]         = perf_meta.get("export_reason", "")
            except Exception:
                pass
        # ────────────────────────────────────────────────────────────────────
        docket_no = meta.get("docket_no", "")
        
        # Get packages from meta
        packages = meta.get("packages", [])
        
        # If no packages in meta, create default empty package
        if not packages:
            packages = [{"name": "", "type": "", "qty": 1, "length": "", "width": "", "height": "", "weight": "", "rate": 0}]
    else:
        # New invoice - default values
        invoice_id = _next_numbered_id(cdb, Invoice.invoice_id, "", extra_filters=[Invoice.company_id == company_id])
        docket_no = _next_awb_number(company_id)
        form_data = {
            "payment_mode": "cash",
            "booking_type": "credit",
            "discount": 0,
            "additional_receivers": [],
        }
        packages = [{"name": "Box", "type": "Box", "qty": 1, "length": "", "width": "", "height": "", "weight": "", "rate": 0}]

    return render_template("booking.html",
                           clients=clients,
                           suppliers=suppliers,
                           form_data=form_data,
                           packages=packages,
                           invoice_id=invoice_id,
                           invoice_date=invoice_date,
                           docket_no=docket_no,
                           is_edit=is_edit,
                           today=str(today_ist()),
                           price_lists=price_lists,
                           client_display_id=client_display_id,
                           invoice=existing_invoice)


@app.route("/booking/edit/<invoice_id>", methods=["GET", "POST"])
@login_required
@require_permission("invoices", "view", method_actions={'POST': 'edit'})
def invoice_edit(invoice_id):
    """GET: render the edit form. POST: save updated line-item invoice."""
    cdb = get_cdb()
    company_id = get_current_company()

    invoice = _first_or_404(cdb.query(Invoice).filter_by(
        invoice_id=invoice_id, company_id=company_id).first())

    clients = cdb.query(Client).filter(
        Client.company_id == company_id,
        ~Client.client_type.in_(["Supplier", "Both", "Cash-Only"])  
    ).all()

    price_lists = cdb.query(PriceList).filter_by(
        company_id=company_id, 
        is_active=True,
        list_type='sales'
    ).all()

    if request.method == "POST":
        # ── Basic fields ────────────────────────────────────────────────────
        client_id_raw = request.form.get("client_id")
        invoice.client_id    = int(client_id_raw) if client_id_raw else None
        invoice.contact_person = request.form.get("contact_person", "")
        invoice.email        = request.form.get("email", "")
        invoice.phone        = request.form.get("phone", "")
        invoice.status       = request.form.get("status", "Draft")
        invoice.terms        = request.form.get("terms", "")

        invoice_date_str = request.form.get("invoice_date")
        if invoice_date_str:
            invoice.date = date.fromisoformat(invoice_date_str)

        due_date_str = request.form.get("due_date")
        invoice.due_date = date.fromisoformat(due_date_str) if due_date_str else None

        # ── Line items: delete old, insert new ──────────────────────────────
        cdb.query(InvoiceItem).filter_by(invoice_id=invoice.id).delete()

        item_codes   = request.form.getlist("item_code[]")
        descriptions = request.form.getlist("description[]")
        qtys         = request.form.getlist("qty[]")
        rates        = request.form.getlist("rate[]")
        discounts    = request.form.getlist("discount[]")

        subtotal = 0.0
        for i, desc in enumerate(descriptions):
            if not desc or not desc.strip():
                continue
            qty      = float(qtys[i])      if i < len(qtys)      and qtys[i]      else 0.0
            rate     = float(rates[i])     if i < len(rates)     and rates[i]     else 0.0
            discount = float(discounts[i]) if i < len(discounts) and discounts[i] else 0.0
            line_amt = qty * rate * (1 - discount / 100)
            subtotal += line_amt

            cdb.add(InvoiceItem(
                invoice_id    = invoice.id,
                code          = item_codes[i] if i < len(item_codes) else "",
                description   = desc.strip(),
                qty           = qty,
                rate          = rate,
                discount      = discount,
            ))

        tax_amount  = round(subtotal * 0.18, 2)
        grand_total = round(subtotal + tax_amount, 2)

        invoice.subtotal    = round(subtotal, 2)
        invoice.tax_amount  = tax_amount
        invoice.grand_total = grand_total
        invoice.balance     = round(grand_total - (invoice.paid_amount or 0), 2)

        cdb.commit()

        try:
            from tasks import send_invoice_update_notification_async
            send_invoice_update_notification_async(company_id=company_id, invoice_id=invoice_id)
        except Exception as e:
            print(f"[whatsapp] could not queue update-notification for {invoice_id}: {e}")

        flash(f"Invoice {invoice_id} updated successfully!", "success")
        return redirect(url_for("invoice_list"))

    # ── GET: build items list for the template ───────────────────────────────
    items = cdb.query(InvoiceItem).filter_by(invoice_id=invoice.id).all()
    today    = str(today_ist())
    due_date = str((today_ist() + timedelta(days=30)))
    can_edit = has_permission("invoices", "edit")

    return render_template("booking_edit.html",
                           invoice=invoice,
                           clients=clients,
                           items=items,
                           today=today,
                           due_date=due_date,
                           price_lists=price_lists,
                           can_edit=can_edit)

@app.route("/booking/customer/update", methods=["POST"])
@login_required
@require_permission("invoices", "edit")
def invoice_customer_update():
    """Update an existing customer invoice"""
    cdb = get_cdb()
    company_id = get_current_company()
    edit_invoice_id = request.form.get("edit_invoice_id")
    
    # Find the existing invoice
    invoice = cdb.query(Invoice).filter_by(invoice_id=edit_invoice_id, company_id=company_id).first()
    if not invoice:
        flash("Invoice not found")
        return redirect(url_for("invoice_list"))
    
    price_lists = cdb.query(PriceList).filter_by(
        company_id=company_id, 
        is_active=True
    ).all()

    # Parse the existing terms JSON
    try:
        old_meta = json.loads(invoice.terms) if invoice.terms else {}
    except:
        old_meta = {}

    # Needed early now: the stock-reconciliation block (below, after the
    # invoice fields are saved) has to know the OLD awb/docket_no so it can
    # find and reverse this booking's original StockPurchaseHistory rows
    # before reapplying the current package quantities.
    old_docket_no = (old_meta.get("docket_no", "") or "").strip()

    # ── Basic fields ──────────────────────────────────────────────────────────
    client_id_raw = request.form.get("customer_id")
    client_id = int(client_id_raw) if client_id_raw else None
    invoice_date = request.form.get("invoice_date") or str(today_ist())
    docket_no = request.form.get("docket_no", "")
    action = request.form.get("action", "final")

    # ── AWB/docket uniqueness — the field is readonly in the UI, but the
    # server never trusts the client. Someone editing two tabs, resubmitting
    # a stale form, or hitting this endpoint directly could still send a
    # docket_no that belongs to a different invoice. Exclude this invoice
    # itself (it's allowed to keep its own docket_no unchanged). ──────────────
    dupe_invoice_id = _docket_no_in_use(
        cdb, company_id, docket_no, exclude_invoice_id=edit_invoice_id
    )
    if dupe_invoice_id:
        duplicate_awb = docket_no
        docket_no = _next_awb_number(company_id)
        flash(f"AWB {duplicate_awb} was already used on invoice {dupe_invoice_id} — "
          f"this invoice was automatically assigned {docket_no} instead.")

    # ── Charges & totals ──────────────────────────────────────────────────────
    freight_weight = float(request.form.get("freight_weight", 0) or 0)
    freight_rate   = float(request.form.get("freight_rate_per_kg", 0) or 0)
    # freight_billing_weight is the rounded rate-card slab weight the rate
    # lookup matched (booking.html's applyRateToFreight sets it alongside
    # freight_rate_per_kg). freight_weight itself is the actual/display
    # weight and must NOT be used for billing math — see booking.html fix.
    # Falls back to freight_weight when no rate lookup ran (rate typed in
    # manually), where weight and rate already agree.
    freight_billing_weight = float(request.form.get("freight_billing_weight", 0) or 0) or freight_weight
    freight        = round(freight_billing_weight * freight_rate, 2)
    fuel = float(request.form.get("fuel_surcharge", 0) or 0)
    other = float(request.form.get("other_charges", 0) or 0)
    discount = float(request.form.get("discount_amount", 0) or 0)
    base = freight + fuel + other
    co = Company.query.filter_by(company_id=company_id).first()
    apply_gst = co.is_gst_registered if (co and hasattr(co, 'is_gst_registered')) else True
    shipper_state  = request.form.get("shipper_state", "")
    receiver_state = request.form.get("receiver_state", "")
    payment_mode   = request.form.get("payment_mode", "cash")
    booking_type   = request.form.get("booking_type", "credit")

    # ── Resale Charges ──────────────────────────────────────────────────────────
    has_resale = request.form.get("resale_active") == "true"
    resale_amount = float(request.form.get("resale_amount", 0) or 0)
    resale_reason = request.form.get("resale_reason", "").strip()
    resale_date_str = request.form.get("resale_date")
    resale_notes = request.form.get("resale_notes", "").strip()

    if has_resale and resale_amount > 0:
        resale_date = date.fromisoformat(resale_date_str) if resale_date_str else today_ist()
    else:
        resale_amount = 0
        resale_date = None
        resale_reason = None
        resale_notes = None

    # ── GST: proper CGST/SGST vs IGST split (based on shipper/receiver state)
    # plus round-off to the nearest rupee, instead of a flat 18% figure. ──────
    # Discount comes off before tax, same reasoning as invoice_customer_save.
    taxable_base = max(0, base + resale_amount - discount)
    gst_calc = compute_invoice_gst(taxable_base, apply_gst, shipper_state, receiver_state)
    gst = gst_calc["gst_total"]
    resale_gst = 0  # resale GST is now folded into the single gst_calc split above
    grand_total = gst_calc["grand_total"]
    booking_amount_paid_form = float(request.form.get("amount_paid", 0) or 0)
    # This form's "amount_paid" is only the payment collected at booking time
    # (cash/UPI on the booking form). It does NOT know about anything applied
    # afterwards through the separate Receipts module (receipt_save), which
    # increments invoice.paid_amount/decrements invoice.balance independently.
    # Blindly overwriting paid_amount with this form value would silently
    # erase any such receipt every time the booking is re-saved. Preserve it:
    # whatever paid_amount has grown by beyond what was recorded here last
    # time (old_meta's stored amount_paid) came from a receipt, and carries
    # forward on top of the current form's booking-time amount.
    old_booking_paid = float(old_meta.get("amount_paid", 0) or 0)
    receipts_applied = max(0.0, (invoice.paid_amount or 0) - old_booking_paid)
    # Whatever the booking-time figure went UP by since the last save is a
    # payment collected just now, on this edit — it needs its own Cash/Bank
    # transaction (recorded below, after invoice fields are saved), same as
    # invoice_customer_save does on create. Without this, editing a booking
    # to add/raise amount_paid changed invoice.paid_amount but never showed
    # up anywhere in the Receipts module.
    booking_payment_delta = round(booking_amount_paid_form - old_booking_paid, 2)
    amount_paid = booking_amount_paid_form + receipts_applied
    balance = round(grand_total - amount_paid, 2)

    # ── Credit limit check (edit path) ───────────────────────────────────────
    # This was missing entirely on update — invoice_customer_save had it,
    # invoice_customer_update didn't, so editing a booking to push a client
    # over their limit went through regardless of the "block" setting.
    # Same note as the save route: booking.html confirms via popup before
    # this request is sent, so this is a backstop and only flashes on block.
    if action != "draft" and client_id:
        _client_for_limit = cdb.query(Client).filter_by(id=client_id, company_id=company_id).first()
        _limit_ok, _limit_msg = _check_credit_limit(co, _client_for_limit, grand_total)
        if not _limit_ok:
            flash(_limit_msg, "danger")
            return redirect(url_for("invoice_edit", invoice_id=edit_invoice_id))

    # ── Payment info ─────────────────────────────────────────────────────────
    upi_app = request.form.get("upi_app", "")
    upi_ref = request.form.get("upi_ref", "")
    cheque_no = request.form.get("cheque_no", "")
    cheque_date = request.form.get("cheque_date", "")
    cheque_bank = request.form.get("cheque_bank", "")

    # ── Status ────────────────────────────────────────────────────────────────
    if action == "draft":
        status = "Draft"
    elif balance <= 0:
        status = "Paid"
    elif amount_paid > 0:
        status = "Partial"
    else:
        status = "Pending"

    notes = request.form.get("notes", "")
    
    # ── Process Packages ─────────────────────────────────────────────────────
        # ── Process Packages ─────────────────────────────────────────────────────
    pkg_names = request.form.getlist("pkg_name[]")
    pkg_types = request.form.getlist("pkg_type[]")
    pkg_units = request.form.getlist("pkg_unit[]")
    pkg_qtys = request.form.getlist("pkg_qty[]")
    pkg_l = request.form.getlist("pkg_l[]")
    pkg_w = request.form.getlist("pkg_w[]")
    pkg_h = request.form.getlist("pkg_h[]")
    pkg_wt = request.form.getlist("pkg_wt[]")
    pkg_division = request.form.getlist("pkg_division[]")
    pkg_discount = request.form.getlist("pkg_discount[]")
    pkg_discwt = request.form.getlist("pkg_discwt[]")
    pkg_volwt = request.form.getlist("pkg_volwt[]")
    pkg_chgwt = request.form.getlist("pkg_chgwt[]")
    pkg_rates = request.form.getlist("pkg_rate[]")
    
    # ── Get cash client ID if this is a cash booking ──────────────────────
    cash_client_id = None
    if booking_type == "cash":
        cash_shipper_name = (request.form.get("shipper_name", "") or "").strip()
        if cash_shipper_name:
            cash_client = _get_or_create_cash_client(cdb, company_id, cash_shipper_name)
            if cash_client:
                cash_client_id = cash_client.id
    
    packages_data = []
    for i in range(len(pkg_names)):
        if pkg_names[i] and pkg_names[i].strip():
            packages_data.append({
                "name": pkg_names[i],
                "type": pkg_types[i] if i < len(pkg_types) else "",
                "unit": pkg_units[i] if i < len(pkg_units) else "cm",
                "qty": float(pkg_qtys[i] or 1) if pkg_qtys[i] else 1,
                "length": float(pkg_l[i] or 0) if i < len(pkg_l) else 0,
                "width": float(pkg_w[i] or 0) if i < len(pkg_w) else 0,
                "height": float(pkg_h[i] or 0) if i < len(pkg_h) else 0,
                "weight": float(pkg_wt[i] or 0) if i < len(pkg_wt) else 0,
                "division": float(pkg_division[i] or 5000) if i < len(pkg_division) and pkg_division[i] else 5000,
                "discount": float(pkg_discount[i] or 0) if i < len(pkg_discount) and pkg_discount[i] else 0,
                "discount_wt": float(pkg_discwt[i] or 0) if i < len(pkg_discwt) and pkg_discwt[i] else 0,
                "vol_weight": float(pkg_volwt[i] or 0) if i < len(pkg_volwt) and pkg_volwt[i] else 0,
                "chg_weight": float(pkg_chgwt[i] or 0) if i < len(pkg_chgwt) and pkg_chgwt[i] else 0,
                "rate": float(pkg_rates[i] or 0) if i < len(pkg_rates) else 0,
            })
    
    # ── Get primary stock info for manifest sync ──
    # BUGFIX: cash bookings store their StockItem rows under the per-shipper
    # cash client id (see client_id_for_stock further below / invoice_new),
    # not under the raw form client_id (which is None for cash). Querying
    # with client_id=client_id always missed the row for cash bookings, so
    # primary_stock_id stayed None -> new ManifestEntry rows created during
    # this edit got stock_item_id=None -> manifest_generate_company silently
    # skipped deducting stock for them.
    primary_stock_id = None
    primary_stock_name = None
    primary_item_type = "Box"
    _client_id_for_primary_stock = cash_client_id if booking_type == "cash" else client_id
    if packages_data and packages_data[0].get("name"):
        primary_stock_name = packages_data[0]["name"]
        stock_item = cdb.query(StockItem).filter_by(
            company_id=company_id,
            name=primary_stock_name,
            client_id=_client_id_for_primary_stock
        ).first()
        if stock_item:
            primary_stock_id = stock_item.id
            primary_item_type = stock_item.item_type or stock_item.category or "Box"

    # ════════════════════════════════════════════════════════════════════════
    # ║  FIELD PERMISSION CHECKS - INSERT HERE                              ║
    # ════════════════════════════════════════════════════════════════════════
    
    # Get current user's field permissions
    user = get_current_user()
    role = user.get("role", "employee")
    user_id = user.get("user_id")
    
    # Get field permissions
    field_perms = get_field_permissions(role, user_id, company_id, cdb)
    
    # Define which fields can be edited
    def can_edit_field(field_group):
        return field_perms.get(field_group, {}).get("edit", False)
    
    # ── 1. PACKAGES - Check if user can edit package fields ──────────────
    if not can_edit_field('invoice_packages'):
        # User cannot edit ANY package fields - use existing values
        old_packages = old_meta.get("packages", [])
        # Keep the old packages data exactly as it was
        packages_data = old_packages
    
    # ── 2. ACTUAL WEIGHT - Check if user can edit actual weight ──────────
    elif not can_edit_field('invoice_packages_actual_weight'):
        # User can edit packages but NOT actual weight
        # Preserve the actual weight from the existing invoice
        old_packages = old_meta.get("packages", [])
        for i, pkg in enumerate(packages_data):
            if i < len(old_packages):
                # Keep the old weight, but allow other fields to be updated
                pkg["weight"] = old_packages[i].get("weight", pkg.get("weight", 0))
            else:
                # If there are more packages than before, use 0 as fallback
                pkg["weight"] = 0
    
    # ── 3. CHARGES - Check if user can edit freight/charges ──────────────
    if not can_edit_field('invoice_charges'):
        # Use existing freight/charges values from old_meta
        freight = old_meta.get("freight", freight)
        fuel = old_meta.get("fuel", fuel)
        other = old_meta.get("other", other)
        discount = old_meta.get("discount", discount)
        freight_weight = old_meta.get("freight_weight", freight_weight)
        freight_rate = old_meta.get("freight_rate_per_kg", freight_rate)
        freight_billing_weight = old_meta.get("freight_billing_weight", freight_billing_weight)
        # Recalculate base with preserved values
        base = freight + fuel + other
    
    # ── 4. PERFORMA ITEMS - Check if user can edit performa items ────────
    if not can_edit_field('invoice_performa'):
        # Preserve existing performa items from old_meta
        perf_items_from_form = []  # Clear any submitted performa data
        # We'll use the old values later when building the meta
        # The form data will be ignored for performa items
        form_data_perf_items = request.form.getlist("perf_desc[]")  # Still read but won't be used
        # The old performa items will be used from old_meta
    
    # ── 5. RESALE CHARGES - Check if user can edit resale ────────────────
    if not can_edit_field('invoice_resale'):
        # Preserve existing resale data from old_meta
        resale_data = old_meta.get("resale", {})
        if resale_data:
            has_resale = True
            resale_amount = resale_data.get("amount", 0)
            resale_reason = resale_data.get("reason", "")
            resale_date_str = resale_data.get("date", "")
            resale_notes = resale_data.get("notes", "")
            if resale_date_str:
                try:
                    resale_date = date.fromisoformat(resale_date_str)
                except:
                    resale_date = None
        else:
            has_resale = False
            resale_amount = 0
            resale_reason = None
            resale_date = None
            resale_notes = None
    
    # ── 6. SERVICE DETAILS - Check if user can edit service fields ──────
    if not can_edit_field('invoice_service'):
        # Preserve existing service details from old_meta
        # These will be used when building shipment_meta
        service_fields = [
            'destination', 'shipment_type', 'mode', 'vendor',
            'courier_company_id', 'carrier', 'tracking_number',
            'carrier_ref', 'origin', 'pickup_date', 'departure_time',
            'expected_delivery', 'comments'
        ]
        for field in service_fields:
            if field in old_meta:
                # Override the form value with the old one
                if field == 'destination':
                    request.form.get("destination", old_meta.get('destination', ''))
                # Continue for other fields
    
    # ── 7. SENDER ADDRESS - Check if user can edit sender fields ──────────
    if not can_edit_field('invoice_sender'):
        # Preserve existing sender address details
        sender_fields = [
            'shipper_address1', 'shipper_address2', 'shipper_city',
            'shipper_state', 'shipper_pincode', 'shipper_country',
            'shipper_doc_type', 'shipper_doc_no', 'client_code'
        ]
        # The values will be used from old_meta when building shipment_meta
    
    # ── 8. RECEIVER ADDRESS - Check if user can edit receiver fields ──────
    if not can_edit_field('invoice_receiver'):
        # Preserve existing receiver address details
        receiver_fields = [
            'receiver_name', 'receiver_company', 'receiver_phone',
            'receiver_address1', 'receiver_address2', 'receiver_city',
            'receiver_state', 'receiver_pincode', 'receiver_country',
            'receiver_doc_type', 'receiver_doc_no'
        ]
        # The values will be used from old_meta when building shipment_meta

    # ════════════════════════════════════════════════════════════════════════
    # ║  END OF FIELD PERMISSION CHECKS                                     ║
    # ════════════════════════════════════════════════════════════════════════

    # ── GST calculation ──────────────────────────────────────────────────────
    taxable_base = max(0, base + resale_amount - discount)
    gst_calc = compute_invoice_gst(taxable_base, apply_gst, shipper_state, receiver_state)
    gst = gst_calc["gst_total"]
    resale_gst = 0
    grand_total = gst_calc["grand_total"]
    booking_amount_paid_form = float(request.form.get("amount_paid", 0) or 0)
    old_booking_paid = float(old_meta.get("amount_paid", 0) or 0)
    receipts_applied = max(0.0, (invoice.paid_amount or 0) - old_booking_paid)
    booking_payment_delta = round(booking_amount_paid_form - old_booking_paid, 2)
    amount_paid = booking_amount_paid_form + receipts_applied
    balance = round(grand_total - amount_paid, 2)

    # ── Credit limit check (edit path) ───────────────────────────────────────
    if action != "draft" and client_id:
        _client_for_limit = cdb.query(Client).filter_by(id=client_id, company_id=company_id).first()
        _limit_ok, _limit_msg = _check_credit_limit(co, _client_for_limit, grand_total)
        if not _limit_ok:
            flash(_limit_msg, "danger")
            return redirect(url_for("invoice_edit", invoice_id=edit_invoice_id))

    # ── Payment info ─────────────────────────────────────────────────────────
    upi_app = request.form.get("upi_app", "")
    upi_ref = request.form.get("upi_ref", "")
    cheque_no = request.form.get("cheque_no", "")
    cheque_date = request.form.get("cheque_date", "")
    cheque_bank = request.form.get("cheque_bank", "")

    # ── Status ────────────────────────────────────────────────────────────────
    if action == "draft":
        status = "Draft"
    elif balance <= 0:
        status = "Paid"
    elif amount_paid > 0:
        status = "Partial"
    else:
        status = "Pending"

    notes = request.form.get("notes", "")
    
    # ── Build shipment_meta with all fields (respecting permissions) ────────
    # For fields that the user doesn't have permission to edit, use old_meta values
    
    def get_field_safe(field_name, default=""):
        """Get field value from form if user has permission, otherwise from old_meta"""
        # Determine which permission group this field belongs to
        field_permission_map = {
            'shipper_name': 'invoice_customer',
            'shipper_contact_name': 'invoice_customer',
            'customer_phone': 'invoice_customer',
            'shipper_address1': 'invoice_sender',
            'shipper_address2': 'invoice_sender',
            'shipper_city': 'invoice_sender',
            'shipper_state': 'invoice_sender',
            'shipper_pincode': 'invoice_sender',
            'shipper_country': 'invoice_sender',
            'shipper_doc_type': 'invoice_sender',
            'shipper_doc_no': 'invoice_sender',
            'client_code': 'invoice_sender',
            'receiver_name': 'invoice_receiver',
            'receiver_company': 'invoice_receiver',
            'receiver_phone': 'invoice_receiver',
            'receiver_address1': 'invoice_receiver',
            'receiver_address2': 'invoice_receiver',
            'receiver_city': 'invoice_receiver',
            'receiver_state': 'invoice_receiver',
            'receiver_pincode': 'invoice_receiver',
            'receiver_country': 'invoice_receiver',
            'receiver_doc_type': 'invoice_receiver',
            'receiver_doc_no': 'invoice_receiver',
            'destination': 'invoice_service',
            'shipment_type': 'invoice_service',
            'mode': 'invoice_service',
            'vendor': 'invoice_service',
            'courier_company_id': 'invoice_service',
            'carrier': 'invoice_service',
            'tracking_number': 'invoice_service',
            'carrier_ref': 'invoice_service',
            'origin': 'invoice_service',
            'pickup_date': 'invoice_service',
            'departure_time': 'invoice_service',
            'expected_delivery': 'invoice_service',
            'comments': 'invoice_service',
        }
        
        perm_group = field_permission_map.get(field_name)
        if perm_group and can_edit_field(perm_group):
            # User has permission - use form value
            return request.form.get(field_name, default)
        else:
            # User doesn't have permission - use old value
            return old_meta.get(field_name, default)
    
    # Update shipment metadata
    shipment_meta = json.dumps({
        "docket_no": docket_no,
        "shipper_name": request.form.get("shipper_name", ""),
        "shipper_contact_name": request.form.get("shipper_contact_name", ""),
        "shipper_address1": request.form.get("shipper_address1", ""),  
        "shipper_address2": request.form.get("shipper_address2", ""),  
        "shipper_city": request.form.get("shipper_city", ""),  
        "shipper_state": request.form.get("shipper_state", ""),  
        "shipper_pincode": request.form.get("shipper_pincode", ""),  
        "shipper_country": request.form.get("shipper_country", "India"),
        "shipper_doc_type": request.form.get("shipper_doc_type", ""),
        "shipper_doc_no": request.form.get("shipper_doc_no", ""),
        "client_code": request.form.get("client_code", ""),
        "receiver_name": request.form.get("receiver_name", ""),
        "receiver_company": request.form.get("receiver_company", ""),
        "receiver_phone": request.form.get("receiver_phone", ""),
        "receiver_address1": request.form.get("receiver_address1", ""),  
        "receiver_address2": request.form.get("receiver_address2", ""),  
        "receiver_city": request.form.get("receiver_city", ""),  
        "receiver_state": request.form.get("receiver_state", ""),  
        "receiver_pincode": request.form.get("receiver_pincode", ""),  
        "receiver_country": request.form.get("receiver_country", "India"),
        "receiver_doc_type": request.form.get("receiver_doc_type", ""),
        "receiver_doc_no": request.form.get("receiver_doc_no", ""),
        "destination": request.form.get("destination", ""),
        "shipment_type": request.form.get("shipment_type", ""),
        "vendor": request.form.get("vendor", ""),
        "mode": request.form.get("mode", ""),
        "courier_company_id": request.form.get("courier_company_id", ""),
        "carrier": request.form.get("carrier", ""),
        "tracking_number": request.form.get("tracking_number", ""),
        "carrier_ref": request.form.get("carrier_ref", ""),
        "origin": request.form.get("origin", "India"),
        "pickup_date": request.form.get("pickup_date", ""),
        "departure_time": request.form.get("departure_time", ""),
        "expected_delivery": request.form.get("expected_delivery", ""),
        "comments": request.form.get("comments", ""),
        "payment_mode": payment_mode,
        "booking_type": booking_type,
        "upi_app": upi_app,
        "upi_ref": upi_ref,
        "cheque_no": cheque_no,
        "cheque_date": cheque_date,
        "cheque_bank": cheque_bank,
        "freight": freight,
        "freight_weight": freight_weight,
        "freight_rate_per_kg": freight_rate,
        "freight_billing_weight": freight_billing_weight,
        "fuel": fuel,
        "other": other,
        "discount": discount,
        "other_charges_reason": request.form.get("other_charges_reason", ""),
        "gst": gst,
        "cgst": gst_calc["cgst"],
        "sgst": gst_calc["sgst"],
        "igst": gst_calc["igst"],
        "is_interstate": gst_calc["is_interstate"],
        "round_off": gst_calc["round_off"],
        "amount_paid": amount_paid,
        "packages": packages_data,
        "resale": {
            "amount": resale_amount,
            "gst": resale_gst if has_resale else 0,
            "reason": resale_reason,
            "date": resale_date.strftime("%Y-%m-%d") if resale_date else "",
            "notes": resale_notes,
            "added_by": get_current_user().get("email")
        } if has_resale and resale_amount > 0 else None
    })
    
    new_carrier_ref = (request.form.get("carrier_ref") or "").strip()

    meta_dict = json.loads(shipment_meta)
    if booking_type == "cash":
        meta_dict["shipper_aadhar_front_file"] = save_shipper_id_doc(request.files.get("shipper_aadhar_front_file"), edit_invoice_id, "aadhar_front", old_meta.get("shipper_aadhar_front_file"))
        meta_dict["shipper_aadhar_back_file"]  = save_shipper_id_doc(request.files.get("shipper_aadhar_back_file"),  edit_invoice_id, "aadhar_back",  old_meta.get("shipper_aadhar_back_file"))
        meta_dict["shipper_pan_front_file"]    = save_shipper_id_doc(request.files.get("shipper_pan_front_file"),    edit_invoice_id, "pan_front",    old_meta.get("shipper_pan_front_file"))
        meta_dict["shipper_pan_back_file"]     = save_shipper_id_doc(request.files.get("shipper_pan_back_file"),     edit_invoice_id, "pan_back",     old_meta.get("shipper_pan_back_file"))
        # Legacy single-file fields (pre front/back) — keep whatever was there
        # so old bookings edited today don't silently lose their one scan.
        meta_dict["shipper_aadhar_file"] = old_meta.get("shipper_aadhar_file", "")
        meta_dict["shipper_pan_file"]    = old_meta.get("shipper_pan_file", "")
    shipment_meta = json.dumps(meta_dict)

    # Credit bookings must be tied to a client, or the pending balance below
    # never gets attached to anyone's outstanding ledger. Cash/UPI walking
    # customers are fine with no client — they're not carrying a balance.
    if action != "draft" and payment_mode == "credit" and not client_id:
        flash("Credit bookings require a customer to be selected.", "error")
        return redirect(url_for("invoice_customer_new"))

    # Update invoice fields
    invoice.client_id = client_id
    invoice.date = date.fromisoformat(invoice_date)
    invoice.status = status
    invoice.contact_person = request.form.get("shipper_contact_name", "")
    invoice.phone = request.form.get("customer_phone", "")
    invoice.subtotal = base
    invoice.tax_amount = gst
    invoice.grand_total = grand_total
    invoice.terms = shipment_meta
    invoice.email = notes
    invoice.paid_amount = amount_paid
    invoice.balance = balance
    invoice.has_resale = has_resale and resale_amount > 0
    invoice.resale_charges = resale_amount
    invoice.resale_reason = resale_reason
    invoice.resale_date = resale_date
    invoice.resale_notes = resale_notes

    # ── Record any NEW payment collected at this edit as a receipt ──────────
    # Mirrors the "RECORD PAYMENT IN CASH IN HAND OR BANK ACCOUNT" block in
    # invoice_customer_save, but only for booking_payment_delta — the amount
    # the booking-time amount_paid went UP by on this save. Without this,
    # raising amount_paid on an edit updated invoice.paid_amount but never
    # created anything in Receipts & Payments / the debtor statement / the
    # client ledger, unlike a payment collected at initial booking. A drop
    # in the booking-time figure (booking_payment_delta <= 0) is treated as
    # a data-entry correction, not a real cash movement, and isn't recorded.
    if booking_payment_delta > 0.01:
        transaction_date = date.fromisoformat(invoice_date)

        # Same party_name resolution as invoice_customer_save — this is what
        # the Receipts history, debtor statement, and client ledger filter on.
        _pay_party_name = get_party_name(
            client_id=client_id,
            form=request.form,
            fallback_name=request.form.get("shipper_name", "").strip() or None
        )

        if payment_mode == "cash":
            cdb.add(CashTransaction(
                company_id=company_id,
                type="income",
                date=transaction_date,
                category="Receipt",
                description=f"Payment received for invoice {edit_invoice_id} - Booking edit",
                amount=booking_payment_delta,
                reference=edit_invoice_id,
                notes="Payment via Cash from customer (added on booking edit)",
                party_name=_pay_party_name,
                created_by=get_current_user().get("email")
            ))
        elif payment_mode == "online":
            bank_account = cdb.query(BankAccount).filter_by(
                company_id=company_id, status='Active'
            ).first()
            if not bank_account:
                bank_account = BankAccount(
                    company_id=company_id,
                    bank_name="Default Bank Account",
                    account_name="Sales Receipts",
                    account_number="SALES001",
                    ifsc_code="DEFAULT0001",
                    branch="Main Branch",
                    opening_balance=0,
                    balance=booking_payment_delta,
                    status='Active',
                    created_at=datetime.utcnow()
                )
                cdb.add(bank_account)
                cdb.flush()
            else:
                bank_account.balance += booking_payment_delta
                bank_account.updated_at = datetime.utcnow()

            cdb.add(BankTransaction(
                bank_account_id=bank_account.id,
                company_id=company_id,
                type="credit",
                date=transaction_date,
                description=f"Payment received for invoice {edit_invoice_id} - via {upi_app or 'Online'} (booking edit)",
                amount=booking_payment_delta,
                reference=upi_ref or edit_invoice_id,
                transaction_mode="Online",
                notes=f"UPI App: {upi_app}, Ref: {upi_ref} (added on booking edit)",
                party_name=_pay_party_name,
                created_by=get_current_user().get("email")
            ))
        elif payment_mode == "cheque":
            bank_account = cdb.query(BankAccount).filter_by(
                company_id=company_id, status='Active'
            ).first()
            if not bank_account:
                bank_account = BankAccount(
                    company_id=company_id,
                    bank_name=cheque_bank or "Cheque Account",
                    account_name="Cheque Receipts",
                    account_number="CHEQ001",
                    ifsc_code="CHEQ0001",
                    branch="Main Branch",
                    opening_balance=0,
                    balance=booking_payment_delta,
                    status='Active',
                    created_at=datetime.utcnow()
                )
                cdb.add(bank_account)
                cdb.flush()
            else:
                bank_account.balance += booking_payment_delta
                bank_account.updated_at = datetime.utcnow()

            cdb.add(BankTransaction(
                bank_account_id=bank_account.id,
                company_id=company_id,
                type="credit",
                date=transaction_date,
                description=f"Cheque payment received for invoice {edit_invoice_id} (booking edit)",
                amount=booking_payment_delta,
                reference=cheque_no or edit_invoice_id,
                transaction_mode="Cheque",
                notes=f"Cheque No: {cheque_no}, Bank: {cheque_bank}, Date: {cheque_date} (added on booking edit)",
                party_name=_pay_party_name,
                created_by=get_current_user().get("email")
            ))

    if booking_payment_delta > 0.01 and client_id:
        client_for_payment = cdb.query(Client).filter_by(id=client_id, company_id=company_id).first()
        if client_for_payment:
            client_for_payment.last_payment = today_ist()        

    cdb.commit()

    # ── Update any customer invoices that contain this booking ──
    try:
        updated_invoices = update_customer_invoice_from_booking(cdb, company_id, invoice.id)
        if updated_invoices:
            flash(f"Customer invoice(s) {', '.join(str(i) for i in updated_invoices)} updated to reflect booking changes.", "info")
    except Exception as e:
        print(f"[customer-invoice-update] failed to update parent invoices: {e}")
    
    # ── Reconcile inventory against this booking's package quantities ────────
    # CRITICAL FIX: Check if this AWB has already been dispatched via manifest
    # If the manifest is already generated, we should NOT reverse the stock
    # because the stock was already deducted when the manifest was generated.
    # Only reverse stock if the AWB is NOT yet dispatched (still Pending).
    ship_source      = (request.form.get("shipper_city") or request.form.get("origin") or "India")
    ship_destination = request.form.get("destination", "")
    
    # ── Get cash client ID for stock matching ──────────────────────────────
    cash_client_id_for_stock = None
    if booking_type == "cash":
        cash_shipper_name = (request.form.get("shipper_name", "") or "").strip()
        if cash_shipper_name:
            cash_client = _get_or_create_cash_client(cdb, company_id, cash_shipper_name)
            if cash_client:
                cash_client_id_for_stock = cash_client.id
    
    # Determine which client_id to use for stock operations
    if booking_type == "cash":
        client_id_for_stock = cash_client_id_for_stock
    else:
        client_id_for_stock = client_id
    
    if old_docket_no:
        # Check if this AWB has any Generated manifest entries
        manifest_entry_exists = cdb.query(ManifestEntry).join(
            CompanyManifest, ManifestEntry.manifest_id == CompanyManifest.id
        ).filter(
            ManifestEntry.docket_no == old_docket_no,
            CompanyManifest.company_id == company_id,
            ManifestEntry.status == 'Generated'
        ).first()
        
        if not manifest_entry_exists:
            # Only reverse stock if NOT already dispatched via manifest
            old_history_rows = cdb.query(StockPurchaseHistory).filter_by(awb_no=old_docket_no).all()
            for h in old_history_rows:
                stock = cdb.query(StockItem).filter_by(id=h.stock_item_id).first()
                if stock:
                    stock.quantity = (stock.quantity or 0) - (h.quantity or 0)
                cdb.delete(h)
        else:
            # AWB already dispatched - skip stock reversal but still update other data
            flash(
                f"Note: AWB {old_docket_no} has already been dispatched via manifest. "
                f"Stock will not be adjusted, but other booking details have been updated.",
                "info"
            )

    # Now add stock back for the updated packages - but ONLY if not already dispatched
    # Check if the NEW docket_no is already dispatched
    new_docket_dispatched = False
    if docket_no:
        new_manifest_entry = cdb.query(ManifestEntry).join(
            CompanyManifest, ManifestEntry.manifest_id == CompanyManifest.id
        ).filter(
            ManifestEntry.docket_no == docket_no,
            CompanyManifest.company_id == company_id,
            ManifestEntry.status == 'Generated'
        ).first()
        if new_manifest_entry:
            new_docket_dispatched = True

    # Only add stock if the AWB is NOT already dispatched
    if not new_docket_dispatched:
        for pkg in packages_data:
            item_name = (pkg["name"] or "").strip()
            if not item_name:
                continue
            qty, rate, pkg_type = pkg["qty"], pkg["rate"], (pkg["type"] or "Box")

            # Match by name + client_id (cash uses cash client ID, credit uses regular client ID)
            stock_filters = dict(
                company_id=company_id, 
                name=item_name, 
                client_id=client_id_for_stock
            )
            existing_item = cdb.query(StockItem).filter_by(**stock_filters).first()

            if existing_item:
                existing_item.quantity = (existing_item.quantity or 0) + qty
                existing_item.last_updated = today_ist()
                if rate > 0:
                    existing_item.unit_price = rate
                    existing_item.purchase_rate = rate
                stock_item_id = existing_item.id
                gst_percent = existing_item.gst_percent or 0
            else:
                new_code = _next_numbered_id(cdb, StockItem.code, "PKG-", extra_filters=[StockItem.company_id == company_id])
                new_item = StockItem(
                    company_id=company_id,
                    code=new_code,
                    name=item_name,
                    category="Packaging",
                    item_type=pkg_type,
                    client_id=client_id_for_stock,  # ← Now uses cash_client_id for cash bookings
                    shipper_name=None,  # ← No longer needed for cash bookings
                    quantity=qty,
                    unit="pcs",
                    unit_price=rate,
                    purchase_rate=rate,
                    reorder_level=0,
                    gst_percent=18,
                    hsn="",
                    last_updated=today_ist(),
                )
                cdb.add(new_item)
                cdb.flush()
                stock_item_id = new_item.id
                gst_percent = 18

            cdb.add(StockPurchaseHistory(
                stock_item_id=stock_item_id,
                purchase_invoice_id=None,
                reference=edit_invoice_id,
                quantity=qty,
                purchase_rate=rate,
                gst_percent=gst_percent,
                purchase_date=date.fromisoformat(invoice_date),
                awb_no=docket_no,
                source=ship_source,
                destination=ship_destination,
                length=pkg["length"], width=pkg["width"], height=pkg["height"], weight=pkg["weight"],
            ))
    else:
        # AWB already dispatched - we still need to update stock metadata but NOT quantity
        # Just update the existing StockPurchaseHistory entries with new metadata
        existing_history_rows = cdb.query(StockPurchaseHistory).filter_by(awb_no=docket_no).all()
        for h in existing_history_rows:
            # Update metadata only, keep quantity the same
            h.source = ship_source
            h.destination = ship_destination
            # Don't change quantity or purchase_rate

    cdb.commit()

    # ── Auto-generate / repair the purchase invoice line for this booking ────
    # Previously this only happened on the initial save (invoice_customer_save).
    # Any booking whose first successful save landed here instead -- e.g. a
    # stale edit_invoice_id, a resubmitted form, or a draft finalized via
    # this route -- never got a purchase line and there was no warning shown.
    # Calling the same shared helper here means (a) edits keep the purchase
    # line's docket/carrier/rate in sync, and (b) simply re-saving a booking
    # that's missing its purchase line (via Edit -> Save) now repairs it.
    _sync_auto_purchase_invoice_line(
        cdb, company_id, request.form, packages_data,
        freight_weight, apply_gst, gst_calc,
        invoice_date, docket_no, edit_invoice_id, invoice.id, action,
    )
    cdb.commit()

    # ── AWB/docket_no sync — mirrors the carrier_ref sync below. If this
    # booking's AWB changed (auto-bump on a duplicate-AWB save is currently
    # the only way it can, since the field is readonly in the UI), push the
    # new number onto every purchase-side record that was auto-generated
    # from this booking, so "Purchase" doesn't keep showing the old AWB.
    # (old_docket_no is now computed earlier in this function.)
    if docket_no and docket_no != old_docket_no:
        try:
            linked_items = cdb.query(PurchaseInvoiceItem).filter_by(
                source_invoice_id=invoice.id
            ).all()

            # Fallback: same reasoning as the carrier_ref fallback below —
            # items created before source_invoice_id was wired up, or via
            # the manual purchase-entry screen, won't be linked yet. Match
            # them by the AWB they still carry (the *old* one, since we
            # haven't renumbered them yet) and backfill the link.
            if not linked_items and old_docket_no:
                linked_items = cdb.query(PurchaseInvoiceItem).filter_by(
                    docket_no=old_docket_no
                ).all()
                for pi_item in linked_items:
                    if not pi_item.source_invoice_id:
                        pi_item.source_invoice_id = invoice.id

            purchase_invoice_ids = set()
            for pi_item in linked_items:
                pi_item.docket_no = docket_no
                if pi_item.purchase_invoice_id:
                    purchase_invoice_ids.add(pi_item.purchase_invoice_id)

            # StockPurchaseHistory carries its own awb_no copy, not linked to
            # the booking directly — reach it via the purchase invoice(s) we
            # just found, matched on the old AWB so we don't touch unrelated
            # history rows on the same purchase invoice.
            if purchase_invoice_ids and old_docket_no:
                cdb.query(StockPurchaseHistory).filter(
                    StockPurchaseHistory.purchase_invoice_id.in_(purchase_invoice_ids),
                    StockPurchaseHistory.awb_no == old_docket_no,
                ).update({"awb_no": docket_no}, synchronize_session=False)

            cdb.commit()
        except Exception as e:
            cdb.rollback()
            print(f"[purchase-sync] could not update docket_no on linked purchase item for {invoice.invoice_id}: {e}")

    # ── Carrier/courier name sync — this is the actual field behind the
    # "supplier" column on the purchase screen (PurchaseInvoiceItem.courier_name
    # is set from this same "carrier" field at creation time — see
    # invoice_customer_save — but was never re-synced on edit, so changing
    # the carrier on a booking, e.g. DHL -> DPD, left the linked purchase
    # invoice item showing the old courier indefinitely).
    new_carrier = (request.form.get("carrier") or "").strip()
    old_carrier = (old_meta.get("carrier", "") or "").strip()
    if new_carrier and new_carrier != old_carrier:
        try:
            carrier_linked_items = cdb.query(PurchaseInvoiceItem).filter_by(
                source_invoice_id=invoice.id
            ).all()

            # Same fallback reasoning as the docket_no/carrier_ref syncs above:
            # match by the AWB this invoice carries right now (post any
            # renumbering already applied earlier in this same request).
            if not carrier_linked_items and docket_no:
                carrier_linked_items = cdb.query(PurchaseInvoiceItem).filter_by(
                    docket_no=docket_no
                ).all()
                for pi_item in carrier_linked_items:
                    if not pi_item.source_invoice_id:
                        pi_item.source_invoice_id = invoice.id

            for pi_item in carrier_linked_items:
                pi_item.courier_name = new_carrier

            cdb.commit()
        except Exception as e:
            cdb.rollback()
            print(f"[purchase-sync] could not update courier_name on linked purchase item for {invoice.invoice_id}: {e}")

    # ── Courier COMPANY change — this is the one that actually matters for
    # billing. "Carrier" (DHL/DPD) is just a label; "Courier Company"
    # (courier_company_id) is the Supplier the purchase bill belongs to.
    # Purchase invoices are one-per-(supplier, date) — see
    # invoice_customer_save — so this isn't a relabel, it's moving the line
    # (and its money) off the old supplier's bill and onto the new
    # supplier's bill for the same date, creating that bill if needed.
    new_courier_company_id = (request.form.get("courier_company_id") or "").strip()
    old_courier_company_id = (old_meta.get("courier_company_id", "") or "").strip()
    if new_courier_company_id and new_courier_company_id != old_courier_company_id:
        try:
            company_linked_items = cdb.query(PurchaseInvoiceItem).filter_by(
                source_invoice_id=invoice.id
            ).all()
            if not company_linked_items and docket_no:
                company_linked_items = cdb.query(PurchaseInvoiceItem).filter_by(
                    docket_no=docket_no
                ).all()
                for pi_item in company_linked_items:
                    if not pi_item.source_invoice_id:
                        pi_item.source_invoice_id = invoice.id

            new_supplier = cdb.query(Supplier).filter_by(
                id=int(new_courier_company_id), company_id=company_id
            ).first() if new_courier_company_id.isdigit() else None

            if new_supplier:
                for pi_item in company_linked_items:
                    old_pi = pi_item.purchase_invoice
                    if not old_pi or old_pi.supplier_id == new_supplier.id:
                        continue  # already on the correct bill

                    # Refuse to split a bill that's already been paid against.
                    # Moving money off a bill with a recorded payment would
                    # leave paid_amount > grand_total on the old bill — flag
                    # it for manual handling instead of guessing.
                    if (old_pi.paid_amount or 0) > 0:
                        flash(
                            f"Courier company changed, but purchase bill {old_pi.invoice_id} "
                            f"already has a payment recorded — the line for AWB "
                            f"{pi_item.docket_no or ''} was NOT moved automatically. "
                            f"Move it manually on the Purchases screen.",
                            "warning",
                        )
                        continue

                    target_pi = cdb.query(PurchaseInvoice).filter_by(
                        company_id=company_id,
                        supplier_id=new_supplier.id,
                        date=old_pi.date,
                    ).first()
                    if not target_pi:
                        new_pi_id = _next_numbered_id(
                            cdb, PurchaseInvoice.invoice_id,
                            "PURCHASE-INV-" + datetime.now().strftime("%Y%m%d") + "-"
                        )
                        target_pi = PurchaseInvoice(
                            invoice_id=new_pi_id,
                            company_id=company_id,
                            supplier_id=new_supplier.id,
                            invoice_number=None,
                            date=old_pi.date,
                            subtotal=0, tax_amount=0, grand_total=0,
                            paid_amount=0, balance=0, status="Pending",
                            created_at=datetime.utcnow(),
                        )
                        cdb.add(target_pi)
                        cdb.flush()

                    item_gst = (pi_item.cgst_amount or 0) + (pi_item.sgst_amount or 0) + (pi_item.igst_amount or 0)
                    item_total = pi_item.total_amount or 0

                    # NOTE: cgst/sgst/igst split is carried over as-is, not
                    # recalculated for the new supplier's state — if the two
                    # couriers are in different states that split may need a
                    # manual correction on the moved line.
                    old_pi.subtotal    = max(0, (old_pi.subtotal or 0) - (pi_item.taxable_value or 0))
                    old_pi.tax_amount  = max(0, (old_pi.tax_amount or 0) - item_gst)
                    old_pi.grand_total = max(0, (old_pi.grand_total or 0) - item_total)
                    old_pi.balance     = max(0, (old_pi.grand_total or 0) - (old_pi.paid_amount or 0))
                    old_supplier = cdb.get(Supplier, old_pi.supplier_id)
                    if old_supplier:
                        old_supplier.payable = max(0, (old_supplier.payable or 0) - item_total)

                    if pi_item.docket_no:
                        cdb.query(StockPurchaseHistory).filter(
                            StockPurchaseHistory.purchase_invoice_id == old_pi.id,
                            StockPurchaseHistory.awb_no == pi_item.docket_no,
                        ).update({"purchase_invoice_id": target_pi.id}, synchronize_session=False)

                    pi_item.purchase_invoice_id = target_pi.id

                    target_pi.subtotal    = (target_pi.subtotal or 0) + (pi_item.taxable_value or 0)
                    target_pi.tax_amount  = (target_pi.tax_amount or 0) + item_gst
                    target_pi.grand_total = (target_pi.grand_total or 0) + item_total
                    target_pi.balance     = max(0, (target_pi.grand_total or 0) - (target_pi.paid_amount or 0))
                    new_supplier.payable  = (new_supplier.payable or 0) + item_total

                    flash(
                        f"Courier company changed — purchase line for AWB "
                        f"{pi_item.docket_no or ''} moved from {old_pi.invoice_id} "
                        f"to {target_pi.invoice_id} ({new_supplier.name}).",
                        "info",
                    )

            cdb.commit()
        except Exception as e:
            cdb.rollback()
            print(f"[purchase-sync] could not move purchase item to new courier company for {invoice.invoice_id}: {e}")

    # ── Manifest entry sync — ManifestEntry has no link back to the booking
    # (no source_invoice_id equivalent), only docket_no. This used to be a
    # hand-rolled block that only synced docket_no/courier_name, which is why
    # editing a booking's box count (e.g. 20 -> 2) never updated
    # ManifestEntry.boxes or CompanyManifest.total_boxes — they stayed at
    # whatever was set on creation, forever. _sync_auto_manifest_entry()
    # already had the correct delta math for exactly this case (it's used by
    # invoice_customer_save on re-save); this route just never called it.
    # Calling it here means docket_no, courier_name, AND boxes all get
    # corrected together, immediately, on every edit.
    try:
        total_boxes_edit = int(sum(p["qty"] for p in packages_data)) or 1
        _sync_auto_manifest_entry(
            cdb, company_id, request.form.get("shipper_name", ""),
            new_carrier or (old_meta.get("carrier", "") or ""), action,
            invoice_date, docket_no, edit_invoice_id, total_boxes_edit,
            primary_stock_id=primary_stock_id,
            primary_stock_name=primary_stock_name,
            item_type=primary_item_type,
            old_docket_no=old_docket_no,
            booking_type=booking_type,
        )
        cdb.commit()
    except Exception as e:
        cdb.rollback()
        print(f"[purchase-sync] could not sync ManifestEntry for {invoice.invoice_id}: {e}")

    # ── Receipts party_name sync — cash/walk-in bookings have no per-shipper
    # Client row, so the typed name only lives as a plain string on each
    # CashTransaction/BankTransaction.party_name, stamped at the moment that
    # transaction was created. Only a NEW payment collected on this edit
    # (booking_payment_delta block above) ever got the fresh name — any
    # transaction already recorded on a previous save kept whatever name was
    # typed back then, so renaming the customer left old Receipts entries
    # showing the stale name forever. Refresh all of them, keyed by this
    # booking's invoice_id, whenever the name actually changed.
    if booking_type == "cash":
        new_shipper_name = (request.form.get("shipper_name", "") or "").strip()
        old_shipper_name = (old_meta.get("shipper_name", "") or "").strip()
        if new_shipper_name and new_shipper_name != old_shipper_name:
            try:
                cdb.query(CashTransaction).filter_by(
                    company_id=company_id, reference=edit_invoice_id
                ).update({"party_name": new_shipper_name}, synchronize_session=False)
                cdb.query(BankTransaction).filter_by(
                    company_id=company_id, reference=edit_invoice_id
                ).update({"party_name": new_shipper_name}, synchronize_session=False)
                cdb.commit()
            except Exception as e:
                cdb.rollback()
                print(f"[purchase-sync] could not refresh receipt party_name for {invoice.invoice_id}: {e}")

    old_carrier_ref = (old_meta.get("carrier_ref", "") or "").strip()
    if new_carrier_ref and new_carrier_ref != old_carrier_ref:
        try:
            updated_count = cdb.query(PurchaseInvoiceItem).filter_by(
                source_invoice_id=invoice.id
            ).update({"carrier_ref": new_carrier_ref})

            # Fallback: no item is linked via source_invoice_id — either it was
            # created through the manual purchase-entry screen (which never sets
            # source_invoice_id), or it was auto-generated before that link was
            # wired up correctly. Match by docket_no instead, update it, and
            # backfill source_invoice_id so future edits hit the fast path above.
            if not updated_count and docket_no:
                fallback_items = cdb.query(PurchaseInvoiceItem).filter_by(
                    docket_no=docket_no
                ).all()
                for pi_item in fallback_items:
                    pi_item.carrier_ref = new_carrier_ref
                    if not pi_item.source_invoice_id:
                        pi_item.source_invoice_id = invoice.id

            cdb.commit()
        except Exception as e:
            cdb.rollback()
            print(f"[purchase-sync] could not update carrier_ref on linked purchase item for {invoice.invoice_id}: {e}")

    old_tracking_number = (old_meta.get("tracking_number", "") or "").strip()
    new_tracking_number = (request.form.get("tracking_number") or "").strip()
    if new_tracking_number and new_tracking_number != old_tracking_number:
        try:
            from tasks import send_tracking_update_notification_async
            send_tracking_update_notification_async(
                company_id=company_id,
                invoice_id=invoice.invoice_id,
                carrier=request.form.get("carrier", ""),
                tracking_number=new_tracking_number,
            )
        except Exception as e:
            print(f"[whatsapp] could not queue tracking-update notification for {invoice.invoice_id}: {e}")

    # ── Save Performa Invoice items (linked Estimate) ───────────────────────────
    # This block was missing here entirely — invoice_customer_save (create) had it,
    # invoice_customer_update (edit) did not, so edits to performa items never persisted.
    perf_descs   = request.form.getlist("perf_desc[]")
    perf_boxes   = request.form.getlist("perf_box[]")
    perf_hsns    = request.form.getlist("perf_hsn[]")
    perf_units   = request.form.getlist("perf_unit[]")
    perf_witems  = request.form.getlist("perf_weight_item[]")
    perf_qtys   = request.form.getlist("perf_qty[]")
    perf_rates  = request.form.getlist("perf_rate[]")
    perf_weight = request.form.get("perf_weight", "0.00").strip()
    perf_ref    = request.form.get("perf_reference", "").strip()
    perf_inv_no   = request.form.get("performa_invoice_no", "").strip()
    perf_inv_date = request.form.get("performa_invoice_date", "").strip()
    perf_export_reason = request.form.get("export_reason", "").strip()
    if perf_export_reason == "Other":
        _perf_export_reason_other = request.form.get("export_reason_other", "").strip()
        if _perf_export_reason_other:
            perf_export_reason = _perf_export_reason_other

    perf_items = []
    perf_subtotal = 0.0
    for i in range(len(perf_descs)):
        desc = (perf_descs[i] or "").strip()
        if not desc:
            continue
        qty  = float(perf_qtys[i])  if i < len(perf_qtys)  and perf_qtys[i]  else 0.0
        rate = float(perf_rates[i]) if i < len(perf_rates) and perf_rates[i] else 0.0
        perf_subtotal += qty * rate
        perf_items.append({
            "description": desc,
            "box": perf_boxes[i] if i < len(perf_boxes) else "",
            "hsn": perf_hsns[i] if i < len(perf_hsns) else "",
            "unit": perf_units[i] if i < len(perf_units) and perf_units[i] else "PCS",
            "weight": float(perf_witems[i] or 0) if i < len(perf_witems) and perf_witems[i] else 0,
            "qty": qty,
            "rate": rate,
        })

    def _fmt_addr_pi(a1, a2, city, state, pin, country):
        return ", ".join(p for p in [a1, a2, city, state, pin, country] if p)

    perf_terms = json.dumps({
        "docket_no":        docket_no,
        "linked_invoice_id": invoice.invoice_id,
        "shipper_name":     request.form.get("shipper_name", ""),
        "shipper_phone":    request.form.get("customer_phone", ""),
        "shipper_address1": request.form.get("shipper_address1", ""),
        "shipper_address2": request.form.get("shipper_address2", ""),
        "shipper_city":     request.form.get("shipper_city", ""),
        "shipper_state":    request.form.get("shipper_state", ""),
        "shipper_pincode":  request.form.get("shipper_pincode", ""),
        "shipper_country":  request.form.get("shipper_country", "India"),
        "shipper_address":  _fmt_addr_pi(
            request.form.get("shipper_address1",""), request.form.get("shipper_address2",""),
            request.form.get("shipper_city",""), request.form.get("shipper_state",""),
            request.form.get("shipper_pincode",""), request.form.get("shipper_country",""),
        ),
        "receiver_name":    request.form.get("receiver_name", ""),
        "receiver_phone":   request.form.get("receiver_phone", ""),
        "receiver_company": "",
        "receiver_address1": request.form.get("receiver_address1", ""),
        "receiver_address2": request.form.get("receiver_address2", ""),
        "receiver_city":    request.form.get("receiver_city", ""),
        "receiver_state":   request.form.get("receiver_state", ""),
        "receiver_pincode": request.form.get("receiver_pincode", ""),
        "receiver_country": request.form.get("receiver_country", "India"),
        "receiver_address": _fmt_addr_pi(
            request.form.get("receiver_address1",""), request.form.get("receiver_address2",""),
            request.form.get("receiver_city",""), request.form.get("receiver_state",""),
            request.form.get("receiver_pincode",""), request.form.get("receiver_country",""),
        ),
        "destination":  request.form.get("destination", ""),
        "weight":       perf_weight,
        "reference":    perf_ref,
        "invoice_no":   perf_inv_no,
        "invoice_date": perf_inv_date,
        "export_reason": perf_export_reason,
        "line_items":   perf_items,
        "dimensions":   [],
    })

    existing_est = cdb.query(Estimate).filter_by(
        company_id=company_id
    ).filter(
        Estimate.terms.like(f'%"linked_invoice_id": "{invoice.invoice_id}"%')
    ).first()

    if perf_items:
        if existing_est:
            existing_est.client_id      = client_id
            existing_est.date           = date.fromisoformat(invoice_date)
            existing_est.status         = "Paid"
            existing_est.contact_person = request.form.get("shipper_contact_name", "")
            existing_est.phone          = request.form.get("customer_phone", "")
            existing_est.subtotal       = perf_subtotal
            existing_est.grand_total    = perf_subtotal
            existing_est.tax_amount     = 0
            existing_est.terms          = perf_terms
            cdb.query(EstimateItem).filter_by(estimate_id=existing_est.id).delete()
            for item in perf_items:
                cdb.add(EstimateItem(
                    estimate_id=existing_est.id,
                    description=item["description"],
                    qty=item["qty"],
                    rate=item["rate"],
                    discount=0,
                ))
        else:
            est_id = _next_numbered_id(cdb, Estimate.estimate_id, "SHIP-" + datetime.now().strftime("%Y%m%d") + "-", extra_filters=[Estimate.company_id == company_id])
            est = Estimate(
                estimate_id    = est_id,
                company_id     = company_id,
                client_id      = client_id,
                date           = date.fromisoformat(invoice_date),
                status         = "Paid",
                contact_person = request.form.get("shipper_contact_name", ""),
                phone          = request.form.get("customer_phone", ""),
                subtotal       = perf_subtotal,
                grand_total    = perf_subtotal,
                tax_amount     = 0,
                terms          = perf_terms,
            )
            cdb.add(est)
            cdb.flush()
            for item in perf_items:
                cdb.add(EstimateItem(
                    estimate_id  = est.id,
                    description  = item["description"],
                    qty          = item["qty"],
                    rate         = item["rate"],
                    discount     = 0,
                ))
    elif existing_est:
        # All performa rows were cleared in the edit form — remove the stale Estimate.
        cdb.query(EstimateItem).filter_by(estimate_id=existing_est.id).delete()
        cdb.delete(existing_est)

    cdb.commit()

    flash(f"Customer invoice {invoice.invoice_id} updated successfully!")
    return redirect(url_for("invoice_list"))

@app.route("/booking/view/<invoice_id>")
@login_required
@require_permission("invoices", "view")
def invoice_view(invoice_id):
    cdb = get_cdb()
    company_id = get_current_company()
    inv        = _first_or_404(cdb.query(Invoice).filter_by(invoice_id=invoice_id, company_id=company_id).first())

    # Resolve customer name & phone
    if inv.client_obj:
        customer_name  = inv.client_obj.name
        customer_phone = inv.client_obj.phone or inv.phone or ""
        client_display_id = inv.client_obj.client_id or ""
    else:
        customer_name  = inv.contact_person or "—"
        customer_phone = inv.phone or ""
        client_display_id = ""

    total    = inv.grand_total or 0.0
    subtotal = inv.subtotal    or 0.0
    tax      = inv.tax_amount  or 0.0

    # Resale/return charges
    resale_charges = getattr(inv, 'resale_charges', 0) or 0
    resale_gst     = round(resale_charges * 0.18, 2)
    resale_total   = resale_charges + resale_gst

    # Derive paid / balance / tab-status from DB status
    db_status = (inv.status or "").lower()
    if db_status == "paid":
        paid       = total
        balance    = 0.0
        tab_status = "paid"
    elif db_status == "partial":
        paid       = subtotal
        balance    = total - paid
        tab_status = "partial"
    else:
        paid       = 0.0
        balance    = total
        tab_status = "pending"

    # Normalize line items
    items = []
    for li in inv.items:
        qty      = li.qty      or 0.0
        rate     = li.rate     or 0.0
        discount = li.discount or 0.0
        items.append({
            "code":     li.code        or "",
            "desc":     li.description or "",
            "qty":      qty,
            "rate":     rate,
            "discount": discount,
            "amount":   qty * rate * (1 - discount / 100),
        })

    # Unpack shipment metadata stored as JSON in inv.terms
    meta = {}
    if inv.terms:
        try:
            meta = json.loads(inv.terms)
        except (ValueError, TypeError):
            meta = {}

    # ── Build complete invoice dict with ALL fields ──
    invoice = {
        "id":               inv.invoice_id,
        "client_display_id": client_display_id,
        "date":             inv.date,
        "due_date":         inv.due_date,
        "status":           tab_status,
        "customer_name":    customer_name,
        "customer_phone":   customer_phone,
        "subtotal":         subtotal,
        "tax":              tax,
        "total":            total,
        "paid":             paid,
        "balance":          balance,
        "bill_type":        "credit",
        "items":            items,
        "related_orders":   [],
        
        # ── SHIPPER / CONSIGNOR FIELDS ──
        "docket_no":        meta.get("docket_no", inv.invoice_id),
        "shipper_name":     meta.get("shipper_name", inv.contact_person or ""),
        "shipper_contact_name": meta.get("shipper_contact_name", ""),
        "shipper_address1": meta.get("shipper_address1", meta.get("shipper_address", "")),
        "shipper_address2": meta.get("shipper_address2", ""),
        "shipper_city":     meta.get("shipper_city", ""),
        "shipper_state":    meta.get("shipper_state", ""),
        "shipper_pincode":  meta.get("shipper_pincode", ""),
        "shipper_country":  meta.get("shipper_country", "India"),
        "shipper_doc_type": meta.get("shipper_doc_type", ""),
        "shipper_doc_no":   meta.get("shipper_doc_no", ""),
        "client_code":      meta.get("client_code", ""),
        
        # ── RECEIVER / CONSIGNEE FIELDS ──
        "receiver_name":    meta.get("receiver_name", ""),
        "receiver_company": meta.get("receiver_company", ""),
        "receiver_phone":   meta.get("receiver_phone", ""),
        "receiver_address1": meta.get("receiver_address1", meta.get("receiver_address", "")),
        "receiver_address2": meta.get("receiver_address2", ""),
        "receiver_city":    meta.get("receiver_city", ""),
        "receiver_state":   meta.get("receiver_state", ""),
        "receiver_pincode": meta.get("receiver_pincode", ""),
        "receiver_country": meta.get("receiver_country", "India"),
        "receiver_doc_type": meta.get("receiver_doc_type", ""),
        "receiver_doc_no":  meta.get("receiver_doc_no", ""),
        
        # ── SHIPMENT / SERVICE FIELDS ──
        "destination":      meta.get("destination", ""),
        "origin":           meta.get("origin", "India"),
        "shipment_type":    meta.get("shipment_type", ""),
        "mode":             meta.get("mode", ""),
        "carrier":          meta.get("carrier", ""),
        "tracking_number":  meta.get("tracking_number", ""),
        "carrier_ref":      meta.get("carrier_ref", ""),
        "vendor":           meta.get("vendor", ""),
        "product":          meta.get("shipment_type", ""),
        
        # ── CHARGES ──
        "payment_mode":     meta.get("payment_mode", "credit"),
        "upi_app":          meta.get("upi_app", ""),
        "transaction_id":   meta.get("upi_ref", ""),
        "cheque_no":        meta.get("cheque_no", ""),
        "cheque_bank":      meta.get("cheque_bank", ""),
        "freight":          meta.get("freight", subtotal),
        "freight_weight":   meta.get("freight_weight", 0),
        "freight_rate_per_kg": meta.get("freight_rate_per_kg", 0),
        "fuel_charge":      meta.get("fuel", 0),
        "other_charges":    meta.get("other", 0),
        "other_charges_reason": meta.get("other_charges_reason", ""),
        "discount":         meta.get("discount", 0),
        "booking_type":     meta.get("booking_type", "credit"),
        "notes":            inv.email or "",
        
        # ── PACKAGES ──
        "packages":         meta.get("packages", []),
        
        # ── RESALE ──
        "has_resale":       getattr(inv, 'has_resale', False),
        "resale_charges":   resale_charges,
        "resale_gst":       resale_gst,
        "resale_total":     resale_total,
        "resale_reason":    getattr(inv, 'resale_reason', '') or '',
        "resale_date":      getattr(inv, 'resale_date', None),
        
        # ── ID DOCUMENTS ──
        "shipper_aadhar_file": meta.get("shipper_aadhar_file", ""),
        "shipper_pan_file":    meta.get("shipper_pan_file", ""),
        
        # ── PERFORMA ──
        "performa_items":    [],
        "perf_weight":       "",
        "perf_reference":    "",
        "performa_invoice_no": "",
        "performa_invoice_date": "",
        "export_reason":     "",
    }

    # ── ID documents: credit bookings read from the linked Client's own record ──
    if inv.client_obj:
        cl = inv.client_obj
        id_docs = {
            "aadhar_front": (cl.aadhar_front_file, "client_docs") if cl.aadhar_front_file else None,
            "aadhar_back":  (cl.aadhar_back_file,  "client_docs") if cl.aadhar_back_file  else None,
            "pan_front":    (cl.pan_front_file,    "client_docs") if cl.pan_front_file    else None,
            "pan_back":     (cl.pan_back_file,     "client_docs") if cl.pan_back_file     else None,
        }
        id_docs_source = "client"
    else:
        id_docs = {
            "aadhar_front": (meta.get("shipper_aadhar_front_file"), "invoice_docs") if meta.get("shipper_aadhar_front_file") else None,
            "aadhar_back":  (meta.get("shipper_aadhar_back_file"),  "invoice_docs") if meta.get("shipper_aadhar_back_file")  else None,
            "pan_front":    (meta.get("shipper_pan_front_file"),    "invoice_docs") if meta.get("shipper_pan_front_file")    else None,
            "pan_back":     (meta.get("shipper_pan_back_file"),     "invoice_docs") if meta.get("shipper_pan_back_file")     else None,
        }
        id_docs_source = "booking"
        # Legacy fallback
        if not id_docs["aadhar_front"] and meta.get("shipper_aadhar_file"):
            id_docs["aadhar_front"] = (meta.get("shipper_aadhar_file"), "invoice_docs")
        if not id_docs["pan_front"] and meta.get("shipper_pan_file"):
            id_docs["pan_front"] = (meta.get("shipper_pan_file"), "invoice_docs")

    invoice["id_docs"] = id_docs
    invoice["id_docs_source"] = id_docs_source
    invoice["has_id_docs"] = any(id_docs.values())
    
    # ── Load linked Performa Invoice items ──
    linked_est = cdb.query(Estimate).filter_by(company_id=company_id).filter(
        Estimate.terms.like(f'%"linked_invoice_id": "{inv.invoice_id}"%')
    ).first()
    if linked_est and linked_est.terms:
        try:
            perf_meta = json.loads(linked_est.terms)
            invoice["performa_items"] = perf_meta.get("line_items", [])
            invoice["perf_weight"]    = perf_meta.get("weight", "")
            invoice["perf_reference"] = perf_meta.get("reference", "")
            invoice["performa_invoice_no"]   = perf_meta.get("invoice_no", "")
            invoice["performa_invoice_date"] = perf_meta.get("invoice_date", "")
            invoice["export_reason"]         = perf_meta.get("export_reason", "")
        except Exception:
            pass

    # ── Derive pieces and chargeable weight ──
    pkg_list = invoice["packages"] or []
    invoice["pieces"] = sum((p.get("qty") or 1) for p in pkg_list) if pkg_list else 1
    pkg_weight_total = sum(
        max(p.get("weight") or 0, p.get("vol_weight") or 0) * (p.get("qty") or 1)
        for p in pkg_list
    )
    invoice["weight"] = pkg_weight_total if pkg_weight_total > 0 else invoice.get("freight_weight", 0)

    pkg_actual_weight_total = sum(
        (p.get("weight") or 0) * (p.get("qty") or 1)
        for p in pkg_list
    )
    pkg_discount_wt_total = sum(
        (p.get("discount_wt") or 0) * (p.get("qty") or 1)
        for p in pkg_list
    )
    invoice["actual_weight"] = (
        max(pkg_actual_weight_total - pkg_discount_wt_total, 0)
        if pkg_actual_weight_total > 0 else invoice.get("freight_weight", 0)
    )

    invoice["clone_url"] = url_for("invoice_clone", invoice_id=inv.invoice_id)

    return render_template("booking_view.html", invoice=invoice)

def _generate_temp_password(length=10):
    # Avoid visually ambiguous chars (0/O, 1/l/I) since this gets read off-screen and typed by hand
    alphabet = "ABCDEFGHJKMNPQRSTUVWXYZabcdefghjkmnpqrstuvwxyz23456789"
    return "".join(secrets.choice(alphabet) for _ in range(length))


@app.route("/company/reset-user-password/<email>", methods=["POST"])
@login_required
@owner_required
def reset_user_password(email):
    """Reset one person's password everywhere they have access — mirrors
    delete_company_user's pattern of walking every one of the owner's
    companies, since the same email can have a separate CompanyUser row
    (and separate password_hash) in each company's own database.
    """
    owner_email = get_current_user().get("email", "").strip().lower()
    owner_companies = get_owner_companies(owner_email)
    email = email.strip().lower()

    temp_password = _generate_temp_password()
    new_hash = hash_password(temp_password)

    reset_in = []
    full_name = None
    for c in owner_companies:
        _cdb = get_customer_session(c.company_id)
        user = _cdb.query(CompanyUser).filter_by(company_id=c.company_id, email=email).first()
        if user and user.role != "owner" and user.is_active:
            user.password_hash = new_hash
            _cdb.commit()
            reset_in.append(c.company_name)
            full_name = full_name or user.full_name

    if reset_in:
        flash(
            f"Password for {full_name or email} reset across: {', '.join(reset_in)}. "
            f"Temporary password: {temp_password} — share this securely, it will not be shown again.",
            "success"
        )
    else:
        flash("Could not reset password — no active non-owner account found for this email.", "error")
    return redirect(url_for("company_settings"))


@app.route("/booking/pdf/<invoice_id>")
def invoice_pdf(invoice_id):
    """Customer Invoice PDF — uses xhtml2pdf"""
    from xhtml2pdf import pisa
    import io

    token = request.args.get("token")
    if token:
        token_company_id, token_invoice_id = verify_pdf_token(token)
        if not token_company_id or token_invoice_id != invoice_id:
            abort(404)
        company_id = token_company_id
        cdb = get_customer_session(company_id, db_session=db.session)
    else:
        if "user" not in session:
            flash("Please login to continue")
            return redirect(url_for("login"))
        company_id = get_current_company()
        cdb = get_cdb()

    # Get invoice data
    inv = _first_or_404(cdb.query(Invoice).filter_by(invoice_id=invoice_id, company_id=company_id).first())

    # ─── Build invoice dict ──────────────────────────────────────────────
    if inv.client_obj:
        customer_name = inv.client_obj.name
        customer_phone = inv.client_obj.phone or inv.phone or ""
        client_display_id = inv.client_obj.client_id or ""
    else:
        customer_name = inv.contact_person or "—"
        customer_phone = inv.phone or ""
        client_display_id = ""

    total = inv.grand_total or 0.0
    subtotal = inv.subtotal or 0.0
    tax = inv.tax_amount or 0.0

    meta = {}
    if inv.terms:
        try:
            meta = json.loads(inv.terms)
        except (ValueError, TypeError):
            meta = {}

    # Build invoice dict (same shape as invoice_view)
    invoice = {
        "id": inv.invoice_id,
        "client_display_id": client_display_id,
        "date": inv.date,
        "due_date": inv.due_date,
        "status": inv.status or "Pending",
        "customer_name": customer_name,
        "customer_phone": customer_phone,
        "subtotal": subtotal,
        "tax": tax,
        "total": total,
        "paid": inv.paid_amount or 0,
        "balance": inv.balance or 0,
        "docket_no": meta.get("docket_no", inv.invoice_id),
        "shipper_name": meta.get("shipper_name", inv.contact_person or ""),
        "shipper_address1": meta.get("shipper_address1", ""),
        "shipper_address2": meta.get("shipper_address2", ""),
        "shipper_city": meta.get("shipper_city", ""),
        "shipper_state": meta.get("shipper_state", ""),
        "shipper_pincode": meta.get("shipper_pincode", ""),
        "shipper_country": meta.get("shipper_country", "India"),
        "receiver_name": meta.get("receiver_name", ""),
        "receiver_phone": meta.get("receiver_phone", ""),
        "receiver_address1": meta.get("receiver_address1", ""),
        "receiver_address2": meta.get("receiver_address2", ""),
        "receiver_city": meta.get("receiver_city", ""),
        "receiver_state": meta.get("receiver_state", ""),
        "receiver_pincode": meta.get("receiver_pincode", ""),
        "receiver_country": meta.get("receiver_country", "India"),
        "destination": meta.get("destination", ""),
        "origin": meta.get("origin", "India"),
        "shipment_type": meta.get("shipment_type", ""),
        "mode": meta.get("mode", ""),
        "carrier": meta.get("carrier", ""),
        "carrier_ref": meta.get("carrier_ref", ""),
        "payment_mode": meta.get("payment_mode", "credit"),
        "freight": meta.get("freight", subtotal),
        "freight_weight": meta.get("freight_weight", 0),
        "freight_rate_per_kg": meta.get("freight_rate_per_kg", 0),
        "fuel_charge": meta.get("fuel", 0),
        "other_charges": meta.get("other", 0),
        "discount": meta.get("discount", 0),
        "notes": inv.email or "",
        "packages": meta.get("packages", []),
        "pieces": sum((p.get("qty") or 1) for p in meta.get("packages", [])) if meta.get("packages") else 1,
        "weight": sum(
            max(p.get("weight") or 0, p.get("vol_weight") or 0) * (p.get("qty") or 1)
            for p in meta.get("packages", [])
        ),
        "vendor": meta.get("vendor", ""),
        "product": meta.get("shipment_type", ""),
    }

    # Get linked performa items if any
    linked_est = cdb.query(Estimate).filter_by(company_id=company_id).filter(
        Estimate.terms.like(f'%"linked_invoice_id": "{inv.invoice_id}"%')
    ).first()
    if linked_est and linked_est.terms:
        try:
            perf_meta = json.loads(linked_est.terms)
            invoice["performa_items"] = perf_meta.get("line_items", [])
            invoice["perf_weight"] = perf_meta.get("weight", "")
            invoice["perf_reference"] = perf_meta.get("reference", "")
            invoice["performa_invoice_no"] = perf_meta.get("invoice_no", "")
            invoice["performa_invoice_date"] = perf_meta.get("invoice_date", "")
            invoice["export_reason"] = perf_meta.get("export_reason", "")
        except Exception:
            pass

    company = Company.query.filter_by(company_id=company_id).first()

    # ─── Render HTML ──────────────────────────────────────────────────────
    html_content = render_template(
        "booking_pdf.html",
        invoice=invoice,
        company=company,
        company_logo_url=url_for('static', filename=f'company_logos/{company.logo_filename}', _external=True) if company and company.logo_filename else None,
        is_gst_registered=company.is_gst_registered if company else True,
        today=today_ist().strftime("%d %b %Y"),
    )

    # ─── Convert to PDF ──────────────────────────────────────────────────
    pdf_file = io.BytesIO()
    pisa_status = pisa.CreatePDF(html_content, dest=pdf_file, encoding='UTF-8')

    if pisa_status.err:
        if token:
            abort(500)
        flash(f"PDF generation error: {pisa_status.err}")
        return redirect(url_for("invoice_view", invoice_id=invoice_id))

    pdf_file.seek(0)

    return send_file(
        pdf_file,
        as_attachment=True,
        download_name=f"Invoice_{invoice_id}.pdf",
        mimetype="application/pdf"
    )

# ─────────────────────────────────────────────────────────────────────────────
# ── Resale / Return Charges Routes ──────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────

@app.route("/booking/<invoice_id>/resale-charges", methods=["GET", "POST"])
@login_required
@require_permission("invoices", "view", method_actions={'POST': 'edit'})
def invoice_resale_charges(invoice_id):
    """Add return/resale charges to an existing invoice"""
    cdb = get_cdb()
    company_id = get_current_company()
    
    invoice = cdb.query(Invoice).filter_by(invoice_id=invoice_id, company_id=company_id).first()
    if not invoice:
        flash("Invoice not found")
        return redirect(url_for("invoice_list"))
    
    if request.method == "POST":
        action = request.form.get("action", "add")
        
        if action == "add":
            resale_amount = float(request.form.get("resale_amount", 0) or 0)
            resale_reason = request.form.get("resale_reason", "").strip()
            resale_date_str = request.form.get("resale_date")
            resale_notes = request.form.get("resale_notes", "").strip()
            
            if resale_amount <= 0:
                flash("Please enter a valid resale charge amount")
                return redirect(url_for("invoice_resale_charges", invoice_id=invoice_id))
            
            if not resale_reason:
                flash("Please select or enter a reason for the resale charge")
                return redirect(url_for("invoice_resale_charges", invoice_id=invoice_id))
            
            # Calculate GST on resale charges (if company is GST registered)
            co = Company.query.filter_by(company_id=company_id).first()
            apply_gst = co.is_gst_registered if (co and hasattr(co, 'is_gst_registered')) else True
            gst_on_resale = round(resale_amount * 0.18, 2) if apply_gst else 0.0
            
            # Update invoice with resale charges
            invoice.has_resale = True
            invoice.resale_charges = resale_amount
            invoice.resale_reason = resale_reason
            invoice.resale_date = date.fromisoformat(resale_date_str) if resale_date_str else today_ist()
            invoice.resale_notes = resale_notes
            
            # Update totals - add resale amount to existing totals
            old_grand_total = invoice.grand_total or 0
            old_tax_amount = invoice.tax_amount or 0
            
            # Add resale amount to subtotal (or you can add separately)
            invoice.subtotal = (invoice.subtotal or 0) + resale_amount
            invoice.tax_amount = (invoice.tax_amount or 0) + gst_on_resale
            invoice.grand_total = (invoice.grand_total or 0) + resale_amount + gst_on_resale
            
            # Update balance
            invoice.balance = (invoice.balance or 0) + resale_amount + gst_on_resale
            
            # Update status if balance > 0
            if invoice.balance > 0:
                invoice.status = "Partial" if invoice.status == "Paid" else invoice.status
            
            # Store resale details in terms JSON for easy retrieval
            try:
                meta = json.loads(invoice.terms) if invoice.terms else {}
            except:
                meta = {}
            
            meta["resale"] = {
                "amount": resale_amount,
                "gst": gst_on_resale,
                "reason": resale_reason,
                "date": (date.fromisoformat(resale_date_str) if resale_date_str else today_ist()).strftime("%Y-%m-%d"),
                "notes": resale_notes,
                "added_by": get_current_user().get("email")
            }
            invoice.terms = json.dumps(meta)
            
            # Create a cash transaction for the resale charge
            # (this is a NEW charge, so it's income)
            cash_txn = CashTransaction(
                company_id=company_id,
                type="income",
                date=date.fromisoformat(resale_date_str) if resale_date_str else today_ist(),
                category="Resale Charges",
                description=f"Resale charge for invoice {invoice_id}: {resale_reason}",
                amount=resale_amount + gst_on_resale,
                reference=invoice_id,
                notes=f"Resale charge - {resale_reason}\n{resale_notes}",
                created_by=get_current_user().get("email")
            )
            cdb.add(cash_txn)
            
            # Update client pending balance
            client = cdb.query(Client).filter_by(id=invoice.client_id, company_id=company_id).first()
            if client and hasattr(client, "pending"):
                client.pending = (client.pending or 0) + resale_amount + gst_on_resale
            
            cdb.commit()
            flash(f"✅ Resale charge of ₹{resale_amount:,.2f} (+ GST ₹{gst_on_resale:,.2f}) added to invoice {invoice_id}")
            
        elif action == "remove":
            # Remove resale charges from invoice
            if invoice.has_resale:
                # Restore original totals (subtract resale charges)
                try:
                    meta = json.loads(invoice.terms) if invoice.terms else {}
                    resale_data = meta.get("resale", {})
                    resale_amount = resale_data.get("amount", 0)
                    resale_gst = resale_data.get("gst", 0)
                    
                    # Subtract from totals
                    invoice.subtotal = max(0, (invoice.subtotal or 0) - resale_amount)
                    invoice.tax_amount = max(0, (invoice.tax_amount or 0) - resale_gst)
                    invoice.grand_total = max(0, (invoice.grand_total or 0) - resale_amount - resale_gst)
                    invoice.balance = max(0, (invoice.balance or 0) - resale_amount - resale_gst)
                    
                    # Update client pending balance
                    client = cdb.query(Client).filter_by(id=invoice.client_id, company_id=company_id).first()
                    if client and hasattr(client, "pending"):
                        client.pending = max(0, (client.pending or 0) - resale_amount - resale_gst)
                    
                    # Remove resale data from terms
                    meta.pop("resale", None)
                    invoice.terms = json.dumps(meta) if meta else None
                    
                    invoice.has_resale = False
                    invoice.resale_charges = 0
                    invoice.resale_reason = None
                    invoice.resale_date = None
                    invoice.resale_notes = None
                    
                    # Recalculate status
                    if invoice.balance <= 0:
                        invoice.status = "Paid"
                    elif invoice.paid_amount > 0:
                        invoice.status = "Partial"
                    
                    cdb.commit()
                    flash(f"✅ Resale charges removed from invoice {invoice_id}")
                except Exception as e:
                    cdb.rollback()
                    flash(f"Error removing resale charges: {str(e)}", "error")
            else:
                flash("No resale charges found on this invoice", "warning")
        
        return redirect(url_for("invoice_view", invoice_id=invoice_id))
    
    # GET - show the form
    # Parse existing terms to see if there's already resale data
    resale_data = None
    try:
        meta = json.loads(invoice.terms) if invoice.terms else {}
        resale_data = meta.get("resale")
    except:
        pass
    
    return render_template("booking_resale.html", 
                         invoice=invoice, 
                         resale_data=resale_data,
                         today=str(today_ist()))

# ─────────────────────────────────────────────────────────────────────────────
# ── Customer Invoice (Shipment) ───────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────

AWB_PREFIX   = "AHL"
AWB_START    = 81000          # first number: AHL81000
AWB_COUNTER_KEY = "awb_last" # we store the last-used counter in a tiny helper


"""def _next_awb_number(company_id: int) -> str:
    
    # Count how many customer invoices already have an AHL docket number
    existing_count = (
        cdb.query(Invoice)
        .filter(
            Invoice.company_id == company_id,
            Invoice.terms.like("AWB:AHL%"),   # we embed the AWB in terms for storage
        )
        .count()
    )
    # Alternatively, just count all customer-type invoices for this company
    # (simpler and still gapless)
    cust_count = (
        cdb.query(Invoice)
        .filter(
            Invoice.company_id == company_id,
            Invoice.invoice_id.like("CUST-%"),
        )
        .count()
    )
    seq = AWB_START + cust_count
    return f"{AWB_PREFIX}{seq}"""
# Replace the _next_awb_number function in app.py (around line 650)

def _next_awb_number(company_id):
    """Generate next sequential AWB using company-specific prefix + start.
       Thread-safe: uses database MAX() on a dedicated column.
    """
    from platform_models import Company as PlatformCompany
    from sqlalchemy import func, and_
    import re
    
    co = PlatformCompany.query.filter_by(company_id=company_id).first()
    prefix = "AHL" if (co is None or co.awb_prefix is None) else co.awb_prefix
    awb_start = (co.awb_start if co else None) or 81000

    cdb = get_cdb()
    
    # Method 1: Query using SQL MAX on the extracted number
    # This is the most reliable way - get the highest number directly from the column
    # We use a SQL expression to extract the numeric part after the prefix
    
    # First, get all docket numbers from invoice terms
    rows = (
        cdb.query(Invoice.terms)
        .filter(Invoice.company_id == company_id)
        .filter(Invoice.terms.isnot(None))
        .all()
    )
    
    max_seq = awb_start - 1
    pattern = re.compile(rf'{re.escape(prefix)}(\d+)')
    
    for (terms,) in rows:
        try:
            if terms:
                # Parse JSON safely
                if isinstance(terms, str):
                    meta = json.loads(terms)
                    dno = meta.get("docket_no", "")
                    if dno and dno.startswith(prefix):
                        match = pattern.search(dno)
                        if match:
                            num = int(match.group(1))
                            if num > max_seq:
                                max_seq = num
        except (ValueError, TypeError, json.JSONDecodeError):
            continue
    
    # Also check the docket_no field if it exists as a direct column
    # (some invoices might have docket_no stored directly)
    try:
        # Check if there's a direct docket_no column we can use
        # This is a fallback in case some invoices store docket_no differently
        from sqlalchemy import inspect
        inspector = inspect(cdb.bind)
        columns = [c['name'] for c in inspector.get_columns('invoices')]
        
        if 'docket_no' in columns:
            direct_rows = (
                cdb.query(Invoice.docket_no)
                .filter(Invoice.company_id == company_id)
                .filter(Invoice.docket_no.isnot(None))
                .filter(Invoice.docket_no != '')
                .all()
            )
            for (dno,) in direct_rows:
                if dno and dno.startswith(prefix):
                    match = pattern.search(dno)
                    if match:
                        num = int(match.group(1))
                        if num > max_seq:
                            max_seq = num
    except Exception:
        pass
    
    # Generate the next number
    next_seq = max_seq + 1
    new_awb = f"{prefix}{next_seq}"
    
    # Safety check - verify this AWB doesn't already exist in the database
    # This catches any race condition where two requests generated the same number
    existing = (
        cdb.query(Invoice)
        .filter(Invoice.company_id == company_id)
        .filter(Invoice.terms.like(f'%"{new_awb}"%'))
        .first()
    )
    
    # Also check direct docket_no column if it exists
    try:
        from sqlalchemy import inspect
        inspector = inspect(cdb.bind)
        columns = [c['name'] for c in inspector.get_columns('invoices')]
        if 'docket_no' in columns:
            existing_direct = (
                cdb.query(Invoice)
                .filter(Invoice.company_id == company_id)
                .filter(Invoice.docket_no == new_awb)
                .first()
            )
            if existing_direct:
                existing = existing_direct
    except Exception:
        pass
    
    # If we found an existing AWB with the same number, increment until we find a free one
    retry_count = 0
    while existing and retry_count < 100:
        next_seq += 1
        new_awb = f"{prefix}{next_seq}"
        existing = (
            cdb.query(Invoice)
            .filter(Invoice.company_id == company_id)
            .filter(Invoice.terms.like(f'%"{new_awb}"%'))
            .first()
        )
        # Also check direct docket_no column
        try:
            from sqlalchemy import inspect
            inspector = inspect(cdb.bind)
            columns = [c['name'] for c in inspector.get_columns('invoices')]
            if 'docket_no' in columns:
                existing_direct = (
                    cdb.query(Invoice)
                    .filter(Invoice.company_id == company_id)
                    .filter(Invoice.docket_no == new_awb)
                    .first()
                )
                if existing_direct:
                    existing = existing_direct
        except Exception:
            pass
        retry_count += 1
    
    return new_awb


def _check_credit_limit(company, client, new_bill_amount):
    """
    Checks a client's credit limit against (current outstanding + this new
    bill). Applies regardless of cash/credit booking type, per how Ibrahim
    wants it — this is a "total exposure" check, not a receivables-only one.

    Returns (allowed: bool, message: str or None).
      - allowed=False  -> caller MUST block the save (company is in "block"
                           mode and the limit would be exceeded).
      - allowed=True + message -> proceed, but flash a warning to the user.
      - allowed=True + no message -> nothing to say, limit not exceeded (or
                           no limit set / no client / no company).
    """
    if not client or not company:
        return True, None
    limit = client.credit_limit or 0
    if limit <= 0:
        return True, None  # 0 / unset credit_limit == unlimited, matches how it's used everywhere else in this app
    outstanding = client.pending or 0
    projected = outstanding + (new_bill_amount or 0)
    if projected <= limit:
        return True, None
    action = (getattr(company, "credit_limit_action", "warn") or "warn").strip().lower()
    msg = (
        f"{client.name}'s credit limit is \u20b9{limit:,.2f}. Outstanding "
        f"\u20b9{outstanding:,.2f} + this bill \u20b9{(new_bill_amount or 0):,.2f} = "
        f"\u20b9{projected:,.2f}, which exceeds the limit."
    )
    if action == "block":
        return False, msg
    return True, msg


@app.route("/booking/customer/check-credit-limit", methods=["POST"])
@login_required
@require_permission("invoices", "create")
def invoice_customer_check_credit_limit():
    """
    AJAX pre-check called from booking.html right before the Generate
    button actually submits. Lets the page show the credit-limit message
    as a confirm() popup at click time instead of as a flash message that
    only shows up after the invoice is already saved and the page has
    redirected to /invoice/list.

    This does NOT save anything — it's read-only. The real, authoritative
    check still runs server-side in invoice_customer_save /
    invoice_customer_update; this endpoint can be skipped, spoofed, or
    fail without compromising that.
    """
    cdb = get_cdb()
    company_id = get_current_company()
    client_id_raw = request.form.get("client_id")
    try:
        amount = float(request.form.get("amount", 0) or 0)
    except (TypeError, ValueError):
        amount = 0.0

    if not client_id_raw:
        return jsonify({"blocked": False, "message": None})

    try:
        client_id_val = int(client_id_raw)
    except (TypeError, ValueError):
        return jsonify({"blocked": False, "message": None})

    client = cdb.query(Client).filter_by(id=client_id_val, company_id=company_id).first()
    co = Company.query.filter_by(company_id=company_id).first()
    allowed, message = _check_credit_limit(co, client, amount)
    return jsonify({"blocked": not allowed, "message": message})


def _docket_no_in_use(cdb, company_id, docket_no, exclude_invoice_id=None):
    """
    Server-side uniqueness check, run at SAVE time (not form-render time).

    _next_awb_number() only ever runs on GET, when the invoice form is
    rendered. The value it returns then sits in an editable <input> on the
    page indefinitely (multiple tabs, a tab left open, etc). Whatever is in
    that field when the form is POSTed is trusted verbatim by
    invoice_customer_save() / invoice_new() with no re-check — that's how
    two invoices end up with the same AWB. This closes that gap: call it
    right before an insert/update and reject the submit if the docket_no is
    already attached to a *different* invoice for this company.

    Returns the invoice_id already using this docket_no, or None if free.
    """
    docket_no = (docket_no or "").strip()
    if not docket_no:
        return None

    q = cdb.query(Invoice).filter(
        Invoice.company_id == company_id,
        Invoice.terms.like(f'%"docket_no": "{docket_no}"%'),
    )
    if exclude_invoice_id:
        q = q.filter(Invoice.invoice_id != exclude_invoice_id)
    dupe = q.first()
    if dupe:
        return dupe.invoice_id

    # Some invoices may also carry docket_no as a real column (see
    # _next_awb_number's dynamic column check) — cover that path too.
    try:
        from sqlalchemy import inspect as _inspect
        inspector = _inspect(cdb.bind)
        columns = [c['name'] for c in inspector.get_columns('invoices')]
        if 'docket_no' in columns:
            q2 = cdb.query(Invoice).filter(
                Invoice.company_id == company_id,
                Invoice.docket_no == docket_no,
            )
            if exclude_invoice_id:
                q2 = q2.filter(Invoice.invoice_id != exclude_invoice_id)
            dupe2 = q2.first()
            if dupe2:
                return dupe2.invoice_id
    except Exception:
        pass

    return None


_FIX_DUPES_TEMPLATE = """
<!doctype html><html><head><title>Duplicate AWB cleanup</title>
<style>
body{font-family:system-ui,sans-serif;padding:24px;color:#111827;background:#F9FAFB;}
table{border-collapse:collapse;width:100%;max-width:900px;background:#fff;}
th,td{border:1px solid #E5E7EB;padding:8px 12px;text-align:left;font-size:14px;}
th{background:#F3F4F6;}
.old{color:#B91C1C;text-decoration:line-through;}
.new{color:#15803D;font-weight:700;}
.warn{background:#FEF3C7;border:1px solid #F59E0B;padding:12px;border-radius:6px;margin-bottom:16px;max-width:900px;}
.btn{display:inline-block;margin-top:16px;padding:10px 18px;background:#DC2626;color:#fff;border:none;border-radius:6px;font-weight:700;cursor:pointer;}
</style></head><body>
<h2>Duplicate AWB cleanup — dry run</h2>
<div class="warn">
This only <b>previews</b> changes — nothing is saved yet. It keeps the
<b>earliest-created</b> invoice on each clashing AWB and reassigns every
newer invoice in that group to the next free AWB.<br><br>
<b>Check before confirming:</b> if a physical AWB label/document already
went out with the old number for any of these, renumbering here will make
the system disagree with the paper. Only click "Apply" once you've verified
these newer invoices haven't actually shipped under the old number.
</div>
{% if changes %}
<table>
<tr><th>Invoice</th><th>Created</th><th>Current AWB</th><th>New AWB</th></tr>
{% for c in changes %}
<tr>
  <td>{{ c.invoice_id }}</td>
  <td>{{ c.created_at }}</td>
  <td class="old">{{ c.old_awb }}</td>
  <td class="new">{{ c.new_awb }}</td>
</tr>
{% endfor %}
</table>
<form method="POST">
  <input type="hidden" name="confirm" value="yes">
  <button class="btn" type="submit" onclick="return confirm('Renumber {{ changes|length }} invoice(s)? This cannot be undone automatically.');">
    Apply — renumber {{ changes|length }} invoice(s)
  </button>
</form>
{% else %}
<p>No duplicate AWBs found for this company. Nothing to fix.</p>
{% endif %}
</body></html>
"""

# ─────────────────────────────────────────────────────────────────────────────
# ── Customer Invoice (Aggregate Bookings) ────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────

def _get_next_customer_invoice_number(cdb, company_id, invoice_type="credit"):
    """
    Generate next customer invoice number based on type:
    - Credit: CR-001, CR-002, etc.
    - Cash: CS-001, CS-002, etc.
    """
    prefix = "CR-" if invoice_type == "credit" else "CS-"
    existing = cdb.query(CustomerInvoice).filter(
        CustomerInvoice.company_id == company_id,
        CustomerInvoice.invoice_number.like(f"{prefix}%")
    ).all()
    
    max_num = 0
    for inv in existing:
        if inv.invoice_number and inv.invoice_number.startswith(prefix):
            try:
                num = int(inv.invoice_number.split("-")[-1])
                if num > max_num:
                    max_num = num
            except (ValueError, IndexError):
                continue
    return f"{prefix}{max_num + 1:03d}"


@app.route("/customer-invoices")
@login_required
@require_permission("invoices", "view")
def customer_invoice_list():
    """List all customer invoices"""
    cdb = get_cdb()
    company_id = get_current_company()
    
    status_filter = request.args.get("status", "All")
    query = cdb.query(CustomerInvoice).filter_by(company_id=company_id)
    
    if status_filter != "All":
        query = query.filter_by(status=status_filter)
    
    invoices = query.order_by(CustomerInvoice.created_at.desc()).all()
    
    # Calculate totals
    total_amount = sum(inv.grand_total for inv in invoices)
    total_paid = sum(inv.paid_amount for inv in invoices)
    total_balance = sum(inv.balance for inv in invoices)
    
    return render_template("customer_invoice_list.html",
                         invoices=invoices,
                         total_amount=total_amount,
                         total_paid=total_paid,
                         total_balance=total_balance,
                         current_status=status_filter,
                         active='customer_invoices')


@app.route("/customer-invoices/new")
@login_required
@require_permission("invoices", "create")
def customer_invoice_new():
    """Create a new customer invoice - select bookings to include"""
    cdb = get_cdb()
    company_id = get_current_company()
    
    # Get all clients - INCLUDE Cash-Only clients for the dropdown
    clients = cdb.query(Client).filter_by(company_id=company_id).order_by(Client.name).all()
    
    # Get selected client filter
    client_filter = request.args.get("client_id", type=int)
    from_date = request.args.get("from_date", "")
    to_date = request.args.get("to_date", "")
    
    # Build query for available bookings
    existing_customer_invoices = cdb.query(CustomerInvoice).filter_by(company_id=company_id).all()
    used_booking_ids = set()
    for ci in existing_customer_invoices:
        if ci.booking_ids_json:
            try:
                ids = json.loads(ci.booking_ids_json)
                used_booking_ids.update(ids)
            except:
                pass
    
    # Get bookings - DON'T filter by client_id for cash bookings
    query = cdb.query(Invoice).filter(
        Invoice.company_id == company_id,
        Invoice.status.notin_(['Void', 'Draft']),
        ~Invoice.id.in_(used_booking_ids) if used_booking_ids else True
    )
    
    # If a regular client is selected (not cash-only), filter by client_id
    selected_client_name = None
    is_cash_client = False
    
    if client_filter:
        selected_client = cdb.query(Client).filter_by(id=client_filter, company_id=company_id).first()
        if selected_client:
            if selected_client.client_type == "Cash-Only":
                # For cash clients, we'll filter by name after loading
                is_cash_client = True
                selected_client_name = selected_client.name
            else:
                # Regular client - filter by client_id
                query = query.filter(Invoice.client_id == client_filter)
    
    if from_date:
        try:
            from_dt = date.fromisoformat(from_date)
            query = query.filter(Invoice.date >= from_dt)
        except:
            pass
    
    if to_date:
        try:
            to_dt = date.fromisoformat(to_date)
            query = query.filter(Invoice.date <= to_dt)
        except:
            pass
    
    bookings = query.order_by(Invoice.date.desc()).all()
    
    # Build booking data with shipment info
    booking_data = []
    credit_count = 0
    cash_count = 0
    
    for inv in bookings:
        meta = {}
        if inv.terms:
            try:
                meta = json.loads(inv.terms)
            except:
                pass
        
        # Determine booking type from meta
        booking_type = meta.get("booking_type", "credit")
        
        # Get client name and ID based on booking type
        if booking_type == "cash":
            # Cash booking: use shipper_name from meta
            client_name = meta.get("shipper_name", "Walk-in")
            client_id = None
            
            # If a cash client is selected, only include matching shipper_name
            if is_cash_client and client_name != selected_client_name:
                continue
        else:
            if is_cash_client:
                continue
            if inv.client_obj:
                client_name = inv.client_obj.name
                client_id = inv.client_id
            else:
                client_name = meta.get("shipper_name", inv.contact_person or "Unknown")
                client_id = None
        
        packages = meta.get("packages", [])
        total_weight = sum(p.get("weight", 0) * p.get("qty", 1) for p in packages) or meta.get("freight_weight", 0)
        freight = meta.get("freight", inv.subtotal or 0)
        gst = meta.get("gst", inv.tax_amount or 0)
        total = inv.grand_total or 0
        
        if booking_type == "credit":
            credit_count += 1
        else:
            cash_count += 1
        
        # Get receiver name
        receiver_name = meta.get("receiver_name", "") or inv.contact_person or ""
        
        booking_data.append({
            "id": inv.id,
            "invoice_id": inv.invoice_id,
            "docket_no": meta.get("docket_no", ""),
            "date": inv.date,
            "client_id": client_id,  # This will be None for cash bookings
            "client_name": client_name,  # This displays in the Customer column
            "receiver_name": receiver_name,
            "destination": meta.get("destination", ""),
            "carrier": meta.get("carrier", ""),
            "carrier_ref": meta.get("carrier_ref", ""),
            "weight": total_weight,
            "amount": freight,
            "gst": gst,
            "total": total,
            "selected": False,
            "booking_type": booking_type
        })
    
    from datetime import timedelta
    today = today_ist()
    default_invoice_date = today.isoformat()
    default_due_date = (today + timedelta(days=30)).isoformat()
    
    credit_number = _get_next_customer_invoice_number(cdb, company_id, "credit")
    cash_number = _get_next_customer_invoice_number(cdb, company_id, "cash")
    
    default_type = "credit"
    default_invoice_number = credit_number
    
    # Get company info
    company = Company.query.filter_by(company_id=company_id).first()
    
    # Get selected client details
    selected_client = None
    if client_filter:
        selected_client = cdb.query(Client).filter_by(id=client_filter, company_id=company_id).first()
    
    return render_template("customer_invoice_form.html",
                         clients=clients,
                         bookings=booking_data,
                         client_filter=client_filter,
                         from_date=from_date,
                         to_date=to_date,
                         active='customer_invoices',
                         default_invoice_date=default_invoice_date,
                         default_due_date=default_due_date,
                         credit_number=credit_number,
                         cash_number=cash_number,
                         default_invoice_number=default_invoice_number,
                         default_type=default_type,
                         company=company,
                         selected_client=selected_client)  


@app.route("/customer-invoices/create", methods=["POST"])
@login_required
@require_permission("invoices", "create")
def customer_invoice_create():
    """Create a customer invoice from selected bookings"""
    cdb = get_cdb()
    company_id = get_current_company()
    
    client_id = request.form.get("client_id", type=int)
    invoice_date = request.form.get("invoice_date") or str(today_ist())
    due_date = request.form.get("due_date")
    notes = request.form.get("notes", "").strip()
    
    # Get selected invoice type from form
    invoice_type = request.form.get("invoice_type", "credit")
    
    # Get selected booking IDs
    booking_ids = request.form.getlist("booking_ids[]")
    booking_ids = [int(bid) for bid in booking_ids if bid]
    
    if not client_id:
        flash("Please select a customer.", "error")
        return redirect(url_for("customer_invoice_new"))
    
    if not booking_ids:
        flash("Please select at least one booking to include.", "error")
        return redirect(url_for("customer_invoice_new"))
    
    # Get the client
    client = cdb.query(Client).filter_by(id=client_id, company_id=company_id).first()
    if not client:
        flash("Customer not found.", "error")
        return redirect(url_for("customer_invoice_new"))
    
    # Get all selected bookings
    bookings = cdb.query(Invoice).filter(Invoice.id.in_(booking_ids), Invoice.company_id == company_id).all()
    
    if len(bookings) != len(booking_ids):
        flash("Some bookings could not be found.", "error")
        return redirect(url_for("customer_invoice_new"))
    
    # Check if any booking is already used
    existing_customer_invoices = cdb.query(CustomerInvoice).filter_by(company_id=company_id).all()
    used_booking_ids = set()
    for ci in existing_customer_invoices:
        if ci.booking_ids_json:
            try:
                ids = json.loads(ci.booking_ids_json)
                used_booking_ids.update(ids)
            except:
                pass
    
    for inv in bookings:
        if inv.id in used_booking_ids:
            flash(f"Booking {inv.invoice_id} is already included in another customer invoice.", "error")
            return redirect(url_for("customer_invoice_new"))
    
    # Generate invoice number with the selected type
    invoice_number = _get_next_customer_invoice_number(cdb, company_id, invoice_type)
    
    # Calculate totals
    subtotal = 0.0
    tax_total = 0.0
    grand_total = 0.0
    
    # Create customer invoice
    cust_inv = CustomerInvoice(
        invoice_number=invoice_number,
        company_id=company_id,
        client_id=client_id,
        client_name=client.name,
        invoice_date=date.fromisoformat(invoice_date),
        due_date=date.fromisoformat(due_date) if due_date else None,
        invoice_type=invoice_type,
        status="Pending",
        subtotal=0,
        tax_amount=0,
        grand_total=0,
        paid_amount=0,
        balance=0,
        notes=notes,
        created_by=get_current_user().get("email"),
        booking_ids_json=json.dumps(booking_ids)
    )
    cdb.add(cust_inv)
    cdb.flush()
    
    subtotal = 0
    tax_total = 0
    cgst_total = 0
    sgst_total = 0
    igst_total = 0
    grand_total = 0

    # Create items for each booking
    for inv in bookings:
        meta = {}
        if inv.terms:
            try:
                meta = json.loads(inv.terms)
            except:
                pass
        
        packages = meta.get("packages", [])
        total_weight = sum(p.get("weight", 0) * p.get("qty", 1) for p in packages) or meta.get("freight_weight", 0)
        freight = meta.get("freight", inv.subtotal or 0)
        gst = meta.get("gst", inv.tax_amount or 0)
        total = inv.grand_total or 0
        
        # Get GST split — whatever type this booking actually used (IGST for
        # interstate, CGST+SGST for intrastate), carried forward as-is.
        cgst = meta.get("cgst", 0)
        sgst = meta.get("sgst", 0)
        igst = meta.get("igst", 0)
        gst_percent = 18 if (cgst + sgst + igst) > 0 else 0
        
        item_desc = ""
        if packages:
            item_desc = ", ".join(p.get("name", p.get("type", "Box")) for p in packages[:3])
            if len(packages) > 3:
                item_desc += f" + {len(packages) - 3} more"
        else:
            item_desc = f"Freight - {meta.get('shipment_type', 'Logistics')}"
        
        item = CustomerInvoiceItem(
            customer_invoice_id=cust_inv.id,
            booking_invoice_id=inv.id,
            booking_invoice_ref=inv.invoice_id,
            docket_no=meta.get("docket_no", ""),
            receiver_name=meta.get("receiver_name", ""),
            destination=meta.get("destination", ""),
            carrier=meta.get("carrier", ""),
            carrier_ref=meta.get("carrier_ref", ""),
            item_description=item_desc,
            quantity=1,
            weight_kg=total_weight,
            rate_per_kg=freight / total_weight if total_weight > 0 else 0,
            taxable_amount=freight,
            gst_percent=gst_percent,
            cgst_amount=cgst,
            sgst_amount=sgst,
            igst_amount=igst,
            total_amount=total,
            booking_date=inv.date
        )
        cdb.add(item)
        
        subtotal += freight
        tax_total += gst
        cgst_total += cgst
        sgst_total += sgst
        igst_total += igst
        grand_total += total
    
    # Update customer invoice totals
    cust_inv.subtotal = subtotal
    cust_inv.tax_amount = tax_total
    cust_inv.cgst_total = cgst_total
    cust_inv.sgst_total = sgst_total
    cust_inv.igst_total = igst_total
    cust_inv.grand_total = grand_total
    cust_inv.balance = grand_total
    
    cdb.commit()
    
    type_label = "Credit" if invoice_type == "credit" else "Cash"
    flash(f"✅ {type_label} customer invoice {invoice_number} created with {len(bookings)} bookings.", "success")
    
    # Use cust_inv_id instead of invoice_id
    return redirect(url_for("customer_invoice_view", cust_inv_id=cust_inv.id))


@app.route("/customer-invoices/view/<int:cust_inv_id>")
@login_required
@require_permission("invoices", "view")
def customer_invoice_view(cust_inv_id):
    """View a customer invoice"""
    cdb = get_cdb()
    company_id = get_current_company()
    
    cust_inv = cdb.query(CustomerInvoice).filter_by(id=cust_inv_id, company_id=company_id).first()
    if not cust_inv:
        flash("Customer invoice not found.", "error")
        return redirect(url_for("customer_invoice_list"))
    
    items = cdb.query(CustomerInvoiceItem).filter_by(customer_invoice_id=cust_inv.id).all()

    # ── Load the client from the company DB ──
    client = None
    if cust_inv.client_id:
        client = cdb.query(Client).filter_by(id=cust_inv.client_id, company_id=company_id).first()

    company = Company.query.filter_by(company_id=company_id).first()

    return render_template("customer_invoice_view.html",
                        invoice=cust_inv,
                        items=items,
                        client=client,
                        company=company,
                        active='customer_invoices')


@app.route("/customer-invoices/delete/<int:cust_inv_id>", methods=["POST"])
@login_required
@require_permission("invoices", "delete")
def customer_invoice_delete(cust_inv_id):
    """Delete a customer invoice (soft delete - mark as Void)"""
    cdb = get_cdb()
    company_id = get_current_company()
    
    cust_inv = cdb.query(CustomerInvoice).filter_by(id=cust_inv_id, company_id=company_id).first()
    if not cust_inv:
        flash("Customer invoice not found.", "error")
        return redirect(url_for("customer_invoice_list"))
    
    cust_inv.status = "Void"
    cdb.commit()
    
    flash(f"Customer invoice {cust_inv.invoice_number} has been voided.", "success")
    return redirect(url_for("customer_invoice_list"))


@app.route("/customer-invoices/print/<int:cust_inv_id>")
@login_required
@require_permission("invoices", "view")
def customer_invoice_print(cust_inv_id):
    """Print a customer invoice - returns HTML for print or PDF download"""
    from xhtml2pdf import pisa
    import io
    
    cdb = get_cdb()
    company_id = get_current_company()
    
    cust_inv = cdb.query(CustomerInvoice).filter_by(id=cust_inv_id, company_id=company_id).first()
    if not cust_inv:
        flash("Customer invoice not found.", "error")
        return redirect(url_for("customer_invoice_list"))
    
    items = cdb.query(CustomerInvoiceItem).filter_by(customer_invoice_id=cust_inv.id).all()
    company = Company.query.filter_by(company_id=company_id).first()
    
    # ── Load the client from the company DB ──
    client = None
    if cust_inv.client_id:
        client = cdb.query(Client).filter_by(id=cust_inv.client_id, company_id=company_id).first()
    
    # Get action parameter: 'pdf' or 'print'
    action = request.args.get('action', 'pdf')

    # Per-company invoice template choice — falls back to 'classic' for
    # companies that haven't picked one yet (existing rows, new signups
    # before this field was set).
    template_name = f"customer_invoice_pdf_{company.invoice_template or 'classic'}.html" if company else "customer_invoice_pdf_classic.html"

    # Render HTML content
    html_content = render_template(
        template_name,
        invoice=cust_inv,
        items=items,
        client=client,
        company=company,
        company_logo_url=url_for('static', filename=f'company_logos/{company.logo_filename}', _external=True) if company and company.logo_filename else None,
        is_gst_registered=company.is_gst_registered if company else True,
        today=today_ist().strftime("%d %b %Y"),
        action=action
    )
    
    # If action is 'print', return HTML with print styles
    if action == 'print':
        return render_template(
            template_name,
            invoice=cust_inv,
            items=items,
            client=client,
            company=company,
            company_logo_url=url_for('static', filename=f'company_logos/{company.logo_filename}', _external=True) if company and company.logo_filename else None,
            is_gst_registered=company.is_gst_registered if company else True,
            today=today_ist().strftime("%d %b %Y"),
            action='print'
        )
    
    # Default: Generate PDF
    pdf_file = io.BytesIO()
    pisa_status = pisa.CreatePDF(html_content, dest=pdf_file, encoding='UTF-8')
    
    if pisa_status.err:
        flash(f"PDF generation error: {pisa_status.err}", "error")
        return redirect(url_for("customer_invoice_view", cust_inv_id=cust_inv.id))
    
    pdf_file.seek(0)
    
    return send_file(
        pdf_file,
        as_attachment=True,
        download_name=f"Customer_Invoice_{cust_inv.invoice_number}.pdf",
        mimetype="application/pdf"
    )

@app.route("/admin/fix-duplicate-awbs", methods=["GET", "POST"])
@login_required
@require_permission("invoices", "edit")
def fix_duplicate_awbs():
    """
    One-time cleanup for AWBs that were already duplicated by the old bug
    (docket_no was editable on the invoice-edit form with no server-side
    uniqueness check). For every AWB shared by 2+ invoices, keeps the
    earliest-created invoice on that AWB and reassigns every newer one to
    the next free number, using the same prefix/sequence as _next_awb_number.

    GET, or POST without confirm=yes -> dry run only, nothing is written.
    POST with confirm=yes            -> applies the renumbering and commits.
    """
    cdb = get_cdb()
    company_id = get_current_company()

    co = PlatformCompany.query.filter_by(company_id=company_id).first()
    prefix = "AHL" if (co is None or co.awb_prefix is None) else co.awb_prefix
    awb_start = (co.awb_start if co else None) or 81000
    pattern = re.compile(rf'{re.escape(prefix)}(\d+)')

    invoices = (
        cdb.query(Invoice)
        .filter(Invoice.company_id == company_id)
        .filter(Invoice.terms.isnot(None))
        .order_by(Invoice.created_at.asc(), Invoice.id.asc())
        .all()
    )

    groups = {}
    parsed = {}
    max_seq = awb_start - 1
    for inv in invoices:
        try:
            meta = json.loads(inv.terms) if inv.terms else {}
        except (ValueError, TypeError):
            continue
        docket = (meta.get("docket_no") or "").strip()
        if not docket:
            continue
        parsed[inv.invoice_id] = meta
        groups.setdefault(docket, []).append(inv)
        m = pattern.search(docket)
        if m:
            max_seq = max(max_seq, int(m.group(1)))

    used_numbers = set(groups.keys())
    changes = []  # (invoice, old_awb, new_awb)

    for docket, invs in groups.items():
        if len(invs) < 2:
            continue
        # invs is already ordered oldest-first (query order preserved) —
        # the oldest keeps the AWB, everything after it gets renumbered.
        for inv in invs[1:]:
            max_seq += 1
            new_awb = f"{prefix}{max_seq}"
            while new_awb in used_numbers:
                max_seq += 1
                new_awb = f"{prefix}{max_seq}"
            used_numbers.add(new_awb)
            changes.append((inv, docket, new_awb))

    if request.method == "GET" or request.form.get("confirm") != "yes":
        return render_template_string(
            _FIX_DUPES_TEMPLATE,
            changes=[{
                "invoice_id": inv.invoice_id,
                "created_at": inv.created_at,
                "old_awb": old,
                "new_awb": new,
            } for inv, old, new in changes],
        )

    for inv, old_awb, new_awb in changes:
        meta = parsed[inv.invoice_id]
        meta["docket_no"] = new_awb
        inv.terms = json.dumps(meta)
    cdb.commit()

    flash(f"Renumbered {len(changes)} duplicate invoice(s) — each now has its own AWB.")
    return redirect(url_for("invoice_list"))


@app.route("/booking/void/<invoice_id>", methods=["POST"])
@login_required
@require_permission("invoices", "delete")
def invoice_void(invoice_id):
    """
    Void a booking: keeps the Invoice row and ALL of its data (line items,
    terms/shipper/receiver JSON, packages, monetary totals) exactly as they
    were — only the status flips to "Void". Downstream effects the booking
    caused (stock quantity, manifest entries, the auto-generated purchase
    invoice line, client/supplier ledger balances) are reversed.
    """
    company_id = get_current_company()
    if not company_id:
        return redirect(url_for("login"))
    cdb = get_customer_session(company_id)

    inv = cdb.query(Invoice).filter_by(invoice_id=invoice_id, company_id=company_id).first()
    if not inv:
        flash("Invoice not found.", "danger")
        return redirect(url_for("invoice_list"))

    if inv.status == "Void":
        flash(f"{invoice_id} is already void.", "info")
        return redirect(url_for("invoice_view", invoice_id=invoice_id))

    docket_no = _get_awb(inv)

    try:
        # 1) Reverse client outstanding
        if inv.balance and inv.client_id:
            client = cdb.query(Client).filter_by(id=inv.client_id, company_id=company_id).first()
            if client:
                client.pending = max(0, (client.pending or 0) - inv.balance)

        # 2) Reverse manifest entries + REMOVE stock that was deducted
        if docket_no:
            entries = cdb.query(ManifestEntry).join(
                CompanyManifest, ManifestEntry.manifest_id == CompanyManifest.id
            ).filter(
                ManifestEntry.docket_no == docket_no,
                CompanyManifest.company_id == company_id,
            ).all()
            touched_manifest_ids = set()
            for entry in entries:
                if entry.status == "Generated" and entry.stock_item_id and entry.boxes:
                    stock = cdb.query(StockItem).filter_by(
                        id=entry.stock_item_id, company_id=company_id
                    ).first()
                    if stock:
                        # REMOVE the stock that was deducted (add it back)
                        stock.quantity = (stock.quantity or 0) + entry.boxes
                        stock.last_updated = today_ist()
                        # Log this reversal
                        cdb.add(StockPurchaseHistory(
                            stock_item_id=stock.id,
                            purchase_invoice_id=None,
                            quantity=entry.boxes,
                            purchase_rate=0,
                            movement_type="IN",
                            purchase_date=today_ist(),
                            reference=f"VOID-MANIFEST-{invoice_id}",
                            awb_no=docket_no,
                        ))
                if entry.manifest_id:
                    touched_manifest_ids.add(entry.manifest_id)
                cdb.delete(entry)
            cdb.flush()
            for manifest_id in touched_manifest_ids:
                parent_manifest = cdb.query(CompanyManifest).filter_by(id=manifest_id).first()
                if parent_manifest:
                    remaining = cdb.query(ManifestEntry).filter_by(manifest_id=parent_manifest.id).count()
                    if remaining == 0:
                        cdb.delete(parent_manifest)
                    else:
                        _recompute_manifest_status(parent_manifest)

        # 3) Reverse the auto-generated purchase invoice line
        pi_item = cdb.query(PurchaseInvoiceItem).filter_by(source_invoice_id=inv.id).first()
        if pi_item:
            parent_pi = cdb.query(PurchaseInvoice).filter_by(id=pi_item.purchase_invoice_id).first()
            line_total = pi_item.total_amount or 0
            
            # ── REVERSE STOCK DEDUCTION FROM PURCHASE BILL ──────────────────────────
            # When a purchase bill line was created, stock was DEDUCTED (OUT movement).
            # Voiding the booking reverses that: add stock back.
            if pi_item.docket_no:
                # Find the booking invoice's items with stock
                booking_invoice_items = cdb.query(InvoiceItem).filter(
                    InvoiceItem.invoice_id == inv.id,
                    InvoiceItem.stock_item_id.isnot(None)
                ).all()
                
                for inv_item in booking_invoice_items:
                    stock = cdb.query(StockItem).filter_by(
                        id=inv_item.stock_item_id, company_id=company_id
                    ).first()
                    if stock:
                        # Add back the stock quantity (reversal of the OUT movement)
                        stock.quantity = (stock.quantity or 0) + (inv_item.qty or 0)
                        stock.last_updated = today_ist()
                        cdb.add(StockPurchaseHistory(
                            stock_item_id=stock.id,
                            purchase_invoice_id=None,
                            quantity=inv_item.qty or 0,
                            purchase_rate=0,
                            movement_type="IN",
                            purchase_date=today_ist(),
                            reference=f"VOID-PURCHASE-{invoice_id}",
                            awb_no=docket_no,
                        ))
            
            if parent_pi:
                supplier = cdb.query(Supplier).filter_by(
                    id=parent_pi.supplier_id, company_id=company_id
                ).first()
                if supplier:
                    supplier.payable = max(0, (supplier.payable or 0) - line_total)
                parent_pi.subtotal = max(0, (parent_pi.subtotal or 0) - (pi_item.taxable_value or 0))
                parent_pi.grand_total = max(0, (parent_pi.grand_total or 0) - line_total)
                parent_pi.balance = max(0, (parent_pi.balance or 0) - line_total)
            cdb.delete(pi_item)
            cdb.flush()
            if parent_pi:
                remaining_items = cdb.query(PurchaseInvoiceItem).filter_by(
                    purchase_invoice_id=parent_pi.id
                ).count()
                if remaining_items == 0:
                    cdb.delete(parent_pi)

        # 4) REMOVE THE ORIGINAL STOCK THAT WAS ADDED AT BOOKING TIME
        # Find all stock history entries for this booking's docket_no
        stock_history_entries = cdb.query(StockPurchaseHistory).filter(
            StockPurchaseHistory.awb_no == docket_no,
            StockPurchaseHistory.purchase_invoice_id.is_(None)
        ).all()
        
        # If not found by awb_no, try to find by stock_item_id from invoice items
        if not stock_history_entries and docket_no:
            for item in inv.items:
                if item.stock_item_id and item.qty:
                    # Try to find history entry for this stock item with this awb
                    hist = cdb.query(StockPurchaseHistory).filter(
                        StockPurchaseHistory.stock_item_id == item.stock_item_id,
                        StockPurchaseHistory.awb_no == docket_no,
                        StockPurchaseHistory.purchase_invoice_id.is_(None)
                    ).first()
                    if hist:
                        stock_history_entries.append(hist)
        
        # Process found history entries
        for hist in stock_history_entries:
            stock = cdb.query(StockItem).filter_by(
                id=hist.stock_item_id, company_id=company_id
            ).first()
            if stock:
                stock.quantity = max(0, (stock.quantity or 0) - (hist.quantity or 0))
                stock.last_updated = today_ist()
                cdb.delete(hist)
            else:
                print(f"[invoice-void] WARNING: stock_item {hist.stock_item_id} missing, history not reversed for {invoice_id}")
        
        # FALLBACK: If no history entries found, reduce stock directly from invoice items
        if not stock_history_entries:
            for item in inv.items:
                if item.stock_item_id and item.qty:
                    stock = cdb.query(StockItem).filter_by(
                        id=item.stock_item_id, company_id=company_id
                    ).first()
                    if stock:
                        stock.quantity = max(0, (stock.quantity or 0) - (item.qty or 0))
                        stock.last_updated = today_ist()
                        print(f"[invoice-void] FALLBACK: Reduced stock for {stock.name} by {item.qty}")

        # 5) Delete the linked proforma invoice (Estimate)
        linked_est = cdb.query(Estimate).filter_by(company_id=company_id).filter(
            Estimate.terms.like(f'%"linked_invoice_id": "{invoice_id}"%')
        ).first()
        if linked_est:
            cdb.query(EstimateItem).filter_by(estimate_id=linked_est.id).delete()
            cdb.delete(linked_est)

        # 6) Void the invoice itself — status flag only.
        #    Line items, terms (shipper/receiver/service JSON), packages, and every
        #    monetary field are LEFT AS-IS so the booking's full history stays visible
        #    on the invoice view/print. Reports and receivables already filter out
        #    status == "Void" elsewhere, so keeping these numbers doesn't affect totals.
        inv.status = "Void"

        cdb.commit()
        flash(
            f"Booking {invoice_id} voided. Stock, manifest, and ledger entries have been reversed; "
            f"all booking details (items, terms, amounts) remain visible on the invoice for reference.",
            "success",
        )

        try:
            updated_invoices = update_customer_invoice_from_booking(cdb, company_id, inv.id)
            if updated_invoices:
                flash(f"Customer invoice(s) {', '.join(str(i) for i in updated_invoices)} updated to reflect voided booking.", "info")
        except Exception as e:
            print(f"[customer-invoice-update] failed to update parent invoices on void: {e}")
    except Exception as e:
        cdb.rollback()
        print(f"[invoice-void] FAILED for {invoice_id}: {e}")
        flash(f"Could not void {invoice_id}: {e}", "danger")

    return redirect(url_for("invoice_view", invoice_id=invoice_id))

@app.route("/booking/clone/<invoice_id>")
@login_required
@require_permission("invoices", "create")
def invoice_clone(invoice_id):
    """
    Clone an existing booking: copy all its details to a new invoice
    with a fresh AWB number. The cloned invoice starts as a Draft so
    the user can review and make changes before generating.
    """
    cdb = get_cdb()
    company_id = get_current_company()
    
    # Find the source invoice
    source_inv = cdb.query(Invoice).filter_by(invoice_id=invoice_id, company_id=company_id).first()
    if not source_inv:
        flash("Invoice not found", "error")
        return redirect(url_for("invoice_list"))
    
    # Parse the source terms JSON
    try:
        source_meta = json.loads(source_inv.terms) if source_inv.terms else {}
    except (ValueError, TypeError):
        source_meta = {}
    
    # Generate new AWB number
    new_docket_no = _next_awb_number(company_id)
    
    # Generate new invoice ID
    new_invoice_id = _next_numbered_id(cdb, Invoice.invoice_id, "", extra_filters=[Invoice.company_id == company_id])
    
    # Prepare form data for the new invoice (copy all fields from source)
    form_data = {
        "customer_id": source_inv.client_id,
        "customer_phone": source_inv.phone or "",
        "shipper_name": source_meta.get("shipper_name", ""),
        "shipper_contact_name": source_meta.get("shipper_contact_name", source_inv.contact_person or ""),
        "courier_company_id": source_meta.get("courier_company_id", ""),
        "shipper_address1": source_meta.get("shipper_address1", source_meta.get("shipper_address", "")),
        "shipper_address2": source_meta.get("shipper_address2", ""),
        "shipper_city": source_meta.get("shipper_city", ""),
        "shipper_state": source_meta.get("shipper_state", ""),
        "shipper_pincode": source_meta.get("shipper_pincode", ""),
        "shipper_country": source_meta.get("shipper_country", "India"),
        "shipper_doc_type": source_meta.get("shipper_doc_type", ""),
        "shipper_doc_no": source_meta.get("shipper_doc_no", ""),
        "client_code": source_meta.get("client_code", ""),
        "receiver_name": source_meta.get("receiver_name", ""),
        "receiver_company": source_meta.get("receiver_company", ""),
        "receiver_phone": source_meta.get("receiver_phone", ""),
        "receiver_address1": source_meta.get("receiver_address1", source_meta.get("receiver_address", "")),
        "receiver_address2": source_meta.get("receiver_address2", ""),
        "receiver_city": source_meta.get("receiver_city", ""),
        "receiver_state": source_meta.get("receiver_state", ""),
        "receiver_pincode": source_meta.get("receiver_pincode", ""),
        "receiver_country": source_meta.get("receiver_country", "India"),
        "receiver_doc_type": source_meta.get("receiver_doc_type", ""),
        "receiver_doc_no": source_meta.get("receiver_doc_no", ""),
        "destination": source_meta.get("destination", ""),
        "shipment_type": source_meta.get("shipment_type", ""),
        "mode": source_meta.get("mode", ""),
        "carrier": source_meta.get("carrier", ""),
        "tracking_number": source_meta.get("tracking_number", ""),
        "carrier_ref": source_meta.get("carrier_ref", ""),
        "origin": source_meta.get("origin", "India"),
        "pickup_date": source_meta.get("pickup_date", ""),
        "departure_time": source_meta.get("departure_time", ""),
        "expected_delivery": source_meta.get("expected_delivery", ""),
        "comments": source_meta.get("comments", ""),
        "vendor": source_meta.get("vendor", ""),
        "freight": source_meta.get("freight", source_inv.subtotal or 0),
        "fuel": source_meta.get("fuel", 0),
        "other": source_meta.get("other", 0),
        "freight_weight": source_meta.get("freight_weight", 0),
        "freight_rate_per_kg": source_meta.get("freight_rate_per_kg", 0),
        "freight_billing_weight": source_meta.get("freight_billing_weight", 0),
        "other_charges_reason": source_meta.get("other_charges_reason", ""),
        "discount": source_meta.get("discount", 0),
        "payment_mode": source_meta.get("payment_mode", "cash"),
        "booking_type": source_meta.get("booking_type", "credit"),
        "upi_app": source_meta.get("upi_app", ""),
        "upi_ref": source_meta.get("upi_ref", ""),
        "cheque_no": source_meta.get("cheque_no", ""),
        "cheque_date": source_meta.get("cheque_date", ""),
        "cheque_bank": source_meta.get("cheque_bank", ""),
        "notes": source_inv.email or "",
        # Copy packages
        "packages": source_meta.get("packages", []),
        # Copy performa items if any
        "performa_items": [],
        # Copy resale data
        "has_resale": getattr(source_inv, 'has_resale', False),
        "resale_charges": getattr(source_inv, 'resale_charges', 0),
        "resale_reason": getattr(source_inv, 'resale_reason', ''),
        "resale_date": getattr(source_inv, 'resale_date', ''),
        "resale_notes": getattr(source_inv, 'resale_notes', ''),
        # Copy GST settings
        "gst_invoice_flag": source_meta.get("gst_invoice_flag", "no"),
        "csb_type": source_meta.get("csb_type", "CSB 4"),
        "term_of_invoice": source_meta.get("term_of_invoice", "Delivered at Place(DAP)"),
        "export_reason": source_meta.get("export_reason", ""),
        "performa_format": source_meta.get("performa_format", "performainv"),
        "perf_weight": source_meta.get("weight", "0.00"),
        "perf_reference": source_meta.get("reference", ""),
        "performa_invoice_no": source_meta.get("invoice_no", ""),
        "performa_invoice_date": source_meta.get("invoice_date", ""),
        "department_no": source_meta.get("department_no", ""),
    }
    
    # Also copy linked performa invoice items if any
    linked_est = cdb.query(Estimate).filter_by(company_id=company_id).filter(
        Estimate.terms.like(f'%"linked_invoice_id": "{source_inv.invoice_id}"%')
    ).first()
    if linked_est and linked_est.terms:
        try:
            perf_meta = json.loads(linked_est.terms)
            form_data["performa_items"] = perf_meta.get("line_items", [])
            # Also copy these fields from the performa if they exist
            if perf_meta.get("weight"):
                form_data["perf_weight"] = perf_meta.get("weight")
            if perf_meta.get("reference"):
                form_data["perf_reference"] = perf_meta.get("reference")
        except Exception:
            pass
    
    # Prepare packages for the template
    packages = form_data.get("packages", [])
    if not packages:
        packages = [{"name": "Box", "type": "Box", "qty": 1, "length": "", "width": "", "height": "", "weight": "", "rate": 0}]
    
    # Get clients and suppliers for the form
    clients = cdb.query(Client).filter(
        Client.company_id == company_id,
        ~Client.client_type.in_(["Supplier", "Both", "Cash-Only"])
    ).all()
    
    suppliers = cdb.query(Supplier).filter_by(company_id=company_id, status="Active").order_by(Supplier.name).all()
    price_lists = cdb.query(PriceList).filter_by(company_id=company_id, is_active=True, list_type='sales').all()
    
    # Get client display ID
    client_display_id = ""
    if source_inv.client_obj:
        client_display_id = source_inv.client_obj.client_id or ""
    
    # Clone the ID documents from the source invoice (if cash booking)
    if form_data.get("booking_type") == "cash":
        form_data["shipper_aadhar_front_file"] = source_meta.get("shipper_aadhar_front_file", "")
        form_data["shipper_aadhar_back_file"] = source_meta.get("shipper_aadhar_back_file", "")
        form_data["shipper_pan_front_file"] = source_meta.get("shipper_pan_front_file", "")
        form_data["shipper_pan_back_file"] = source_meta.get("shipper_pan_back_file", "")
    
    flash(f"✅ Cloned booking {invoice_id} to new AWB {new_docket_no}. Review and edit before generating.", "success")
    
    return render_template(
        "booking.html",
        clients=clients,
        suppliers=suppliers,
        form_data=form_data,
        packages=packages,
        invoice_id=new_invoice_id,
        invoice_date=str(today_ist()),
        docket_no=new_docket_no,
        is_edit=False,
        today=str(today_ist()),
        price_lists=price_lists,
        client_display_id=client_display_id,
        invoice=None,
    )

@app.route("/company/clear-data", methods=["POST"])
@login_required
@owner_required
def company_clear_data():
    """
    Owner-only: hard-delete selected categories of company data so the
    account can start fresh. Company profile, users/permissions, and
    subscription are never touched by this — only the operational data
    categories the owner explicitly ticks.

    Requires the owner's password AND typing DELETE, since this cannot be
    undone. Deletion order respects FK dependencies (children before
    parents); where a booking's auto-generated purchase-invoice line
    references it via source_invoice_id, that reference is nulled rather
    than left dangling.
    """
    company_id = get_current_company()
    if not company_id:
        return redirect(url_for("login"))
    cdb = get_customer_session(company_id)
    user = get_current_user()

    categories = request.form.getlist("categories")
    password = request.form.get("password", "")
    confirm_text = request.form.get("confirm_text", "")

    if confirm_text.strip().upper() != "DELETE":
        flash("Type DELETE (exactly) to confirm. Nothing was removed.", "danger")
        return redirect(url_for("company_settings"))

    reg_user = RegisteredUser.query.filter_by(email=user.get("email")).first()
    if not reg_user or not verify_password(password, reg_user.password_hash):
        flash("Incorrect password. Nothing was removed.", "danger")
        return redirect(url_for("company_settings"))

    if not categories:
        flash("No categories were selected. Nothing was removed.", "info")
        return redirect(url_for("company_settings"))

    ALL_CATEGORIES = ["bookings", "proforma", "purchases", "manifests", "stock",
                       "parties", "finance", "price_lists", "whatsapp"]
    if "everything" in categories:
        categories = ALL_CATEGORIES

    removed = []
    try:
        if "bookings" in categories:
            inv_ids = [r.id for r in cdb.query(Invoice.id).filter_by(company_id=company_id)]
            if inv_ids:
                cdb.query(PurchaseInvoiceItem).filter(
                    PurchaseInvoiceItem.source_invoice_id.in_(inv_ids)
                ).update({"source_invoice_id": None}, synchronize_session=False)
                cdb.query(InvoiceItem).filter(InvoiceItem.invoice_id.in_(inv_ids)).delete(synchronize_session=False)
            cdb.query(Invoice).filter_by(company_id=company_id).delete(synchronize_session=False)
            removed.append("Bookings")

        if "proforma" in categories:
            est_ids = [r.id for r in cdb.query(Estimate.id).filter_by(company_id=company_id)]
            if est_ids:
                cdb.query(EstimateItem).filter(EstimateItem.estimate_id.in_(est_ids)).delete(synchronize_session=False)
            cdb.query(Estimate).filter_by(company_id=company_id).delete(synchronize_session=False)
            removed.append("Proforma Invoices")

        if "purchases" in categories:
            pi_ids = [r.id for r in cdb.query(PurchaseInvoice.id).filter_by(company_id=company_id)]
            cdb.query(PurchasePayment).filter_by(company_id=company_id).delete(synchronize_session=False)
            if pi_ids:
                cdb.query(PurchaseInvoiceItem).filter(
                    PurchaseInvoiceItem.purchase_invoice_id.in_(pi_ids)
                ).delete(synchronize_session=False)
                cdb.query(StockPurchaseHistory).filter(
                    StockPurchaseHistory.purchase_invoice_id.in_(pi_ids)
                ).delete(synchronize_session=False)
            cdb.query(PurchaseInvoice).filter_by(company_id=company_id).delete(synchronize_session=False)
            removed.append("Purchase Invoices")

        if "manifests" in categories:
            man_ids = [r.id for r in cdb.query(CompanyManifest.id).filter_by(company_id=company_id)]
            if man_ids:
                cdb.query(ManifestEntry).filter(ManifestEntry.manifest_id.in_(man_ids)).delete(synchronize_session=False)
            cdb.query(CompanyManifest).filter_by(company_id=company_id).delete(synchronize_session=False)
            removed.append("Manifests")

        if "stock" in categories:
            stock_ids = [r.id for r in cdb.query(StockItem.id).filter_by(company_id=company_id)]
            if stock_ids:
                cdb.query(StockPurchaseHistory).filter(
                    StockPurchaseHistory.stock_item_id.in_(stock_ids)
                ).delete(synchronize_session=False)
            cdb.query(StockItem).filter_by(company_id=company_id).delete(synchronize_session=False)
            removed.append("Stock")

        if "parties" in categories:
            sup_ids = [r.id for r in cdb.query(Supplier.id).filter_by(company_id=company_id)]
            if sup_ids:
                cdb.query(SupplierBrand).filter(SupplierBrand.supplier_id.in_(sup_ids)).delete(synchronize_session=False)
            cdb.query(Supplier).filter_by(company_id=company_id).delete(synchronize_session=False)
            cdb.query(Client).filter_by(company_id=company_id).delete(synchronize_session=False)
            removed.append("Clients & Suppliers")

        if "finance" in categories:
            loan_ids = [r.id for r in cdb.query(Loan.id).filter_by(company_id=company_id)]
            if loan_ids:
                cdb.query(LoanRepayment).filter(LoanRepayment.loan_id.in_(loan_ids)).delete(synchronize_session=False)
            cdb.query(Cheque).filter_by(company_id=company_id).delete(synchronize_session=False)
            cdb.query(BankTransaction).filter_by(company_id=company_id).delete(synchronize_session=False)
            cdb.query(BankAccount).filter_by(company_id=company_id).delete(synchronize_session=False)
            cdb.query(Loan).filter_by(company_id=company_id).delete(synchronize_session=False)
            cdb.query(CashTransaction).filter_by(company_id=company_id).delete(synchronize_session=False)
            cdb.query(Expense).filter_by(company_id=company_id).delete(synchronize_session=False)
            removed.append("Cash, Bank, Loans & Expenses")

        if "price_lists" in categories:
            cdb.query(RateLookup).filter_by(company_id=company_id).delete(synchronize_session=False)
            cdb.query(PriceList).filter_by(company_id=company_id).delete(synchronize_session=False)
            removed.append("Price Lists")

        if "whatsapp" in categories:
            cdb.query(WhatsAppLog).filter_by(company_id=company_id).delete(synchronize_session=False)
            removed.append("WhatsApp Logs")

        cdb.commit()
        flash(
            f"Cleared: {', '.join(removed)}. Company profile, users and permissions were not touched.",
            "success",
        )
    except Exception as e:
        cdb.rollback()
        print(f"[clear-data] FAILED for company {company_id}: {e}")
        flash(f"Clear-data failed and was rolled back: {e}", "danger")

    return redirect(url_for("company_settings"))


@app.route("/booking/customer")
@login_required
@require_permission("invoices", "view")
def invoice_customer_new():
    """Show the blank customer / shipment invoice form."""
    cdb = get_cdb()
    company_id = get_current_company()
    clients = cdb.query(Client).filter(
        Client.company_id == company_id,
        ~Client.client_type.in_(["Supplier", "Both", "Cash-Only"])  # Exclude suppliers
    ).all()

    price_lists = cdb.query(PriceList).filter_by(
        company_id=company_id, 
        is_active=True,
        list_type='sales'
    ).all()

    suppliers = cdb.query(Supplier).filter_by(company_id=company_id, status="Active").order_by(Supplier.name).all()

    # Auto-generate invoice ID
    invoice_id = _next_numbered_id(cdb, Invoice.invoice_id, "", extra_filters=[Invoice.company_id == company_id])
    docket_no  = _next_awb_number(company_id)

    return render_template(
        "booking.html",
        clients=clients,
        suppliers=suppliers,
        invoice_id=invoice_id,
        docket_no=docket_no,
        today=str(today_ist()),
        form_data={},
        stock_items_json=json.dumps([{
            "code":     s.code,
            "name":     s.name,
            "unit":     s.unit or "pcs",
            "quantity": s.quantity,
        } for s in cdb.query(StockItem).filter_by(company_id=company_id).order_by(StockItem.name).all()]),
        price_lists=price_lists
    )


@app.route("/booking/customer/save", methods=["POST"])
@login_required
@require_permission("invoices", "create")
def invoice_customer_save():
    """Save a customer / shipment invoice submitted from booking.html."""
    cdb = get_cdb()
    company_id = get_current_company()

    # ← insert here, before anything else
    submit_token = request.form.get("submit_token")
    if submit_token:
        existing_invoice = cdb.query(Invoice).filter_by(
            company_id=company_id,
            submit_token=submit_token
        ).first()
        if existing_invoice:
            flash("This booking was already submitted — duplicate request ignored.")
            return redirect(url_for("invoice_list"))

    # ── Basic fields ──────────────────────────────────────────────────────────
    client_id_raw  = request.form.get("customer_id")
    client_id      = int(client_id_raw) if client_id_raw else None
    invoice_date   = request.form.get("invoice_date") or str(today_ist())
    docket_no      = request.form.get("docket_no", "")
    action         = request.form.get("action", "final")

    # ── AWB/docket uniqueness — this is a CREATE-only route, so no invoice
    # to exclude. If the docket_no shown on the form got claimed by someone
    # else while this tab was open, stop here instead of writing a dupe. ──
    dupe_invoice_id = _docket_no_in_use(cdb, company_id, docket_no)
    if dupe_invoice_id:
        old_docket_no = docket_no
        docket_no = _next_awb_number(company_id)
        flash(f"AWB {old_docket_no} was already used on invoice {dupe_invoice_id} — "
              f"this invoice was automatically assigned {docket_no} instead.")

    # ── Charges & totals ──────────────────────────────────────────────────────
    freight_weight = float(request.form.get("freight_weight", 0) or 0)
    freight_rate   = float(request.form.get("freight_rate_per_kg", 0) or 0)
    # Rounded rate-card slab weight the rate lookup matched (see booking.html's
    # applyRateToFreight/calcFreight). Not used for the money calc here — this
    # route trusts freight_amount, already computed correctly client-side —
    # but it's persisted so a later edit-load has it to recompute freight from
    # correctly instead of falling back to the actual weight.
    freight_billing_weight = float(request.form.get("freight_billing_weight", 0) or 0) or freight_weight
    freight        = float(request.form.get("freight_amount", 0) or 0)
    fuel           = float(request.form.get("fuel_surcharge",  0) or 0)
    other          = float(request.form.get("other_charges",   0) or 0)
    discount       = float(request.form.get("discount_amount", 0) or 0)
    base           = freight + fuel + other
    co             = Company.query.filter_by(company_id=company_id).first()
    apply_gst      = co.is_gst_registered if (co and hasattr(co, 'is_gst_registered')) else True
    shipper_state  = request.form.get("shipper_state", "")
    receiver_state = request.form.get("receiver_state", "")
    amount_paid    = float(request.form.get("amount_paid", 0) or 0)

    # ── Payment info ─────────────────────────────────────────────────────────
    payment_mode   = request.form.get("payment_mode", "cash")
    booking_type   = request.form.get("booking_type", "credit")
    upi_app        = request.form.get("upi_app", "")
    upi_ref        = request.form.get("upi_ref", "")
    cheque_no      = request.form.get("cheque_no", "")
    cheque_date    = request.form.get("cheque_date", "")
    cheque_bank    = request.form.get("cheque_bank", "")

    # ── Resale Charges ──────────────────────────────────────────────────────────
    has_resale = request.form.get("resale_active") == "true"
    resale_amount = float(request.form.get("resale_amount", 0) or 0)
    resale_reason = request.form.get("resale_reason", "").strip()
    resale_date_str = request.form.get("resale_date")
    resale_notes = request.form.get("resale_notes", "").strip()

    if has_resale and resale_amount > 0:
        resale_date = date.fromisoformat(resale_date_str) if resale_date_str else today_ist()
    else:
        resale_amount = 0
        resale_date = None
        resale_reason = None
        resale_notes = None

    # ── GST: proper CGST/SGST vs IGST split (based on shipper/receiver state)
    # plus round-off to the nearest rupee, instead of a flat 18% figure. ──────
    # Discount is taken off before tax — it reduces what the customer is
    # actually being charged for, so it shouldn't be taxed. Clamped at 0 so a
    # discount bigger than the freight+charges can't flip the invoice negative.
    taxable_base = max(0, base + resale_amount - discount)
    gst_calc = compute_invoice_gst(taxable_base, apply_gst, shipper_state, receiver_state)
    gst = gst_calc["gst_total"]
    resale_gst = 0  # resale GST is now folded into the single gst_calc split above
    grand_total = gst_calc["grand_total"]
    balance = round(grand_total - amount_paid, 2)

    # ── Status ────────────────────────────────────────────────────────────────
    if action == "draft":
        status = "Draft"
    elif balance <= 0:
        status = "Paid"
    elif amount_paid > 0:
        status = "Partial"
    else:
        status = "Pending"

    # Credit bookings must be tied to a client, or the pending balance below
    # never gets attached to anyone's outstanding ledger. Cash/UPI walking
    # customers are fine with no client — they're not carrying a balance.
    if action != "draft" and payment_mode == "credit" and not client_id:
        flash("Credit bookings require a customer to be selected.", "error")
        return redirect(url_for("invoice_customer_new"))

    # ── Credit limit check (customer invoices, cash bookings included) ────────
    # booking.html asks the user to confirm this via a popup BEFORE this
    # request is even sent (see /invoice/customer/check-credit-limit), so on
    # a normal submit this is just a backstop. We only flash here in the
    # "block" case — flashing the "warn" case too would just resurface the
    # message on /invoice/list after the redirect below, which is the bug
    # this replaces.
    if action != "draft" and client_id:
        _client_for_limit = cdb.query(Client).filter_by(id=client_id, company_id=company_id).first()
        _limit_ok, _limit_msg = _check_credit_limit(co, _client_for_limit, grand_total)
        if not _limit_ok:
            flash(_limit_msg, "danger")
            return redirect(url_for("invoice_customer_new"))

    # ── Generate invoice ID ───────────────────────────────────────────────────
    invoice_id = _next_numbered_id(cdb, Invoice.invoice_id, "", extra_filters=[Invoice.company_id == company_id])

    # ── Shipment / receiver details stored in notes / terms ──────────────────
    notes = request.form.get("notes", "")
    
    # ── Process Packages - ADD TO INVENTORY AND CREATE INVOICE ITEMS ──────────
    pkg_names = request.form.getlist("pkg_name[]")
    pkg_types = request.form.getlist("pkg_type[]")
    pkg_units = request.form.getlist("pkg_unit[]")
    pkg_qtys  = request.form.getlist("pkg_qty[]")
    pkg_l     = request.form.getlist("pkg_l[]")
    pkg_w     = request.form.getlist("pkg_w[]")
    pkg_h     = request.form.getlist("pkg_h[]")
    pkg_wt    = request.form.getlist("pkg_wt[]")
    pkg_division = request.form.getlist("pkg_division[]")
    pkg_discount = request.form.getlist("pkg_discount[]")
    pkg_discwt   = request.form.getlist("pkg_discwt[]")
    pkg_volwt    = request.form.getlist("pkg_volwt[]")
    pkg_chgwt    = request.form.getlist("pkg_chgwt[]")
    pkg_rates = request.form.getlist("pkg_rate[]")
    
    stock_added = []
    stock_warnings = []
    invoice_items_data = []  # Store for creating InvoiceItem records

    # ── Shipment-level detail to stamp onto each package's history row ────────
    ship_source      = (request.form.get("shipper_city") or request.form.get("origin") or "India")
    ship_destination = request.form.get("destination", "")

        # ── Get cash client ID if this is a cash booking ──────────────────────
    cash_client_id = None
    if booking_type == "cash":
        cash_shipper_name = (request.form.get("shipper_name", "") or "").strip()
        if cash_shipper_name:
            cash_client = _get_or_create_cash_client(cdb, company_id, cash_shipper_name)
            if cash_client:
                cash_client_id = cash_client.id

    # ── Process each package ───────────────────────────────────────────────
    for i in range(len(pkg_names)):
        item_name = (pkg_names[i] or "").strip()
        if not item_name:
            continue

        stock_item_id = None
        qty      = float(pkg_qtys[i] or 1) if pkg_qtys[i] else 1
        rate     = float(pkg_rates[i] or 0) if pkg_rates[i] else 0
        pkg_type = (pkg_types[i] if i < len(pkg_types) else "Box") or "Box"

        # Determine which client_id to use for stock matching
        if booking_type == "cash":
            # Use the cash client ID (NOT shipper_name string)
            # This ensures stock is tracked per cash customer
            client_id_for_stock = cash_client_id
            shipper_name_for_stock = None  # Not needed when using client_id
        else:
            # Credit booking - use the regular client ID
            client_id_for_stock = client_id
            shipper_name_for_stock = None

        # Match stock by name + client_id (no shipper_name needed for cash now)
        stock_filters = dict(
            company_id=company_id, 
            name=item_name, 
            client_id=client_id_for_stock
        )
        existing_item = cdb.query(StockItem).filter_by(**stock_filters).first()

        if existing_item:
            stock_item_id = existing_item.id
            existing_item.quantity   += qty
            existing_item.last_updated = today_ist()
            if rate > 0:
                existing_item.unit_price   = rate
                existing_item.purchase_rate = rate
            cdb.add(StockPurchaseHistory(
                stock_item_id=existing_item.id,
                purchase_invoice_id=None,
                reference=invoice_id,
                quantity=qty,
                purchase_rate=rate,
                gst_percent=existing_item.gst_percent or 0,
                purchase_date=date.fromisoformat(invoice_date),
                awb_no=docket_no,
                source=ship_source,
                destination=ship_destination,
                length=float(pkg_l[i] or 0) if i < len(pkg_l) else 0,
                width=float(pkg_w[i] or 0) if i < len(pkg_w) else 0,
                height=float(pkg_h[i] or 0) if i < len(pkg_h) else 0,
                weight=float(pkg_wt[i] or 0) if i < len(pkg_wt) else 0,
            ))
        else:
            new_code = _next_numbered_id(cdb, StockItem.code, "PKG-", extra_filters=[StockItem.company_id == company_id])
            new_item = StockItem(
                company_id    = company_id,
                code          = new_code,
                name          = item_name,
                category      = "Packaging",
                item_type     = pkg_type,
                client_id     = client_id_for_stock,  # ← Now uses cash_client_id for cash bookings
                shipper_name  = None,  # ← No longer needed for cash bookings
                quantity      = qty,
                unit          = "pcs",
                unit_price    = rate,
                purchase_rate = rate,
                reorder_level = 0,
                gst_percent   = 18,
                hsn           = "",
                last_updated  = today_ist(),
            )
            cdb.add(new_item)
            cdb.flush()
            cdb.add(StockPurchaseHistory(
                stock_item_id=new_item.id,
                purchase_invoice_id=None,
                reference=invoice_id,
                quantity=qty,
                purchase_rate=rate,
                gst_percent=18,
                purchase_date=date.fromisoformat(invoice_date),
                awb_no=docket_no,
                source=ship_source,
                destination=ship_destination,
                length=float(pkg_l[i] or 0) if i < len(pkg_l) else 0,
                width=float(pkg_w[i] or 0) if i < len(pkg_w) else 0,
                height=float(pkg_h[i] or 0) if i < len(pkg_h) else 0,
                weight=float(pkg_wt[i] or 0) if i < len(pkg_wt) else 0,
            ))
            stock_added.append(f"{qty}× {item_name} (new stock item {new_code})")
            stock_item_id = new_item.id
        
        # Store invoice item data for later creation
        invoice_items_data.append({
            'stock_item_id': stock_item_id,
            'description': item_name,
            'qty': qty,
            'rate': rate,
            'discount': 0
        })
    
    # Collect package data for JSON storage
    packages_data = []
    for i in range(len(pkg_names)):
        if pkg_names[i] and pkg_names[i].strip():
            packages_data.append({
                "name": pkg_names[i],
                "type": pkg_types[i] if i < len(pkg_types) else "",
                "unit": pkg_units[i] if i < len(pkg_units) else "cm",
                "qty": float(pkg_qtys[i] or 1) if pkg_qtys[i] else 1,
                "length": float(pkg_l[i] or 0) if i < len(pkg_l) else 0,
                "width": float(pkg_w[i] or 0) if i < len(pkg_w) else 0,
                "height": float(pkg_h[i] or 0) if i < len(pkg_h) else 0,
                "weight": float(pkg_wt[i] or 0) if i < len(pkg_wt) else 0,
                "division": float(pkg_division[i] or 5000) if i < len(pkg_division) and pkg_division[i] else 5000,
                "discount": float(pkg_discount[i] or 0) if i < len(pkg_discount) and pkg_discount[i] else 0,
                "discount_wt": float(pkg_discwt[i] or 0) if i < len(pkg_discwt) and pkg_discwt[i] else 0,
                "vol_weight": float(pkg_volwt[i] or 0) if i < len(pkg_volwt) and pkg_volwt[i] else 0,
                "chg_weight": float(pkg_chgwt[i] or 0) if i < len(pkg_chgwt) and pkg_chgwt[i] else 0,
                "rate": float(pkg_rates[i] or 0) if i < len(pkg_rates) else 0,
            })
    
    # Pack all extra shipment metadata into the terms field as JSON
    shipment_meta = json.dumps({
        "docket_no":        docket_no,
        "shipper_name":     request.form.get("shipper_name", ""),
        "shipper_contact_name": request.form.get("shipper_contact_name", ""),
        "shipper_address1": request.form.get("shipper_address1", ""),
        "shipper_address2": request.form.get("shipper_address2", ""),
        "shipper_city": request.form.get("shipper_city", ""),
        "shipper_state": request.form.get("shipper_state", ""),
        "shipper_pincode": request.form.get("shipper_pincode", ""),
        "shipper_country": request.form.get("shipper_country", "India"),
        "shipper_doc_type": request.form.get("shipper_doc_type", ""),
        "shipper_doc_no": request.form.get("shipper_doc_no", ""),
        "client_code": request.form.get("client_code", ""),
        "receiver_name": request.form.get("receiver_name", ""),
        "receiver_company": request.form.get("receiver_company", ""),
        "receiver_phone": request.form.get("receiver_phone", ""),
        "receiver_address1": request.form.get("receiver_address1", ""),
        "receiver_address2": request.form.get("receiver_address2", ""),
        "receiver_city": request.form.get("receiver_city", ""),
        "receiver_state": request.form.get("receiver_state", ""),
        "receiver_pincode": request.form.get("receiver_pincode", ""),
        "receiver_country": request.form.get("receiver_country", "India"),
        "receiver_doc_type": request.form.get("receiver_doc_type", ""),
        "receiver_doc_no": request.form.get("receiver_doc_no", ""),
        "destination":      request.form.get("destination", ""),
        "shipment_type":    request.form.get("shipment_type", ""),
        "mode":             request.form.get("mode", ""),
        "courier_company_id": request.form.get("courier_company_id", ""),
        "carrier":          request.form.get("carrier", ""),
        "tracking_number": request.form.get("tracking_number", ""),
        "carrier_ref":      request.form.get("carrier_ref", ""),
        "origin":           request.form.get("origin", "India"),
        "pickup_date":      request.form.get("pickup_date", ""),
        "departure_time":   request.form.get("departure_time", ""),
        "expected_delivery":request.form.get("expected_delivery", ""),
        "comments":         request.form.get("comments", ""),
        "vendor":           request.form.get("vendor", ""),
        "payment_mode":     payment_mode,
        "booking_type":     booking_type,
        "upi_app":          upi_app,
        "upi_ref":          upi_ref,
        "cheque_no":        cheque_no,
        "cheque_date":      cheque_date,
        "cheque_bank":      cheque_bank,
        "freight":          freight,
        "freight_weight":   freight_weight,
        "freight_rate_per_kg": freight_rate,
        "freight_billing_weight": freight_billing_weight,
        "fuel":             fuel,
        "other":            other,
        "discount":         discount,
        "gst":              gst,
        "cgst":             gst_calc["cgst"],
        "sgst":             gst_calc["sgst"],
        "igst":             gst_calc["igst"],
        "is_interstate":    gst_calc["is_interstate"],
        "round_off":        gst_calc["round_off"],
        "amount_paid":      amount_paid,
        "packages":         packages_data,
        "resale": {
        "amount": resale_amount,
        "gst": resale_gst if has_resale else 0,
        "reason": resale_reason,
        "date": resale_date.strftime("%Y-%m-%d") if resale_date else "",
        "notes": resale_notes,
        "added_by": get_current_user().get("email")
    } if has_resale and resale_amount > 0 else None
    })

    # CREATE INVOICE
    inv = Invoice(
        invoice_id     = invoice_id,
        company_id     = company_id,
        client_id      = client_id,
        date           = date.fromisoformat(invoice_date),
        status         = status,
        contact_person = request.form.get("shipper_contact_name", ""),
        phone          = request.form.get("customer_phone", ""),
        subtotal       = base,
        tax_amount     = gst,
        grand_total    = grand_total,
        terms          = shipment_meta,
        email          = notes,
        paid_amount    = amount_paid,
        balance        = balance,
        submit_token   = submit_token,
    )
    cdb.add(inv)
    try:
        cdb.commit()
    except IntegrityError:
        # The real guard: two parallel requests both passed the check above,
        # both tried to insert, the unique constraint let exactly one through.
        cdb.rollback()
        flash("This booking was already submitted — duplicate request ignored.")
        return redirect(url_for("invoice_list"))
    cdb.flush()  # Get the invoice ID

    if booking_type == "cash":
        aadhar_front = save_shipper_id_doc(request.files.get("shipper_aadhar_front_file"), inv.invoice_id, "aadhar_front")
        aadhar_back  = save_shipper_id_doc(request.files.get("shipper_aadhar_back_file"),  inv.invoice_id, "aadhar_back")
        pan_front    = save_shipper_id_doc(request.files.get("shipper_pan_front_file"),    inv.invoice_id, "pan_front")
        pan_back     = save_shipper_id_doc(request.files.get("shipper_pan_back_file"),     inv.invoice_id, "pan_back")
        if aadhar_front or aadhar_back or pan_front or pan_back:
            meta_dict = json.loads(inv.terms) if inv.terms else {}
            meta_dict["shipper_aadhar_front_file"] = aadhar_front or ""
            meta_dict["shipper_aadhar_back_file"]  = aadhar_back or ""
            meta_dict["shipper_pan_front_file"]    = pan_front or ""
            meta_dict["shipper_pan_back_file"]     = pan_back or ""
            inv.terms = json.dumps(meta_dict)

    # CREATE INVOICE ITEMS (THIS IS WHAT WAS MISSING!)
    for item_data in invoice_items_data:
        inv_item = InvoiceItem(
            invoice_id    = inv.id,
            stock_item_id = item_data['stock_item_id'],
            code          = f"PKG-{item_data['stock_item_id']}",
            description   = item_data['description'],
            qty           = item_data['qty'],
            rate          = item_data['rate'],
            discount      = item_data['discount']
        )
        cdb.add(inv_item)

    # ── RECORD PAYMENT IN CASH IN HAND OR BANK ACCOUNT ──────────────────────────
    if amount_paid > 0:
        transaction_date = date.fromisoformat(invoice_date)

        # party_name is what the Receipts history, the debtor statement, and
        # the client ledger all filter/match on (see debtor_statement,
        # _build_client_ledger, receipt_new). Without it these transactions
        # are invisible to every one of those views even though the invoice
        # itself shows as paid. Fall back to the walk-in shipper name for
        # cash bookings that have no client_id.
        _pay_party_name = get_party_name(
            client_id=client_id,
            form=request.form,
            fallback_name=request.form.get("shipper_name", "").strip() or None
        )

        if payment_mode == "cash":
            cash_txn = CashTransaction(
                company_id=company_id,
                type="income",
                date=transaction_date,
                # Must be "Receipt", not "Sales" — the Receipts page history
                # and every statement/ledger query filter on category
                # in ("Receipt", "Adjustment"); "Sales" matched nothing.
                category="Receipt",
                description=f"Payment received for invoice {invoice_id} - Customer Invoice",
                amount=amount_paid,
                reference=invoice_id,
                notes=f"Payment via Cash from customer",
                party_name=_pay_party_name,
                created_by=get_current_user().get("email")
            )
            cdb.add(cash_txn)
        elif payment_mode == "online":
            bank_account = cdb.query(BankAccount).filter_by(
                company_id=company_id, 
                status='Active'
            ).first()
            if not bank_account:
                bank_account = BankAccount(
                    company_id=company_id,
                    bank_name="Default Bank Account",
                    account_name="Sales Receipts",
                    account_number="SALES001",
                    ifsc_code="DEFAULT0001",
                    branch="Main Branch",
                    opening_balance=0,
                    balance=amount_paid,
                    status='Active',
                    created_at=datetime.utcnow()
                )
                cdb.add(bank_account)
                cdb.flush()
            else:
                bank_account.balance += amount_paid
                bank_account.updated_at = datetime.utcnow()
            
            bank_txn = BankTransaction(
                bank_account_id=bank_account.id,
                company_id=company_id,
                type="credit",
                date=transaction_date,
                description=f"Payment received for invoice {invoice_id} - via {upi_app or 'Online'}",
                amount=amount_paid,
                reference=upi_ref or invoice_id,
                transaction_mode="Online",
                notes=f"UPI App: {upi_app}, Ref: {upi_ref}",
                party_name=_pay_party_name,
                created_by=get_current_user().get("email")
            )
            cdb.add(bank_txn)
        elif payment_mode == "cheque":
            bank_account = cdb.query(BankAccount).filter_by(
                company_id=company_id, 
                status='Active'
            ).first()
            if not bank_account:
                bank_account = BankAccount(
                    company_id=company_id,
                    bank_name=cheque_bank or "Cheque Account",
                    account_name="Cheque Receipts",
                    account_number="CHEQ001",
                    ifsc_code="CHEQ0001",
                    branch="Main Branch",
                    opening_balance=0,
                    balance=amount_paid,
                    status='Active',
                    created_at=datetime.utcnow()
                )
                cdb.add(bank_account)
                cdb.flush()
            else:
                bank_account.balance += amount_paid
                bank_account.updated_at = datetime.utcnow()
            
            bank_txn = BankTransaction(
                bank_account_id=bank_account.id,
                company_id=company_id,
                type="credit",
                date=transaction_date,
                description=f"Cheque payment received for invoice {invoice_id}",
                amount=amount_paid,
                reference=cheque_no or invoice_id,
                transaction_mode="Cheque",
                notes=f"Cheque No: {cheque_no}, Bank: {cheque_bank}, Date: {cheque_date}",
                party_name=_pay_party_name,
                created_by=get_current_user().get("email")
            )
            cdb.add(bank_txn)

    # ── Update client pending balance if credit / unpaid ──────────────────────
    if balance > 0 and client_id:
        client = cdb.query(Client).filter_by(id=client_id, company_id=company_id).first()
        if client and hasattr(client, "pending"):
            client.pending = (client.pending or 0) + balance

    # ── Save Performa Invoice items (linked Estimate) ──────────────────────────
    perf_descs   = request.form.getlist("perf_desc[]")
    perf_boxes   = request.form.getlist("perf_box[]")
    perf_hsns    = request.form.getlist("perf_hsn[]")
    perf_units   = request.form.getlist("perf_unit[]")
    perf_witems  = request.form.getlist("perf_weight_item[]")
    perf_qtys   = request.form.getlist("perf_qty[]")
    perf_rates  = request.form.getlist("perf_rate[]")
    perf_weight = request.form.get("perf_weight", "0.00").strip()
    perf_ref    = request.form.get("perf_reference", "").strip()
    perf_inv_no   = request.form.get("performa_invoice_no", "").strip()
    perf_inv_date = request.form.get("performa_invoice_date", "").strip()
    perf_export_reason = request.form.get("export_reason", "").strip()
    if perf_export_reason == "Other":
        _perf_export_reason_other = request.form.get("export_reason_other", "").strip()
        if _perf_export_reason_other:
            perf_export_reason = _perf_export_reason_other

    perf_items = []
    perf_subtotal = 0.0
    for i in range(len(perf_descs)):
        desc = (perf_descs[i] or "").strip()
        if not desc:
            continue
        qty  = float(perf_qtys[i])  if i < len(perf_qtys)  and perf_qtys[i]  else 0.0
        rate = float(perf_rates[i]) if i < len(perf_rates) and perf_rates[i] else 0.0
        perf_subtotal += qty * rate
        perf_items.append({
            "description": desc,
            "box": perf_boxes[i] if i < len(perf_boxes) else "",
            "hsn": perf_hsns[i] if i < len(perf_hsns) else "",
            "unit": perf_units[i] if i < len(perf_units) and perf_units[i] else "PCS",
            "weight": float(perf_witems[i] or 0) if i < len(perf_witems) and perf_witems[i] else 0,
            "qty": qty,
            "rate": rate,
        })

    if perf_items:
        def _fmt_addr_pi(a1, a2, city, state, pin, country):
            return ", ".join(p for p in [a1, a2, city, state, pin, country] if p)

        perf_terms = json.dumps({
            "docket_no":        docket_no,
            "linked_invoice_id": invoice_id,   # links back to the CUST- invoice
            "shipper_name":     request.form.get("shipper_name", ""),
            "shipper_phone":    request.form.get("customer_phone", ""),
            "shipper_address1": request.form.get("shipper_address1", ""),
            "shipper_address2": request.form.get("shipper_address2", ""),
            "shipper_city":     request.form.get("shipper_city", ""),
            "shipper_state":    request.form.get("shipper_state", ""),
            "shipper_pincode":  request.form.get("shipper_pincode", ""),
            "shipper_country":  request.form.get("shipper_country", "India"),
            "shipper_address":  _fmt_addr_pi(
                request.form.get("shipper_address1",""), request.form.get("shipper_address2",""),
                request.form.get("shipper_city",""), request.form.get("shipper_state",""),
                request.form.get("shipper_pincode",""), request.form.get("shipper_country",""),
            ),
            "receiver_name":    request.form.get("receiver_name", ""),
            "receiver_phone":   request.form.get("receiver_phone", ""),
            "receiver_company": "",
            "receiver_address1": request.form.get("receiver_address1", ""),
            "receiver_address2": request.form.get("receiver_address2", ""),
            "receiver_city":    request.form.get("receiver_city", ""),
            "receiver_state":   request.form.get("receiver_state", ""),
            "receiver_pincode": request.form.get("receiver_pincode", ""),
            "receiver_country": request.form.get("receiver_country", "India"),
            "receiver_address": _fmt_addr_pi(
                request.form.get("receiver_address1",""), request.form.get("receiver_address2",""),
                request.form.get("receiver_city",""), request.form.get("receiver_state",""),
                request.form.get("receiver_pincode",""), request.form.get("receiver_country",""),
            ),
            "destination":  request.form.get("destination", ""),
            "weight":       perf_weight,
            "reference":    perf_ref,
            "invoice_no":   perf_inv_no,
            "invoice_date": perf_inv_date,
            "export_reason": perf_export_reason,
            "line_items":   perf_items,
            "dimensions":   [],   # dimensions come from the packages section
        })

        # Check if an Estimate already exists for this docket (edit scenario)
        existing_est = cdb.query(Estimate).filter_by(
            company_id=company_id
        ).filter(
            Estimate.terms.like(f'%"linked_invoice_id": "{invoice_id}"%')
        ).first()

        if existing_est:
            existing_est.client_id      = client_id
            existing_est.date           = date.fromisoformat(invoice_date)
            existing_est.status         = "Paid"
            existing_est.contact_person = request.form.get("shipper_contact_name", "")
            existing_est.phone          = request.form.get("customer_phone", "")
            existing_est.subtotal       = perf_subtotal
            existing_est.grand_total    = perf_subtotal
            existing_est.tax_amount     = 0
            existing_est.terms          = perf_terms
            cdb.query(EstimateItem).filter_by(estimate_id=existing_est.id).delete()
            for item in perf_items:
                cdb.add(EstimateItem(
                    estimate_id=existing_est.id,
                    description=item["description"],
                    qty=item["qty"],
                    rate=item["rate"],
                    discount=0,
                ))
        else:
            est_id = _next_numbered_id(cdb, Estimate.estimate_id, "SHIP-" + datetime.now().strftime("%Y%m%d") + "-", extra_filters=[Estimate.company_id == company_id])
            est = Estimate(
                estimate_id    = est_id,
                company_id     = company_id,
                client_id      = client_id,
                date           = date.fromisoformat(invoice_date),
                status         = "Paid",
                contact_person = request.form.get("shipper_contact_name", ""),
                phone          = request.form.get("customer_phone", ""),
                subtotal       = perf_subtotal,
                grand_total    = perf_subtotal,
                tax_amount     = 0,
                terms          = perf_terms,
            )
            cdb.add(est)
            cdb.flush()
            for item in perf_items:
                cdb.add(EstimateItem(
                    estimate_id  = est.id,
                    description  = item["description"],
                    qty          = item["qty"],
                    rate         = item["rate"],
                    discount     = 0,
                ))

    # ── Auto-generate purchase invoice line from this booking ────────────────
    # Delegates to _sync_auto_purchase_invoice_line() (shared with
    # invoice_customer_update() so edits/re-saves can also create or repair
    # this line instead of only the initial save being able to).
    _sync_auto_purchase_invoice_line(
        cdb, company_id, request.form, packages_data,
        freight_weight, apply_gst, gst_calc,
        invoice_date, docket_no, invoice_id, inv.id, action,
    )

    # ── Auto-create / update Company Manifest from this booking ──────────────
    # Delegates to _sync_auto_manifest_entry() (shared with the manual
    # "Add Missing Purchase Line" repair route, so a repaired booking also
    # lands on the manifest instead of only getting its purchase line back).
    total_boxes_mf = int(sum(it["qty"] for it in invoice_items_data)) or 1
    primary_stock_id   = invoice_items_data[0]["stock_item_id"] if invoice_items_data else None
    primary_stock_name = invoice_items_data[0]["description"]  if invoice_items_data else None
    _sync_auto_manifest_entry(
        cdb, company_id, request.form.get("shipper_name", ""),
        request.form.get("carrier", "").strip(), action,
        invoice_date, docket_no, invoice_id, total_boxes_mf,
        primary_stock_id=primary_stock_id, primary_stock_name=primary_stock_name,
        booking_type=booking_type,
    )

    cdb.commit()
    try:
        from tasks import send_invoice_generate_notification_async
        send_invoice_generate_notification_async(company_id=company_id, invoice_id=invoice_id)
        print(f"[WhatsApp] Notification queued for invoice {invoice_id}")
    except Exception as e:
        print(f"[WhatsApp] Could not queue notification for {invoice_id}: {e}")

    

    try:
        updated_invoices = update_customer_invoice_from_booking(cdb, company_id, inv.id)
        if updated_invoices:
            flash(f"Customer invoice(s) {', '.join(str(i) for i in updated_invoices)} updated to reflect booking changes.", "info")
    except Exception as e:
        print(f"[customer-invoice-update] failed to update parent invoices: {e}")
    # ── Build flash message ───────────────────────────────────────────────────
    msg = f"Customer invoice {invoice_id} (AWB: {docket_no}) saved successfully!"
    if stock_added:
        msg += f" Stock added: {', '.join(stock_added)}."
    if amount_paid > 0:
        msg += f" Payment of ₹{amount_paid:,.2f} recorded via {payment_mode}."
    if balance > 0:
        msg += f" Balance of ₹{balance:,.2f} added to debtors."

    flash(msg)
    return redirect(url_for("invoice_list"))

@app.route("/api/suppliers/list")
@login_required
@require_permission("suppliers", "view")
def api_suppliers_list():
    cdb = get_cdb()
    company_id = get_current_company()
    
    # Query the Supplier table (not Client)
    suppliers = cdb.query(Supplier).filter(
        Supplier.company_id == company_id,
        Supplier.status == "Active"
    ).order_by(Supplier.name).all()
    
    return jsonify([{
        "id": s.id,
        "name": s.name,
        "gst": s.gst_number or "",
        "phone": s.phone or "",
        "contact_person": s.contact_person or ""
    } for s in suppliers])


# ── Suppliers ─────────────────────────────────────────────────────────────────
# Add this block in app.py right after client_delete() and before the Stock section.
# Also add  Supplier  to the customer_models import line at the top of app.py.
# ─────────────────────────────────────────────────────────────────────────────

def _normalize_supplier(s, payable=None):
    """Return a dict whose keys match what suppliers.html / supplier_form.html expect.

    `payable`: pass a live-computed total (opening_balance + Σ(grand_total -
    paid_amount) across purchase invoices) to show real current dues. If
    omitted, falls back to the cached s.payable field."""
    return {
        "id":              s.id,
        "client_name":     s.name,               # alias so ledger_statement.html's entity.client_name works for suppliers too
        "client_id":       s.supplier_id or "—",  # same reason, mirrors _normalize_client's client_id
        "supplier_id":     s.supplier_id or "—",
        "name": s.name,
        "supplier_name":   s.name,
        "supplier_type":   s.supplier_type   or "Business",
        "contact_person":  s.contact_person  or "",
        "phone":           s.phone           or "",
        "alternate_phone": s.alternate_phone or "",
        "email":           s.email           or "",
        "website":         s.website         or "",
        "address_line1":   s.address_line1   or "",
        "address_line2":   s.address_line2   or "",
        "city":            s.city            or "",
        "state":           s.state           or "",
        "pincode":         s.pincode         or "",
        "country":         s.country         or "India",
        "gst_number":      s.gst_number      or "",
        "pan_number":      s.pan_number      or "",
        "aadhar_number":   s.aadhar_number   or "",
        "gst_type":        s.gst_type        or "Regular",
        "credit_limit":    s.credit_limit    or 0.0,
        "credit_days":     s.credit_days     or 30,
        "payable":         payable if payable is not None else (s.payable or 0.0),
        "opening_balance": s.opening_balance or 0.0,
        "last_purchase":   s.last_purchase,
        "status":          s.status          or "Active",
        "notes":           s.notes           or "",
        "created_at":      s.created_at,
    }


@app.route("/suppliers")
@login_required
@require_permission("suppliers", "view")
def supplier_list():
    cdb           = get_cdb()
    company_id    = get_current_company()
    filter_status = request.args.get("status", "All")

    query = cdb.query(Supplier).filter_by(company_id=company_id)
    if filter_status != "All":
        query = query.filter_by(status=filter_status)

    supplier_rows = query.all()
    supplier_ids  = [s.id for s in supplier_rows]

    # Live total — same formula as the creditor statement's closing balance:
    # opening_balance + Σ(grand_total − paid_amount), filtered to invoices
    # since statement_cutoff (a cutoff means opening_balance already nets
    # out everything before it — summing pre-cutoff invoices again on top
    # double-counts, same bug fixed on /clients and /debtors).
    net_by_supplier = dict(
        cdb.query(
            PurchaseInvoice.supplier_id,
            func.sum(PurchaseInvoice.grand_total - PurchaseInvoice.paid_amount)
        ).filter(
            PurchaseInvoice.company_id == company_id,
            PurchaseInvoice.supplier_id.in_(supplier_ids)
        ).group_by(PurchaseInvoice.supplier_id).all()
    ) if supplier_ids else {}

    suppliers = []
    for s in supplier_rows:
        cutoff_date = s.statement_cutoff.date() if s.statement_cutoff else None
        if cutoff_date:
            net = float(cdb.query(func.sum(PurchaseInvoice.grand_total - PurchaseInvoice.paid_amount))
                        .filter(PurchaseInvoice.company_id == company_id,
                                PurchaseInvoice.supplier_id == s.id,
                                PurchaseInvoice.date >= cutoff_date)
                        .scalar() or 0)
        else:
            net = float(net_by_supplier.get(s.id, 0) or 0)
        true_payable = (s.opening_balance or 0) + net
        suppliers.append(_normalize_supplier(s, payable=true_payable))

    return render_template("suppliers.html", suppliers=suppliers, current_status=filter_status)


@app.route("/suppliers/new", methods=["GET", "POST"])
@login_required
@require_permission("suppliers", "view", method_actions={'POST': 'create'})
def supplier_new():
    cdb        = get_cdb()
    company_id = get_current_company()
    if request.method == "POST":
        f   = request.form
        gst = f.get("gst_number", "").strip().upper()

        if gst:
            existing_gst = cdb.query(Supplier).filter_by(
                company_id=company_id, gst_number=gst
            ).first()
            if existing_gst:
                flash(f"GST number {gst} is already registered to supplier '{existing_gst.name}'. Please check and try again.", "error")
                return render_template("supplier_form.html", form_data=f, existing_brands=f.getlist("brand_name[]"))

        company_obj = Company.query.filter_by(company_id=company_id).first()
        supplier_prefix = _company_name_prefix(company_obj.company_name if company_obj else "", from_end=False)
        new_supplier_id = _next_numbered_id(cdb, Supplier.supplier_id, supplier_prefix, extra_filters=[Supplier.company_id == company_id])

        new_supplier = Supplier(
            supplier_id     = new_supplier_id,
            company_id      = company_id,
            name            = f.get("supplier_name", "").strip(),
            supplier_type   = f.get("supplier_type", "Business"),
            contact_person  = f.get("contact_person", "").strip(),
            phone           = f.get("phone", "").strip(),
            alternate_phone = f.get("alternate_phone", "").strip(),
            email           = f.get("email", "").strip().lower(),
            website         = f.get("website", "").strip(),
            address_line1   = f.get("address_line1", "").strip(),
            address_line2   = f.get("address_line2", "").strip(),
            city            = f.get("city", "").strip(),
            state           = f.get("state", "").strip(),
            pincode         = f.get("pincode", "").strip(),
            country         = f.get("country", "India").strip(),
            gst_number      = gst or None,
            pan_number      = f.get("pan_number", "").strip().upper() or None,
            aadhar_number   = f.get("aadhar_number", "").strip() or None,
            gst_type        = f.get("gst_type", "Regular"),
            credit_limit    = float(f.get("credit_limit", 0) or 0),
            credit_days     = int(f.get("credit_days", 30) or 30),
            payable         = float(f.get("opening_balance", 0) or 0),
            opening_balance = float(f.get("opening_balance", 0) or 0),
            status          = f.get("status", "Active") or "Active",
            notes           = f.get("notes", "").strip(),
            created_at      = today_ist(),
        )
        cdb.add(new_supplier)
        cdb.commit()

        brand_names = [b.strip() for b in f.getlist("brand_name[]") if b.strip()]
        seen = set()
        for b in brand_names:
            key = b.lower()
            if key in seen:
                continue
            seen.add(key)
            cdb.add(SupplierBrand(supplier_id=new_supplier.id, brand_name=b))
        if brand_names:
            cdb.commit()

        flash(f"Supplier '{new_supplier.name}' added successfully!")
        return redirect(url_for("supplier_list"))
    return render_template("supplier_form.html", form_data={}, existing_brands=[])

@app.route("/debug/suppliers")
@login_required
def debug_suppliers():
    cdb = get_cdb()
    company_id = get_current_company()
    all_s = cdb.query(Supplier).filter_by(company_id=company_id).all()
    return jsonify([{"id": s.id, "name": s.name, "status": s.status} for s in all_s])

@app.route("/suppliers/<int:supplier_pk>")
@login_required
@require_permission("suppliers", "view")
def supplier_view(supplier_pk):
    cdb        = get_cdb()
    company_id = get_current_company()
    s          = _first_or_404(cdb.query(Supplier).filter_by(id=supplier_pk, company_id=company_id).first())
    supplier   = _normalize_supplier(s)
    purchases  = cdb.query(PurchaseInvoice).filter_by(company_id=company_id, supplier_id=s.id).order_by(PurchaseInvoice.date.desc()).all()
    return render_template("supplier_detail.html", supplier=supplier, purchases=purchases)

def _build_supplier_ledger(cdb, company_id, s, since=None, until=None):
    """Builds the creditor ledger for a supplier. `since` (a datetime) is the
    statement cutoff — only purchase invoices dated ON or AFTER its date are
    included, and the opening line reflects the carried-forward balance as
    of that cutoff. `until` (a date, exclusive) caps an archive at entries
    dated BEFORE today — same reasoning as _build_client_ledger()."""
    since_date = since.date() if since else None
    invoices_q = cdb.query(PurchaseInvoice).filter_by(company_id=company_id, supplier_id=s.id)
    invoices_q = invoices_q.filter(PurchaseInvoice.status.notin_(['Cancelled', 'Void']))
    if since_date:
        invoices_q = invoices_q.filter(PurchaseInvoice.date >= since_date)
    if until:
        invoices_q = invoices_q.filter(PurchaseInvoice.date < until)
    invoices = invoices_q.order_by(PurchaseInvoice.date.asc()).all()

    ledger = []
    running_balance = s.opening_balance or 0.0

    # Opening balance / balance carried forward. Same fix as the client
    # ledger: use the cutoff date, not the supplier's original created_at.
    if running_balance:
        ledger.append({
            "date": since.date() if since else (s.created_at or today_ist()),
            "type": "Balance Carried Forward" if since else "Opening Balance",
            "ref": "—",
            "awb": "", "consignor": "", "consignee": "", "destination": "", "carrier_ref": "", "carrier": "",
            "chrg_wt": 0, "act_wt": 0, "vol_wt": 0,
            "grand_total": 0, "other_charges": 0, "billing_amount": 0,
            "debit": 0,
            "credit": running_balance,
            "balance": running_balance,
            "status": "",
            "id": None,
            "inv_id": None,
        })

    for inv in invoices:
        ship_rows = _purchase_shipment_rows(inv.items)
        grand_total = inv.grand_total or 0
        # Other Charges is captured per line item on the purchase side, so the
        # invoice-level figure show alongside Grand Total is the sum across items.
        other_charges = sum(r["other_charges"] for r in ship_rows)
        running_balance += grand_total
        ledger.append({
            "date": inv.date,
            "type": "Purchase Invoice",
            "ref": inv.invoice_number or inv.invoice_id,
            "awb": ship_rows[0]["awb"],
            "consignor": "",
            "consignee": ship_rows[0]["consignee"],
            "destination": ship_rows[0]["destination"],
            "carrier_ref": ship_rows[0]["carrier_ref"],
            "carrier": ship_rows[0]["carrier"],
            "chrg_wt": ship_rows[0]["chrg_wt"],
            "act_wt": ship_rows[0]["act_wt"],
            "vol_wt": ship_rows[0]["vol_wt"],
            "grand_total": grand_total,
            "other_charges": other_charges,
            "billing_amount": other_charges + grand_total,
            "per_kg": ship_rows[0].get("per_kg", 0),
            "shipments": ship_rows,
            "debit": 0,
            "credit": grand_total,
            "balance": running_balance,
            "status": inv.status,
            "id": inv.id,
            "inv_id": inv.invoice_id,
        })

        if inv.paid_amount and inv.paid_amount > 0:
            running_balance -= inv.paid_amount
            ledger.append({
                "date": inv.date,
                "type": "Payment Made",
                "ref": inv.invoice_number or inv.invoice_id,
                "awb": "", "consignor": "", "consignee": "", "destination": "", "carrier_ref": "", "carrier": "",
                "chrg_wt": 0, "act_wt": 0, "vol_wt": 0,
                "grand_total": 0, "other_charges": 0, "billing_amount": 0,
                "debit": inv.paid_amount,
                "credit": 0,
                "balance": running_balance,
                "status": "",
                "id": inv.id,
                "inv_id": inv.invoice_id,
            })

    total_debit = sum(r["debit"] for r in ledger)
    total_credit = sum(r["credit"] for r in ledger)

    return ledger, total_debit, total_credit, running_balance


@app.route("/suppliers/<int:supplier_pk>/statement")
@login_required
@require_permission("suppliers", "view")
def supplier_statement(supplier_pk):
    """Statement view for a supplier (creditor-style ledger)"""
    cdb = get_cdb()
    company_id = get_current_company()
    s = _first_or_404(cdb.query(Supplier).filter_by(id=supplier_pk, company_id=company_id).first())

    ledger, total_debit, total_credit, running_balance = _build_supplier_ledger(
        cdb, company_id, s, since=s.statement_cutoff)

    archives = (cdb.query(StatementClosing)
                .filter_by(company_id=company_id, entity_type="supplier", entity_id=s.id)
                .order_by(StatementClosing.closed_at.desc())
                .all())

    return render_template("ledger_statement.html",
                           entity=_normalize_supplier(s),
                           company=get_company_by_id(company_id),
                           ledger=ledger,
                           total_debit=total_debit,
                           total_credit=total_credit,
                           closing_balance=running_balance,
                           mode="creditor",
                           nav_active="suppliers",
                           back_url=f"/suppliers/{supplier_pk}",
                           archive_base_url=f"/suppliers/{supplier_pk}",
                           archives=archives,
                           archived=False,
                           today=today_ist().strftime("%d %b %Y"))


@app.route("/suppliers/<int:supplier_pk>/statement/archive/<int:archive_id>")
@login_required
@require_permission("suppliers", "view")
def supplier_statement_archive(supplier_pk, archive_id):
    """Prints a frozen old statement exactly as it looked at the moment the
    payable was cleared/shifted."""
    cdb = get_cdb()
    company_id = get_current_company()
    s = _first_or_404(cdb.query(Supplier).filter_by(id=supplier_pk, company_id=company_id).first())
    archive = _first_or_404(cdb.query(StatementClosing).filter_by(
        id=archive_id, company_id=company_id, entity_type="supplier", entity_id=supplier_pk).first())

    return render_template("ledger_statement.html",
                           entity=_normalize_supplier(s),
                           company=get_company_by_id(company_id),
                           ledger=json.loads(archive.ledger_snapshot or "[]"),
                           total_debit=archive.total_debit,
                           total_credit=archive.total_credit,
                           closing_balance=archive.closing_balance,
                           mode="creditor",
                           nav_active="suppliers",
                           back_url=f"/suppliers/{supplier_pk}/statement",
                           archived=True,
                           archived_at=archive.closed_at,
                           today=today_ist().strftime("%d %b %Y"))

@app.route("/suppliers/<int:supplier_pk>/edit", methods=["GET", "POST"])
@login_required
@require_permission("suppliers", "view", method_actions={'POST': 'edit'})
def supplier_edit(supplier_pk):
    cdb        = get_cdb()
    company_id = get_current_company()
    s          = _first_or_404(cdb.query(Supplier).filter_by(id=supplier_pk, company_id=company_id).first())
    if request.method == "POST":
        f   = request.form
        gst = f.get("gst_number", "").strip().upper()

        if gst:
            existing_gst = cdb.query(Supplier).filter(
                Supplier.company_id == company_id,
                Supplier.gst_number == gst,
                Supplier.id != s.id
            ).first()
            if existing_gst:
                flash(f"GST number {gst} is already registered to supplier '{existing_gst.name}'.", "error")
                return render_template(
                    "supplier_form.html",
                    supplier=_normalize_supplier(s),
                    form_data=f,
                    existing_brands=f.getlist("brand_name[]")
                )

        s.name            = f.get("supplier_name",   s.name).strip()
        s.supplier_type   = f.get("supplier_type",   s.supplier_type)
        s.contact_person  = f.get("contact_person",  s.contact_person  or "").strip()
        s.phone           = f.get("phone",            s.phone           or "").strip()
        s.alternate_phone = f.get("alternate_phone",  s.alternate_phone or "").strip()
        s.email           = f.get("email",            s.email           or "").strip().lower()
        s.website         = f.get("website",          s.website         or "").strip()
        s.address_line1   = f.get("address_line1",    s.address_line1   or "").strip()
        s.address_line2   = f.get("address_line2",    s.address_line2   or "").strip()
        s.city            = f.get("city",             s.city            or "").strip()
        s.state           = f.get("state",            s.state           or "").strip()
        s.pincode         = f.get("pincode",          s.pincode         or "").strip()
        s.country         = f.get("country",          s.country         or "India").strip()
        s.gst_number      = gst or None
        s.pan_number      = f.get("pan_number",       s.pan_number      or "").strip().upper() or None
        s.aadhar_number   = f.get("aadhar_number",    s.aadhar_number   or "").strip() or None
        s.gst_type        = f.get("gst_type",         s.gst_type)
        s.credit_limit    = float(f.get("credit_limit",    s.credit_limit    or 0) or 0)
        s.credit_days     = int(f.get("credit_days",       s.credit_days     or 30) or 30)
        s.opening_balance = float(f.get("opening_balance", s.opening_balance or 0) or 0)
        s.status          = f.get("status", s.status)
        s.notes           = f.get("notes",  s.notes or "").strip()

        # Replace brand list with whatever was submitted (simplest correct
        # behavior for a short add/remove-row UI — no per-row diffing needed)
        cdb.query(SupplierBrand).filter_by(supplier_id=s.id).delete()
        brand_names = [b.strip() for b in f.getlist("brand_name[]") if b.strip()]
        seen = set()
        for b in brand_names:
            key = b.lower()
            if key in seen:
                continue
            seen.add(key)
            cdb.add(SupplierBrand(supplier_id=s.id, brand_name=b))

        cdb.commit()
        flash(f"Supplier '{s.name}' updated successfully!")
        return redirect(url_for("supplier_list"))
    return render_template(
        "supplier_form.html",
        supplier=_normalize_supplier(s),
        form_data={},
        existing_brands=sorted(b.brand_name for b in s.brands)
    )





# /suppliers/<id>/delete  ── kept at the old URL/template link so nothing else
# breaks, but this NO LONGER deletes the supplier row (same reasoning as
# client_delete: invoices are GST records referenced by other FKs). Now
# archives the old statement and clears the payable/statement only.
@app.route("/suppliers/<int:supplier_pk>/delete", methods=["GET", "POST"])
@login_required
@owner_required
def supplier_delete(supplier_pk):
    cdb        = get_cdb()
    company_id = get_current_company()
    s          = _first_or_404(cdb.query(Supplier).filter_by(id=supplier_pk, company_id=company_id).first())
    scope = request.args.get("scope", "till_yesterday")
    if scope not in ("complete", "till_yesterday"):
        scope = "till_yesterday"
    amount = _supplier_close_statement(cdb, company_id, s, action="cleared", scope=scope)
    cdb.commit()
    if amount:
        if scope == "complete":
            flash(f"Payable of ₹{amount:,.2f} cleared for '{s.name}', including today's entries. Old statement archived — supplier record and purchase invoices were kept.")
        else:
            flash(f"Payable of ₹{amount:,.2f} cleared for '{s.name}' up to yesterday. Old statement archived — today's entries remain in the new statement.")
    else:
        flash(f"'{s.name}' had no payable to clear.")
    return redirect(url_for("supplier_list"))


# /suppliers/<id>/shift-to-opening  ── archives the itemised ledger the same
# way as above, but carries the amount forward as a single opening_balance
# figure. Defaults to yesterday, but an explicit ?as_of=YYYY-MM-DD lets the
# user pick an earlier cutoff — anything after that date stays live.
@app.route("/suppliers/<int:supplier_pk>/shift-to-opening", methods=["GET", "POST"])
@login_required
@owner_required
def supplier_shift_to_opening(supplier_pk):
    cdb        = get_cdb()
    company_id = get_current_company()
    s          = _first_or_404(cdb.query(Supplier).filter_by(id=supplier_pk, company_id=company_id).first())
    as_of_date = None
    as_of_raw = request.args.get("as_of")
    if as_of_raw:
        try:
            as_of_date = datetime.strptime(as_of_raw, "%Y-%m-%d").date()
        except ValueError:
            as_of_date = None
    amount = _supplier_close_statement(cdb, company_id, s, action="carried_forward", as_of_date=as_of_date)
    cdb.commit()
    flash(f"₹{amount:,.2f} carried forward as opening balance for '{s.name}', as of "
          f"{(s.statement_cutoff - timedelta(days=1)).strftime('%d %b %Y')}. New statement starts "
          f"{s.statement_cutoff.strftime('%d %b %Y')}; entries from then on stay live.")
    return redirect(url_for("supplier_list"))

@app.route("/api/supplier/<int:supplier_pk>/brands")
@login_required
@require_permission("suppliers", "view")
def api_supplier_brands(supplier_pk):
    """Return this supplier's registered brand/courier names for the purchase form's
    dependent courier dropdown. Empty list means 'no brands registered' — the
    frontend falls back to the general COURIER_OPTIONS list in that case."""
    cdb        = get_cdb()
    company_id = get_current_company()
    s = cdb.query(Supplier).filter_by(id=supplier_pk, company_id=company_id).first()
    if not s:
        return jsonify({"brands": []}), 404
    brands = sorted(b.brand_name for b in s.brands)
    return jsonify({"brands": brands})


@app.route("/api/customers/list")
@login_required
@require_permission("clients", "view")
def api_customers_list():
    """Return list of customers (non-supplier clients) for the purchase form dropdown."""
    cdb = get_cdb()
    company_id = get_current_company()
    customers = cdb.query(Client).filter(
        Client.company_id == company_id,
        ~Client.client_type.in_(["Supplier", "Both", "Cash-Only"])
    ).filter(Client.status == "Active").order_by(Client.name).all()

    return jsonify([{
        "id":       c.id,
        "name":     c.name,
        "phone":    c.phone or "",
        "email":    c.email or "",
        "city":     c.city or "",
        "gst":      c.gst_number or "",
        "address":  c.address_line1 or "",
    } for c in customers])

@app.route("/api/stock/items/by-client/<int:client_id>")
@login_required
@require_permission("stock", "view")
def api_stock_items_by_client(client_id):
    """Return stock items that have been previously invoiced to a specific client.
    If no history found, returns ALL stock items."""
    cdb = get_cdb()
    company_id = get_current_company()

    # Find all stock item IDs that appear in invoices for this client
    linked_stock_ids = db.session.query(InvoiceItem.stock_item_id).join(
        Invoice, InvoiceItem.invoice_id == Invoice.id
    ).filter(
        Invoice.company_id == company_id,
        Invoice.client_id  == client_id,
        InvoiceItem.stock_item_id.isnot(None)
    ).distinct().all()

    stock_ids = [row[0] for row in linked_stock_ids if row[0] is not None]

    if not stock_ids:
        # No history - return ALL stock items
        items = cdb.query(StockItem).filter_by(company_id=company_id).order_by(StockItem.name).all()
    else:
        items = cdb.query(StockItem).filter(
            StockItem.company_id == company_id,
            StockItem.id.in_(stock_ids)
        ).order_by(StockItem.name).all()

    return jsonify([{
        "id":            item.id,
        "code":          item.code or "",
        "name":          item.name,
        "unit":          item.unit or "pcs",
        "quantity":      item.quantity,
        "unit_price":    float(item.unit_price or 0),
        "purchase_rate": float(item.purchase_rate or item.last_purchase_rate or 0),
        "gst_percent":   float(item.gst_percent or 18),
        "hsn":           item.hsn or "",
        "category":      item.category or "",
    } for item in items])


# ─────────────────────────────────────────────────────────────────────────────
# ── Shipper Invoice (estimate.html) ──────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────

def _get_available_dockets(company_id, exclude_estimate_id=None):
    """Return customer invoices that have NOT yet had a Shipper Invoice generated.
    If exclude_estimate_id is provided, include that invoice's docket even if used."""
    cdb = get_cdb()
    used_invoice_ids = set()
    shipper_estimates = cdb.query(Estimate).filter_by(company_id=company_id).all()
    
    for est in shipper_estimates:
        # Skip the current estimate being edited
        if exclude_estimate_id and est.estimate_id == exclude_estimate_id:
            continue
        if est.terms:
            try:
                t = json.loads(est.terms)
                lid = t.get("linked_invoice_id", "")
                if lid:
                    used_invoice_ids.add(lid)
            except (ValueError, TypeError):
                pass

    all_cust = cdb.query(Invoice).filter_by(company_id=company_id).filter(Invoice.invoice_id.like("CUST-%")).order_by(Invoice.date.desc()).all()

    dockets = []
    for inv in all_cust:
        if inv.invoice_id in used_invoice_ids:
            continue
        meta = {}
        if inv.terms:
            try:
                meta = json.loads(inv.terms)
            except (ValueError, TypeError):
                pass
        docket_no = meta.get("docket_no", "")
        if not docket_no:
            continue
        cname = inv.client_obj.name if inv.client_obj else (inv.contact_person or inv.invoice_id)
        dockets.append({
            "invoice_id": inv.invoice_id,
            "docket_no": docket_no,
            "customer_name": cname,
        })
    return dockets

@app.route("/api/docket-info/<docket_no>")
@login_required
@require_permission("manifest", "view")
def api_docket_info(docket_no):
    """Return sender/receiver details for a given AWB/docket number."""
    cdb = get_cdb()
    company_id = get_current_company()
    all_cust = cdb.query(Invoice).filter_by(company_id=company_id).filter(Invoice.invoice_id.like("CUST-%")).all()
    for inv in all_cust:
        meta = {}
        if inv.terms:
            try:
                meta = json.loads(inv.terms)
            except (ValueError, TypeError):
                pass
        if meta.get("docket_no", "") == docket_no:
            cname = inv.client_obj.name if inv.client_obj else (inv.contact_person or "")
            cphone = inv.client_obj.phone if inv.client_obj else (inv.phone or "")

            # Calculate total box weight from packages
            packages = meta.get("packages", [])
            box_weight = sum(
                float(p.get("weight") or 0) * float(p.get("qty") or 1)
                for p in packages
            )
            freight_weight = float(meta.get("freight_weight") or 0)
            total_weight = box_weight if box_weight > 0 else freight_weight

            # Build dimensions list for estimate from invoice packages
            dimensions = [
                {
                    "label": p.get("name") or p.get("type") or "Box",
                    "l": p.get("length") or p.get("l") or "",
                    "w": p.get("width") or p.get("w") or "",
                    "h": p.get("height") or p.get("h") or "",
                    "wt": p.get("weight") or "",
                }
                for p in packages
                if p.get("name") or p.get("type")
            ]

            return jsonify({
                "invoice_id":       inv.invoice_id,
                "client_id":        inv.client_id,
                "shipper_name":     meta.get("shipper_name", cname),
                "shipper_phone":    meta.get("shipper_phone", cphone),
                # split address fields
                "shipper_address1": meta.get("shipper_address1", ""),
                "shipper_address2": meta.get("shipper_address2", ""),
                "shipper_city":     meta.get("shipper_city", ""),
                "shipper_state":    meta.get("shipper_state", ""),
                "shipper_pincode":  meta.get("shipper_pincode", ""),
                "shipper_country":  meta.get("shipper_country", "India"),
                "shipper_address":  meta.get("shipper_address", ""),
                "receiver_name":    meta.get("receiver_name", ""),
                "receiver_phone":   meta.get("receiver_phone", ""),
                "receiver_address1": meta.get("receiver_address1", ""),
                "receiver_address2": meta.get("receiver_address2", ""),
                "receiver_city":    meta.get("receiver_city", ""),
                "receiver_state":   meta.get("receiver_state", ""),
                "receiver_pincode": meta.get("receiver_pincode", ""),
                "receiver_country": meta.get("receiver_country", "India"),
                "receiver_address": meta.get("receiver_address", ""),
                "destination":      meta.get("destination", ""),
                "shipment_type":    meta.get("shipment_type", ""),
                "mode":             meta.get("mode", ""),
                "carrier":          meta.get("carrier", ""),
                # weight / dimensions
                "weight":           str(round(total_weight, 2)),
                "box_weight":       str(round(box_weight, 2)),
                "freight_weight":   str(freight_weight),
                "dimensions":       dimensions,
            })
    return jsonify({"error": "not found"}), 404


@app.route("/api/purchase/awb-list")
@login_required
@require_permission("purchase", "view")
def api_purchase_awb_list():
    """All AWB numbers for this company, for the Purchase Bill AWB dropdown."""
    cdb = get_cdb()
    company_id = get_current_company()
    invoices = (
        cdb.query(Invoice)
        .filter_by(company_id=company_id)
        .filter(Invoice.invoice_id.like("CUST-%"))
        .order_by(Invoice.id.desc())
        .all()
    )
    seen = set()
    result = []
    for inv in invoices:
        try:
            meta = json.loads(inv.terms) if inv.terms else {}
        except (ValueError, TypeError):
            meta = {}
        docket = (meta.get("docket_no") or "").strip()
        if not docket or docket in seen:
            continue
        seen.add(docket)
        result.append({"docket_no": docket, "invoice_id": inv.invoice_id})
    return jsonify(result)


@app.route("/api/purchase/awb-info/<docket_no>")
@login_required
@require_permission("purchase", "view")
def api_purchase_awb_info(docket_no):
    """
    Given an AWB/docket number, return the party (client) name, destination,
    and how many of each packing item (box/envelope/crate) plus weight is
    still un-deducted against that AWB — for the Purchase Bill form.

    'Already deducted' = the sum of weight_kg/quantity already recorded on
    earlier PurchaseInvoiceItem rows carrying this same docket_no.
    """
    cdb = get_cdb()
    company_id = get_current_company()

    inv = (
        cdb.query(Invoice)
        .filter_by(company_id=company_id)
        .filter(Invoice.invoice_id.like("CUST-%"))
        .all()
    )
    target = None
    meta = {}
    for i in inv:
        try:
            m = json.loads(i.terms) if i.terms else {}
        except (ValueError, TypeError):
            m = {}
        if (m.get("docket_no") or "").strip() == docket_no.strip():
            target, meta = i, m
            break

    if not target:
        return jsonify({"error": "AWB not found"}), 404

    party_name = target.client_obj.name if target.client_obj else (target.contact_person or "")
    destination = meta.get("destination", "")

    # Already billed against this AWB on prior purchase bills
    already = (
        cdb.query(
            PurchaseInvoiceItem.description,
            func.sum(PurchaseInvoiceItem.quantity),
            func.sum(PurchaseInvoiceItem.weight_kg),
        )
        .filter(PurchaseInvoiceItem.docket_no == docket_no)
        .group_by(PurchaseInvoiceItem.description)
        .all()
    )
    already_qty = {row[0]: (row[1] or 0) for row in already}
    already_wt = {row[0]: (row[2] or 0) for row in already}

    items = []
    linked_items = [line for line in target.items if line.stock_item_id]
    if linked_items:
        # Weight isn't stored on StockItem/InvoiceItem, so pull it from this
        # invoice's packages metadata (matched by name) the same way the
        # no-linked-items branch below does.
        pkg_weight_by_name = {}
        for pkg in meta.get("packages", []):
            pkg_name = (pkg.get("name") or pkg.get("type") or "").strip()
            if not pkg_name:
                continue
            pkg_qty = float(pkg.get("qty") or 1)
            # Purchase bill weight follows the discounted weight (actual − Disc. Wt),
            # not the chargeable/volumetric weight — a weight discount on the AWB
            # should reduce what we're billed for here, same as booking.html's
            # "Discounted weight" column. No discount entered = just the actual weight.
            pkg_net_wt = max((float(pkg.get("weight") or 0) - float(pkg.get("discount_wt") or 0)), 0.0)
            pkg_weight_by_name[pkg_name] = pkg_weight_by_name.get(pkg_name, 0.0) + (pkg_net_wt * pkg_qty)

        for line in linked_items:
            stock = cdb.query(StockItem).filter_by(id=line.stock_item_id).first()
            if not stock:
                continue
            used = already_qty.get(stock.name, 0)
            used_wt = already_wt.get(stock.name, 0)
            total_wt = pkg_weight_by_name.get(stock.name, 0.0)
            items.append({
                "stock_item_id": stock.id,
                "name": stock.name,
                "unit": stock.unit or "pcs",
                "available_qty": max(0, float(line.qty) - used),
                "weight_kg": max(0.0, total_wt - used_wt),
            })
    else:
        for pkg in meta.get("packages", []):
            pkg_name = (pkg.get("name") or pkg.get("type") or "").strip()
            if not pkg_name:
                continue
            qty = float(pkg.get("qty") or 1)
            # Same discounted-weight rule as above: (actual − Disc. Wt), never the
            # chargeable/volumetric weight.
            net_unit_wt = max((float(pkg.get("weight") or 0) - float(pkg.get("discount_wt") or 0)), 0.0)
            weight = net_unit_wt * qty
            stock = (
                cdb.query(StockItem)
                .filter(StockItem.company_id == company_id, StockItem.name == pkg_name)
                .first()
            ) or (
                cdb.query(StockItem)
                .filter(StockItem.company_id == company_id, StockItem.name.ilike(f"%{pkg_name}%"))
                .first()
            )
            used_qty = already_qty.get(pkg_name, 0)
            used_wt = already_wt.get(pkg_name, 0)
            items.append({
                "stock_item_id": stock.id if stock else None,
                "name": pkg_name,
                "unit": stock.unit if stock else "pcs",
                "available_qty": max(0, qty - used_qty),
                "weight_kg": max(0.0, weight - used_wt),
            })

    return jsonify({
        "docket_no": docket_no,
        "invoice_id": target.invoice_id,
        "party_name": party_name,
        "destination": destination,
        "carrier_suggested": meta.get("carrier", ""),
        "carrier_ref": meta.get("carrier_ref", ""),
        "items": items,
    })


# ── Estimate to Booking Conversion ──────────────────────────────────────────

@app.route("/estimate/new", methods=["GET", "POST"])
@login_required
@require_permission("estimates", "view", method_actions={'POST': 'create'})
def estimate_new():
    """Create a new estimate (quotation/proforma)"""
    cdb = get_cdb()
    company_id = get_current_company()
    company = Company.query.filter_by(company_id=company_id).first()
    clients = cdb.query(Client).filter(
        Client.company_id == company_id,
        ~Client.client_type.in_(["Supplier", "Both", "Cash-Only"])
    ).all()
    
    suppliers = cdb.query(Supplier).filter_by(
        company_id=company_id, 
        status="Active"
    ).order_by(Supplier.name).all()
    
    price_lists = cdb.query(PriceList).filter_by(
        company_id=company_id,
        is_active=True,
        list_type='sales'
    ).all()

    edit_id = request.args.get("edit")
    existing = cdb.query(Estimate).filter_by(estimate_id=edit_id, company_id=company_id).first() if edit_id else None
    
    if request.method == "POST":
        # Parse form data
        booking_type = request.form.get("booking_type", "credit")
        client_id_raw = request.form.get("customer_id")
        client_id = int(client_id_raw) if client_id_raw else None
        estimate_date = request.form.get("estimate_date") or str(today_ist())
        valid_until = request.form.get("valid_until")
        reference = request.form.get("reference", "").strip()
        
        # Handle cash/walk-in booking - get shipper name from manual input
        shipper_name = request.form.get("shipper_name", "")
        if booking_type == "cash":
            manual_name = request.form.get("shipper_name_cash", "").strip()
            if manual_name:
                shipper_name = manual_name
            client_id = None
        else:
            if client_id:
                client = cdb.query(Client).filter_by(id=client_id, company_id=company_id).first()
                if client:
                    shipper_name = client.name
        
        # Get packages — same pkg_*[] fields as booking.html's Packages section
        pkg_names = request.form.getlist("pkg_name[]")
        pkg_types = request.form.getlist("pkg_type[]")
        pkg_units = request.form.getlist("pkg_unit[]")
        pkg_qtys = request.form.getlist("pkg_qty[]")
        pkg_l = request.form.getlist("pkg_l[]")
        pkg_w = request.form.getlist("pkg_w[]")
        pkg_h = request.form.getlist("pkg_h[]")
        pkg_wt = request.form.getlist("pkg_wt[]")
        pkg_division = request.form.getlist("pkg_division[]")
        pkg_discount = request.form.getlist("pkg_discount[]")
        pkg_discwt = request.form.getlist("pkg_discwt[]")
        pkg_volwt = request.form.getlist("pkg_volwt[]")
        pkg_chgwt = request.form.getlist("pkg_chgwt[]")
        pkg_rates = request.form.getlist("pkg_rate[]")

        # Consignee / receiver
        receiver_name = request.form.get("receiver_name", "").strip()
        receiver_company = request.form.get("receiver_company", "").strip()
        receiver_phone = request.form.get("receiver_phone", "").strip()
        receiver_address1 = request.form.get("receiver_address1", "").strip()
        receiver_address2 = request.form.get("receiver_address2", "").strip()
        receiver_city = request.form.get("receiver_city", "").strip()
        receiver_state = request.form.get("receiver_state", "").strip()
        receiver_pincode = request.form.get("receiver_pincode", "").strip()
        receiver_country = request.form.get("receiver_country", "India").strip()

        # Shipper address
        shipper_address1 = request.form.get("address1", "").strip()
        shipper_address2 = request.form.get("address2", "").strip()
        shipper_city = request.form.get("city", "").strip()
        shipper_state = request.form.get("state", "").strip()
        shipper_pincode = request.form.get("pincode", "").strip()
        shipper_country = request.form.get("country", "India").strip()
        contact_person = request.form.get("contact_person", "").strip()
        customer_phone = request.form.get("customer_phone", "").strip()

        # Service details
        destination = request.form.get("destination", "").strip()
        mode = request.form.get("mode", "").strip()
        carrier = request.form.get("carrier", "").strip()
        courier_company_id = request.form.get("courier_company_id", "").strip()
        vendor = request.form.get("vendor", "").strip()
        pickup_date = request.form.get("pickup_date", str(today_ist()))
        expected_delivery = request.form.get("expected_delivery", "").strip()
        tracking_number = request.form.get("tracking_number", "").strip()

        freight_weight = float(request.form.get("freight_weight", 0) or 0)
        freight_rate = float(request.form.get("freight_rate", 0) or 0)
        freight_amount = float(request.form.get("freight_amount", 0) or 0)
        other_charges = float(request.form.get("other_charges", 0) or 0)
        discount = float(request.form.get("discount", 0) or 0)
        notes = request.form.get("notes", "").strip()
        action = request.form.get("action", "final")
        
        # Calculate item subtotal — packages ARE the line items now (Description
        # field removed so the estimate's Packages section is identical to
        # booking.html's, and edit-reload reads back from the same "packages"
        # shape both screens share).
        line_items = []
        packages_data = []
        item_subtotal = 0.0
        for i in range(len(pkg_names)):
            name = (pkg_names[i] or "").strip()
            if not name:
                continue
            qty = float(pkg_qtys[i] or 1) if i < len(pkg_qtys) and pkg_qtys[i] else 1
            rate = float(pkg_rates[i] or 0) if i < len(pkg_rates) and pkg_rates[i] else 0
            weight = float(pkg_wt[i] or 0) if i < len(pkg_wt) and pkg_wt[i] else 0
            length = float(pkg_l[i] or 0) if i < len(pkg_l) and pkg_l[i] else 0
            width = float(pkg_w[i] or 0) if i < len(pkg_w) and pkg_w[i] else 0
            height = float(pkg_h[i] or 0) if i < len(pkg_h) and pkg_h[i] else 0
            division = float(pkg_division[i] or 5000) if i < len(pkg_division) and pkg_division[i] else 5000
            discount = float(pkg_discount[i] or 0) if i < len(pkg_discount) and pkg_discount[i] else 0
            discount_wt = float(pkg_discwt[i] or 0) if i < len(pkg_discwt) and pkg_discwt[i] else 0
            vol_weight = float(pkg_volwt[i] or 0) if i < len(pkg_volwt) and pkg_volwt[i] else 0
            chg_weight = float(pkg_chgwt[i] or 0) if i < len(pkg_chgwt) and pkg_chgwt[i] else 0
            amount = qty * rate
            item_subtotal += amount
            line_items.append({
                "description": name,
                "hsn": "",
                "qty": qty,
                "weight": weight,
                "unit": "PCS",
                "rate": rate,
                "amount": amount,
                "length": length,
                "width": width,
                "height": height,
            })
            packages_data.append({
                "name": name,
                "type": pkg_types[i] if i < len(pkg_types) and pkg_types[i] else "Box",
                "unit": pkg_units[i] if i < len(pkg_units) and pkg_units[i] else "cm",
                "qty": qty,
                "length": length,
                "width": width,
                "height": height,
                "weight": weight,
                "division": division,
                "discount": discount,
                "discount_wt": discount_wt,
                "vol_weight": vol_weight,
                "chg_weight": chg_weight,
                "rate": rate,
            })
        
        # Calculate totals with GST
        base = item_subtotal + freight_amount + other_charges - discount
        apply_gst = company.is_gst_registered if (company and hasattr(company, 'is_gst_registered')) else True
        gst_calc = compute_invoice_gst(base, apply_gst, shipper_state, receiver_state)
        grand_total = gst_calc["grand_total"]
        
        if action == "draft":
            status = "Draft"
        else:
            status = "Unpaid"

        # Build terms JSON
        terms_data = json.dumps({
            "booking_type": booking_type,
            "reference": reference,
            "shipper_name": shipper_name,
            "shipper_address1": shipper_address1,
            "shipper_address2": shipper_address2,
            "shipper_city": shipper_city,
            "shipper_state": shipper_state,
            "shipper_pincode": shipper_pincode,
            "shipper_country": shipper_country,
            "customer_phone": customer_phone,
            "contact_person": contact_person,
            "receiver_name": receiver_name,
            "receiver_company": receiver_company,
            "receiver_phone": receiver_phone,
            "receiver_address1": receiver_address1,
            "receiver_address2": receiver_address2,
            "receiver_city": receiver_city,
            "receiver_state": receiver_state,
            "receiver_pincode": receiver_pincode,
            "receiver_country": receiver_country,
            "destination": destination,
            "mode": mode,
            "carrier": carrier,
            "courier_company_id": courier_company_id,
            "vendor": vendor,
            "pickup_date": pickup_date,
            "expected_delivery": expected_delivery,
            "tracking_number": tracking_number,
            "freight_weight": freight_weight,
            "freight_rate": freight_rate,
            "freight_amount": freight_amount,
            "other_charges": other_charges,
            "discount": discount,
            "other_reason": request.form.get("other_reason", ""),
            "gst": gst_calc["gst_total"],
            "cgst": gst_calc["cgst"],
            "sgst": gst_calc["sgst"],
            "igst": gst_calc["igst"],
            "is_interstate": gst_calc["is_interstate"],
            "round_off": gst_calc["round_off"],
            "notes": notes,
            "line_items": line_items,
            "packages": packages_data,
            "weight": freight_weight,
        })
        
        # Check if editing existing
        edit_estimate_id = request.form.get("edit_estimate_id")
        
        if edit_estimate_id:
            est = cdb.query(Estimate).filter_by(estimate_id=edit_estimate_id, company_id=company_id).first()
            if est:
                est.client_id = client_id
                est.date = date.fromisoformat(estimate_date)
                est.valid_until = date.fromisoformat(valid_until) if valid_until else None
                est.status = status
                est.contact_person = contact_person or shipper_name
                est.phone = customer_phone or est.phone or ""
                est.subtotal = item_subtotal + freight_amount + other_charges
                est.tax_amount = gst_calc["gst_total"]
                est.grand_total = grand_total
                est.terms = terms_data
                est.email = notes
                
                # Update items
                cdb.query(EstimateItem).filter_by(estimate_id=est.id).delete()
                for item in line_items:
                    cdb.add(EstimateItem(
                        estimate_id=est.id,
                        description=item["description"],
                        qty=item["qty"],
                        rate=item["rate"],
                        discount=0,
                    ))
                
                cdb.commit()
                flash(f"Estimate {est.estimate_id} updated successfully!")
                return redirect(url_for("estimate_view", estimate_id=est.estimate_id))
        
        # Create new estimate
        estimate_id = _next_numbered_id(cdb, Estimate.estimate_id, "EST-" + datetime.now().strftime("%Y%m%d") + "-", 
                                        extra_filters=[Estimate.company_id == company_id])
        
        est = Estimate(
            estimate_id=estimate_id,
            company_id=company_id,
            client_id=client_id,
            date=date.fromisoformat(estimate_date),
            valid_until=date.fromisoformat(valid_until) if valid_until else None,
            status=status,
            contact_person=contact_person or shipper_name,
            phone=customer_phone or "",
            subtotal=item_subtotal + freight_amount + other_charges,
            tax_amount=gst_calc["gst_total"],
            grand_total=grand_total,
            terms=terms_data,
            email=notes,
        )
        cdb.add(est)
        cdb.flush()
        
        for item in line_items:
            cdb.add(EstimateItem(
                estimate_id=est.id,
                description=item["description"],
                qty=item["qty"],
                rate=item["rate"],
                discount=0,
            ))
        
        cdb.commit()
        flash(f"Estimate {estimate_id} created successfully!")
        return redirect(url_for("estimate_view", estimate_id=estimate_id))
    
    # ============================================================
    # GET - render form - FIXED to properly load existing data
    # ============================================================
    form_data = {}
    estimate_date = str(today_ist())
    valid_until = str(today_ist() + timedelta(days=30))
    is_edit = False
    estimate_id = None
    packages = []  # IMPORTANT: Initialize packages for the template
    
    if existing:
        is_edit = True
        estimate_date = existing.date.strftime('%Y-%m-%d')
        valid_until = existing.valid_until.strftime('%Y-%m-%d') if existing.valid_until else ''
        estimate_id = existing.estimate_id
        
        # Parse terms
        try:
            meta = json.loads(existing.terms) if existing.terms else {}
        except:
            meta = {}
        
        # ============================================================
        # CRITICAL FIX: Get line items with ALL data from meta
        # ============================================================
        if meta.get("line_items"):
            line_items = meta.get("line_items", [])
        else:
            line_items = []
            for item in existing.items:
                line_items.append({
                    "description": item.description or "",
                    "hsn": "",
                    "qty": item.qty or 0,
                    "weight": 0,
                    "length": 0,
                    "width": 0,
                    "height": 0,
                    "unit": "PCS",
                    "rate": item.rate or 0,
                    "amount": (item.qty or 0) * (item.rate or 0)
                })
        
        # ============================================================
        # CRITICAL FIX: Build packages data for the template
        # ============================================================
        packages = []
        if meta.get("packages"):
            packages = meta.get("packages", [])
        else:
            # Build from line items
            for item in line_items:
                packages.append({
                    "name": item.get("description", "Box"),
                    "type": "Box",
                    "unit": "cm",
                    "qty": item.get("qty", 1),
                    "length": item.get("length", 0),
                    "width": item.get("width", 0),
                    "height": item.get("height", 0),
                    "weight": item.get("weight", 0),
                    "division": 5000,
                    "discount": 0,
                    "discount_wt": 0,
                    "vol_weight": 0,
                    "chg_weight": item.get("weight", 0),
                    "rate": item.get("rate", 0),
                })
        
        # If still no packages, create a default one
        if not packages:
            packages = [{"name": "Box", "type": "Box", "qty": 1, "length": "", "width": "", "height": "", "weight": "", "rate": 0}]
        
        # Get client name for display
        client_name = ""
        if existing.client_id:
            client = cdb.query(Client).filter_by(id=existing.client_id, company_id=company_id).first()
            if client:
                client_name = client.name
        
        # ============================================================
        # CRITICAL FIX: Build complete form_data with ALL fields
        # ============================================================
        form_data = {
            "booking_type": meta.get("booking_type", "credit"),
            "status": existing.status or "Draft",
            "customer_id": existing.client_id,
            "customer_name": client_name,
            "customer_phone": meta.get("customer_phone", existing.phone or ""),
            "contact_person": meta.get("contact_person", existing.contact_person or ""),
            "shipper_name": meta.get("shipper_name", client_name or existing.contact_person or ""),
            "reference": meta.get("reference", ""),
            "address1": meta.get("shipper_address1", ""),
            "address2": meta.get("shipper_address2", ""),
            "city": meta.get("shipper_city", ""),
            "state": meta.get("shipper_state", ""),
            "pincode": meta.get("shipper_pincode", ""),
            "country": meta.get("shipper_country", "India"),
            "receiver_name": meta.get("receiver_name", ""),
            "receiver_company": meta.get("receiver_company", ""),
            "receiver_phone": meta.get("receiver_phone", ""),
            "receiver_address1": meta.get("receiver_address1", ""),
            "receiver_address2": meta.get("receiver_address2", ""),
            "receiver_city": meta.get("receiver_city", ""),
            "receiver_state": meta.get("receiver_state", ""),
            "receiver_pincode": meta.get("receiver_pincode", ""),
            "receiver_country": meta.get("receiver_country", "India"),
            "destination": meta.get("destination", ""),
            "mode": meta.get("mode", ""),
            "carrier": meta.get("carrier", ""),
            "courier_company_id": meta.get("courier_company_id", ""),
            "vendor": meta.get("vendor", ""),
            "pickup_date": meta.get("pickup_date", str(today_ist())),
            "expected_delivery": meta.get("expected_delivery", ""),
            "tracking_number": meta.get("tracking_number", ""),
            # ============================================================
            # CRITICAL FIX: Load freight and charges data
            # ============================================================
            "freight_weight": meta.get("freight_weight", 0),
            "freight_rate": meta.get("freight_rate", 0),
            "freight_amount": meta.get("freight_amount", 0),
            "other_charges": meta.get("other_charges", 0),
            "discount": meta.get("discount", 0),
            "other_reason": meta.get("other_reason", ""),
            "notes": existing.email or "",
            "items": line_items,
        }
    else:
        # New estimate - default packages
        packages = [{"name": "Box", "type": "Box", "qty": 1, "length": "", "width": "", "height": "", "weight": "", "rate": 0}]
        form_data = {
            "booking_type": "credit",
            "items": [],
        }
    
    is_gst_registered = company.is_gst_registered if (company and hasattr(company, 'is_gst_registered')) else True
    
    return render_template("estimate_form.html",
                         suppliers=suppliers,
                         clients=clients,
                         form_data=form_data,
                         packages=packages,  # Pass packages to template
                         estimate_date=estimate_date,
                         valid_until=valid_until,
                         estimate_id=estimate_id if is_edit else None,
                         is_edit=is_edit,
                         today=str(today_ist()),
                         company=company,
                         is_gst_registered=is_gst_registered,
                         price_lists=price_lists)


@app.route("/estimate/list")
@login_required
@require_permission("estimates", "view")
def estimate_list():
    cdb = get_cdb()
    company_id = get_current_company()
    filter_status = request.args.get("status", "All")
    query = cdb.query(Estimate).filter_by(company_id=company_id)
    if filter_status != "All":
        query = query.filter_by(status=filter_status)
    raw = query.order_by(Estimate.date.desc()).all()

    estimates = []
    for est in raw:
        meta = {}
        if est.terms:
            try:
                meta = json.loads(est.terms)
            except (ValueError, TypeError):
                meta = {}
        
        # Check if this estimate has been converted
        converted_to = meta.get("converted_to_invoice", "")
        
        estimates.append({
            "id": est.estimate_id,
            "date": est.date.strftime("%d %b %Y") if est.date else "—",
            "valid_until": est.valid_until.strftime("%d %b %Y") if est.valid_until else "—",
            "status": est.status or "Draft",
            "grand_total": est.grand_total or 0,
            "subtotal": est.subtotal or 0,
            "tax_amount": est.tax_amount or 0,
            "client_name": est.client_obj.name if est.client_obj else (est.contact_person or "—"),
            "phone": est.phone or "",
            "destination": meta.get("destination", ""),
            "reference": meta.get("reference", ""),
            "weight": meta.get("freight_weight", 0) or 0,
            "converted_to": converted_to,
        })

    return render_template("estimate_list.html", 
                         estimates=estimates, 
                         current_status=filter_status)

@app.route("/estimate/view/<estimate_id>")
@login_required
@require_permission("estimates", "view")
def estimate_view(estimate_id):
    """Show a single estimate — the print-ready proforma plus the on-screen
    summary that estimate_view.html renders. This was previously missing
    entirely; every redirect in the estimate flow (create, update, convert)
    points here."""
    cdb = get_cdb()
    company_id = get_current_company()

    est = cdb.query(Estimate).filter_by(estimate_id=estimate_id, company_id=company_id).first()
    if not est:
        abort(404)

    meta = {}
    if est.terms:
        try:
            meta = json.loads(est.terms)
        except (ValueError, TypeError):
            meta = {}

    # Line items — terms JSON has the full picture (hsn/weight/dimensions);
    # EstimateItem rows are the fallback for estimates saved before that
    # was captured.
    if meta.get("line_items"):
        raw_items = meta.get("line_items", [])
        line_items = [{
            "desc": li.get("description", ""),
            "qty": li.get("qty", 0) or 0,
            "rate": li.get("rate", 0) or 0,
            "amount": li.get("amount", (li.get("qty", 0) or 0) * (li.get("rate", 0) or 0)),
        } for li in raw_items]
    else:
        line_items = [{
            "desc": item.description or "",
            "qty": item.qty or 0,
            "rate": item.rate or 0,
            "amount": (item.qty or 0) * (item.rate or 0),
        } for item in est.items]

    # Unified packages view (description + qty + weight + dimensions + rate) —
    # mirrors the Packages table on estimate_form.html so the view page shows
    # exactly what was quoted, not just a bare line-item list.
    if meta.get("line_items"):
        packages = []
        for li in meta.get("line_items", []):
            qty = li.get("qty", 0) or 0
            wt = li.get("weight", 0) or 0
            l = li.get("length", 0) or 0
            w = li.get("width", 0) or 0
            h = li.get("height", 0) or 0
            vol_wt = (l * w * h / 5000) * qty if (l and w and h) else 0
            act_wt = wt * qty
            packages.append({
                "desc": li.get("description", ""),
                "qty": qty,
                "weight": wt,
                "length": l,
                "width": w,
                "height": h,
                "chg_weight": max(act_wt, vol_wt),
                "rate": li.get("rate", 0) or 0,
                "amount": li.get("amount", qty * (li.get("rate", 0) or 0)),
            })
    else:
        packages = [{
            "desc": p["desc"], "qty": p["qty"], "weight": 0, "length": 0, "width": 0,
            "height": 0, "chg_weight": 0, "rate": p["rate"], "amount": p["amount"],
        } for p in line_items]

    company = Company.query.filter_by(company_id=company_id).first()

    estimate_ctx = {
        "id": est.estimate_id,
        "date": est.date.strftime("%d %b %Y") if est.date else "—",
        "valid_until": est.valid_until.strftime("%d %b %Y") if est.valid_until else "—",
        "status": est.status or "Draft",
        "client_name": est.client_obj.name if est.client_obj else (est.contact_person or "—"),
        "contact_person": est.contact_person or "",
        "phone": est.phone or "",
        "shipper_address": meta.get("shipper_address", ""),
        "destination": meta.get("destination", ""),
        "mode": meta.get("mode", ""),
        "shipment_type": meta.get("mode", "") or "Courier",
        "carrier": meta.get("carrier", ""),
        "reference": meta.get("reference", ""),
        # An estimate has no real AWB — that's only assigned on conversion
        # (see /estimate/convert). Once converted, the docket lives on the
        # Invoice record and shows on its own booking view, not here.
        "docket_no": "",
        "receiver_name": meta.get("receiver_name", ""),
        "receiver_company": meta.get("receiver_company", ""),
        "receiver_phone": meta.get("receiver_phone", ""),
        "receiver_address": meta.get("receiver_address", ""),
        "weight": meta.get("weight") or meta.get("freight_weight") or 0,
        "dimensions": meta.get("dimensions", []),
        "line_items": line_items,
        "packages": packages,
        "subtotal": est.subtotal or 0,
        "tax_amount": est.tax_amount or 0,
        "grand_total": est.grand_total or 0,
        "email": est.email or "",
        "converted_to": meta.get("converted_to_invoice", ""),
    }

    return render_template(
        "estimate_view.html",
        estimate=estimate_ctx,
        company_name=company.company_name if company else None,
    )


@app.route("/estimate/edit/<estimate_id>")
@login_required
@require_permission("estimates", "edit")
def estimate_edit(estimate_id):
    """Edit an existing estimate"""
    return redirect(url_for("estimate_new", edit=estimate_id))


@app.route("/estimate/update", methods=["POST"])
@login_required
@require_permission("estimates", "edit")
def estimate_update():
    """Update an existing estimate - redirects to estimate_new for processing"""
    estimate_id = request.form.get("edit_estimate_id")
    if estimate_id:
        return redirect(url_for("estimate_new", edit=estimate_id))
    flash("No estimate specified to update.")
    return redirect(url_for("estimate_list"))


@app.route("/estimate/convert/<estimate_id>", methods=["POST"])
@login_required
@require_permission("estimates", "create")
def estimate_convert_to_booking(estimate_id):
    """
    Convert an estimate to a full booking invoice.
    Creates a new booking with a fresh AWB, copies all data, and marks the estimate as converted.
    """
    cdb = get_cdb()
    company_id = get_current_company()
    
    # Find the estimate
    est = cdb.query(Estimate).filter_by(estimate_id=estimate_id, company_id=company_id).first()
    if not est:
        flash("Estimate not found.", "error")
        return redirect(url_for("estimate_list"))
    
    if est.status == "Paid" or est.status == "Void":
        flash("This estimate has already been paid or voided and cannot be converted.", "error")
        return redirect(url_for("estimate_view", estimate_id=estimate_id))
    
    try:
        # Parse terms
        meta = {}
        if est.terms:
            try:
                meta = json.loads(est.terms)
            except:
                pass
        
        # Get booking type
        booking_type = meta.get("booking_type", "credit")
        
        # Get shipper name - for cash bookings this comes from the manual entry
        shipper_name = meta.get("shipper_name", est.contact_person or "")
        
        # Get client - only for credit bookings
        client_id = est.client_id
        if booking_type == "cash":
            # For cash bookings, we don't use a client_id
            client_id = None
        
        # Get line items from estimate
        line_items = []
        for item in est.items:
            line_items.append({
                "description": item.description or "",
                "qty": item.qty or 0,
                "rate": item.rate or 0,
            })
        
        # Generate new AWB
        docket_no = _next_awb_number(company_id)
        
        # Generate new invoice ID
        invoice_id = _next_numbered_id(cdb, Invoice.invoice_id, "", extra_filters=[Invoice.company_id == company_id])
        
        # Get client (for credit bookings)
        client = None
        if client_id:
            client = cdb.query(Client).filter_by(id=client_id, company_id=company_id).first()
        
        # ============================================================
        # Extract ALL fields from the estimate meta
        # ============================================================
        
        # Shipper address fields
        shipper_contact_name = meta.get("contact_person", est.contact_person or "")
        customer_phone = meta.get("customer_phone", est.phone or "")
        
        shipper_address1 = meta.get("shipper_address1", "")
        shipper_address2 = meta.get("shipper_address2", "")
        shipper_city = meta.get("shipper_city", "")
        shipper_state = meta.get("shipper_state", "")
        shipper_pincode = meta.get("shipper_pincode", "")
        shipper_country = meta.get("shipper_country", "India")
        
        # Receiver address fields
        receiver_name = meta.get("receiver_name", "")
        receiver_company = meta.get("receiver_company", "")
        receiver_phone = meta.get("receiver_phone", "")
        receiver_address1 = meta.get("receiver_address1", "")
        receiver_address2 = meta.get("receiver_address2", "")
        receiver_city = meta.get("receiver_city", "")
        receiver_state = meta.get("receiver_state", "")
        receiver_pincode = meta.get("receiver_pincode", "")
        receiver_country = meta.get("receiver_country", "India")
        
        # Service details
        destination = meta.get("destination", "")
        mode = meta.get("mode", "")
        carrier = meta.get("carrier", "")
        courier_company_id = meta.get("courier_company_id", "")
        vendor = meta.get("vendor", "")
        reference = meta.get("reference", "")
        pickup_date = meta.get("pickup_date", str(today_ist()))
        expected_delivery = meta.get("expected_delivery", "")
        tracking_number = meta.get("tracking_number", "")
        
        # Charges
        freight_weight = meta.get("freight_weight", 0) or sum(p.get("weight", 0) * p.get("qty", 1) for p in meta.get("packages", []))
        freight_rate = meta.get("freight_rate", 0)
        freight = meta.get("freight_amount", 0) or (freight_weight * freight_rate)
        other = meta.get("other_charges", 0)
        discount = meta.get("discount", 0)
        other_reason = meta.get("other_reason", "")
        base = freight + other - discount
        
        # Get GST status
        co = Company.query.filter_by(company_id=company_id).first()
        apply_gst = co.is_gst_registered if (co and hasattr(co, 'is_gst_registered')) else True
        
        # GST calculation
        gst_calc = compute_invoice_gst(base, apply_gst, shipper_state, receiver_state)
        gst = gst_calc["gst_total"]
        grand_total = gst_calc["grand_total"]
        
        # ============================================================
        # Build packages data
        # ============================================================
        packages_data = []
        # First try to get packages from meta (estimate form packages)
        if meta.get("packages"):
            packages_data = meta.get("packages", [])
        else:
            # Fall back to building from line items
            for item in est.items:
                # Try to find dimensions in meta line_items
                item_weight = 0
                item_length = 0
                item_width = 0
                item_height = 0
                for li in meta.get("line_items", []):
                    if li.get("description") == item.description:
                        item_weight = li.get("weight", 0) or 0
                        item_length = li.get("length", 0) or 0
                        item_width = li.get("width", 0) or 0
                        item_height = li.get("height", 0) or 0
                        break
                
                packages_data.append({
                    "name": item.description or "Package",
                    "type": "Box",
                    "unit": "cm",
                    "qty": item.qty or 1,
                    "length": item_length,
                    "width": item_width,
                    "height": item_height,
                    "weight": item_weight,
                    "division": 5000,
                    "discount": 0,
                    "discount_wt": 0,
                    "vol_weight": 0,
                    "chg_weight": item_weight,
                    "rate": item.rate or 0,
                })
        
        # ============================================================
        # Build complete shipment metadata
        # ============================================================
        shipment_meta = json.dumps({
            "docket_no": docket_no,
            "booking_type": booking_type,
            "shipper_name": shipper_name,
            "shipper_contact_name": shipper_contact_name,
            "customer_phone": customer_phone,
            "shipper_address1": shipper_address1,
            "shipper_address2": shipper_address2,
            "shipper_city": shipper_city,
            "shipper_state": shipper_state,
            "shipper_pincode": shipper_pincode,
            "shipper_country": shipper_country,
            "receiver_name": receiver_name,
            "receiver_company": receiver_company,
            "receiver_phone": receiver_phone,
            "receiver_address1": receiver_address1,
            "receiver_address2": receiver_address2,
            "receiver_city": receiver_city,
            "receiver_state": receiver_state,
            "receiver_pincode": receiver_pincode,
            "receiver_country": receiver_country,
            "destination": destination,
            "shipment_type": "DOC",
            "mode": mode,
            "carrier": carrier,
            "courier_company_id": courier_company_id,
            "carrier_ref": reference,
            "vendor": vendor,
            "origin": "India",
            "pickup_date": pickup_date,
            "expected_delivery": expected_delivery,
            "tracking_number": tracking_number,
            "payment_mode": "credit",
            "freight": freight,
            "freight_weight": freight_weight,
            "freight_rate_per_kg": freight_rate,
            "freight_billing_weight": freight_weight,
            "fuel": 0,
            "other": other,
            "discount": discount,
            "other_charges_reason": other_reason,
            "gst": gst,
            "cgst": gst_calc["cgst"],
            "sgst": gst_calc["sgst"],
            "igst": gst_calc["igst"],
            "is_interstate": gst_calc["is_interstate"],
            "round_off": gst_calc["round_off"],
            "amount_paid": 0,
            "packages": packages_data,
            "reference": reference,
            "converted_from_estimate": estimate_id,
            "notes": meta.get("notes", ""),
        })
        
        # Create the invoice
        invoice = Invoice(
            invoice_id=invoice_id,
            company_id=company_id,
            client_id=client_id,  # Will be None for cash bookings
            date=today_ist(),
            status="Pending",
            contact_person=shipper_contact_name or shipper_name,
            phone=customer_phone or est.phone or "",
            subtotal=max(0, base),
            tax_amount=gst,
            grand_total=grand_total,
            terms=shipment_meta,
            email=meta.get("notes", ""),
            paid_amount=0,
            balance=grand_total,
        )
        cdb.add(invoice)
        cdb.flush()
        
        # Create invoice items from estimate items
        for item in est.items:
            inv_item = InvoiceItem(
                invoice_id=invoice.id,
                description=item.description or "",
                qty=item.qty or 0,
                rate=item.rate or 0,
                discount=0,
            )
            cdb.add(inv_item)
        
        # Update client pending balance (only for credit bookings)
        if client_id and booking_type == "credit":
            client_obj = cdb.query(Client).filter_by(id=client_id, company_id=company_id).first()
            if client_obj:
                client_obj.pending = (client_obj.pending or 0) + grand_total
        
        # Mark estimate as converted/paid
        est.status = "Paid"
        # Store the invoice ID in terms for reference
        try:
            est_meta = json.loads(est.terms) if est.terms else {}
            est_meta["converted_to_invoice"] = invoice_id
            est.terms = json.dumps(est_meta)
        except:
            pass
        
        # ── Auto-generate purchase line ──
        # Create a mock form dict for the purchase line helper
        mock_form = {
            "courier_company_id": courier_company_id,
            "carrier": carrier,
            "destination": destination,
            "shipper_name": shipper_name,
            "receiver_name": receiver_name,
            "carrier_ref": reference,
        }
        
        _sync_auto_purchase_invoice_line(
            cdb, company_id, mock_form, packages_data,
            freight_weight, apply_gst, gst_calc,
            str(today_ist()), docket_no, invoice_id, invoice.id, "final"
        )
        
        # ── Create manifest entry ──
        total_boxes = sum(p.get("qty", 1) for p in packages_data) or 1
        primary_stock_name = packages_data[0].get("name") if packages_data else None
        _sync_auto_manifest_entry(
            cdb, company_id,
            shipper_name,
            carrier, "final",
            str(today_ist()), docket_no, invoice_id, total_boxes,
            primary_stock_id=None,
            primary_stock_name=primary_stock_name,
            booking_type=booking_type
        )
        
        cdb.commit()
        
        # Send WhatsApp notification
        try:
            from tasks import send_invoice_generate_notification_async
            send_invoice_generate_notification_async(company_id=company_id, invoice_id=invoice_id)
        except Exception as e:
            print(f"[whatsapp] could not queue notification for {invoice_id}: {e}")
        
        flash(f"✅ Estimate {estimate_id} converted to booking {invoice_id} (AWB: {docket_no})!", "success")
        return redirect(url_for("invoice_view", invoice_id=invoice_id))
        
    except Exception as e:
        cdb.rollback()
        print(f"[estimate-convert] ERROR: {e}")
        import traceback
        traceback.print_exc()
        flash(f"Error converting estimate: {str(e)}", "error")
        return redirect(url_for("estimate_view", estimate_id=estimate_id))


@login_required
@owner_required
def estimate_delete(estimate_id):
    """Delete an estimate (soft delete - mark as Void)"""
    cdb = get_cdb()
    company_id = get_current_company()
    
    est = cdb.query(Estimate).filter_by(estimate_id=estimate_id, company_id=company_id).first()
    if not est:
        flash("Estimate not found.", "error")
        return redirect(url_for("estimate_list"))
    
    est.status = "Void"
    cdb.commit()
    
    flash(f"Estimate {estimate_id} has been voided.", "success")
    return redirect(url_for("estimate_list"))

# ── Manifest List ─────────────────────────────────────────────────────────────
@app.route('/manifest/list')
@login_required
@require_permission("manifest", "view")
def manifest_list():
    company_id = get_current_company()
    if not company_id:
        return redirect(url_for('login'))
    cdb = get_customer_session(company_id)

    from_date   = request.args.get('from_date')
    to_date     = request.args.get('to_date')
    shipper_id  = request.args.get('shipper_id')
    courier     = request.args.get('courier', '').strip()
    supplier_id = request.args.get('supplier_id', '').strip()
    status_filter = request.args.get('status', '').strip()

    q = cdb.query(CompanyManifest).filter_by(company_id=company_id)

    if from_date:
        try:
            q = q.filter(CompanyManifest.date >= date.fromisoformat(from_date))
        except ValueError:
            pass
    if to_date:
        try:
            q = q.filter(CompanyManifest.date <= date.fromisoformat(to_date))
        except ValueError:
            pass
    if shipper_id:
        q = q.filter(CompanyManifest.shipper_client_id == int(shipper_id))

    # Status, courier, and supplier all filter at the ENTRY level, not the
    # manifest level. CompanyManifest.status is an aggregate field and
    # doesn't tell you which individual entries are Generated/Pending, so
    # filtering on it (or filtering courier/supplier separately) let a
    # manifest qualify without its qualifying entry being the one shown.
    # One combined exists() means: this manifest has at least one entry
    # that matches every active filter at once — same entry, not just
    # "some entry matches status" and "some other entry matches courier".
    from sqlalchemy import exists, and_, or_, func

    entry_conditions = [ManifestEntry.manifest_id == CompanyManifest.id]

    if status_filter == 'generated':
        entry_conditions.append(ManifestEntry.status == 'Generated')
    elif status_filter == 'pending':
        entry_conditions.append(ManifestEntry.status != 'Generated')

    selected_supplier_name = None
    match_names_lower = None
    if supplier_id:
        # Matches on supplier name itself PLUS any registered brand names.
        # Exact (case-insensitive) match — NOT a substring match. Using
        # ilike with '%name%' wildcards meant selecting "DHL" also pulled
        # in every manifest whose courier_name merely *contained* "DHL"
        # as a fragment (e.g. "DHL Express", "DHL International").
        sup = cdb.query(Supplier).filter_by(id=int(supplier_id), company_id=company_id).first()
        if sup:
            selected_supplier_name = sup.name
            match_names = [sup.name] + [b.brand_name for b in sup.brands]
            match_names_lower = set(n.strip().lower() for n in match_names if n)
            entry_conditions.append(
                or_(*[func.lower(func.trim(ManifestEntry.courier_name)) == n.strip().lower() for n in match_names if n])
            )
    elif courier:
        entry_conditions.append(ManifestEntry.courier_name.ilike(f'%{courier}%'))

    if status_filter or supplier_id or courier:
        q = q.filter(exists().where(and_(*entry_conditions)).correlate(CompanyManifest))

    manifests = q.order_by(CompanyManifest.date.desc(), CompanyManifest.id.desc()).all()

    # Which manifests QUALIFY (filtered above) is not the same as which
    # ENTRIES on those manifests match the filter — a manifest can have
    # entries from several couriers/statuses. Build a per-manifest filtered
    # entry list so the template only renders entries matching ALL active
    # filters, not every entry in the manifest group.
    def _entry_matches(e):
        ok = True
        if status_filter == 'generated':
            ok = ok and e.status == 'Generated'
        elif status_filter == 'pending':
            ok = ok and e.status != 'Generated'
        if match_names_lower is not None:
            ok = ok and (e.courier_name.strip().lower() in match_names_lower)
        elif courier:
            ok = ok and (courier.lower() in e.courier_name.strip().lower())
        return ok

    if status_filter or supplier_id or courier:
        entries_by_manifest = {
            m.id: [e for e in m.entries if _entry_matches(e)] for m in manifests
        }
    else:
        entries_by_manifest = {m.id: m.entries for m in manifests}

    # Carrier ref (and other booking-side shipment data) lives on the
    # customer invoice, not on ManifestEntry itself — same lookup used on
    # the print pages, keyed off each entry's docket_no.
    shipment_data = {}
    for m in manifests:
        for entry in entries_by_manifest[m.id]:
            shipment_data[entry.id] = _manifest_entry_shipment_data(cdb, company_id, entry.docket_no)

    clients      = cdb.query(Client).filter_by(company_id=company_id, status='Active').filter(Client.client_type != 'Cash-Only').order_by(Client.name).all()
    total_boxes  = sum(sum(e.boxes for e in entries_by_manifest[m.id]) for m in manifests)
    courier_set  = set()
    for m in manifests:
        for e in entries_by_manifest[m.id]:
            courier_set.add(e.courier_name.strip().lower())
    unique_couriers = len(courier_set)

    # NEW — all suppliers show in filter, not just ones with brand rows
    suppliers_with_brands = (
        cdb.query(Supplier)
        .filter_by(company_id=company_id)
        .order_by(Supplier.name)
        .all()
    )

    # Map courier_name -> parent supplier name.
    # Includes both registered SupplierBrand entries AND the supplier's own name.
    brand_to_supplier = {}
    for sup in suppliers_with_brands:
        # The supplier name itself matches (e.g. "Blue Dart Aviation" in ManifestEntry.courier_name)
        brand_to_supplier[sup.name.strip().lower()] = sup.name
        for b in sup.brands:
            brand_to_supplier[b.brand_name.strip().lower()] = sup.name

    from collections import defaultdict
    grouped_manifests = defaultdict(list)
    for m in manifests:
        grouped_manifests[str(m.date)].append(m)
    date_keys = list(grouped_manifests.keys())

    return render_template(
        'manifest_list.html',
        manifests=manifests,
        entries_by_manifest=entries_by_manifest,
        shipment_data=shipment_data,
        clients=clients,
        from_date=from_date,
        to_date=to_date,
        shipper_id=shipper_id,
        courier=courier,
        supplier_id=supplier_id,
        selected_supplier_name=selected_supplier_name,
        suppliers_with_brands=suppliers_with_brands,
        brand_to_supplier=brand_to_supplier,
        total_manifests=len(manifests),
        total_boxes=total_boxes,
        unique_couriers=unique_couriers,
        grouped_manifests=grouped_manifests,
        date_keys=date_keys,
    )


# ── Manifest Create Form ───────────────────────────────────────────────────────
@app.route('/manifest/create')
@login_required
@require_permission("manifest", "view")
def manifest_create():
    company_id = get_current_company()
    if not company_id:
        return redirect(url_for('login'))
    cdb = get_customer_session(company_id)

    clients     = cdb.query(Client).filter_by(company_id=company_id, status='Active').filter(Client.client_type != 'Cash-Only').order_by(Client.name).all()
    stock_items = cdb.query(StockItem).filter_by(company_id=company_id).order_by(StockItem.name).all()

    # Generate next manifest ID
    last = cdb.query(CompanyManifest).filter_by(company_id=company_id)\
               .order_by(CompanyManifest.id.desc()).first()
    next_num   = (last.id + 1) if last else 1
    manifest_id = f"MFT-{next_num:04d}"

    return render_template(
        'manifest_form.html',
        edit_mode=False,
        manifest_id=manifest_id,
        clients=clients,
        stock_items=stock_items,
        today=today_ist().isoformat(),
    )

@app.route('/manifest/shipper-dockets/<int:client_id>')
@login_required
@require_permission("manifest", "view")
def shipper_last_dockets(client_id):
    company_id = get_current_company()
    if not company_id:
        return jsonify([])
    cdb = get_customer_session(company_id)

    invoices = (
        cdb.query(Invoice)
        .filter_by(company_id=company_id, client_id=client_id)
        .filter(Invoice.invoice_id.like('CUST-%'))
        .order_by(Invoice.id.desc())
        .all()
    )

    result = []
    seen = set()

    for inv in invoices:
        try:
            meta = json.loads(inv.terms) if inv.terms else {}
        except Exception:
            meta = {}

        docket = meta.get('docket_no', '').strip()
        if not docket or docket in seen:
            continue
        seen.add(docket)

        stock_items = []

        # ── Path 1: invoice_items has stock_item_id linked (ideal) ──
        linked_items = [line for line in inv.items if line.stock_item_id]
        if linked_items:
            for line in linked_items:
                stock = cdb.query(StockItem).filter_by(id=line.stock_item_id).first()
                if not stock:
                    continue
                already_used = (
                    cdb.query(func.sum(ManifestEntry.boxes))
                    .join(CompanyManifest, ManifestEntry.manifest_id == CompanyManifest.id)
                    .filter(
                        CompanyManifest.company_id == company_id,
                        ManifestEntry.docket_no == docket,
                        ManifestEntry.stock_item_id == stock.id
                    )
                    .scalar() or 0
                )
                available = max(0, int(line.qty) - int(already_used))
                stock_items.append({
                    'id':       stock.id,
                    'name':     stock.name,
                    'code':     stock.code or '',
                    'quantity': available,
                    'unit':     stock.unit or 'pcs'
                })

        # ── Path 2: no invoice_items — read packages from terms JSON ──
        else:
            packages = meta.get('packages', [])
            for pkg in packages:
                # packages may use 'name', 'type', or both
                pkg_name = (pkg.get('name') or pkg.get('type') or '').strip()
                pkg_qty  = float(pkg.get('qty') or 1)
                if not pkg_name:
                    continue

                # Match stock by exact name first, then partial
                stock = (
                    cdb.query(StockItem)
                    .filter(StockItem.company_id == company_id,
                            StockItem.name == pkg_name)
                    .first()
                ) or (
                    cdb.query(StockItem)
                    .filter(StockItem.company_id == company_id,
                            StockItem.name.ilike(f'%{pkg_name}%'))
                    .first()
                )

                if not stock:
                    continue

                # Avoid duplicates — sum qty if same stock appears twice
                existing = next((s for s in stock_items if s['id'] == stock.id), None)
                if existing:
                    existing['quantity'] += pkg_qty
                    continue

                already_used = (
                    cdb.query(func.sum(ManifestEntry.boxes))
                    .join(CompanyManifest, ManifestEntry.manifest_id == CompanyManifest.id)
                    .filter(
                        CompanyManifest.company_id == company_id,
                        ManifestEntry.docket_no == docket,
                        ManifestEntry.stock_item_id == stock.id
                    )
                    .scalar() or 0
                )
                available = max(0, int(pkg_qty) - int(already_used))
                stock_items.append({
                    'id':       stock.id,
                    'name':     stock.name,
                    'code':     stock.code or '',
                    'quantity': available,
                    'unit':     stock.unit or 'pcs'
                })

        result.append({
            'docket_id':   inv.id,
            'docket_no':   docket,
            'invoice_id':  inv.invoice_id,
            'date':        inv.date.strftime('%d %b %Y') if inv.date else '',
            'stock_items': stock_items
        })

    return jsonify(result)
    
@app.route('/manifest/invoice-packages/<int:client_id>/<docket_no>')
@login_required
@require_permission("manifest", "view")
def invoice_packages(client_id, docket_no):
    company_id = get_current_company()
    if not company_id:
        return jsonify({})
    cdb = get_customer_session(company_id)

    invoices = (
        cdb.query(Invoice)
        .filter_by(company_id=company_id, client_id=client_id)
        .all()
    )
    for inv in invoices:
        try:
            meta = json.loads(inv.terms) if inv.terms else {}
        except Exception:
            meta = {}
        if meta.get('docket_no', '').strip() == docket_no.strip():
            packages = meta.get('packages', [])
            # Aggregate by type
            summary = {}
            for p in packages:
                t = (p.get('type') or p.get('name') or 'Box').strip()
                q = float(p.get('qty') or 1)
                summary[t] = summary.get(t, 0) + q
            return jsonify({
                'invoice_id': inv.invoice_id,
                'date': inv.date.strftime('%d %b %Y') if inv.date else '',
                'packages': [{'type': k, 'qty': int(v)} for k, v in summary.items()]
            })
    return jsonify({'packages': []})

# ── Expenses ──────────────────────────────────────────────────────────────────
EXPENSE_CATEGORIES = [
    "Rent", "Electricity", "Internet", "Salaries", "Fuel",
    "Office Supplies", "Maintenance", "Travel", "Food & Refreshments",
    "Marketing", "Courier Charges", "Bank Charges", "Misc", "Others",
]

@app.route("/expenses")
@login_required
@require_permission("expenses", "view")
def expenses():
    company_id = get_current_company()
    if not company_id:
        return redirect(url_for('login'))
    cdb = get_customer_session(company_id)

    from_date = request.args.get("from_date", today_ist().replace(day=1).isoformat())
    to_date   = request.args.get("to_date",   today_ist().isoformat())

    try:
        fd = date.fromisoformat(from_date)
        td = date.fromisoformat(to_date)
    except ValueError:
        fd = today_ist().replace(day=1)
        td = today_ist()

    rows = (
        cdb.query(Expense)
        .filter(
            Expense.company_id == company_id,
            Expense.date >= fd,
            Expense.date <= td,
        )
        .order_by(Expense.date.desc(), Expense.id.desc())
        .all()
    )

    total = sum(e.amount for e in rows)

    # Category breakdown for chart
    cat_totals = {}
    for e in rows:
        cat_totals[e.category] = cat_totals.get(e.category, 0) + e.amount

    return render_template(
        "expenses.html",
        expenses=rows,
        total=total,
        cat_totals=cat_totals,
        categories=EXPENSE_CATEGORIES,
        from_date=from_date,
        to_date=to_date,
        today_str=today_ist().isoformat(),
        active="expenses",
    )


@app.route("/expenses/add", methods=["GET", "POST"])
@login_required
@require_permission("expenses", "view", method_actions={'POST': 'create'})
def add_expense():
    company_id = get_current_company()
    if not company_id:
        return redirect(url_for('login'))
    cdb = get_customer_session(company_id)
    user = get_current_user()

    if request.method == "POST":
        try:
            expense_date = date.fromisoformat(request.form.get("date", today_ist().isoformat()))
            category = request.form.get("category", "Misc")
            description = request.form.get("description", "").strip()
            amount = float(request.form.get("amount", 0))
            payment_mode = request.form.get("payment_mode", "Cash")
            reference = request.form.get("reference", "").strip()
            
            if amount <= 0:
                flash("Amount must be greater than 0.", "error")
                return redirect(url_for("expenses"))

            # ── 1. CREATE EXPENSE RECORD ──
            exp = Expense(
                company_id=company_id,
                date=expense_date,
                category=category,
                description=description,
                amount=amount,
                payment_mode=payment_mode,
                reference=reference,
                created_by=user.get("full_name", user.get("email")),
            )
            cdb.add(exp)
            cdb.flush()  # Get expense ID

            # ── 2. DEDUCT FROM CASH IN HAND OR BANK ACCOUNT ──
            if payment_mode.lower() == "cash":
                # Deduct from Cash in Hand
                cash_txn = CashTransaction(
                    company_id=company_id,
                    type="expense",
                    date=expense_date,
                    category=category,
                    description=f"Expense: {description}",
                    amount=amount,
                    reference=reference or f"EXP-{exp.id}",
                    notes=f"Expense recorded - {category}",
                    party_name="Expense",
                    created_by=user.get("full_name", user.get("email"))
                )
                cdb.add(cash_txn)
                
            elif payment_mode.lower() in ["bank transfer", "online", "upi", "cheque"]:
                # Deduct from Bank Account
                if not reference:
                    flash("Reference/Transaction ID is required for bank payments.", "error")
                    cdb.rollback()
                    return redirect(url_for("expenses"))
                
                # Find a bank account to deduct from
                # Try to find the specific bank account from reference
                bank_account = None
                
                # First, try to find a bank account matching the reference or bank name
                if reference:
                    # Look for bank account by reference or bank name
                    bank_accounts = cdb.query(BankAccount).filter_by(company_id=company_id, status='Active').all()
                    for acc in bank_accounts:
                        if reference.lower() in acc.account_number.lower() or reference.lower() in acc.bank_name.lower():
                            bank_account = acc
                            break
                
                # If no specific bank found, use the first active bank account
                if not bank_account:
                    bank_account = cdb.query(BankAccount).filter_by(company_id=company_id, status='Active').first()
                    if not bank_account:
                        flash("No active bank account found. Please add a bank account first.", "error")
                        cdb.rollback()
                        return redirect(url_for("expenses"))
                
                # Check if sufficient balance
                if bank_account.balance < amount:
                    flash(f"Insufficient balance in {bank_account.bank_name} - {bank_account.account_name}. Available: ₹{bank_account.balance:,.2f}", "error")
                    cdb.rollback()
                    return redirect(url_for("expenses"))
                
                # Deduct from bank account
                bank_txn = BankTransaction(
                    bank_account_id=bank_account.id,
                    company_id=company_id,
                    type="debit",
                    date=expense_date,
                    description=f"Expense: {description}",
                    amount=amount,
                    reference=reference or f"EXP-{exp.id}",
                    transaction_mode=payment_mode.title(),
                    notes=f"Expense recorded - {category}",
                    party_name="Expense",
                    created_by=user.get("full_name", user.get("email"))
                )
                cdb.add(bank_txn)
                bank_account.balance -= amount
                bank_account.updated_at = datetime.utcnow()
                
            else:
                # Unknown payment mode - treat as cash
                cash_txn = CashTransaction(
                    company_id=company_id,
                    type="expense",
                    date=expense_date,
                    category=category,
                    description=f"Expense: {description}",
                    amount=amount,
                    reference=reference or f"EXP-{exp.id}",
                    notes=f"Expense recorded - {category} (via {payment_mode})",
                    party_name="Expense",
                    created_by=user.get("full_name", user.get("email"))
                )
                cdb.add(cash_txn)

            cdb.commit()
            flash(f"✅ Expense of ₹{amount:,.2f} recorded successfully and deducted from {'Cash' if payment_mode.lower() == 'cash' else bank_account.bank_name if bank_account else 'Bank'}.", "success")

        except Exception as e:
            cdb.rollback()
            flash(f"Error recording expense: {str(e)}", "error")
            print(f"[expense-error] {e}")
            import traceback
            traceback.print_exc()

        return redirect(url_for("expenses"))

    return redirect(url_for("expenses"))


@app.route("/expenses/delete/<int:expense_id>", methods=["POST"])
@login_required
@owner_required
def delete_expense(expense_id):
    company_id = get_current_company()
    if not company_id:
        return redirect(url_for('login'))
    cdb = get_customer_session(company_id)
    
    exp = cdb.query(Expense).filter_by(id=expense_id, company_id=company_id).first()
    if not exp:
        flash("Expense not found.", "error")
        return redirect(url_for("expenses"))
    
    try:
        payment_mode = exp.payment_mode.lower() if exp.payment_mode else "cash"
        amount = exp.amount
        expense_date = exp.date
        
        # ── REVERSE THE DEDUCTION ──
        if payment_mode == "cash":
            # Reverse cash deduction - add back to Cash in Hand
            cash_txn = CashTransaction(
                company_id=company_id,
                type="income",  # Reverse: add back
                date=expense_date,
                category="Expense Reversal",
                description=f"Reversal: {exp.description}",
                amount=amount,
                reference=f"REV-EXP-{exp.id}",
                notes=f"Expense deletion reversal - {exp.category}",
                party_name="Expense",
                created_by=get_current_user().get('email')
            )
            cdb.add(cash_txn)
            
        elif payment_mode in ["bank transfer", "online", "upi", "cheque"]:
            # Reverse bank deduction - find the associated transaction
            bank_txn = cdb.query(BankTransaction).filter(
                BankTransaction.company_id == company_id,
                BankTransaction.type == "debit",
                BankTransaction.reference == str(exp.id)
            ).first()
            
            if bank_txn and bank_txn.bank_account:
                # Reverse the bank transaction
                reverse_txn = BankTransaction(
                    bank_account_id=bank_txn.bank_account_id,
                    company_id=company_id,
                    type="credit",  # Reverse: add back
                    date=expense_date,
                    description=f"Reversal: {exp.description}",
                    amount=amount,
                    reference=f"REV-EXP-{exp.id}",
                    transaction_mode=payment_mode.title(),
                    notes=f"Expense deletion reversal - {exp.category}",
                    party_name="Expense",
                    created_by=get_current_user().get('email')
                )
                cdb.add(reverse_txn)
                bank_txn.bank_account.balance += amount
                bank_txn.bank_account.updated_at = datetime.utcnow()
            else:
                # Fallback: try to find by description
                bank_txn = cdb.query(BankTransaction).filter(
                    BankTransaction.company_id == company_id,
                    BankTransaction.type == "debit",
                    BankTransaction.description.like(f"%{exp.description}%"),
                    BankTransaction.amount == amount
                ).first()
                if bank_txn and bank_txn.bank_account:
                    reverse_txn = BankTransaction(
                        bank_account_id=bank_txn.bank_account_id,
                        company_id=company_id,
                        type="credit",
                        date=expense_date,
                        description=f"Reversal: {exp.description}",
                        amount=amount,
                        reference=f"REV-EXP-{exp.id}",
                        transaction_mode=payment_mode.title(),
                        notes=f"Expense deletion reversal - {exp.category}",
                        party_name="Expense",
                        created_by=get_current_user().get('email')
                    )
                    cdb.add(reverse_txn)
                    bank_txn.bank_account.balance += amount
                    bank_txn.bank_account.updated_at = datetime.utcnow()
                else:
                    # No bank transaction found - add as cash reversal as fallback
                    cash_txn = CashTransaction(
                        company_id=company_id,
                        type="income",
                        date=expense_date,
                        category="Expense Reversal",
                        description=f"Reversal (fallback): {exp.description}",
                        amount=amount,
                        reference=f"REV-EXP-{exp.id}",
                        notes=f"Expense deletion reversal - no bank txn found",
                        party_name="Expense",
                        created_by=get_current_user().get('email')
                    )
                    cdb.add(cash_txn)

        # Delete the expense
        cdb.delete(exp)
        cdb.commit()
        
        flash(f"✅ Expense deleted and cash/bank restored.", "success")
        
    except Exception as e:
        cdb.rollback()
        flash(f"Error deleting expense: {str(e)}", "error")
        print(f"[expense-delete-error] {e}")
        import traceback
        traceback.print_exc()
    
    return redirect(url_for("expenses"))


@app.route("/api/expenses-summary")
@login_required
@require_permission("expenses", "view")
def api_expenses_summary():
    company_id = get_current_company()
    if not company_id:
        return redirect(url_for('login'))
    cdb = get_customer_session(company_id)

    from_date = request.args.get("from_date", today_ist().replace(day=1).isoformat())
    to_date   = request.args.get("to_date",   today_ist().isoformat())

    try:
        fd = date.fromisoformat(from_date)
        td = date.fromisoformat(to_date)
    except ValueError:
        fd = today_ist().replace(day=1)
        td = today_ist()

    rows = cdb.query(Expense).filter(
        Expense.company_id == company_id,
        Expense.date >= fd,
        Expense.date <= td,
    ).all()

    cat_totals = {}
    for e in rows:
        cat_totals[e.category] = cat_totals.get(e.category, 0) + e.amount

    return jsonify({
        "total": sum(e.amount for e in rows),
        "count": len(rows),
        "by_category": cat_totals,
    })


# ── Manifest Save (POST) ───────────────────────────────────────────────────────
@app.route('/manifest/save', methods=['POST'])
@login_required
@require_permission("manifest", "create")
def manifest_save():
    company_id = get_current_company()
    if not company_id:
        return redirect(url_for('login'))
    cdb = get_customer_session(company_id)

    manifest_id       = request.form.get('manifest_id', '').strip()
    manifest_date_s   = request.form.get('manifest_date')
    shipper_client_id = int(request.form.get('shipper_client_id', 0))
    notes             = request.form.get('notes', '').strip()

    courier_names   = request.form.getlist('courier_name[]')
    boxes_list      = request.form.getlist('boxes[]')
    docket_nos      = request.form.getlist('docket_no[]')
    docket_ids      = request.form.getlist('docket_id[]')
    stock_item_ids  = request.form.getlist('stock_item_id[]')
    entry_notes     = request.form.getlist('entry_notes[]')

    # Build valid entries
    entries_data = []
    total_boxes = 0
    for i, cn in enumerate(courier_names):
        cn = cn.strip()
        bx = int(boxes_list[i]) if i < len(boxes_list) and boxes_list[i] else 0
        if cn and bx > 0:
            sid_raw = stock_item_ids[i] if i < len(stock_item_ids) else ''
            entries_data.append({
                'courier_name':   cn,
                'boxes':          bx,
                'docket_no':      docket_nos[i].strip() if i < len(docket_nos) else '',
                'docket_id':      int(docket_ids[i]) if i < len(docket_ids) and docket_ids[i] else None,
                'stock_item_id':  int(sid_raw) if sid_raw else None,
                'notes':          entry_notes[i].strip() if i < len(entry_notes) else '',
            })
            total_boxes += bx

    if not entries_data:
        flash('Add at least one courier row with boxes > 0.', 'danger')
        return redirect(url_for('manifest_create'))

    # Get shipper name
    shipper = cdb.query(Client).filter_by(id=shipper_client_id, company_id=company_id).first()
    if not shipper:
        flash('Shipper not found.', 'danger')
        return redirect(url_for('manifest_create'))

    try:
        manifest_date = date.fromisoformat(manifest_date_s)
    except (ValueError, TypeError):
        manifest_date = today_ist()

    # Create manifest header
    manifest = CompanyManifest(
        manifest_id=manifest_id,
        company_id=company_id,
        date=manifest_date,
        shipper_client_id=shipper_client_id,
        shipper_client_name=shipper.name,
        total_boxes=total_boxes,
        notes=notes or None,
        created_by=session.get('user', {}).get('email', ''),
    )
    cdb.add(manifest)
    cdb.flush()  # get manifest.id

    # NOTE: stock is no longer touched here. Stock now moves OUT only when a
    # Purchase Bill (courier bill against an AWB) is saved — see /purchase/new.
    # Manifest is now a pure record of which courier each AWB/box went to.
    for ed in entries_data:
        stock_name = None
        stock_type = 'Box'
        if ed['stock_item_id']:
            stock = cdb.query(StockItem).filter_by(id=ed['stock_item_id']).first()
            if stock:
                stock_name = stock.name
                stock_type = stock.item_type or stock.category or 'Box'

        # One row PER BOX, not one row for the whole qty. Same manifest_id,
        # same courier/docket/stock on every row from this booking — only
        # `boxes` (always 1 here) and later `status` differ per row. This is
        # what lets you select just 2 of 3 boxes to generate and leave the
        # 3rd untouched.
        for _ in range(ed['boxes']):  # ← Loop per box
            entry = ManifestEntry(
                manifest_id=manifest.id,
                courier_name=ed['courier_name'],
                boxes=1,  # ← Always 1
                docket_no=ed['docket_no'] or None,
                docket_id=ed['docket_id'],
                stock_item_id=ed['stock_item_id'],
                stock_item_name=stock_name,
                notes=ed['notes'] or None,
                item_type=stock_type,
                status='Pending',
            )
            cdb.add(entry)

    cdb.commit()

    # Notify the shipper (sender) that their AWB/docket numbers are booked —
    # fire-and-forget, same pattern as the invoice notification.
    try:
        from tasks import send_manifest_notification_async
        send_manifest_notification_async(company_id=company_id, manifest_db_id=manifest.id)
    except Exception as e:
        print(f"[whatsapp] could not queue manifest notification for {manifest_id}: {e}")

    flash(f'Manifest {manifest_id} saved.', 'success')
    return redirect(url_for('manifest_list'))


# ── Manifest View ──────────────────────────────────────────────────────────────
@app.route('/manifest/view/<int:manifest_db_id>')
@login_required
@require_permission("manifest", "view")
def manifest_view(manifest_db_id):
    company_id = get_current_company()
    if not company_id:
        return redirect(url_for('login'))
    cdb = get_customer_session(company_id)

    manifest = cdb.query(CompanyManifest).filter_by(
        id=manifest_db_id, company_id=company_id
    ).first()
    if not manifest:
        flash('Manifest not found.', 'danger')
        return redirect(url_for('manifest_list'))

    return render_template('manifest_view.html', manifest=manifest)


# ── Manifest Print (courier-handover document layout) ──────────────────────────
@app.route('/manifest/print/<int:manifest_db_id>')
@login_required
@require_permission("manifest", "view")
def manifest_print(manifest_db_id):
    """
    Printable manifest document matching the physical handover-sheet format
    couriers/coloaders expect (FROM/TO/DATE header block + per-AWB table).

    ManifestEntry itself only tracks courier_name/boxes/docket_no. The
    per-shipment charge weight, actual weight, L/B/H, volumetric weight,
    destination and receiver shown below are pulled live from the matching
    customer invoice's terms JSON via _manifest_entry_shipment_data(),
    keyed off each entry's docket_no. Coloader and payment type still have
    nowhere to come from and stay "—".
    """
    company_id = get_current_company()
    if not company_id:
        return redirect(url_for('login'))
    cdb = get_customer_session(company_id)

    manifest = cdb.query(CompanyManifest).filter_by(
        id=manifest_db_id, company_id=company_id
    ).first()
    if not manifest:
        flash('Manifest not found.', 'danger')
        return redirect(url_for('manifest_list'))

    # Registered platform company (Magnustic, Al Hammad, etc.) — this is the
    # actual "FROM" party handing the boxes to the courier, not the shipper
    # client whose stock happened to be in the box.
    reg_company = Company.query.filter_by(company_id=company_id).first()
    from_company_name = reg_company.company_name if reg_company else ''

    # Map courier_name -> parent supplier/company name, same lookup used on
    # the manifest list page, so the "Company" column stays consistent.
    suppliers = cdb.query(Supplier).filter_by(company_id=company_id).all()
    brand_to_supplier = {}
    for sup in suppliers:
        brand_to_supplier[sup.name.strip().lower()] = sup.name
        for b in sup.brands:
            brand_to_supplier[b.brand_name.strip().lower()] = sup.name

    shipment_data = {
        entry.id: _manifest_entry_shipment_data(cdb, company_id, entry.docket_no)
        for entry in manifest.entries
    }
    total_weight = sum(
        (ship.get('charge_weight') or ship.get('actual_weight') or 0)
        for ship in shipment_data.values() if ship
    )

    return render_template(
        'manifest_print.html',
        manifest=manifest,
        brand_to_supplier=brand_to_supplier,
        shipment_data=shipment_data,
        from_company_name=from_company_name,
        total_weight=total_weight,
    )


@app.route('/manifest/print/day/<date_str>')
@login_required
@require_permission("manifest", "view")
def manifest_print_day(date_str):
    """
    Printable manifest document for ALL manifests on a given date, combined
    into a single handover-sheet-style table — same visual format as
    manifest_print.html, but spanning every manifest logged that day instead
    of just one.
    """
    company_id = get_current_company()
    if not company_id:
        return redirect(url_for('login'))
    cdb = get_customer_session(company_id)

    try:
        target_date = date.fromisoformat(date_str)
    except ValueError:
        flash('Invalid date.', 'danger')
        return redirect(url_for('manifest_list'))

    supplier_id = request.args.get('supplier_id', '').strip()
    status_filter = request.args.get('status', '').strip() 

    q = cdb.query(CompanyManifest).filter_by(company_id=company_id, date=target_date)

    # --- FIX: Filter based on entry status, not manifest status ---
    from sqlalchemy import exists, and_

    if status_filter == 'generated':
        # Show ONLY manifests that have at least one generated entry
        # AND only show the generated entries themselves
        q = q.filter(
            exists().where(
                and_(
                    ManifestEntry.manifest_id == CompanyManifest.id,
                    ManifestEntry.status == 'Generated'
                )
            ).correlate(CompanyManifest)
        )
    elif status_filter == 'pending':
        # Show manifests with NO generated entries
        q = q.filter(
            ~exists().where(
                and_(
                    ManifestEntry.manifest_id == CompanyManifest.id,
                    ManifestEntry.status == 'Generated'
                )
            ).correlate(CompanyManifest)
        )

    # match_names_lower stays None when no supplier filter is active, so the
    # entries_by_manifest comprehension below can use one condition either way.
    match_names_lower = None
    if supplier_id:
        sup = cdb.query(Supplier).filter_by(id=int(supplier_id), company_id=company_id).first()
        if sup:
            match_names = [sup.name] + [b.brand_name for b in sup.brands]
            match_names_lower = set(n.strip().lower() for n in match_names if n)
            from sqlalchemy import or_
            q = q.join(ManifestEntry).filter(
                or_(*[func.lower(func.trim(ManifestEntry.courier_name)) == n for n in match_names_lower])
            )

    manifests = q.order_by(CompanyManifest.id.asc()).all()

    # --- Filter out entries that are NOT generated when printing ---
    # Do NOT do `manifest.entries = [...]` here. CompanyManifest.entries has
    # cascade="all, delete-orphan" — reassigning it marks the excluded
    # (Pending) rows as orphans, and the very next query on this session
    # (the Supplier query a few lines down) autoflushes and PERMANENTLY
    # DELETES those Pending entries from the database. Build a separate
    # lookup instead and never touch the real relationship.
    #
    # The supplier_id filter above joins ManifestEntry to decide which
    # MANIFESTS to include, but a single manifest can hold entries from
    # several couriers (e.g. MFT-0043 had both IM-FEDEX and SKY-SELF rows).
    # Matching one entry pulled the whole manifest in, and this dict used to
    # grab every Generated entry on it regardless of courier. Re-apply the
    # same courier-name match here, per entry, so only the selected
    # supplier's own rows end up on the printed sheet.
    entries_by_manifest = {
        m.id: [
            e for e in m.entries
            if e.status == 'Generated'
            and (match_names_lower is None or (e.courier_name or '').strip().lower() in match_names_lower)
        ]
        for m in manifests
    }
    manifests = [m for m in manifests if entries_by_manifest[m.id]]

    if not manifests:
        flash('No generated manifests found for that date.', 'danger')
        return redirect(url_for('manifest_list'))

    # Same courier -> parent supplier lookup used everywhere else
    suppliers = cdb.query(Supplier).filter_by(company_id=company_id).all()
    brand_to_supplier = {}
    for sup in suppliers:
        brand_to_supplier[sup.name.strip().lower()] = sup.name
        for b in sup.brands:
            brand_to_supplier[b.brand_name.strip().lower()] = sup.name

    # Separate object-keyed lookup so the "TO," handover block can show
    # address/phone — brand_to_supplier above stays name-only since other
    # templates depend on it being a plain string.
    supplier_by_brand = {}
    for sup in suppliers:
        supplier_by_brand[sup.name.strip().lower()] = sup
        for b in sup.brands:
            supplier_by_brand[b.brand_name.strip().lower()] = sup

    # "TO," only makes sense when this sheet is being handed to ONE courier.
    # If every entry on it maps to the same supplier, show that supplier;
    # if it spans more than one (multiple companies selected/combined),
    # leave it blank rather than guess which one it means.
    to_supplier = None
    if supplier_id:
        to_supplier = cdb.query(Supplier).filter_by(id=int(supplier_id), company_id=company_id).first()
    else:
        resolved = None
        ambiguous = False
        for m in manifests:
            for entry in entries_by_manifest[m.id]:
                key = (entry.courier_name or '').strip().lower()
                sup_obj = supplier_by_brand.get(key)
                if sup_obj is None:
                    continue
                if resolved is None:
                    resolved = sup_obj
                elif resolved.id != sup_obj.id:
                    ambiguous = True
                    break
            if ambiguous:
                break
        to_supplier = None if ambiguous else resolved

    # Registered platform company (Magnustic, Al Hammad, etc.) — the "FROM"
    # party on the printed sheet
    reg_company = Company.query.filter_by(company_id=company_id).first()
    from_company_name = reg_company.company_name if reg_company else ''

    # Recalculate total boxes from ONLY generated entries
    total_boxes = sum(sum(e.boxes for e in entries_by_manifest[m.id]) for m in manifests)

    shipment_data = {}
    for m in manifests:
        for entry in entries_by_manifest[m.id]:
            shipment_data[entry.id] = _manifest_entry_shipment_data(cdb, company_id, entry.docket_no)

    total_weight = sum(
        (ship.get('charge_weight') or ship.get('actual_weight') or 0)
        for ship in shipment_data.values() if ship
    )

    return render_template(
        'manifest_print_day.html',
        manifests=manifests,
        entries_by_manifest=entries_by_manifest,
        target_date=target_date,
        total_boxes=total_boxes,
        from_company_name=from_company_name,
        to_supplier=to_supplier,
        total_weight=total_weight,
        shipment_data=shipment_data,
        brand_to_supplier=brand_to_supplier,
    )


@app.route('/manifest/print/selected')
@login_required
@require_permission("manifest", "view")
def manifest_print_selected():
    """
    Same combined handover-sheet format as manifest_print_day.html, but for an
    explicit, arbitrary set of manifest IDs (?ids=12,13,14) instead of every
    manifest on one date.
    """
    company_id = get_current_company()
    if not company_id:
        return redirect(url_for('login'))
    cdb = get_customer_session(company_id)

    ids_param = request.args.get('ids', '').strip()
    try:
        ids = [int(x) for x in ids_param.split(',') if x.strip()]
    except ValueError:
        ids = []
    if not ids:
        flash('No manifests selected to print.', 'danger')
        return redirect(url_for('manifest_list'))

    manifests = cdb.query(CompanyManifest).filter(
        CompanyManifest.id.in_(ids), CompanyManifest.company_id == company_id
    ).order_by(CompanyManifest.id.asc()).all()
    
    if not manifests:
        flash('Manifest not found.', 'danger')
        return redirect(url_for('manifest_list'))

    # --- FILTER: Keep only generated entries ---
    # Same rule as manifest_print_day: never reassign manifest.entries
    # (cascade="all, delete-orphan" would delete the excluded Pending rows
    # for real on the next autoflush). Use a separate lookup instead.
    entries_by_manifest = {
        m.id: [e for e in m.entries if e.status == 'Generated'] for m in manifests
    }
    manifests = [m for m in manifests if entries_by_manifest[m.id]]

    if not manifests:
        flash('No generated entries found in the selected manifests.', 'danger')
        return redirect(url_for('manifest_list'))

    suppliers = cdb.query(Supplier).filter_by(company_id=company_id).all()
    brand_to_supplier = {}
    for sup in suppliers:
        brand_to_supplier[sup.name.strip().lower()] = sup.name
        for b in sup.brands:
            brand_to_supplier[b.brand_name.strip().lower()] = sup.name

    # Same object-keyed lookup and single-supplier-or-blank rule as
    # manifest_print_day: only show "TO," when every selected entry maps to
    # the same courier's supplier record; blank it out if multiple
    # companies are represented in this selection.
    supplier_by_brand = {}
    for sup in suppliers:
        supplier_by_brand[sup.name.strip().lower()] = sup
        for b in sup.brands:
            supplier_by_brand[b.brand_name.strip().lower()] = sup

    to_supplier = None
    resolved = None
    ambiguous = False
    for m in manifests:
        for entry in entries_by_manifest[m.id]:
            key = (entry.courier_name or '').strip().lower()
            sup_obj = supplier_by_brand.get(key)
            if sup_obj is None:
                continue
            if resolved is None:
                resolved = sup_obj
            elif resolved.id != sup_obj.id:
                ambiguous = True
                break
        if ambiguous:
            break
    to_supplier = None if ambiguous else resolved

    reg_company = Company.query.filter_by(company_id=company_id).first()
    from_company_name = reg_company.company_name if reg_company else ''

    total_boxes = sum(sum(e.boxes for e in entries_by_manifest[m.id]) for m in manifests)

    shipment_data = {}
    for m in manifests:
        for entry in entries_by_manifest[m.id]:
            shipment_data[entry.id] = _manifest_entry_shipment_data(cdb, company_id, entry.docket_no)

    total_weight = sum(
        (ship.get('charge_weight') or ship.get('actual_weight') or 0)
        for ship in shipment_data.values() if ship
    )

    return render_template(
        'manifest_print_day.html',
        manifests=manifests,
        entries_by_manifest=entries_by_manifest,
        target_date=None,
        total_boxes=total_boxes,
        from_company_name=from_company_name,
        to_supplier=to_supplier,
        total_weight=total_weight,
        shipment_data=shipment_data,
        brand_to_supplier=brand_to_supplier,
    )


def _recompute_manifest_status(manifest):
    """
    Manifest.status is a DERIVED summary of its rows' individual status —
    it's for filtering/display in manifest_list only, never the source of
    truth for what's been dispatched. That's entry.status, per box row.
    """
    entries = manifest.entries
    if not entries:
        return
    generated_count = sum(1 for e in entries if e.status == 'Generated')
    if generated_count == 0:
        manifest.status = 'Pending'
        manifest.stock_deducted = False
    elif generated_count == len(entries):
        manifest.status = 'Generated'
        manifest.stock_deducted = True
        if not manifest.generated_at:
            manifest.generated_at = datetime.utcnow()
    else:
        manifest.status = 'Partial'
        manifest.stock_deducted = True


@app.route('/manifest/generate/company', methods=['POST'])
@login_required
@require_permission("manifest", "edit")
def manifest_generate_company():
    """Generate all checked entries for a specific company"""
    company_id = get_current_company()
    if not company_id:
        return redirect(url_for('login'))
    cdb = get_customer_session(company_id)
    
    # Get company name and entry IDs
    company_name = request.form.get('company_name', '').strip()
    entry_ids_param = request.form.get('entry_ids', '').strip()
    
    if not company_name:
        flash('Company name is required.', 'danger')
        return redirect(url_for('manifest_list'))
    
    # Parse entry IDs (comma-separated)
    entry_ids = []
    if entry_ids_param:
        for part in entry_ids_param.split(','):
            part = part.strip()
            if part:
                try:
                    entry_ids.append(int(part))
                except ValueError:
                    pass
    
    if not entry_ids:
        flash('Select at least one box to generate for this company.', 'danger')
        return redirect(url_for('manifest_list'))
    
    # Get entries and verify they belong to the right company
    entries = cdb.query(ManifestEntry).join(
        CompanyManifest, ManifestEntry.manifest_id == CompanyManifest.id
    ).filter(
        ManifestEntry.id.in_(entry_ids),
        CompanyManifest.company_id == company_id
    ).all()
    
    if not entries:
        flash('No matching entries found.', 'danger')
        return redirect(url_for('manifest_list'))
    
    # Verify all entries belong to the same company (parent supplier)
    from sqlalchemy import func
    suppliers = cdb.query(Supplier).filter_by(company_id=company_id).all()
    brand_to_supplier = {}
    for sup in suppliers:
        brand_to_supplier[sup.name.strip().lower()] = sup.name
        for b in sup.brands:
            brand_to_supplier[b.brand_name.strip().lower()] = sup.name
    
    target_company = None
    for entry in entries:
        parent = brand_to_supplier.get(entry.courier_name.strip().lower())
        if parent:
            if target_company is None:
                target_company = parent
            elif target_company != parent:
                flash('All selected entries must belong to the same company.', 'danger')
                return redirect(url_for('manifest_list'))
    
    user_email = session.get('user', {}).get('email', '')
    generated_count = 0
    no_stock_link_count = 0
    touched_manifests = {}
    today = today_ist()
    today_manifest_by_shipper = {}

    for entry in entries:
        if entry.status == 'Generated':
            continue
        
        if entry.stock_item_id and entry.boxes:
            stock = cdb.query(StockItem).filter_by(
                id=entry.stock_item_id, company_id=company_id
            ).first()
            if stock:
                stock.quantity = (stock.quantity or 0) - entry.boxes
                stock.last_updated = today_ist()
                cdb.add(StockPurchaseHistory(
                    stock_item_id=stock.id,
                    purchase_invoice_id=None,
                    quantity=-entry.boxes,
                    purchase_rate=0,
                    movement_type="OUT",
                    purchase_date=today_ist(),
                    reference=entry.manifest.manifest_id,
                    awb_no=entry.docket_no,
                ))
        else:
            no_stock_link_count += 1
        
        entry.status = 'Generated'
        entry.generated_at = datetime.utcnow()
        entry.generated_by = user_email
        generated_count += 1

        # ── Move onto TODAY's manifest ────────────────────────────────────
        # A manifest is dated to the booking (invoice_date), so a box booked
        # on the 6th and only generated today (the 12th) would otherwise sit
        # forever on the 6th's manifest. manifest_print_company, and the
        # "today" grouping expected on manifest_list, both key off
        # CompanyManifest.date == today — so a Generated entry left behind
        # on a back-dated manifest is invisible to both. Move it onto (or
        # create) today's manifest for the same shipper — same pattern as
        # the shipper-mismatch move above, but keyed on date instead.
        old_manifest = entry.manifest
        if old_manifest.date != today:
            shipper_id = old_manifest.shipper_client_id
            target_manifest = today_manifest_by_shipper.get(shipper_id)
            if target_manifest is None:
                target_manifest = cdb.query(CompanyManifest).filter_by(
                    company_id=company_id,
                    shipper_client_id=shipper_id,
                    date=today,
                ).first()
            if target_manifest is None:
                last_mf = cdb.query(CompanyManifest).filter_by(company_id=company_id) \
                              .order_by(CompanyManifest.id.desc()).first()
                target_manifest = CompanyManifest(
                    manifest_id=f"MFT-{(last_mf.id + 1) if last_mf else 1:04d}",
                    company_id=company_id,
                    date=today,
                    shipper_client_id=shipper_id,
                    shipper_client_name=old_manifest.shipper_client_name,
                    total_boxes=0,
                    notes=f"Auto-created on generate from {old_manifest.manifest_id}",
                    created_by=user_email,
                )
                cdb.add(target_manifest)
                cdb.flush()
            today_manifest_by_shipper[shipper_id] = target_manifest

            entry.manifest_id = target_manifest.id
            touched_manifests[old_manifest.id] = old_manifest
            touched_manifests[target_manifest.id] = target_manifest
        else:
            touched_manifests[entry.manifest_id] = entry.manifest

    cdb.flush()
    for mid, manifest in touched_manifests.items():
        current_entries = cdb.query(ManifestEntry).filter_by(manifest_id=mid).all()
        if not current_entries:
            cdb.delete(manifest)
            continue
        manifest.total_boxes = len(current_entries)
        _recompute_manifest_status(manifest)
    
    cdb.commit()
    
    if no_stock_link_count:
        flash(f'{no_stock_link_count} box(es) had no linked stock item.', 'info')
    
    flash(f'{generated_count} box(es) generated for {target_company}!', 'success')
    return redirect(url_for('manifest_list'))

@app.route('/manifest/print/company/<company_name>')
@login_required
@require_permission("manifest", "view")
def manifest_print_company(company_name):
    """
    Print all GENERATED entries for a specific company across all manifests for today.
    """
    company_id = get_current_company()
    if not company_id:
        return redirect(url_for('login'))
    cdb = get_customer_session(company_id)
    
    date_param = request.args.get('date', '').strip()
    if date_param:
        try:
            today = date.fromisoformat(date_param)
        except ValueError:
            flash('Invalid date.', 'danger')
            return redirect(url_for('manifest_list'))
    else:
        today = today_ist()
    
    # Find the supplier
    supplier = cdb.query(Supplier).filter(
        Supplier.company_id == company_id,
        func.lower(Supplier.name) == func.lower(company_name)
    ).first()
    
    if not supplier:
        # Try to find by brand
        brand = cdb.query(SupplierBrand).join(Supplier).filter(
            Supplier.company_id == company_id,
            func.lower(SupplierBrand.brand_name) == func.lower(company_name)
        ).first()
        if brand:
            supplier = brand.supplier
    
    if not supplier:
        flash(f'Company "{company_name}" not found.', 'danger')
        return redirect(url_for('manifest_list'))
    
    # Get all manifests for today
    manifests = cdb.query(CompanyManifest).filter_by(
        company_id=company_id,
        date=today
    ).all()
    
    if not manifests:
        flash(f'No manifests found for today ({today.strftime("%d %b %Y")}).', 'danger')
        return redirect(url_for('manifest_list'))
    
    # Build entries_by_manifest with only generated entries for this supplier
    entries_by_manifest = {}
    total_boxes = 0
    for manifest in manifests:
        entries = cdb.query(ManifestEntry).filter(
            ManifestEntry.manifest_id == manifest.id,
            ManifestEntry.status == 'Generated'
        ).all()
        
        # Filter entries that belong to this supplier
        supplier_entries = []
        for entry in entries:
            # Check if this courier belongs to this supplier
            from sqlalchemy import func as _func
            brand_match = cdb.query(SupplierBrand).join(Supplier).filter(
                Supplier.company_id == company_id,
                _func.lower(SupplierBrand.brand_name) == _func.lower(entry.courier_name.strip())
            ).first()
            
            if brand_match and brand_match.supplier_id == supplier.id:
                supplier_entries.append(entry)
            elif _func.lower(entry.courier_name.strip()) == _func.lower(supplier.name):
                supplier_entries.append(entry)
        
        if supplier_entries:
            entries_by_manifest[manifest.id] = supplier_entries
            total_boxes += sum(e.boxes for e in supplier_entries)
    
    manifests = [m for m in manifests if m.id in entries_by_manifest]
    
    if not manifests:
        flash(f'No generated entries found for {company_name} today.', 'danger')
        return redirect(url_for('manifest_list'))
    
    # Get shipment data
    shipment_data = {}
    for m in manifests:
        for entry in entries_by_manifest[m.id]:
            shipment_data[entry.id] = _manifest_entry_shipment_data(cdb, company_id, entry.docket_no)
    
    # Brand to supplier lookup
    suppliers = cdb.query(Supplier).filter_by(company_id=company_id).all()
    brand_to_supplier = {}
    for sup in suppliers:
        brand_to_supplier[sup.name.strip().lower()] = sup.name
        for b in sup.brands:
            brand_to_supplier[b.brand_name.strip().lower()] = sup.name
    
    reg_company = Company.query.filter_by(company_id=company_id).first()
    from_company_name = reg_company.company_name if reg_company else ''
    
    total_weight = sum(
        (ship.get('charge_weight') or ship.get('actual_weight') or 0)
        for ship in shipment_data.values() if ship
    )
    
    return render_template(
        'manifest_print_day.html',
        manifests=manifests,
        entries_by_manifest=entries_by_manifest,
        target_date=today,
        total_boxes=total_boxes,
        from_company_name=from_company_name,
        to_supplier=supplier,
        total_weight=total_weight,
        shipment_data=shipment_data,
        brand_to_supplier=brand_to_supplier,
    )

# ── Manifest Edit Form ─────────────────────────────────────────────────────────
@app.route('/manifest/edit/<int:manifest_db_id>')
@login_required
@require_permission("manifest", "edit")
def manifest_edit(manifest_db_id):
    company_id = get_current_company()
    if not company_id:
        return redirect(url_for('login'))
    cdb = get_customer_session(company_id)

    manifest = cdb.query(CompanyManifest).filter_by(
        id=manifest_db_id, company_id=company_id
    ).first()
    if not manifest:
        flash('Manifest not found.', 'danger')
        return redirect(url_for('manifest_list'))

    if any(e.status == 'Generated' for e in manifest.entries):
        flash('This manifest has boxes already dispatched — it can no longer be edited as a whole. '
              'Generate/dispatch remaining boxes individually from the list, or delete the manifest.', 'danger')
        return redirect(url_for('manifest_list'))

    clients     = cdb.query(Client).filter_by(company_id=company_id, status='Active').filter(Client.client_type != 'Cash-Only').order_by(Client.name).all()
    stock_items = cdb.query(StockItem).filter_by(company_id=company_id).order_by(StockItem.name).all()

    return render_template(
        'manifest_form.html',
        edit_mode=True,
        manifest=manifest,
        manifest_id=manifest.manifest_id,
        clients=clients,
        stock_items=stock_items,
        today=today_ist().isoformat(),
    )


# ── Manifest Update (POST) ─────────────────────────────────────────────────────
@app.route('/manifest/update/<int:manifest_db_id>', methods=['POST'])
@login_required
@require_permission("manifest", "edit")
def manifest_update(manifest_db_id):
    company_id = get_current_company()
    if not company_id:
        return redirect(url_for('login'))
    cdb = get_customer_session(company_id)

    manifest = cdb.query(CompanyManifest).filter_by(
        id=manifest_db_id, company_id=company_id
    ).first()
    if not manifest:
        flash('Manifest not found.', 'danger')
        return redirect(url_for('manifest_list'))

    if any(e.status == 'Generated' for e in manifest.entries):
        flash('This manifest has boxes already dispatched — it can no longer be edited as a whole.', 'danger')
        return redirect(url_for('manifest_list'))

    manifest_date_s  = request.form.get('manifest_date')
    new_stock_item_id= int(request.form.get('stock_item_id', 0))
    notes            = request.form.get('notes', '').strip()

    courier_names = request.form.getlist('courier_name[]')
    boxes_list    = request.form.getlist('boxes[]')
    docket_nos    = request.form.getlist('docket_no[]')
    entry_notes   = request.form.getlist('entry_notes[]')

    entries_data = []
    new_total = 0
    for i, cn in enumerate(courier_names):
        cn = cn.strip()
        bx = int(boxes_list[i]) if i < len(boxes_list) else 0
        if cn and bx > 0:
            entries_data.append({
                'courier_name': cn,
                'boxes': bx,
                'docket_no': docket_nos[i].strip() if i < len(docket_nos) else '',
                'notes': entry_notes[i].strip() if i < len(entry_notes) else '',
            })
            new_total += bx

    old_total = manifest.total_boxes

    # NOTE: stock is no longer adjusted here — see /purchase/new.

    # Update manifest header
    try:
        manifest.date = date.fromisoformat(manifest_date_s)
    except (ValueError, TypeError):
        pass
    manifest.stock_item_id = new_stock_item_id
    manifest.total_boxes   = new_total
    manifest.notes         = notes or None

    # Replace entries — safe here because we already blocked this route above
    # if any row was Generated, so every row being deleted is still Pending.
    for e in list(manifest.entries):
        cdb.delete(e)
    for ed in entries_data:
        for _ in range(ed['boxes']):
            entry = ManifestEntry(
                manifest_id=manifest.id,
                courier_name=ed['courier_name'],
                boxes=1,
                docket_no=ed['docket_no'] or None,
                notes=ed['notes'] or None,
                status='Pending',
            )
            cdb.add(entry)

    cdb.commit()
    flash(f'Manifest updated.', 'success')
    return redirect(url_for('manifest_list'))


# ── Manifest Delete ────────────────────────────────────────────────────────────
@app.route('/manifest/delete/<int:manifest_db_id>')
@login_required
@owner_required
def manifest_delete(manifest_db_id):
    company_id = get_current_company()
    if not company_id:
        return redirect(url_for('login'))
    cdb = get_customer_session(company_id)

    manifest = cdb.query(CompanyManifest).filter_by(
        id=manifest_db_id, company_id=company_id
    ).first()
    if not manifest:
        flash('Manifest not found.', 'danger')
        return redirect(url_for('manifest_list'))

    # NOTE: stock is no longer restored here — manifest doesn't touch stock anymore.

    cdb.delete(manifest)
    cdb.commit()
    flash(f'Manifest {manifest.manifest_id} deleted.', 'success')
    return redirect(url_for('manifest_list'))

# ─────────────────────────────────────────────────────────────────────────────
# ── Super Admin ───────────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────

@app.route("/migrations")
@login_required
@super_admin_required
def migrations():
    """Migration panel — list all past migrations across all company DBs."""
    from platform_models import Company
    from db_router import _engine_cache, _get_or_create

    companies = Company.query.filter_by(is_active=True).all()
    history   = []   # list of dicts for the template

    for company in companies:
        try:
            engine = _engine_cache.get(company.company_id)
            if engine is None:
                _get_or_create(company.company_id)
                engine = _engine_cache[company.company_id]

            _ensure_migration_table(engine)

            with engine.connect() as conn:
                rows = conn.execute(
                    text("SELECT label, status, error_msg, applied_at, applied_by FROM schema_migrations ORDER BY applied_at DESC LIMIT 50")
                ).fetchall()

            for row in rows:
                history.append({
                    "company_id":   company.company_id,
                    "company_name": company.company_name,
                    "label":        row[0],
                    "status":       row[1],
                    "error_msg":    row[2],
                    "applied_at":   row[3],
                    "applied_by":   row[4],
                })
        except Exception as e:
            history.append({
                "company_id":   company.company_id,
                "company_name": company.company_name,
                "label":        "— could not read history —",
                "status":       "error",
                "error_msg":    str(e),
                "applied_at":   None,
                "applied_by":   None,
            })

    # Sort all history by applied_at desc
    history.sort(key=lambda x: x["applied_at"] or datetime.min, reverse=True)

    return render_template(
        "migrations.html",
        active="migrations",
        companies=companies,
        history=history,
    )


    return jsonify({"results": results, "summary": summary})

@app.route("/migrations/run", methods=["POST"])
@login_required
@super_admin_required
def run_migration():
    """
    Execute SQL on customer databases only.
    Platform database changes must be done manually.
    """
    from platform_models import Company
    from db_router import _engine_cache, _get_or_create

    data = request.get_json()
    label = (data.get("label") or "").strip()
    sql = (data.get("sql") or "").strip()
    target = data.get("target", "all")
    dry_run = data.get("dry_run", False)
    user_email = get_current_user().get("email", "unknown")

    if not label:
        return jsonify({"error": "Migration label is required"}), 400
    if not sql:
        return jsonify({"error": "SQL is required"}), 400

    # Determine which companies to target
    if target == "all":
        companies = Company.query.filter_by(is_active=True).all()
    else:
        companies = Company.query.filter_by(company_id=target, is_active=True).all()

    results = []

    for company in companies:
        company_id = company.company_id
        result = {
            "company_id": company_id,
            "company_name": company.company_name,
            "status": None,
            "message": "",
            "skipped": False,
        }

        try:
            engine = _engine_cache.get(company_id)
            if engine is None:
                _get_or_create(company_id)
                engine = _engine_cache[company_id]

            _ensure_migration_table(engine)

            # Skip if already applied successfully
            if _already_applied(engine, label):
                result["status"] = "skipped"
                result["message"] = "Already applied — skipped"
                result["skipped"] = True
                results.append(result)
                continue

            if dry_run:
                result["status"] = "dry_run"
                result["message"] = "Dry run — SQL not executed"
                results.append(result)
                continue

            # Run the SQL
            with engine.connect() as conn:
                statements = [s.strip() for s in sql.split(";") if s.strip()]
                for stmt in statements:
                    conn.execute(text(stmt))
                conn.commit()

            _log_migration(engine, label, sql, "success", None, user_email)
            result["status"] = "success"
            result["message"] = "Applied successfully"

        except Exception as e:
            err = str(e)
            try:
                _log_migration(engine, label, sql, "failed", err, user_email)
            except Exception:
                pass
            result["status"] = "failed"
            result["message"] = err

        results.append(result)

    summary = {
        "total": len(results),
        "success": sum(1 for r in results if r["status"] == "success"),
        "skipped": sum(1 for r in results if r["status"] == "skipped"),
        "failed": sum(1 for r in results if r["status"] == "failed"),
        "dry_run": sum(1 for r in results if r["status"] == "dry_run"),
    }

    return jsonify({"results": results, "summary": summary})


def _ensure_migration_table(engine):
    """Create migration history table in customer database if it doesn't exist"""
    with engine.connect() as conn:
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS schema_migrations (
                id INT AUTO_INCREMENT PRIMARY KEY,
                label VARCHAR(200) NOT NULL,
                sql_executed TEXT,
                status VARCHAR(20) NOT NULL DEFAULT 'pending',
                error_msg TEXT,
                applied_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                applied_by VARCHAR(100)
            )
        """))
        conn.commit()


def _already_applied(engine, label):
    """Check if migration already applied to this customer DB"""
    with engine.connect() as conn:
        result = conn.execute(
            text("SELECT COUNT(*) FROM schema_migrations WHERE label = :label AND status = 'success'"),
            {"label": label}
        ).scalar()
        return result > 0


def _log_migration(engine, label, sql, status, error_msg, applied_by):
    """Log migration to customer database history table"""
    with engine.connect() as conn:
        conn.execute(
            text("""
                INSERT INTO schema_migrations (label, sql_executed, status, error_msg, applied_by)
                VALUES (:label, :sql, :status, :error_msg, :applied_by)
            """),
            {
                "label": label,
                "sql": sql,
                "status": status,
                "error_msg": error_msg,
                "applied_by": applied_by
            }
        )
        conn.commit()

@app.route("/migrations/history")
@login_required
@super_admin_required
def migration_history_all():
    """
    Return combined migration history across ALL active companies as JSON.
    Called by the super_admin page when the Migrations tab is opened.
    """
    from platform_models import Company
    from db_router import _engine_cache, _get_or_create

    companies = Company.query.filter_by(is_active=True).all()
    all_rows  = []

    for company in companies:
        company_id = company.company_id
        try:
            engine = _engine_cache.get(company_id)
            if engine is None:
                _get_or_create(company_id)
                engine = _engine_cache[company_id]

            _ensure_migration_table(engine)

            with engine.connect() as conn:
                rows = conn.execute(
                    text("SELECT label, sql_executed, status, error_msg, applied_at, applied_by FROM schema_migrations ORDER BY applied_at DESC LIMIT 100")
                ).fetchall()

            for r in rows:
                all_rows.append({
                    "company_id":   company_id,
                    "company_name": company.company_name,
                    "label":        r[0],
                    "sql":          r[1],
                    "status":       r[2],
                    "error_msg":    r[3],
                    "applied_at":   r[4].strftime("%d %b %Y %H:%M") if r[4] else "",
                    "applied_by":   r[5],
                    "_sort_key":    r[4].isoformat() if r[4] else "",
                })
        except Exception as e:
            pass  # Skip companies whose DB is unreachable

    # Sort newest first
    all_rows.sort(key=lambda x: x["_sort_key"], reverse=True)
    for row in all_rows:
        del row["_sort_key"]

    return jsonify({"history": all_rows})


@app.route("/migrations/history/<company_id>")
@login_required
@super_admin_required
def migration_history(company_id):
    """Return migration history for one specific company as JSON."""
    from db_router import _engine_cache, _get_or_create

    engine = _engine_cache.get(company_id)
    if engine is None:
        _get_or_create(company_id)
        engine = _engine_cache[company_id]

    _ensure_migration_table(engine)

    with engine.connect() as conn:
        rows = conn.execute(
            text("SELECT label, sql_executed, status, error_msg, applied_at, applied_by FROM schema_migrations ORDER BY applied_at DESC")
        ).fetchall()

    history = [
        {
            "label":      r[0],
            "sql":        r[1],
            "status":     r[2],
            "error_msg":  r[3],
            "applied_at": r[4].strftime("%Y-%m-%d %H:%M:%S") if r[4] else "",
            "applied_by": r[5],
        }
        for r in rows
    ]

    return jsonify({"history": history})

@app.route("/admin/dashboard")
@login_required
@super_admin_required
def admin_dashboard():
    stats = {
        "total_companies":  Company.query.count(),
        "total_users":      0,  # Will calculate differently
        "active_companies": Company.query.filter_by(is_active=True).count(),
        "monthly_revenue":  0,
    }
    
    # Calculate total users across all companies
    companies = Company.query.all()
    for company in companies:
        try:
            cdb = get_customer_session(company.company_id, db_session=db.session)
            user_count = cdb.query(CompanyUser).count()
            stats["total_users"] += user_count
            from db_router import close_customer_session
            close_customer_session(company.company_id)
        except Exception:
            pass
    
    plan_distribution = {}
    for c in companies:
        plan_distribution[c.subscription_plan] = plan_distribution.get(c.subscription_plan, 0) + 1

    # Clients registered by super admin who haven't created their company yet
    all_owners = RegisteredUser.query.filter_by(role="owner").all()
    pending_clients = [u for u in all_owners if not u.has_company]

    stats["pending_setup"] = len(pending_clients)
    stats["total_clients"] = len(all_owners)

    # Users tab: every registered user + how many companies they own
    all_users = RegisteredUser.query.order_by(RegisteredUser.created_at.desc()).all()
    users_data = []
    for u in all_users:
        user_companies = [c for c in companies if c.owner_email == u.email]
        users_data.append({
            'user': u,
            'company_count': len(user_companies),
            'companies': user_companies
        })

    return render_template("super_admin.html",
                           stats=stats,
                           companies=companies,
                           pending_clients=pending_clients,
                           users_data=users_data,
                           total_users=len(all_users),
                           plans=get_all_plans(),
                           plan_distribution=plan_distribution,
                           today=today_ist())


@app.route("/admin/companies")
@login_required
@super_admin_required
def admin_companies():
    return render_template("admin_companies.html", companies=Company.query.all())


@app.route("/admin/company/<company_id>")
@login_required
@super_admin_required
def admin_company_detail(company_id):
    cdb = get_cdb()
    company = get_company_by_id(company_id)
    users   = cdb.query(CompanyUser).filter_by(company_id=company_id).all()
    return render_template("admin_company_detail.html",
                           company=company, users=users, plans=get_all_plans())


@app.route("/admin/company/<company_id>/update-plan", methods=["POST"])
@login_required
@super_admin_required
def admin_update_company_plan(company_id):
    plan_id = request.form.get("plan")
    company = get_company_by_id(company_id)
    plan    = SubscriptionPlan.query.get(plan_id)
    if company and plan:
        company.subscription_plan     = plan.id
        company.max_companies_allowed = plan.max_companies
        company.max_users_per_company = plan.max_users
        db.session.commit()
        flash(f"Company plan updated to {plan.name}")
    return redirect(url_for("admin_company_detail", company_id=company_id))


@app.route("/admin/company/<company_id>/renew", methods=["POST"])
@login_required
@super_admin_required
def admin_renew_company(company_id):
    company = get_company_by_id(company_id)
    if not company:
        flash("Company not found")
        return redirect(url_for("super_admin_dashboard"))

    # Renew from today, or from the current expiry if it's still in the future
    # (renewing early shouldn't cost the company days they already paid for).
    start_from = company.subscription_end if (company.subscription_end and company.subscription_end > today_ist()) else today_ist()
    company.subscription_start = company.subscription_start or today_ist()
    company.subscription_end = start_from + timedelta(days=365)
    company.is_active = True
    db.session.commit()
    flash(f"{company.company_name} renewed until {company.subscription_end.strftime('%d %b %Y')}")
    return redirect(url_for("admin_company_detail", company_id=company_id))


@app.route("/admin/company/<company_id>/toggle-status", methods=["POST"])
@login_required
@super_admin_required
def admin_toggle_company_status(company_id):
    company = get_company_by_id(company_id)
    if company:
        company.is_active = not company.is_active
        db.session.commit()
        status = "activated" if company.is_active else "suspended"
        flash(f"Company {status}")
    return redirect(url_for("admin_company_detail", company_id=company_id))


@app.route("/admin/company/<company_id>/edit", methods=["POST"])
@login_required
@super_admin_required
def admin_edit_company(company_id):
    company = get_company_by_id(company_id)
    if not company:
        return jsonify({"error": "Company not found"}), 404
    
    try:
        new_name = request.form.get("company_name", company.company_name).strip()
        new_gst = request.form.get("gst_number", "").strip().upper()
        if not new_gst:
            new_gst = None
        
        # Check if another active company has this name (same owner)
        if new_name != company.company_name:
            existing = Company.query.filter(
                func.lower(Company.company_name) == func.lower(new_name),
                Company.owner_email == company.owner_email,
                Company.is_active == True,
                Company.company_id != company_id
            ).first()
            if existing:
                return jsonify({"error": f"Company name '{new_name}' is already taken by another active company."}), 400
        
        # ── Check if another active company has this GST number ──
        if new_gst and new_gst != company.gst_number:
            existing_gst = Company.query.filter(
                func.lower(Company.gst_number) == func.lower(new_gst),
                Company.is_active == True,
                Company.company_id != company_id
            ).first()
            if existing_gst:
                return jsonify({"error": f"GST number '{new_gst}' is already registered to another active company."}), 400
        
        company.company_name = new_name
        company.phone = request.form.get("phone", company.phone)
        company.address = request.form.get("address", company.address)
        company.gst_number = new_gst

        # Max users / max companies are only hand-editable for companies on the
        # "custom" plan — every other plan's limits come from admin_update_company_plan()
        # and get overwritten there, so letting them drift here would just cause
        # the table value and the plan value to silently disagree.
        if company.subscription_plan == "custom":
            max_users = request.form.get("max_users_per_company", "").strip()
            if max_users:
                if not max_users.isdigit():
                    return jsonify({"error": "Max users must be a whole number."}), 400
                company.max_users_per_company = max_users

            max_companies = request.form.get("max_companies_allowed", "").strip()
            if max_companies:
                if not max_companies.isdigit():
                    return jsonify({"error": "Max companies must be a whole number."}), 400
                company.max_companies_allowed = max_companies

        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500
    
    return jsonify({"success": True})

@app.route("/admin/user/<user_id>/delete", methods=["POST"])
@login_required
@super_admin_required
def admin_delete_user(user_id):
    """
    Delete a registered user (owner account) from the platform.
    This removes the user and ALL their companies.
    WARNING: This is permanent and cannot be undone.
    """
    user = RegisteredUser.query.filter_by(user_id=user_id).first()
    if not user:
        return jsonify({"error": "User not found"}), 404
    
    # Prevent super admin from deleting themselves
    current_user = get_current_user()
    if user.email == current_user.get("email"):
        return jsonify({"error": "You cannot delete your own account"}), 400
    
    try:
        from platform_models import CompanyWhatsAppConfig, BackupRecord, BackupSchedule
        
        # Get all companies owned by this user
        companies = Company.query.filter_by(owner_email=user.email).all()
        company_names = [c.company_name for c in companies]
        
        # Delete each company's related data
        for company in companies:
            # Delete WhatsApp configs
            CompanyWhatsAppConfig.query.filter_by(company_id=company.company_id).delete()
            # Delete backup records
            BackupRecord.query.filter_by(company_id=company.company_id).delete()
            # Delete backup schedules
            BackupSchedule.query.filter_by(company_id=company.company_id).delete()
            # Delete the company
            db.session.delete(company)
        
        # Delete the user
        db.session.delete(user)
        db.session.commit()
        
        return jsonify({
            "success": True, 
            "message": f"User '{user.full_name}' and their {len(companies)} company(ies) have been deleted.",
            "deleted_companies": company_names
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

@app.route("/admin/users")
@login_required
@super_admin_required
def admin_users():
    """Users are now a tab inside /admin/dashboard (usersTab) rather than a
    standalone page — kept as a redirect so any old links/bookmarks still work."""
    return redirect(url_for('admin_dashboard'))

@app.route("/admin/company/<company_id>/delete", methods=["POST"])
@login_required
@super_admin_required
def admin_delete_company(company_id):
    """
    Delete a company at the platform level. Removes the Company row and its
    platform-side dependents (WhatsApp config, backup records/schedules).
    Does NOT drop the company's own customer database — left in place so it
    can be recovered or archived manually if needed.
    """
    company = get_company_by_id(company_id)
    if not company:
        return jsonify({"error": "Company not found"}), 404

    try:
        from platform_models import BackupRecord, BackupSchedule, CompanyWhatsAppConfig

        CompanyWhatsAppConfig.query.filter_by(company_id=company_id).delete()
        BackupRecord.query.filter_by(company_id=company_id).delete()
        BackupSchedule.query.filter_by(company_id=company_id).delete()

        db.session.delete(company)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

    return jsonify({"success": True})


@app.route("/admin/register-client", methods=["GET", "POST"])
@login_required
@super_admin_required
def register_client():
    """
    Super admin creates a client account directly: full name, address, phone,
    email/login, password, plan, and payment status. No Company is created
    here — the client sets that up themselves on first login
    (see onboard_company).
    """
    if request.method == "POST":
        full_name = request.form.get("full_name", "").strip()
        address   = request.form.get("address", "").strip()
        phone     = request.form.get("phone", "").strip()
        email     = request.form.get("email", "").strip().lower()
        password  = request.form.get("password", "")
        plan_key  = request.form.get("subscription_plan", "starter")

        payment_status = request.form.get("payment_status", "pending")
        amount_total   = request.form.get("amount_total", "").strip()
        amount_paid    = request.form.get("amount_paid", "0").strip()

        if not full_name or not email or not password:
            flash("Full name, email and password are required", "error")
            return redirect(url_for("register_client"))

        if RegisteredUser.query.filter_by(email=email).first():
            flash("An account with this email already exists", "error")
            return redirect(url_for("register_client"))

        if len(password) < 8:
            flash("Password must be at least 8 characters", "error")
            return redirect(url_for("register_client"))

        plan_obj = SubscriptionPlan.query.get(plan_key) or SubscriptionPlan.query.order_by(SubscriptionPlan.id).first()
        if not plan_obj:
            flash("No subscription plans are configured. Add a plan before registering clients.", "error")
            return redirect(url_for("register_client"))

        try:
            amount_total_val = float(amount_total) if amount_total else None
        except ValueError:
            amount_total_val = None
        try:
            amount_paid_val = float(amount_paid) if amount_paid else 0.0
        except ValueError:
            amount_paid_val = 0.0

        # Keep status consistent with the amounts actually entered
        if amount_total_val and amount_paid_val >= amount_total_val:
            payment_status = "paid"
        elif amount_paid_val > 0:
            payment_status = "partial"
        else:
            payment_status = "pending"

        user_id = generate_next_user_id()

        # ── NEW: Handle custom plan fields ──────────────────────────────────
        custom_max_companies = None
        custom_max_users = None
        
        if plan_key == "custom":
            custom_max_companies = request.form.get("custom_max_companies", "").strip()
            custom_max_users = request.form.get("custom_max_users", "").strip()
            
            # Validate custom fields
            if not custom_max_companies or not custom_max_users:
                flash("Please enter custom Max Companies and Max Users for the Custom plan.", "error")
                return redirect(url_for("register_client"))
            
            try:
                custom_max_companies = int(custom_max_companies)
                custom_max_users = int(custom_max_users)
                if custom_max_companies < 1 or custom_max_users < 1:
                    flash("Max Companies and Max Users must be at least 1.", "error")
                    return redirect(url_for("register_client"))
            except ValueError:
                flash("Max Companies and Max Users must be valid numbers.", "error")
                return redirect(url_for("register_client"))

        new_user = RegisteredUser(
            user_id=user_id,
            email=email,
            password_hash=hash_password(password),
            full_name=full_name,
            phone=phone,
            address=address,
            role="owner",
            subscription_plan=plan_obj.id,
            created_at=today_ist(),
            is_active=True,
            payment_status=payment_status,
            amount_total=amount_total_val,
            amount_paid=amount_paid_val,
            registered_by=get_current_user().get("email"),
            registered_at=datetime.utcnow(),
            custom_max_companies=custom_max_companies,
            custom_max_users=custom_max_users,
        )
        db.session.add(new_user)
        db.session.commit()

        flash(
            f"Client '{full_name}' registered. They can log in with {email} — "
            f"they'll be asked to set up their company profile on first login.",
            "success"
        )
        return redirect(url_for("admin_dashboard"))

    return render_template("register_client.html", plans=get_all_plans())


# ─────────────────────────────────────────────────────────────────────────────
# ── Employee Management ───────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────
@app.route("/employees")
@login_required
@owner_required
def employee_list():
    cdb = get_cdb()
    company_id = get_current_company()
    employees  = cdb.query(CompanyUser).filter_by(company_id=company_id).all()
    return render_template("employees.html", employees=employees)


@app.route("/employees/add", methods=["GET", "POST"])
@login_required
@owner_required
def employee_add():
    cdb = get_cdb()
    company_id = get_current_company()
    can_add, msg = check_company_limit(company_id, "user")
    if not can_add:
        flash(msg)
        return redirect(url_for("employee_list"))

    if request.method == "POST":
        email    = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        emp_id    = _next_numbered_id(cdb, CompanyUser.user_id, "EMP")
        new_emp   = CompanyUser(
            user_id=emp_id, company_id=company_id, email=email,
            password_hash=hash_password(password),
            full_name=request.form.get("full_name", ""),
            role=request.form.get("role", "employee"),
            department=request.form.get("department", ""),
            phone=request.form.get("phone", ""),
            is_active=True, created_at=today_ist(),
        )
        cdb.add(new_emp)
        cdb.commit()
        flash("Employee added!")
        return redirect(url_for("employee_list"))
    return render_template("employee_form.html")


@app.route("/employees/toggle/<user_id>", methods=["POST"])
@login_required
@owner_required
def employee_toggle(user_id):
    cdb = get_cdb()
    company_id = get_current_company()
    emp        = _first_or_404(cdb.query(CompanyUser).filter_by(user_id=user_id, company_id=company_id).first())
    emp.is_active = not emp.is_active
    cdb.commit()
    flash(f"Employee {'activated' if emp.is_active else 'deactivated'}.")
    return redirect(url_for("employee_list"))


# ─────────────────────────────────────────────────────────────────────────────
# ── Product Lookup API ────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────
@app.route("/api/product/<code>")
@login_required
@require_permission("stock", "view")
def api_product_lookup(code):
    cdb = get_cdb()
    company_id = get_current_company()
    code_clean = code.strip().upper()
    item = cdb.query(StockItem).filter_by(company_id=company_id, code=code_clean).first()
    if not item:
        item = cdb.query(StockItem).filter(
            StockItem.company_id == company_id,
            StockItem.name.ilike(f"%{code_clean}%")
        ).first()
    if not item:
        return jsonify({"found": False, "message": f"No product found for '{code}'"}), 404
    return jsonify({
        "found": True, "code": item.code, "name": item.name,
        "rate": item.unit_price, "unit": item.unit or "pcs",
        "category": item.category or "", "stock": item.quantity,
        "hsn": item.hsn or "",
        "low_stock": item.quantity <= item.reorder_level,
    }), 200


@app.route("/api/products/search")
@login_required
@require_permission("stock", "view")
def api_products_search():
    cdb = get_cdb()
    company_id = get_current_company()
    q = request.args.get("q", "").strip().upper()
    if not q:
        return jsonify({"results": []})
    items = cdb.query(StockItem).filter(
        StockItem.company_id == company_id,
        db.or_(StockItem.code.ilike(f"%{q}%"), StockItem.name.ilike(f"%{q}%"))
    ).limit(8).all()
    return jsonify({"results": [{
        "code": s.code, "name": s.name, "rate": s.unit_price,
        "unit": s.unit or "pcs", "stock": s.quantity, "hsn": s.hsn or "",
    } for s in items]})


# ============================================
# BANK ACCOUNTS & FINANCE ROUTES
# ============================================

# ============================================
# CASH IN HAND ROUTES
# ============================================

@app.route("/cash-in-hand")
@login_required
@require_permission("cash", "view")
def cash_in_hand():
    """Cash in hand tracking"""
    cdb = get_cdb()
    company_id = get_current_company()
    
    # Get filter parameters
    from_date_str = request.args.get('from_date', '')
    to_date_str = request.args.get('to_date', '')
    filter_type = request.args.get('type', 'all')
    
    # Set default dates (last 30 days)
    if not from_date_str:
        from_date = today_ist() - timedelta(days=30)
    else:
        from_date = date.fromisoformat(from_date_str)
    
    if not to_date_str:
        to_date = today_ist()
    else:
        to_date = date.fromisoformat(to_date_str)
    
    # Build query
    query = cdb.query(CashTransaction).filter(
        CashTransaction.company_id == company_id,
        CashTransaction.date >= from_date,
        CashTransaction.date <= to_date
    )
    
    if filter_type != 'all':
        query = query.filter(CashTransaction.type == filter_type)
    
    transactions = query.order_by(CashTransaction.date.desc()).all()
    
    # Calculate totals
    total_inflow = sum(t.amount for t in transactions if t.type == 'income')
    total_outflow = sum(t.amount for t in transactions if t.type == 'expense')
    
    # Calculate current balance (all time)
    all_income = cdb.query(CashTransaction).filter_by(company_id=company_id, type='income').all()
    all_expense = cdb.query(CashTransaction).filter_by(company_id=company_id, type='expense').all()
    current_balance = sum(t.amount for t in all_income) - sum(t.amount for t in all_expense)
    
    # Format transactions for template
    running_balance = 0
    all_transactions = cdb.query(CashTransaction).filter_by(company_id=company_id).order_by(CashTransaction.date.asc()).all()
    
    # Create a dict of running balances
    balance_map = {}
    for t in all_transactions:
        if t.type == 'income':
            running_balance += t.amount
        else:
            running_balance -= t.amount
        balance_map[t.id] = running_balance
    
    transactions_list = []
    for t in transactions:
        transactions_list.append({
            'id': t.id,
            'date': t.date.strftime('%d %b %Y'),
            'type': t.type,
            'category': t.category,
            'description': t.description,
            'amount': t.amount,
            'reference': t.reference or '',
            'notes': t.notes or '',
            'balance_after': balance_map.get(t.id, 0)
        })
    
    return render_template("cash_in_hand.html",
                         active='cash_in_hand',
                         current_balance=current_balance,
                         total_inflow=total_inflow,
                         total_outflow=total_outflow,
                         transactions=transactions_list,
                         from_date=from_date.strftime('%Y-%m-%d'),
                         to_date=to_date.strftime('%Y-%m-%d'),
                         today=today_ist().strftime('%Y-%m-%d'))


@app.route("/api/cash-transaction/save", methods=["POST"])
@login_required
@require_permission("cash", "create")
def save_cash_transaction():
    """Save a cash transaction"""
    company_id = get_current_company()
    data = request.get_json()
    
    try:
        transaction = CashTransaction(
            company_id=company_id,
            type=data.get('type'),
            date=date.fromisoformat(data.get('date')),
            category=data.get('category'),
            description=data.get('description'),
            amount=data.get('amount'),
            reference=data.get('reference', ''),
            notes=data.get('notes', ''),
            created_by=get_current_user().get('email')
        )
        cdb.add(transaction)
        cdb.commit()
        
        return jsonify({'success': True, 'message': 'Transaction saved successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 400


@app.route("/api/cash-transaction/delete/<int:txn_id>", methods=["DELETE"])
@login_required
@owner_required
def delete_cash_transaction(txn_id):
    """Delete a cash transaction"""
    cdb = get_cdb()
    company_id = get_current_company()
    transaction = cdb.query(CashTransaction).filter_by(id=txn_id, company_id=company_id).first()
    
    if not transaction:
        return jsonify({'success': False, 'message': 'Transaction not found'}), 404
    
    try:
        cdb.delete(transaction)
        cdb.commit()
        return jsonify({'success': True, 'message': 'Transaction deleted'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 400

# ============================================
# BANK ACCOUNTS ROUTES
# ============================================

@app.route("/bank-accounts")
@login_required
@require_permission("bank", "view")
def bank_accounts():
    """Bank Accounts management page"""
    cdb = get_cdb()
    company_id = get_current_company()
    bank_accounts = cdb.query(BankAccount).filter_by(company_id=company_id, status='Active').all()
    inactive_accounts = cdb.query(BankAccount).filter_by(company_id=company_id, status='Inactive').all()
    inactive_txn_counts = {
        acc.id: cdb.query(BankTransaction).filter_by(bank_account_id=acc.id).count()
        for acc in inactive_accounts
    }
    
    # Calculate total balance
    total_balance = sum(acc.balance for acc in bank_accounts)
    
    return render_template("bank_accounts.html", 
                         active='bank_accounts',
                         bank_accounts=bank_accounts,
                         inactive_accounts=inactive_accounts,
                         inactive_txn_counts=inactive_txn_counts,
                         total_balance=total_balance)


@app.route("/bank-accounts/add", methods=["POST"])
@login_required
@require_permission("bank", "create")
def add_bank_account():
    """Add a new bank account"""
    cdb = get_cdb()
    company_id = get_current_company()
    
    bank_name = request.form.get("bank_name", "").strip()
    account_name = request.form.get("account_name", "").strip()
    account_number = request.form.get("account_number", "").strip()
    ifsc_code = request.form.get("ifsc_code", "").strip()
    branch = request.form.get("branch", "").strip()
    opening_balance = float(request.form.get("balance", 0) or 0)
    
    if not bank_name or not account_name or not account_number:
        flash("Bank Name, Account Name, and Account Number are required!")
        return redirect(url_for("bank_accounts"))
    
    # Check if account number already exists for this company
    existing = cdb.query(BankAccount).filter_by(company_id=company_id, account_number=account_number).first()
    if existing:
        flash(f"Account number {account_number} already exists!")
        return redirect(url_for("bank_accounts"))
    
    new_account = BankAccount(
        company_id=company_id,
        bank_name=bank_name,
        account_name=account_name,
        account_number=account_number,
        ifsc_code=ifsc_code,
        branch=branch,
        opening_balance=opening_balance,
        balance=opening_balance,
        status='Active',
        created_at=datetime.utcnow()
    )
    
    cdb.add(new_account)
    cdb.flush()  # assigns new_account.id before it's referenced below

    # Add opening balance transaction if opening_balance > 0
    if opening_balance > 0:
        opening_txn = BankTransaction(
            bank_account_id=new_account.id,
            company_id=company_id,
            type='credit',
            date=today_ist(),
            description=f"Opening Balance for {bank_name} - {account_name}",
            amount=opening_balance,
            reference="Opening Balance",
            transaction_mode="Cash",
            created_by=get_current_user().get('email')
        )
        cdb.add(opening_txn)
    
    cdb.commit()
    flash(f"Bank account {bank_name} - {account_name} added successfully!")
    return redirect(url_for("bank_accounts"))


@app.route("/bank-accounts/<int:account_id>/transactions")
@login_required
@require_permission("bank", "view")
def bank_transactions(account_id):
    """View transactions for a specific bank account"""
    cdb = get_cdb()
    company_id = get_current_company()
    account = _first_or_404(cdb.query(BankAccount).filter_by(id=account_id, company_id=company_id).first())
    
    # Get filter parameters
    from_date_str = request.args.get('from_date', '')
    to_date_str = request.args.get('to_date', '')
    txn_type = request.args.get('type', 'all')
    
    # Set default dates (last 30 days)
    if not from_date_str:
        from_date = today_ist() - timedelta(days=30)
    else:
        from_date = date.fromisoformat(from_date_str)
    
    if not to_date_str:
        to_date = today_ist()
    else:
        to_date = date.fromisoformat(to_date_str)
    
    # Build query
    query = cdb.query(BankTransaction).filter(
        BankTransaction.bank_account_id == account_id,
        BankTransaction.company_id == company_id,
        BankTransaction.date >= from_date,
        BankTransaction.date <= to_date
    )
    
    if txn_type != 'all':
        query = query.filter(BankTransaction.type == txn_type)
    
    transactions = query.order_by(BankTransaction.date.desc()).all()
    
    # Calculate totals
    total_credits = sum(t.amount for t in transactions if t.type == 'credit')
    total_debits = sum(t.amount for t in transactions if t.type == 'debit')
    
    return render_template("bank_transactions.html",
                         active='bank_accounts',
                         account=account,
                         transactions=transactions,
                         total_credits=total_credits,
                         total_debits=total_debits,
                         from_date=from_date.strftime('%Y-%m-%d'),
                         to_date=to_date.strftime('%Y-%m-%d'),
                         today=today_ist().strftime('%Y-%m-%d'))


@app.route("/bank-accounts/<int:account_id>/add-transaction", methods=["POST"])
@login_required
@require_permission("bank", "create")
def add_bank_transaction(account_id):
    """Add a transaction to a bank account"""
    cdb = get_cdb()
    company_id = get_current_company()
    account = _first_or_404(cdb.query(BankAccount).filter_by(id=account_id, company_id=company_id).first())
    
    txn_type = request.form.get("type")
    date_str = request.form.get("date")
    description = request.form.get("description", "").strip()
    amount = float(request.form.get("amount", 0))
    reference = request.form.get("reference", "").strip()
    transaction_mode = request.form.get("transaction_mode", "Transfer")
    notes = request.form.get("notes", "").strip()
    
    if not description or amount <= 0:
        flash("Description and valid amount are required!")
        return redirect(url_for("bank_transactions", account_id=account_id))
    
    # Create transaction
    transaction = BankTransaction(
        bank_account_id=account.id,
        company_id=company_id,
        type=txn_type,
        date=date.fromisoformat(date_str) if date_str else today_ist(),
        description=description,
        amount=amount,
        reference=reference,
        transaction_mode=transaction_mode,
        notes=notes,
        created_by=get_current_user().get('email')
    )
    cdb.add(transaction)
    
    # Update account balance
    if txn_type == 'credit':
        account.balance += amount
    else:
        account.balance -= amount
    
    account.updated_at = datetime.utcnow()
    
    cdb.commit()
    flash(f"{'Deposit' if txn_type == 'credit' else 'Withdrawal'} of ₹{amount:,.2f} recorded successfully!")
    return redirect(url_for("bank_transactions", account_id=account_id))


@app.route("/bank-accounts/<int:account_id>/delete", methods=["GET", "POST"])
@login_required
@owner_required
def delete_bank_account(account_id):
    """Delete a bank account (soft delete by setting status to Inactive)"""
    cdb = get_cdb()
    company_id = get_current_company()
    account = _first_or_404(cdb.query(BankAccount).filter_by(id=account_id, company_id=company_id).first())
    
    # Soft delete - just mark as inactive
    account.status = 'Inactive'
    cdb.commit()
    
    flash(f"Bank account {account.bank_name} - {account.account_name} has been deactivated.")
    return redirect(url_for("bank_accounts"))


@app.route("/bank-accounts/<int:account_id>/reactivate", methods=["GET", "POST"])
@login_required
@owner_required
def reactivate_bank_account(account_id):
    """Reactivate a previously deactivated bank account"""
    cdb = get_cdb()
    company_id = get_current_company()
    account = _first_or_404(cdb.query(BankAccount).filter_by(id=account_id, company_id=company_id).first())

    account.status = 'Active'
    cdb.commit()

    flash(f"Bank account {account.bank_name} - {account.account_name} has been reactivated.")
    return redirect(url_for("bank_accounts"))


@app.route("/bank-accounts/<int:account_id>/delete-permanent", methods=["POST"])
@login_required
@owner_required
def delete_bank_account_permanent(account_id):
    """Permanently delete a bank account. Only allowed when the account is
    already Inactive and has zero transactions, so a real ledger with history
    can never be wiped by accident."""
    cdb = get_cdb()
    company_id = get_current_company()
    account = _first_or_404(cdb.query(BankAccount).filter_by(id=account_id, company_id=company_id).first())

    if account.status != 'Inactive':
        flash("Deactivate this account before deleting it permanently.")
        return redirect(url_for("bank_accounts"))

    txn_count = cdb.query(BankTransaction).filter_by(bank_account_id=account.id).count()
    name = f"{account.bank_name} - {account.account_name}"
    cdb.delete(account)  # cascade="all, delete-orphan" also removes its BankTransaction rows
    cdb.commit()

    if txn_count > 0:
        flash(f"Bank account {name} and its {txn_count} transaction(s) have been permanently deleted.")
    else:
        flash(f"Bank account {name} has been permanently deleted.")
    return redirect(url_for("bank_accounts"))


@app.route("/bank-accounts/<int:account_id>/transfer", methods=["POST"])
@login_required
@require_permission("bank", "edit")
def bank_transfer(account_id):
    """Transfer money between bank accounts"""
    cdb = get_cdb()
    company_id = get_current_company()
    from_account = _first_or_404(cdb.query(BankAccount).filter_by(id=account_id, company_id=company_id).first())
    
    to_account_id = request.form.get("to_account_id", type=int)
    amount = float(request.form.get("amount", 0))
    date_str = request.form.get("date")
    description = request.form.get("description", "").strip()
    reference = request.form.get("reference", "").strip()
    
    to_account = cdb.query(BankAccount).filter_by(id=to_account_id, company_id=company_id).first()
    
    if not to_account:
        flash("Destination account not found!")
        return redirect(url_for("bank_transactions", account_id=account_id))
    
    if amount <= 0:
        flash("Amount must be greater than 0!")
        return redirect(url_for("bank_transactions", account_id=account_id))
    
    if from_account.balance < amount:
        flash(f"Insufficient balance in {from_account.bank_name} - {from_account.account_name}!")
        return redirect(url_for("bank_transactions", account_id=account_id))
    
    txn_date = date.fromisoformat(date_str) if date_str else today_ist()
    
    # Debit transaction from source account
    debit_txn = BankTransaction(
        bank_account_id=from_account.id,
        company_id=company_id,
        type='debit',
        date=txn_date,
        description=f"Transfer to {to_account.bank_name} - {to_account.account_name}: {description}" if description else f"Transfer to {to_account.bank_name} - {to_account.account_name}",
        amount=amount,
        reference=reference,
        transaction_mode="Transfer",
        notes=f"Transfer from {from_account.bank_name} to {to_account.bank_name}",
        created_by=get_current_user().get('email'),
        party_name=f"{from_account.bank_name} - Transfer"  # ← ADD THIS
    )
    cdb.add(debit_txn)
    from_account.balance -= amount
    
    # Credit transaction to destination account
    credit_txn = BankTransaction(
        bank_account_id=to_account.id,
        company_id=company_id,
        type='credit',
        date=txn_date,
        description=f"Transfer from {from_account.bank_name} - {from_account.account_name}: {description}" if description else f"Transfer from {from_account.bank_name} - {from_account.account_name}",
        amount=amount,
        reference=reference,
        transaction_mode="Transfer",
        notes=f"Transfer from {from_account.bank_name} to {to_account.bank_name}",
        created_by=get_current_user().get('email'),
        party_name=f"{to_account.bank_name} - Transfer"
    )
    cdb.add(credit_txn)
    to_account.balance += amount
    
    from_account.updated_at = datetime.utcnow()
    to_account.updated_at = datetime.utcnow()
    
    cdb.commit()
    flash(f"Transferred ₹{amount:,.2f} from {from_account.bank_name} to {to_account.bank_name} successfully!")
    return redirect(url_for("bank_transactions", account_id=account_id))

@app.route("/admin/repair-party-names", methods=["POST"])
@login_required
@owner_required
def repair_party_names():
    """Fix all existing transactions with inconsistent party_name values."""
    cdb = get_cdb()
    company_id = get_current_company()
    
    # Get all clients
    clients = cdb.query(Client).filter_by(company_id=company_id).all()
    client_name_map = {c.name.lower(): c.name for c in clients}
    
    # Fix CashTransactions
    cash_txns = cdb.query(CashTransaction).filter_by(company_id=company_id).all()
    cash_repaired = 0
    for txn in cash_txns:
        if txn.party_name:
            txn_lower = txn.party_name.lower()
            # Try to match with a client
            for client_lower, client_name in client_name_map.items():
                if client_lower in txn_lower or txn_lower in client_lower:
                    if txn.party_name != client_name:
                        txn.party_name = client_name
                        cash_repaired += 1
                    break
    
    # Fix BankTransactions
    bank_txns = cdb.query(BankTransaction).filter_by(company_id=company_id).all()
    bank_repaired = 0
    for txn in bank_txns:
        if txn.party_name:
            txn_lower = txn.party_name.lower()
            for client_lower, client_name in client_name_map.items():
                if client_lower in txn_lower or txn_lower in client_lower:
                    if txn.party_name != client_name:
                        txn.party_name = client_name
                        bank_repaired += 1
                    break
    
    # Fix Suppliers
    suppliers = cdb.query(Supplier).filter_by(company_id=company_id).all()
    supplier_name_map = {s.name.lower(): s.name for s in suppliers}
    
    # Fix CashTransactions for suppliers
    for txn in cash_txns:
        if txn.party_name:
            txn_lower = txn.party_name.lower()
            for sup_lower, sup_name in supplier_name_map.items():
                if sup_lower in txn_lower or txn_lower in sup_lower:
                    if txn.party_name != sup_name:
                        txn.party_name = sup_name
                        cash_repaired += 1
                    break
    
    # Fix BankTransactions for suppliers
    for txn in bank_txns:
        if txn.party_name:
            txn_lower = txn.party_name.lower()
            for sup_lower, sup_name in supplier_name_map.items():
                if sup_lower in txn_lower or txn_lower in sup_lower:
                    if txn.party_name != sup_name:
                        txn.party_name = sup_name
                        bank_repaired += 1
                    break
    
    cdb.commit()
    
    flash(
        f"Repaired {cash_repaired} cash transactions and {bank_repaired} bank transactions. "
        f"Party names are now consistent with client/supplier records.",
        "success"
    )
    return redirect(url_for("company_settings"))

@app.route("/cheques")
@login_required
@require_permission("cheques", "view")
def cheques():
    """Cheque register — record and track cheques received/paid"""
    cdb        = get_cdb()
    company_id = get_current_company()

    all_clients   = cdb.query(Client).filter_by(company_id=company_id).order_by(Client.name).all()
    all_suppliers = cdb.query(Supplier).filter_by(company_id=company_id).order_by(Supplier.name).all()
    bank_accounts = cdb.query(BankAccount).filter_by(company_id=company_id, status='Active').all()

    direction   = request.args.get("direction", "")     # '', 'received', 'paid'
    status      = request.args.get("status", "")        # '', 'Pending', 'Cleared', 'Bounced', 'Cancelled'
    date_from_s = request.args.get("date_from", "")
    date_to_s   = request.args.get("date_to", "")
    date_from   = date.fromisoformat(date_from_s) if date_from_s else None
    date_to     = date.fromisoformat(date_to_s) if date_to_s else None

    q = cdb.query(Cheque).filter_by(company_id=company_id)
    if direction:
        q = q.filter(Cheque.direction == direction)
    if status:
        q = q.filter(Cheque.status == status)
    if date_from:
        q = q.filter(Cheque.cheque_date >= date_from)
    if date_to:
        q = q.filter(Cheque.cheque_date <= date_to)
    cheque_rows = q.order_by(Cheque.cheque_date.desc(), Cheque.id.desc()).all()

    cheque_list = [{
        "id":           c.id,
        "direction":    c.direction,
        "party_name":   c.party_name,
        "cheque_no":    c.cheque_no,
        "cheque_date":  c.cheque_date.strftime("%d %b %Y") if c.cheque_date else "",
        "bank_name":    c.bank_name or "—",
        "amount":       c.amount,
        "narration":    c.narration or "",
        "status":       c.status,
        "cleared_date": c.cleared_date.strftime("%d %b %Y") if c.cleared_date else "",
        "bill_ref":     (c.invoice.invoice_id if c.invoice else
                          (c.purchase_invoice.invoice_number or c.purchase_invoice.invoice_id) if c.purchase_invoice else None),
    } for c in cheque_rows]

    pending_received = sum(c.amount for c in cheque_rows if c.direction == "received" and c.status == "Pending")
    pending_paid     = sum(c.amount for c in cheque_rows if c.direction == "paid" and c.status == "Pending")
    cleared_received = sum(c.amount for c in cheque_rows if c.direction == "received" and c.status == "Cleared")
    cleared_paid     = sum(c.amount for c in cheque_rows if c.direction == "paid" and c.status == "Cleared")
    bounced_count    = sum(1 for c in cheque_rows if c.status == "Bounced")

    # ── Bill-wise + total pending per party, for the "select a party" step ──
    # Reuses the same helpers the Receipts/Payments pages already use, so a
    # client's outstanding invoices (and a supplier's outstanding purchase
    # bills) show up identically here — one source of truth instead of a
    # second copy of the balance logic.
    invoices_json          = _build_invoices_json(company_id, all_clients, _outstanding_invoices_for_client)
    purchase_invoices_json = _build_invoices_json(company_id, all_suppliers, _outstanding_invoices_for_supplier)
    client_pending_json    = json.dumps({str(c.id): (c.pending or 0) for c in all_clients})
    supplier_payable_json  = json.dumps({str(s.id): (s.payable or 0) for s in all_suppliers})

    # ── "Yet to receive / yet to pay" lists ──────────────────────────────────
    debtors   = [d for d in _debtor_summary(company_id) if d["total_pending"] > 0]
    creditors = [c for c in _creditor_summary(company_id) if c["total_pending"] > 0]

    return render_template(
        "cheques.html",
        active='cheques',
        clients=all_clients,
        suppliers=all_suppliers,
        bank_accounts=bank_accounts,
        cheques=cheque_list,
        direction=direction,
        status=status,
        date_from=date_from_s,
        date_to=date_to_s,
        pending_received=pending_received,
        pending_paid=pending_paid,
        cleared_received=cleared_received,
        cleared_paid=cleared_paid,
        bounced_count=bounced_count,
        invoices_json=invoices_json,
        purchase_invoices_json=purchase_invoices_json,
        client_pending_json=client_pending_json,
        supplier_payable_json=supplier_payable_json,
        debtors=debtors,
        creditors=creditors,
        today=str(today_ist()),
    )


@app.route("/cheques/save", methods=["POST"])
@login_required
@require_permission("cheques", "create")
def cheque_save():
    """Record a new cheque (received or paid) — starts life as Pending"""
    cdb        = get_cdb()
    company_id = get_current_company()

    direction   = request.form.get("direction", "received")
    party_type  = request.form.get("party_type", "")
    party_id    = request.form.get("party_id", type=int)
    party_name = get_party_name(
        client_id=party_id if party_type == "client" else None,
        supplier_id=party_id if party_type == "supplier" else None,
        form=request.form,
        fallback_name=request.form.get("party_name", "").strip()
    )
    cheque_no   = request.form.get("cheque_no", "").strip()
    cheque_date_s = request.form.get("cheque_date")
    bank_name   = request.form.get("bank_name", "").strip()
    bank_account_id = request.form.get("bank_account_id", type=int)
    amount      = request.form.get("amount", type=float, default=0)
    narration   = request.form.get("narration", "")

    # ── Apply against a specific bill, or record as a general/advance cheque ──
    # apply_to == "bill" pairs this cheque with one open Invoice (received) or
    # PurchaseInvoice (paid); "advance" leaves both link columns NULL. Either
    # way the cheque itself still starts life as Pending — the linked bill's
    # balance isn't touched until the cheque actually clears (see cheque_clear).
    apply_to        = request.form.get("apply_to", "advance")
    invoice_pk      = request.form.get("invoice_id", type=int)
    purchase_pk     = request.form.get("purchase_invoice_id", type=int)
    linked_invoice_id          = None
    linked_purchase_invoice_id = None

    if not party_name or not cheque_no or amount <= 0 or not cheque_date_s:
        flash("Please fill in party, cheque number, date, and a valid amount.", "error")
        return redirect(url_for("cheques"))

    if apply_to == "bill":
        if direction == "received":
            if not invoice_pk:
                flash("Select which invoice this cheque is against, or switch to a general payment.", "error")
                return redirect(url_for("cheques"))
            bill = cdb.query(Invoice).filter_by(id=invoice_pk, company_id=company_id, client_id=party_id).first()
            if not bill:
                flash("Selected invoice was not found for this client.", "error")
                return redirect(url_for("cheques"))
            if amount > (bill.balance or 0) + 0.01:
                flash(f"Cheque amount (₹{amount:,.2f}) is more than {bill.invoice_id}'s balance (₹{bill.balance or 0:,.2f}).", "error")
                return redirect(url_for("cheques"))
            linked_invoice_id = bill.id
        else:
            if not purchase_pk:
                flash("Select which bill this cheque is against, or switch to a general payment.", "error")
                return redirect(url_for("cheques"))
            bill = cdb.query(PurchaseInvoice).filter_by(id=purchase_pk, company_id=company_id, supplier_id=party_id).first()
            if not bill:
                flash("Selected purchase bill was not found for this supplier.", "error")
                return redirect(url_for("cheques"))
            if amount > (bill.balance or 0) + 0.01:
                flash(f"Cheque amount (₹{amount:,.2f}) is more than bill {bill.invoice_number or bill.invoice_id}'s balance (₹{bill.balance or 0:,.2f}).", "error")
                return redirect(url_for("cheques"))
            linked_purchase_invoice_id = bill.id

    cheque = Cheque(
        company_id=company_id,
        direction=direction,
        party_type=party_type or None,
        party_id=party_id,
        party_name=party_name,
        cheque_no=cheque_no,
        cheque_date=date.fromisoformat(cheque_date_s),
        bank_name=bank_name or None,
        bank_account_id=bank_account_id,
        amount=amount,
        narration=narration,
        status="Pending",
        invoice_id=linked_invoice_id,
        purchase_invoice_id=linked_purchase_invoice_id,
        created_by=get_current_user().get('email'),
    )
    cdb.add(cheque)
    cdb.commit()

    flash(f"Cheque {cheque_no} ({'received from' if direction == 'received' else 'issued to'} {party_name}) recorded as Pending.", "success")
    return redirect(url_for("cheques"))


@app.route("/cheques/<int:cheque_id>/clear", methods=["POST"])
@login_required
@require_permission("cheques", "edit")
def cheque_clear(cheque_id):
    """Mark a cheque as Cleared — this is what actually moves the bank balance"""
    cdb        = get_cdb()
    company_id = get_current_company()

    cheque = cdb.query(Cheque).filter_by(id=cheque_id, company_id=company_id).first()
    if not cheque:
        flash("Cheque not found.", "error")
        return redirect(url_for("cheques"))
    if cheque.status != "Pending":
        flash(f"Only pending cheques can be cleared (this one is {cheque.status}).", "error")
        return redirect(url_for("cheques"))
    if not cheque.bank_account_id:
        flash("Select a bank account for this cheque before clearing it.", "error")
        return redirect(url_for("cheques"))

    bank_account = cdb.query(BankAccount).filter_by(
        id=cheque.bank_account_id, company_id=company_id, status='Active'
    ).first()
    if not bank_account:
        flash("Linked bank account not found or inactive.", "error")
        return redirect(url_for("cheques"))

    cleared_date_s = request.form.get("cleared_date")
    cleared_date   = date.fromisoformat(cleared_date_s) if cleared_date_s else today_ist()

    is_received = cheque.direction == "received"
    bank_txn = BankTransaction(
        bank_account_id=bank_account.id,
        company_id=company_id,
        type="credit" if is_received else "debit",
        date=cleared_date,
        description=f"Cheque {'received from' if is_received else 'paid to'} {cheque.party_name}",
        amount=cheque.amount,
        reference=cheque.cheque_no,
        transaction_mode="Cheque",
        notes=cheque.narration,
        created_by=get_current_user().get('email'),
    )
    cdb.add(bank_txn)
    bank_account.balance += cheque.amount if is_received else -cheque.amount

    cheque.status       = "Cleared"
    cheque.cleared_date = cleared_date
    cdb.flush()
    cheque.bank_txn_id  = bank_txn.id

    # ── Settle the linked bill, if this cheque was applied against one ──────
    # Deliberately happens at CLEAR time, not at save time — a Pending cheque
    # can still bounce, and a bill shouldn't be marked paid off money that
    # hasn't actually landed yet. Same balance/status math as
    # purchase_make_payment() (paid) and receipt_save() (received) use for
    # cash/bank payments, so a cheque settles a bill exactly the same way
    # those do. A general/advance cheque (no invoice_id/purchase_invoice_id)
    # skips this — there's no bill to apply it to.
    if is_received and cheque.invoice_id:
        inv = cdb.query(Invoice).filter_by(id=cheque.invoice_id, company_id=company_id).first()
        if inv:
            apply = min(cheque.amount, inv.balance or 0)
            inv.paid_amount = (inv.paid_amount or 0) + apply
            inv.balance     = max(0, (inv.balance or 0) - apply)
            if inv.balance <= 0:
                inv.status = "Paid"
            elif inv.paid_amount > 0:
                inv.status = "Partial"
            if inv.client_id:
                client = cdb.query(Client).filter_by(id=inv.client_id, company_id=company_id).first()
                if client:
                    client.pending = max(0, (client.pending or 0) - apply)
    elif (not is_received) and cheque.purchase_invoice_id:
        pinv = cdb.query(PurchaseInvoice).filter_by(id=cheque.purchase_invoice_id, company_id=company_id).first()
        if pinv:
            apply = min(cheque.amount, pinv.balance or 0)
            pinv.paid_amount = (pinv.paid_amount or 0) + apply
            pinv.balance     = max(0, (pinv.balance or 0) - apply)
            if pinv.balance <= 0:
                pinv.status = "Paid"
            elif pinv.paid_amount > 0:
                pinv.status = "Partial"
            if pinv.supplier_id:
                supplier = cdb.query(Supplier).filter_by(id=pinv.supplier_id, company_id=company_id).first()
                if supplier:
                    supplier.payable = max(0, (supplier.payable or 0) - apply)

    cdb.commit()

    flash(f"Cheque {cheque.cheque_no} cleared — ₹{cheque.amount:,.2f} {'credited to' if is_received else 'debited from'} {bank_account.bank_name}.", "success")
    return redirect(url_for("cheques"))


@app.route("/cheques/<int:cheque_id>/bounce", methods=["POST"])
@login_required
@require_permission("cheques", "edit")
def cheque_bounce(cheque_id):
    """Mark a pending cheque as Bounced"""
    cdb        = get_cdb()
    company_id = get_current_company()

    cheque = cdb.query(Cheque).filter_by(id=cheque_id, company_id=company_id).first()
    if not cheque:
        flash("Cheque not found.", "error")
        return redirect(url_for("cheques"))
    if cheque.status != "Pending":
        flash(f"Only pending cheques can be marked as bounced (this one is {cheque.status}).", "error")
        return redirect(url_for("cheques"))

    cheque.status = "Bounced"
    cdb.commit()
    flash(f"Cheque {cheque.cheque_no} marked as Bounced.", "success")
    return redirect(url_for("cheques"))


@app.route("/cheques/<int:cheque_id>/cancel", methods=["POST"])
@login_required
@require_permission("cheques", "edit")
def cheque_cancel(cheque_id):
    """Cancel a pending cheque (e.g. entered by mistake)"""
    cdb        = get_cdb()
    company_id = get_current_company()

    cheque = cdb.query(Cheque).filter_by(id=cheque_id, company_id=company_id).first()
    if not cheque:
        flash("Cheque not found.", "error")
        return redirect(url_for("cheques"))
    if cheque.status != "Pending":
        flash(f"Only pending cheques can be cancelled (this one is {cheque.status}).", "error")
        return redirect(url_for("cheques"))

    cheque.status = "Cancelled"
    cdb.commit()
    flash(f"Cheque {cheque.cheque_no} cancelled.", "success")
    return redirect(url_for("cheques"))

# ============================================
# LOAN ACCOUNTS ROUTES
# ============================================

@app.route("/loan-accounts")
@login_required
@require_permission("loans", "view")
def loan_accounts():
    """Loan accounts management"""
    cdb = get_cdb()
    company_id = get_current_company()
    
    # Get all loans
    all_loans = cdb.query(Loan).filter_by(company_id=company_id).all()
    
    # Separate by type
    loans_given = []
    loans_taken = []
    
    for loan in all_loans:
        payments = []
        for payment in loan.repayments:
            payments.append({
                'id': payment.id,
                'date': payment.date.strftime('%d %b %Y'),
                'amount': payment.amount,
                'payment_mode': payment.payment_mode,
                'reference': payment.reference or '',
                'notes': payment.notes or ''
            })
        
        loan_dict = {
            'id': loan.id,
            'type': loan.type,
            'party_name': loan.party_name,
            'borrower_name': loan.party_name,
            'lender_name': loan.party_name,
            'loan_date': loan.loan_date.strftime('%d %b %Y'),
            'amount': loan.amount,
            'remaining_amount': loan.remaining_amount,
            'repaid_amount': loan.repaid_amount,
            'repayment_percentage': loan.repayment_percentage,
            'interest_rate': loan.interest_rate,
            'tenure': loan.tenure,
            'emi_amount': loan.emi_amount,
            'purpose': loan.purpose or '',
            'notes': loan.notes or '',
            'status': loan.status,
            'payments': payments
        }
        
        if loan.type == 'given':
            loans_given.append(loan_dict)
        else:
            loans_taken.append(loan_dict)
    
    # Calculate totals
    total_given = sum(l.amount for l in all_loans if l.type == 'given')
    total_taken = sum(l.amount for l in all_loans if l.type == 'taken')
    total_repaid = sum(l.repaid_amount for l in all_loans)
    
    return render_template("loan_accounts.html",
                         active='loan_accounts',
                         loans_given=loans_given,
                         loans_taken=loans_taken,
                         total_given=total_given,
                         total_taken=total_taken,
                         total_repaid=total_repaid,
                         today=today_ist().strftime('%Y-%m-%d'))


@app.route("/api/loan/save", methods=["POST"])
@login_required
@require_permission("loans", "create")
def save_loan():
    """Save a new loan"""
    cdb = get_cdb()
    company_id = get_current_company()
    data = request.get_json()
    
    try:
        loan = Loan(
            company_id=company_id,
            type=data.get('type'),
            party_name=data.get('party_name'),
            loan_date=date.fromisoformat(data.get('loan_date')),
            amount=data.get('amount'),
            interest_rate=data.get('interest_rate', 0),
            tenure=data.get('tenure', 12),
            emi_amount=data.get('emi_amount', 0),
            purpose=data.get('purpose', ''),
            notes=data.get('notes', ''),
            status='Active',
            created_by=get_current_user().get('email')
        )
        cdb.add(loan)
        cdb.commit()
        
        return jsonify({'success': True, 'message': 'Loan saved successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 400


@app.route("/api/loan/repayment/save", methods=["POST"])
@login_required
@require_permission("loans", "create")
def save_loan_repayment():
    """Save a loan repayment"""
    cdb = get_cdb()
    company_id = get_current_company()
    data = request.get_json()
    
    try:
        loan_id = data.get('loan_id')
        loan = cdb.query(Loan).filter_by(id=loan_id, company_id=company_id).first()
        
        if not loan:
            return jsonify({'success': False, 'message': 'Loan not found'}), 404
        
        repayment = LoanRepayment(
            loan_id=loan.id,
            date=date.fromisoformat(data.get('date')),
            amount=data.get('amount'),
            payment_mode=data.get('payment_mode', 'Cash'),
            reference=data.get('reference', ''),
            notes=data.get('notes', '')
        )
        cdb.add(repayment)
        
        # Update loan status if fully repaid
        if loan.remaining_amount - repayment.amount <= 0:
            loan.status = 'Completed'
        
        cdb.commit()
        
        return jsonify({'success': True, 'message': 'Repayment recorded successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 400

# ============================================
# LEDGER & TRIAL BALANCE ROUTES
# ============================================

@app.route("/ledger")
@login_required
@require_permission("analytics", "view")
def ledger():
    """General Ledger - shows all transactions with filters"""
    cdb = get_cdb()
    company_id = get_current_company()
    
    # Get filter parameters
    from_date_str = request.args.get('from_date', '')
    to_date_str = request.args.get('to_date', '')
    account_type = request.args.get('account_type', 'all')

    # A filter is only "applied" if the user actually submitted one.
    # Bare page load / no query args => full ledger, no date bound.
    filter_applied = bool(from_date_str or to_date_str or (account_type != 'all'))

    from_date = date.fromisoformat(from_date_str) if from_date_str else None
    to_date = date.fromisoformat(to_date_str) if to_date_str else None

    ledger_entries = []

    def _date_filters(date_col):
        conds = []
        if from_date:
            conds.append(date_col >= from_date)
        if to_date:
            conds.append(date_col <= to_date)
        return conds

    # 1. Sales Invoices
    invoices = cdb.query(Invoice).filter(
        Invoice.company_id == company_id,
        *_date_filters(Invoice.date)
    ).order_by(Invoice.date.asc()).all()
    
    for inv in invoices:
        client_name = inv.client_obj.name if inv.client_obj else (inv.contact_person or "Unknown")
        
        # Skip if filtering by account type
        if account_type != 'all' and account_type != 'sales':
            pass
        else:
            ledger_entries.append({
                'date': inv.date,
                'voucher_type': 'Sales Invoice',
                'voucher_no': inv.invoice_id,
                'party_name': client_name,
                'debit': inv.grand_total or 0,
                'credit': 0,
                'balance': 0,  # Will calculate running balance
                'type': 'sales'
            })
            
            # Add payment entries if paid
            paid_amount = (inv.grand_total or 0) - (getattr(inv, 'balance', 0) or 0)
            if paid_amount > 0:
                ledger_entries.append({
                    'date': inv.date,
                    'voucher_type': 'Payment Received',
                    'voucher_no': inv.invoice_id,
                    'party_name': client_name,
                    'debit': 0,
                    'credit': paid_amount,
                    'balance': 0,
                    'type': 'payment_received'
                })
    
    # 2. Purchase Invoices
    purchases = cdb.query(PurchaseInvoice).filter(
        PurchaseInvoice.company_id == company_id,
        *_date_filters(PurchaseInvoice.date)
    ).order_by(PurchaseInvoice.date.asc()).all()
    
    for pur in purchases:
        supplier_name = pur.supplier.name if pur.supplier else "Unknown"
        
        if account_type != 'all' and account_type != 'purchases':
            pass
        else:
            ledger_entries.append({
                'date': pur.date,
                'voucher_type': 'Purchase Invoice',
                'voucher_no': pur.invoice_number or pur.invoice_id,
                'party_name': supplier_name,
                'debit': 0,
                'credit': pur.grand_total or 0,
                'balance': 0,
                'type': 'purchases'
            })
            
            # Add payment entries if paid
            if pur.paid_amount and pur.paid_amount > 0:
                ledger_entries.append({
                    'date': pur.date,
                    'voucher_type': 'Payment Made',
                    'voucher_no': pur.invoice_number or pur.invoice_id,
                    'party_name': supplier_name,
                    'debit': pur.paid_amount,
                    'credit': 0,
                    'balance': 0,
                    'type': 'payment_made'
                })
    
    # 3. Expenses (if any expense table exists - you can add later)
    # 4. Bank transactions (if any bank table exists - you can add later)
    
    # Sort by date
    ledger_entries.sort(key=lambda x: x['date'])
    
    # Calculate running balance
    running_balance = 0
    for entry in ledger_entries:
        running_balance = running_balance + entry['debit'] - entry['credit']
        entry['balance'] = running_balance
    
    # Calculate totals
    total_debits = sum(e['debit'] for e in ledger_entries)
    total_credits = sum(e['credit'] for e in ledger_entries)
    closing_balance = running_balance
    
    return render_template('ledger.html',
                         ledger_entries=ledger_entries,
                         from_date=from_date,
                         to_date=to_date,
                         account_type=account_type,
                         filter_applied=filter_applied,
                         total_debits=total_debits,
                         total_credits=total_credits,
                         closing_balance=closing_balance,
                         active='ledger')


@app.route("/trial-balance")
@login_required
@require_permission("analytics", "view")
def trial_balance():
    """Trial Balance - shows all account balances"""
    cdb = get_cdb()
    company_id = get_current_company()
    
    # Get filter parameter
    as_on_date_str = request.args.get('as_on_date', '')
    
    if not as_on_date_str:
        as_on_date = today_ist()
    else:
        as_on_date = date.fromisoformat(as_on_date_str)
    
    accounts = {}
    
    # 1. Sales/Customers (Debtors)
    clients = cdb.query(Client).filter_by(company_id=company_id).all()
    for client in clients:
        # Calculate outstanding from invoices
        invoices = cdb.query(Invoice).filter_by(company_id=company_id, client_id=client.id).all()
        total_sales = sum(i.grand_total or 0 for i in invoices)
        total_paid = sum((i.grand_total or 0) - (getattr(i, 'balance', 0) or 0) for i in invoices)
        outstanding = total_sales - total_paid
        
        if outstanding != 0:
            accounts[f"Debtors - {client.name}"] = {
                'debit': outstanding if outstanding > 0 else 0,
                'credit': abs(outstanding) if outstanding < 0 else 0
            }
    
    # 2. Suppliers (Creditors)
    suppliers = cdb.query(Client).filter(
        Client.company_id == company_id,
        db.or_(Client.client_type == "Supplier", Client.client_type == "Both")
    ).all()
    
    for supplier in suppliers:
        purchases = cdb.query(PurchaseInvoice).filter_by(company_id=company_id, supplier_id=supplier.id).all()
        total_purchases = sum(p.grand_total or 0 for p in purchases)
        total_paid = sum(p.paid_amount or 0 for p in purchases)
        outstanding = total_purchases - total_paid
        
        if outstanding != 0:
            accounts[f"Creditors - {supplier.name}"] = {
                'debit': 0,
                'credit': outstanding if outstanding > 0 else 0
            }
    
    # 3. Sales Revenue
    all_invoices = cdb.query(Invoice).filter_by(company_id=company_id).all()
    total_revenue = sum(i.grand_total or 0 for i in all_invoices)
    if total_revenue > 0:
        accounts["Sales Revenue"] = {
            'debit': 0,
            'credit': total_revenue
        }
    
    # 4. Purchase Cost
    all_purchases = cdb.query(PurchaseInvoice).filter_by(company_id=company_id).all()
    total_purchase_cost = sum(p.grand_total or 0 for p in all_purchases)
    if total_purchase_cost > 0:
        accounts["Purchase Cost"] = {
            'debit': total_purchase_cost,
            'credit': 0
        }
    
    # 5. Stock/Inventory Value
    stock_items = cdb.query(StockItem).filter_by(company_id=company_id).all()
    total_stock_value = sum((s.purchase_rate or s.unit_price or 0) * s.quantity for s in stock_items)
    if total_stock_value > 0:
        accounts["Inventory"] = {
            'debit': total_stock_value,
            'credit': 0
        }
    
    # 6. GST Collected (from sales)
    total_gst_collected = sum(i.tax_amount or 0 for i in all_invoices)
    if total_gst_collected > 0:
        accounts["GST Collected (Output)"] = {
            'debit': 0,
            'credit': total_gst_collected
        }
    
    # 7. GST Paid (on purchases)
    total_gst_paid = sum(p.tax_amount or 0 for p in all_purchases)
    if total_gst_paid > 0:
        accounts["GST Paid (Input)"] = {
            'debit': total_gst_paid,
            'credit': 0
        }
    
    # Calculate totals
    total_debits = sum(acc['debit'] for acc in accounts.values())
    total_credits = sum(acc['credit'] for acc in accounts.values())
    
    # Convert to list for template
    account_list = [{'name': name, 'debit': data['debit'], 'credit': data['credit']} 
                    for name, data in accounts.items()]
    
    # Sort by name
    account_list.sort(key=lambda x: x['name'])
    
    return render_template('trial_balance.html',
                         accounts=account_list,
                         total_debits=total_debits,
                         total_credits=total_credits,
                         as_on_date=as_on_date,
                         difference=total_debits - total_credits,
                         active='trial_balance')

# ============================================
# REPORTS ROUTES
# ============================================

@app.route("/api/reports/sales-data")
@login_required
@require_permission("analytics", "view")
def api_sales_report_data():
    """API endpoint for sales report data"""
    cdb = get_cdb()
    if not cdb:
        return jsonify({"error": "Could not connect to company database"}), 500
    
    company_id = get_current_company()
    
    from_date_str = request.args.get('from_date', '')
    to_date_str = request.args.get('to_date', '')
    
    if not from_date_str:
        from_date = today_ist().replace(day=1)
    else:
        from_date = date.fromisoformat(from_date_str)
    
    if not to_date_str:
        to_date = today_ist()
    else:
        to_date = date.fromisoformat(to_date_str)
    
    # Get invoices (exclude draft)
    invoices = cdb.query(Invoice).filter(
        Invoice.company_id == company_id,
        Invoice.date >= from_date,
        Invoice.date <= to_date,
        Invoice.status.notin_(['Cancelled', 'Void'])
    ).order_by(Invoice.date.desc()).all()
    
    # Calculate totals
    total_revenue = sum(float(i.grand_total or 0) for i in invoices)
    total_tax = sum(float(i.tax_amount or 0) for i in invoices)
    total_pending = sum(float(getattr(i, 'balance', 0) or 0) for i in invoices)
    total_received = total_revenue - total_pending
    
    # Monthly trend
    monthly_revenue = {}
    for inv in invoices:
        month_key = inv.date.strftime('%b %Y')
        monthly_revenue[month_key] = monthly_revenue.get(month_key, 0) + float(inv.grand_total or 0)
    
    month_labels = list(monthly_revenue.keys())
    monthly_revenue_data = list(monthly_revenue.values())
    
    # Top destinations (from terms JSON)
    destinations = {}
    for inv in invoices:
        meta = {}
        if inv.terms:
            try:
                meta = json.loads(inv.terms)
            except:
                pass
        dest = meta.get('destination', 'Domestic')
        destinations[dest] = destinations.get(dest, 0) + 1
    
    top_destinations = [{'name': k, 'count': v} for k, v in sorted(destinations.items(), key=lambda x: x[1], reverse=True)[:5]]
    
    # Top products from invoice items
    products = {}
    for inv in invoices:
        for item in inv.items:
            name = item.description or item.code or 'Unknown'
            products[name] = products.get(name, 0) + float(item.qty or 0)
    
    top_products = [{'name': k, 'qty': v} for k, v in sorted(products.items(), key=lambda x: x[1], reverse=True)[:5]]
    
    # Top customers
    customers = {}
    for inv in invoices:
        name = inv.client_obj.name if inv.client_obj else (inv.contact_person or 'Unknown')
        customers[name] = customers.get(name, 0) + float(inv.grand_total or 0)
    
    top_customers = [{'name': k, 'amount': v} for k, v in sorted(customers.items(), key=lambda x: x[1], reverse=True)[:5]]
    
    # Status counts
    paid_count = sum(1 for i in invoices if i.status == 'Paid')
    partial_count = sum(1 for i in invoices if i.status == 'Partial')
    pending_count = sum(1 for i in invoices if i.status not in ['Paid', 'Partial'])
    
    # Invoice list for table
    invoice_list = []
    for inv in invoices[:50]:
        meta = {}
        if inv.terms:
            try:
                meta = json.loads(inv.terms)
            except:
                pass
        invoice_list.append({
            'id': inv.invoice_id,
            'date': inv.date.strftime('%d %b %Y'),
            'customer': inv.client_obj.name if inv.client_obj else (inv.contact_person or '—'),
            'destination': meta.get('destination', '—'),
            'subtotal': float(inv.subtotal or 0),
            'tax': float(inv.tax_amount or 0),
            'total': float(inv.grand_total or 0),
            'status': inv.status or 'Pending'
        })
    
    return jsonify({
        'total_revenue': total_revenue,
        'total_tax': total_tax,
        'total_received': total_received,
        'total_pending': total_pending,
        'total_invoices': len(invoices),
        'month_labels': month_labels,
        'monthly_revenue': monthly_revenue_data,
        'top_destinations': top_destinations,
        'top_products': top_products,
        'top_customers': top_customers,
        'paid_count': paid_count,
        'partial_count': partial_count,
        'pending_count': pending_count,
        'invoices': invoice_list
    })


@app.route("/api/reports/purchase-data")
@login_required
@require_permission("analytics", "view")
def api_purchase_report_data():
    """API endpoint for purchase report data"""
    cdb = get_cdb()
    if not cdb:
        return jsonify({"error": "Could not connect to company database"}), 500
    
    company_id = get_current_company()
    
    from_date_str = request.args.get('from_date', '')
    to_date_str = request.args.get('to_date', '')
    
    if not from_date_str:
        from_date = date(2000, 1, 1)   # Show all records by default
    else:
        from_date = date.fromisoformat(from_date_str)
    
    if not to_date_str:
        to_date = today_ist()
    else:
        to_date = date.fromisoformat(to_date_str)

    purchases = cdb.query(PurchaseInvoice).filter(
        PurchaseInvoice.company_id == company_id,
        PurchaseInvoice.date >= from_date,
        PurchaseInvoice.date <= to_date
    ).order_by(PurchaseInvoice.date.desc()).all()
    
    total_amount = sum(float(p.grand_total or 0) for p in purchases)
    total_gst = sum(float(p.tax_amount or 0) for p in purchases)
    total_paid = sum(float(p.paid_amount or 0) for p in purchases)
    total_pending = sum(float(p.balance or 0) for p in purchases)
    
    # Unique supplier count
    supplier_ids = set()
    for p in purchases:
        if p.supplier_id:
            supplier_ids.add(p.supplier_id)
    supplier_count = len(supplier_ids)
    
    # Monthly trend
    monthly_purchases = {}
    for p in purchases:
        month_key = p.date.strftime('%b %Y')
        monthly_purchases[month_key] = monthly_purchases.get(month_key, 0) + float(p.grand_total or 0)
    
    month_labels = list(monthly_purchases.keys())
    monthly_purchases_data = list(monthly_purchases.values())
    
    # Top suppliers
    suppliers = {}
    for p in purchases:
        try:
            name = p.supplier.name if p.supplier else (getattr(p, 'supplier_name', None) or 'Unknown')
        except Exception:
            name = getattr(p, 'supplier_name', None) or 'Unknown'
        suppliers[name] = suppliers.get(name, 0) + float(p.grand_total or 0)
    
    top_suppliers = [{'name': k, 'amount': v} for k, v in sorted(suppliers.items(), key=lambda x: x[1], reverse=True)[:5]]
    
    # Top purchased products
    products = {}
    for p in purchases:
        for item in p.items:
            name = item.description or 'Unknown'
            products[name] = products.get(name, 0) + float(item.quantity or 0)
    
    top_products = [{'name': k, 'qty': v} for k, v in sorted(products.items(), key=lambda x: x[1], reverse=True)[:5]]
    
    # Status counts
    paid_count = sum(1 for p in purchases if p.status == 'Paid')
    partial_count = sum(1 for p in purchases if p.status == 'Partial')
    pending_count = sum(1 for p in purchases if p.status not in ['Paid', 'Partial'])
    
    invoice_list = []
    for p in purchases[:50]:
        try:
            sup_name = p.supplier.name if p.supplier else (getattr(p, 'supplier_name', None) or '—')
        except Exception:
            sup_name = getattr(p, 'supplier_name', None) or '—'
        invoice_list.append({
            'id': p.invoice_id,
            'date': p.date.strftime('%d %b %Y'),
            'supplier': sup_name,
            'subtotal': float(p.subtotal or 0),
            'tax': float(p.tax_amount or 0),
            'total': float(p.grand_total or 0),
            'status': p.status or 'Pending'
        })
    
    return jsonify({
        'total_amount': total_amount,
        'total_gst': total_gst,
        'total_paid': total_paid,
        'total_pending': total_pending,
        'supplier_count': supplier_count,
        'month_labels': month_labels,
        'monthly_purchases': monthly_purchases_data,
        'top_suppliers': top_suppliers,
        'top_products': top_products,
        'paid_count': paid_count,
        'partial_count': partial_count,
        'pending_count': pending_count,
        'invoices': invoice_list
    })

@app.route("/api/reports/stock-data")
@login_required
@require_permission("analytics", "view")
def api_stock_report_data():
    """API endpoint for stock report data - reads directly from StockItem table"""
    cdb = get_cdb()
    if not cdb:
        return jsonify({"error": "Could not connect to company database"}), 500
    
    company_id = get_current_company()
    
    # Get ALL stock items directly from StockItem table
    stock_items = cdb.query(StockItem).filter_by(company_id=company_id).all()
    
    total_items = len(stock_items)
    
    # Calculate total value using purchase_rate or unit_price
    total_value = 0
    in_stock = 0
    low_stock = 0
    out_stock = 0
    
    categories = {}
    stock_list = []
    
    for s in stock_items:
        qty = float(s.quantity or 0)
        price = float(s.purchase_rate or s.unit_price or 0)
        total_value += price * qty
        
        # Status
        reorder = float(s.reorder_level or 10)
        if qty <= 0:
            out_stock += 1
            status = 'out'
            status_label = 'Out of Stock'
        elif qty <= reorder:
            low_stock += 1
            status = 'low'
            status_label = 'Low Stock'
        else:
            in_stock += 1
            status = 'in'
            status_label = 'In Stock'
        
        # Category
        cat = s.category or 'Uncategorized'
        categories[cat] = categories.get(cat, 0) + 1
        
        stock_list.append({
            'code': s.code,
            'name': s.name,
            'category': s.category or '—',
            'quantity': int(qty),
            'price': price,
            'total': price * qty,
            'status': status,
            'status_label': status_label,
            'unit': s.unit or 'pcs',
            'reorder_level': int(reorder)
        })
    
    # Top selling items (from InvoiceItem table - sales data)
    top_selling = {}
    invoices = cdb.query(Invoice).filter_by(company_id=company_id).all()
    for inv in invoices:
        for item in inv.items:
            name = item.description or item.code or 'Unknown'
            top_selling[name] = top_selling.get(name, 0) + float(item.qty or 0)
    
    top_selling_list = [{'name': k, 'qty': v} for k, v in sorted(top_selling.items(), key=lambda x: x[1], reverse=True)[:10]]
    
    return jsonify({
        'total_items': total_items,
        'total_value': total_value,
        'in_stock': in_stock,
        'low_stock': low_stock,
        'out_stock': out_stock,
        'category_count': len(categories),
        'categories': [{'name': k, 'count': v} for k, v in categories.items()],
        'top_selling': top_selling_list,
        'stock_items': stock_list
    })

@app.route("/api/reports/tax-data")
@login_required
@require_permission("analytics", "view")
def api_tax_report_data():
    cdb = get_cdb()
    if not cdb:
        return jsonify({"error": "Could not connect to company database"}), 500
    
    company_id = get_current_company()
    
    from_date_str = request.args.get('from_date', '')
    to_date_str = request.args.get('to_date', '')
    
    if not from_date_str:
        from_date = today_ist().replace(day=1)
    else:
        from_date = date.fromisoformat(from_date_str)
    
    if not to_date_str:
        to_date = today_ist()
    else:
        to_date = date.fromisoformat(to_date_str)
    
    sales = cdb.query(Invoice).filter(
        Invoice.company_id == company_id,
        Invoice.date >= from_date,
        Invoice.date <= to_date,
        Invoice.status.notin_(['Cancelled', 'Void'])
    ).all()
    
    purchases = cdb.query(PurchaseInvoice).filter(
        PurchaseInvoice.company_id == company_id,
        PurchaseInvoice.date >= from_date,
        PurchaseInvoice.date <= to_date
    ).all()
    
    output_gst = sum(float(i.tax_amount or 0) for i in sales)
    input_gst = sum(float(p.tax_amount or 0) for p in purchases)
    net_gst = output_gst - input_gst
    
    total_sales = sum(float(i.grand_total or 0) for i in sales)
    effective_rate = (net_gst / total_sales * 100) if total_sales > 0 else 0
    
    # Monthly GST
    monthly_gst = {}
    for inv in sales:
        month_key = inv.date.strftime('%b %Y')
        monthly_gst[month_key] = monthly_gst.get(month_key, 0) + float(inv.tax_amount or 0)
    
    # HSN Summary
    hsn_summary = []
    hsn_dict = {}
    for inv in sales:
        for item in inv.items:
            hsn = (item.code or 'Other')[:6] if item.code else 'Other'
            if hsn not in hsn_dict:
                hsn_dict[hsn] = {'hsn': hsn, 'description': item.description or '', 'quantity': 0, 'value': 0, 'rate': 18, 'cgst': 0, 'sgst': 0, 'total': 0}
            qty = float(item.qty or 0)
            rate = float(item.rate or 0)
            amount = qty * rate
            gst = amount * 0.18
            hsn_dict[hsn]['quantity'] += qty
            hsn_dict[hsn]['value'] += amount
            hsn_dict[hsn]['cgst'] += gst / 2
            hsn_dict[hsn]['sgst'] += gst / 2
            hsn_dict[hsn]['total'] += gst
    hsn_summary = list(hsn_dict.values())
    
    return jsonify({
        'output_gst': output_gst,
        'input_gst': input_gst,
        'net_gst': net_gst,
        'effective_rate': round(effective_rate, 2),
        'month_labels': list(monthly_gst.keys()),
        'monthly_gst': list(monthly_gst.values()),
        'cgst': output_gst / 2,
        'sgst': output_gst / 2,
        'igst': 0,
        'hsn_summary': hsn_summary
    })


@app.route("/api/reports/financial-data")
@login_required
@require_permission("analytics", "view")
def api_financial_report_data():
    """API endpoint for financial report data"""
    cdb = get_cdb()
    if not cdb:
        return jsonify({"error": "Could not connect to company database"}), 500
    
    company_id = get_current_company()
    
    from_date_str = request.args.get('from_date', '')
    to_date_str = request.args.get('to_date', '')
    
    if not from_date_str:
        from_date = today_ist().replace(day=1)
    else:
        from_date = date.fromisoformat(from_date_str)
    
    if not to_date_str:
        to_date = today_ist()
    else:
        to_date = date.fromisoformat(to_date_str)
    
    # ── INCOME ───────────────────────────────────────────────────────────────
    # 1. Sales Revenue
    sales = cdb.query(Invoice).filter(
        Invoice.company_id == company_id,
        Invoice.date >= from_date,
        Invoice.date <= to_date,
        Invoice.status.notin_(['Cancelled', 'Void'])
    ).all()
    sales_income = sum(float(i.grand_total or 0) for i in sales)
    
    # 2. Other Income (Cash Transactions - income type)
    other_income = cdb.query(CashTransaction).filter(
        CashTransaction.company_id == company_id,
        CashTransaction.type == 'income',
        CashTransaction.date >= from_date,
        CashTransaction.date <= to_date
    ).all()
    other_income_total = sum(t.amount for t in other_income)
    
    total_income = sales_income + other_income_total
    
    # ── EXPENSES ─────────────────────────────────────────────────────────────
    # 1. Cost of Goods Sold (Purchases)
    purchases = cdb.query(PurchaseInvoice).filter(
        PurchaseInvoice.company_id == company_id,
        PurchaseInvoice.date >= from_date,
        PurchaseInvoice.date <= to_date
    ).all()
    purchase_expense = sum(float(p.grand_total or 0) for p in purchases)
    
    # 2. Operating Expenses (from Expense table) ← FIXED!
    operating_expenses = cdb.query(Expense).filter(
        Expense.company_id == company_id,
        Expense.date >= from_date,
        Expense.date <= to_date
    ).all()
    operating_expense_total = sum(e.amount for e in operating_expenses)
    
    # 3. Cash Transaction Expenses (if any - but these should be in Expense table)
    cash_expenses = cdb.query(CashTransaction).filter(
        CashTransaction.company_id == company_id,
        CashTransaction.type == 'expense',
        CashTransaction.date >= from_date,
        CashTransaction.date <= to_date
    ).all()
    cash_expense_total = sum(t.amount for t in cash_expenses)
    
    # Total Expenses = Purchases + Operating Expenses + Cash Expenses
    total_expenses = purchase_expense + operating_expense_total + cash_expense_total
    
    # ── PROFIT ────────────────────────────────────────────────────────────────
    net_profit = total_income - total_expenses
    profit_margin = (net_profit / total_income * 100) if total_income > 0 else 0
    
    # ── MONTHLY BREAKDOWN ────────────────────────────────────────────────────
    monthly_income = {}
    monthly_expenses = {}
    
    # Monthly income from sales
    for inv in sales:
        month_key = inv.date.strftime('%b %Y')
        monthly_income[month_key] = monthly_income.get(month_key, 0) + float(inv.grand_total or 0)
    
    # Monthly expenses from purchases
    for p in purchases:
        month_key = p.date.strftime('%b %Y')
        monthly_expenses[month_key] = monthly_expenses.get(month_key, 0) + float(p.grand_total or 0)
    
    # Monthly expenses from Expense table ← FIXED!
    for e in operating_expenses:
        month_key = e.date.strftime('%b %Y')
        monthly_expenses[month_key] = monthly_expenses.get(month_key, 0) + e.amount
    
    # Monthly expenses from CashTransaction expenses
    for t in cash_expenses:
        month_key = t.date.strftime('%b %Y')
        monthly_expenses[month_key] = monthly_expenses.get(month_key, 0) + t.amount
    
    all_months = set(monthly_income.keys()) | set(monthly_expenses.keys())
    sorted_months = sorted(all_months, key=lambda x: datetime.strptime(x, '%b %Y'))
    
    monthly_profit = {}
    for m in sorted_months:
        monthly_profit[m] = monthly_income.get(m, 0) - monthly_expenses.get(m, 0)
    
    # ── CASH AND BANK BALANCES ──────────────────────────────────────────────
    # Cash balance from CashTransaction
    all_cash_txns = cdb.query(CashTransaction).filter_by(company_id=company_id).all()
    cash_balance = sum(t.amount for t in all_cash_txns if t.type == 'income') - sum(t.amount for t in all_cash_txns if t.type == 'expense')
    
    # Bank balance
    bank_accounts = cdb.query(BankAccount).filter_by(company_id=company_id, status='Active').all()
    bank_balance = sum(acc.balance for acc in bank_accounts)
    
    # ── EXPENSE BREAKDOWN BY CATEGORY ──────────────────────────────────────
    expense_breakdown = {}
    
    # From Expense table ← FIXED!
    for e in operating_expenses:
        expense_breakdown[e.category] = expense_breakdown.get(e.category, 0) + e.amount
    
    # Add purchases as a category
    if purchase_expense > 0:
        expense_breakdown['Purchases (COGS)'] = purchase_expense
    
    # Add cash expenses by category
    for t in cash_expenses:
        cat = t.category or 'Misc'
        expense_breakdown[cat] = expense_breakdown.get(cat, 0) + t.amount
    
    # ── CASH FLOW ENTRIES ────────────────────────────────────────────────────
    cashflow = []
    
    # Income entries
    for inv in sales[:20]:
        cashflow.append({
            'date': inv.date.strftime('%d %b %Y'),
            'type': 'income',
            'category': 'Sales',
            'description': f"Invoice {inv.invoice_id}",
            'amount': float(inv.grand_total or 0),
            'mode': 'Credit'
        })
    
    # Expense entries from Expense table ← FIXED!
    for e in operating_expenses[:20]:
        cashflow.append({
            'date': e.date.strftime('%d %b %Y'),
            'type': 'expense',
            'category': e.category,
            'description': e.description or e.category,
            'amount': e.amount,
            'mode': e.payment_mode or 'Cash'
        })
    
    # Cash transaction expenses
    for t in cash_expenses[:10]:
        cashflow.append({
            'date': t.date.strftime('%d %b %Y'),
            'type': 'expense',
            'category': t.category,
            'description': t.description,
            'amount': t.amount,
            'mode': 'Cash'
        })
    
    cashflow.sort(key=lambda x: x['date'], reverse=True)
    
    return jsonify({
        'total_income': total_income,
        'total_expenses': total_expenses,
        'net_profit': net_profit,
        'profit_margin': round(profit_margin, 2),
        'month_labels': sorted_months,
        'monthly_income': [monthly_income.get(m, 0) for m in sorted_months],
        'monthly_expenses': [monthly_expenses.get(m, 0) for m in sorted_months],
        'monthly_profit': [monthly_profit.get(m, 0) for m in sorted_months],
        'cash_balance': cash_balance,
        'bank_balance': bank_balance,
        'expense_breakdown': expense_breakdown,
        'cashflow': cashflow
    })

@app.route("/reports/profit-loss")
@login_required
@require_permission("analytics", "view")
def profit_loss():
    """Profit & Loss Statement"""
    cdb = get_cdb()
    company_id = get_current_company()
    
    # Get filter parameters
    from_date_str = request.args.get('from_date', '')
    to_date_str = request.args.get('to_date', '')
    period = request.args.get('period', 'custom')
    
    # Set date range based on period
    if period == 'month':
        from_date = today_ist().replace(day=1)
        to_date = today_ist()
    elif period == 'quarter':
        current_month = today_ist().month
        if current_month <= 3:
            from_date = date(today_ist().year, 1, 1)
        elif current_month <= 6:
            from_date = date(today_ist().year, 4, 1)
        elif current_month <= 9:
            from_date = date(today_ist().year, 7, 1)
        else:
            from_date = date(today_ist().year, 10, 1)
        to_date = today_ist()
    elif period == 'year':
        from_date = date(today_ist().year, 1, 1)
        to_date = today_ist()
    else:
        if not from_date_str:
            from_date = today_ist().replace(day=1)
        else:
            from_date = date.fromisoformat(from_date_str)
        
        if not to_date_str:
            to_date = today_ist()
        else:
            to_date = date.fromisoformat(to_date_str)
    
    # INCOME: Sales Revenue
    sales_invoices = cdb.query(Invoice).filter(
        Invoice.company_id == company_id,
        Invoice.date >= from_date,
        Invoice.date <= to_date,
        Invoice.status.notin_(['Cancelled', 'Void'])
    ).all()
    
    total_revenue = sum(i.grand_total or 0 for i in sales_invoices)
    
    # EXPENSES: Purchase Cost
    purchase_invoices = cdb.query(PurchaseInvoice).filter(
        PurchaseInvoice.company_id == company_id,
        PurchaseInvoice.date >= from_date,
        PurchaseInvoice.date <= to_date,
        PurchaseInvoice.status.notin_(['Cancelled', 'Void'])
    ).all()
    
    cost_of_goods_sold = sum(p.grand_total or 0 for p in purchase_invoices)
    
    # GROSS PROFIT
    gross_profit = total_revenue - cost_of_goods_sold
    
    # Calculate other income (cash transactions)
    cash_income = cdb.query(CashTransaction).filter(
        CashTransaction.company_id == company_id,
        CashTransaction.type == 'income',
        CashTransaction.date >= from_date,
        CashTransaction.date <= to_date
    ).all()
    other_income = sum(i.amount for i in cash_income)
    
    # Calculate expenses (cash transactions)
    cash_expenses = cdb.query(CashTransaction).filter(
        CashTransaction.company_id == company_id,
        CashTransaction.type == 'expense',
        CashTransaction.date >= from_date,
        CashTransaction.date <= to_date
    ).all()
    
    # Categorize expenses
    expense_categories = {}
    for exp in cash_expenses:
        if exp.category not in expense_categories:
            expense_categories[exp.category] = 0
        expense_categories[exp.category] += exp.amount
    
    total_expenses = sum(expense_categories.values())
    
    # NET PROFIT
    net_profit = gross_profit + other_income - total_expenses
    
    # Calculate ratios
    gross_margin = (gross_profit / total_revenue * 100) if total_revenue > 0 else 0
    net_margin = (net_profit / total_revenue * 100) if total_revenue > 0 else 0
    
    # Monthly profit trend
    monthly_profit = {}
    all_months = set()
    
    for inv in sales_invoices:
        month_key = inv.date.strftime('%Y-%m')
        all_months.add(month_key)
    
    for pur in purchase_invoices:
        month_key = pur.date.strftime('%Y-%m')
        all_months.add(month_key)
    
    for month in sorted(all_months):
        month_date = datetime.strptime(month, '%Y-%m')
        monthly_profit[month] = {
            'month': month_date.strftime('%b %Y'),
            'revenue': 0,
            'expenses': 0,
            'profit': 0
        }
    
    for inv in sales_invoices:
        month_key = inv.date.strftime('%Y-%m')
        monthly_profit[month_key]['revenue'] += inv.grand_total or 0
    
    for exp in cash_expenses:
        month_key = exp.date.strftime('%Y-%m')
        if month_key in monthly_profit:
            monthly_profit[month_key]['expenses'] += exp.amount
    
    for month in monthly_profit:
        monthly_profit[month]['profit'] = monthly_profit[month]['revenue'] - monthly_profit[month]['expenses']
    
    profit_trend = list(monthly_profit.values())
    
    return render_template("profit_loss.html",
                         active='profit_loss',
                         from_date=from_date,
                         to_date=to_date,
                         period=period,
                         total_revenue=total_revenue,
                         cost_of_goods_sold=cost_of_goods_sold,
                         gross_profit=gross_profit,
                         other_income=other_income,
                         expense_categories=expense_categories,
                         total_expenses=total_expenses,
                         net_profit=net_profit,
                         gross_margin=gross_margin,
                         net_margin=net_margin,
                         profit_trend=profit_trend,
                         today=today_ist())

# ============================================
# SYNC, SHARE & BACKUP ROUTES
# ============================================

@app.route("/sync")
@login_required
def sync_data():
    """Sync data with cloud"""
    company_id = get_current_company()
    return render_template("sync.html", active='sync')


@app.route("/share")
@login_required
def share_data():
    """Share data with others"""
    company_id = get_current_company()
    return render_template("share.html", active='share')

# ============================================
# OTHER PRODUCTS ROUTES
# ============================================

@app.route("/integrations")
@login_required
def integrations():
    """Third-party integrations"""
    company_id = get_current_company()
    return render_template("integrations.html", active='integrations')

@app.route("/addons")
@login_required
def addons():
    """Add-ons marketplace"""
    company_id = get_current_company()
    return render_template("addons.html", active='addons')

# ============================================
# UTILITIES ROUTES
# ============================================

@app.route("/import")
@login_required
def import_data():
    """Import data from files"""
    company_id = get_current_company()
    return render_template("import.html", active='import')

@app.route("/export")
@login_required
def export_data():
    """Export data to files"""
    company_id = get_current_company()
    return render_template("export.html", active='export')

@app.route("/audit-log")
@login_required
def audit_log():
    """View audit logs"""
    company_id = get_current_company()
    return render_template("audit_log.html", active='audit')

# ─────────────────────────────────────────────────────────────────────────────
# ── Profile ───────────────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────
@app.route("/profile")
@login_required
def profile():
    user = get_current_user()
    return render_template("profile.html", user=user)


# ─────────────────────────────────────────────────────────────────────────────
# ── Company Settings ──────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────


@app.route("/company/settings")
@login_required
@owner_required
def company_settings():
    company_id = get_current_company()
    company = get_company_by_id(company_id)
    owner_email = get_current_user().get("email", "").strip().lower()

    # Use customer database session to query CompanyUser (this company only, for the users table)
    cdb = get_cdb()
    if not cdb:
        flash("Could not connect to company database")
        return redirect(url_for("dashboard"))

    users = cdb.query(CompanyUser).filter_by(company_id=company_id).all()

    # All companies this owner has, for the "grant access to" checkbox list
    owner_companies = get_owner_companies(owner_email)

    # Owner-wide seat usage (across ALL of the owner's companies, not just this one)
    owner_user_count, owner_max_users, owner_user_emails = get_owner_user_stats(owner_email)

    # For each distinct user email under this owner, which companies + role do they have?
    # { email: {"full_name":..., "companies": [{"company_id":, "company_name":, "role":}, ...]} }
    user_access_map = {}
    for c in owner_companies:
        try:
            _cdb = get_customer_session(c.company_id)
            for u in _cdb.query(CompanyUser).filter_by(is_active=True).all():
                key = (u.email or "").strip().lower()
                if not key:
                    continue
                entry = user_access_map.setdefault(key, {"full_name": u.full_name, "companies": []})
                entry["companies"].append({
                    "company_id": c.company_id,
                    "company_name": c.company_name,
                    "role": u.role,
                })
        except Exception as e:
            print(f"⚠  Could not read users for {c.company_id}: {e}")

    # Fix: Convert plans to a dictionary with proper structure
    plans = {}
    for p in SubscriptionPlan.query.all():
        plans[p.id] = {
            "name":          p.name,
            "price":         p.price,
            "max_companies": p.max_companies,
            "max_users":     p.max_users,
            "features":      p.features.split(",") if p.features else [],
        }

    current_plan = plans.get(company.subscription_plan) if company else None

    # ── Permission matrix data for the "Access" tab ──────────────────────────
    role_permissions = {
        role: perms_module.get_effective_permissions(
            role, company_id, None, cdb, CompanyRolePermission, CompanyUser
        )
        for role in ("employee", "accountant", "manager")
    }
    # Per-user overrides: only non-owner users get a row here
    user_permissions = {}
    for u in users:
        if u.role in ("owner", "super_admin"):
            continue
        user_permissions[u.user_id] = perms_module.get_effective_permissions(
            u.role, company_id, u.user_id, cdb, CompanyRolePermission, CompanyUser
        )

    # ── Field-level permissions for the "Access" tab ──────────────────────────
    # Import INVOICE_FIELDS and DEFAULT_FIELD_PERMISSIONS from permissions module
    from permissions import INVOICE_FIELDS, DEFAULT_FIELD_PERMISSIONS, get_field_permissions
    
    # Get field permissions for each role (role defaults)
    role_field_permissions = {
        role: get_field_permissions(role, None, company_id, cdb)
        for role in ("employee", "accountant", "manager")
    }
    
    # Get field permissions for each user (user overrides)
    user_field_permissions = {}
    for u in users:
        if u.role in ("owner", "super_admin"):
            continue
        user_field_permissions[u.user_id] = get_field_permissions(
            u.role, u.user_id, company_id, cdb
        )

    # ── API Keys ──────────────────────────────────────────────────────────────
    from platform_models import CompanyApiKey
    api_keys = CompanyApiKey.query.filter_by(company_id=company_id, is_active=True).order_by(CompanyApiKey.created_at.desc()).all()

    return render_template("company_settings.html",
                           company=company,
                           users=users,
                           plans=plans,
                           current_plan=current_plan,
                           owner_companies=owner_companies,
                           owner_user_count=owner_user_count,
                           owner_max_users=owner_max_users,
                           user_access_map=user_access_map,
                           allowed_roles=ALLOWED_COMPANY_ROLES,
                           perm_modules=perms_module.MODULES,
                           perm_actions=perms_module.ACTIONS,
                           perm_labels=perms_module.MODULE_LABELS,
                           role_permissions=role_permissions,
                           user_permissions=user_permissions,
                           invoice_fields=INVOICE_FIELDS,  # ← ADDED
                           role_field_permissions=role_field_permissions,  # ← ADDED
                           user_field_permissions=user_field_permissions,  # ← ADDED
                           api_keys=api_keys,  # ← ADDED
                           )

def _read_permission_matrix_from_form():
    matrix = {}
    for module in perms_module.MODULES:
        matrix[module] = {
            action: (request.form.get(f"perm__{module}__{action}") == "on")
            for action in perms_module.ACTIONS
        }
    return matrix


@app.route("/company/permissions/role/<role>", methods=["POST"])
@login_required
@owner_required
def save_role_permissions(role):
    if role not in ("employee", "accountant", "manager"):
        flash("Invalid role")
        return redirect(url_for("company_settings"))
    company_id = get_current_company()
    cdb = get_customer_session(company_id)
    row = cdb.query(CompanyRolePermission).filter_by(company_id=company_id, role=role).first()
    if not row:
        row = CompanyRolePermission(company_id=company_id, role=role)
        cdb.add(row)
    row.permissions_json = json.dumps(_read_permission_matrix_from_form())
    row.updated_at = datetime.utcnow()
    cdb.commit()
    flash(f"{role.title()} access updated")
    return redirect(url_for("company_settings"))


@app.route("/company/permissions/user/<user_id>", methods=["POST"])
@login_required
@owner_required
def save_user_permissions(user_id):
    company_id = get_current_company()
    cdb = get_customer_session(company_id)
    cu = cdb.query(CompanyUser).filter_by(user_id=user_id, company_id=company_id).first()
    if not cu:
        flash("User not found")
        return redirect(url_for("company_settings"))
    if cu.role in ("owner", "super_admin"):
        flash("Owner access can't be limited this way")
        return redirect(url_for("company_settings"))
    cu.permission_overrides = json.dumps(_read_permission_matrix_from_form())
    cdb.commit()
    flash(f"Access updated for {cu.full_name}")
    return redirect(url_for("company_settings"))

# Add to app.py

@app.route("/settings/whatsapp", methods=["GET", "POST"])
@login_required
@owner_required
def whatsapp_settings():
    from whatsapp_service import encrypt_secret
    from platform_models import WhatsAppTemplate
    import json as _json

    company_id = get_current_company()
    company = get_company_by_id(company_id)
    if not company:
        flash("Company not found")
        return redirect(url_for("dashboard"))

    # (template_key, form field prefix)
    EVENTS = [
        ("invoice_created", "tpl_invoice_created"),
        ("invoice_updated", "tpl_invoice_updated"),
        ("tracking_number_updated", "tpl_tracking_number_updated"),
    ]

    if request.method == "POST":
        raw_key = request.form.get("whatsapp_api_key", "").strip()
        provider = request.form.get("whatsapp_provider", "").strip()
        base_url = request.form.get("whatsapp_base_url", "").strip()
        
        # Save API Key (encrypted)
        if raw_key:
            company.whatsapp_api_key = encrypt_secret(raw_key)
            company.whatsapp_enabled = True
        elif not company.whatsapp_api_key:
            company.whatsapp_enabled = False
        
        # Save provider
        if provider:
            company.whatsapp_provider = provider
        
        # Save Base URL (NEW!)
        if base_url:
            company.whatsapp_base_url = base_url
        elif not base_url and company.whatsapp_provider == 'mobicomm':
            # Set default for MobiCOMM
            company.whatsapp_base_url = 'https://api.dovesoft.io/REST/directApi/message'
        
        # Save Business Number
        business_no = request.form.get("whatsapp_business_no", "").strip()
        if business_no:
            company.whatsapp_business_no = business_no

        # Save Templates — one WhatsAppTemplate row per event. template_name
        # is a free-text Meta/provider template name (must match the approved
        # name exactly). The variable list is fixed (not user-editable) —
        # order here = param order sent to the provider (slot 1, slot 2, ...):
        #   1. receiver_name -> client.name
        #   2. docket_no     -> invoice.docket_no
        #   3. date          -> invoice.date
        #   4. phone         -> company.phone
        HARDCODED_VARS = [
            "{{ receiver.name }}",
            "{{ invoice.docket_no }}",
            "{{ invoice.date }}",
            "{{ company.phone }}",
        ]
        for event_key, prefix in EVENTS:
            tpl_name = request.form.get(f"{prefix}_name", "").strip()

            tpl = WhatsAppTemplate.query.filter_by(
                company_id=company_id, template_key=event_key
            ).first()

            if not tpl_name:
                # Blank name = "not configured" for this event. Deactivate
                # rather than delete, so re-adding the name later doesn't
                # require re-typing anything.
                if tpl:
                    tpl.is_active = False
                continue

            if not tpl:
                tpl = WhatsAppTemplate(company_id=company_id, template_key=event_key, template_name=tpl_name)
                db.session.add(tpl)

            tpl.template_name = tpl_name
            tpl.variables_json = _json.dumps(HARDCODED_VARS)
            tpl.param_count = len(HARDCODED_VARS)
            tpl.is_active = True

        db.session.commit()
        flash("WhatsApp Connect settings saved.")
        return redirect(url_for("whatsapp_settings"))

    # GET: load existing per-event template name to pre-fill the form
    existing = {}
    for event_key, prefix in EVENTS:
        tpl = WhatsAppTemplate.query.filter_by(
            company_id=company_id, template_key=event_key, is_active=True
        ).first()
        existing[event_key] = {
            "prefix": prefix,
            "template_name": tpl.template_name if tpl else "",
        }

    return render_template(
        "settings_whatsapp.html",
        company=company,
        has_key=bool(company.whatsapp_api_key),
        active="whatsapp_settings",
        events=existing,
    )


@app.route("/settings/whatsapp/disconnect", methods=["POST"])
@login_required
@owner_required
def whatsapp_disconnect():
    company_id = get_current_company()
    company = get_company_by_id(company_id)
    if company:
        company.whatsapp_api_key = None
        company.whatsapp_enabled = False
        company.whatsapp_provider = None
        company.whatsapp_base_url = None
        db.session.commit()
        flash("WhatsApp disconnected.")
    return redirect(url_for("whatsapp_settings"))


@app.route("/settings/whatsapp/test", methods=["POST"])
@login_required
@owner_required
def whatsapp_test():
    from whatsapp_service import _send_whatsapp_template as send_whatsapp_template, decrypt_secret
    from datetime import datetime

    company_id = get_current_company()
    company = get_company_by_id(company_id)
    
    if not company or not company.whatsapp_enabled:
        flash("WhatsApp is not configured.")
        return redirect(url_for("whatsapp_settings"))

    test_phone = request.form.get("test_phone", "").strip()
    if not test_phone:
        flash("Please enter a phone number.")
        return redirect(url_for("whatsapp_settings"))

    # Format phone
    test_phone = ''.join(filter(str.isdigit, test_phone))
    if len(test_phone) == 10:
        test_phone = "91" + test_phone

    # Send test using configured template
    template_name = company.whatsapp_template_generate or "alhamamd1001"
    
    result = send_whatsapp_template(
        company=company,
        to_number=test_phone,
        template_name=template_name,
        params=[
            "TEST123",  # docket
            datetime.now().strftime("%d-%b-%Y"),  # date
            company.phone or "9876543210",  # phone
        ]
    )
    
    if result.get('success'):
        flash(f"✅ Test message sent successfully to {test_phone}!")
    else:
        flash(f"❌ Test failed: {result.get('error')}")
    
    return redirect(url_for("whatsapp_settings"))

@app.route("/company/update-info", methods=["POST"])
@login_required
@owner_required
def update_company_info():
    company_id = get_current_company()
    company    = get_company_by_id(company_id)
    if company:
        company.company_name = request.form.get("company_name", company.company_name).strip()
        company.address      = request.form.get("address",      company.address)
        company.phone        = request.form.get("phone",        company.phone)
        company.mobile       = request.form.get("mobile",       company.mobile)
        company.slogan       = request.form.get("slogan",       company.slogan)
        company.website      = request.form.get("website",      company.website)
        company.email        = request.form.get("email",        company.email)
        company.extra_info   = request.form.get("extra_info",   company.extra_info)

        # ── Public tracking page slug ──
        # Normalize so "Acme Logistics!" -> "acme-logistics" instead of rejecting
        # the input outright — most owners won't type a URL-safe string unprompted.
        raw_slug = request.form.get("public_slug", "").strip()
        if raw_slug:
            slug = re.sub(r"[^a-z0-9-]", "-", raw_slug.lower())
            slug = re.sub(r"-+", "-", slug).strip("-")
            existing = Company.query.filter(
                Company.public_slug == slug, Company.id != company.id
            ).first()
            if existing:
                flash(f"The slug '{slug}' is already used by another company — pick a different one. "
                      f"Everything else on this form was still saved.")
            else:
                company.public_slug = slug
        # Blank means "explicitly clear it" — an empty tracking link is not useful,
        # so clearing just disables the public page rather than leaving a broken link live.
        elif "public_slug" in request.form:
            company.public_slug = None

        # ── Address Visibility (per print format) ──
        company.show_address_customer_invoice = "show_address_customer_invoice" in request.form
        company.show_address_awb_invoice = "show_address_awb_invoice" in request.form
        company.show_address_performa_invoice = "show_address_performa_invoice" in request.form
        company.show_address_box_label = "show_address_box_label" in request.form
        company.show_address_shipping_label = "show_address_shipping_label" in request.form
        company.show_manifest_checkboxes = "show_manifest_checkboxes" in request.form

        # ── Logo upload ──
        logo_file = request.files.get("logo")
        if logo_file and logo_file.filename:
            if allowed_logo_file(logo_file.filename):
                ext = logo_file.filename.rsplit('.', 1)[1].lower()
                new_filename = f"{company.company_id}.{ext}"
                # Remove any old logo with a different extension so stale files don't pile up
                if getattr(company, 'logo_filename', None) and company.logo_filename != new_filename:
                    old_path = os.path.join(LOGO_UPLOAD_FOLDER, company.logo_filename)
                    if os.path.exists(old_path):
                        os.remove(old_path)
                logo_file.save(os.path.join(LOGO_UPLOAD_FOLDER, new_filename))
                company.logo_filename = new_filename
            else:
                flash("Logo must be a PNG, JPG, or WEBP file.")

        is_gst = request.form.get("is_gst_registered", "1") == "1"
        company.is_gst_registered = is_gst
        company.gst_number = request.form.get('gst_number', '').strip() if is_gst else None

        # Blank is a valid, explicit choice here (means "no AWB prefix"),
        # so we do NOT fall back to the old value like before.
        if "awb_prefix" in request.form:
            company.awb_prefix = request.form.get("awb_prefix", "").strip().upper()
        try:
            company.awb_start = int(request.form.get("awb_start", company.awb_start))
        except (ValueError, TypeError):
            pass

        # ── Credit limit behaviour: 'warn' (flash only) or 'block' (refuse the save) ──
        _credit_action = request.form.get("credit_limit_action", "warn")
        _credit_action = _credit_action if _credit_action in ("warn", "block") else "warn"
        company.credit_limit_action = _credit_action

        # ── Customer invoice PDF template ──
        _invoice_template = request.form.get("invoice_template", "classic")
        _invoice_template = _invoice_template if _invoice_template in ("classic", "modern", "minimal", "tally_style") else "classic"
        company.invoice_template = _invoice_template

        db.session.commit()

        # ── Verify it actually persisted ──────────────────────────────────────
        # Setting an attribute on an ORM object and committing succeeds even
        # if that attribute isn't a real mapped column — SQLAlchemy just
        # drops it silently, no error. That would look EXACTLY like this
        # setting "not working": the dropdown shows your choice right after
        # saving (because the in-memory object still has it), but it never
        # actually reached the database, so every later request reads the
        # old value back. Forcing a fresh SELECT here catches that instead
        # of pretending the save worked.
        db.session.expire(company, ["credit_limit_action"])
        _persisted_action = getattr(company, "credit_limit_action", None)
        if _persisted_action != _credit_action:
            flash(
                f"Credit limit setting did not save — chose '{_credit_action}' but the "
                f"database still has '{_persisted_action}'. This looks like a missing "
                f"'credit_limit_action' column on the Company table, not a form bug. "
                f"Check platform_models.py / run the migration.",
                "danger",
            )

        # Keep session in sync
        if "user" in session:
            session["user"]["company_name"] = company.company_name
            session.modified = True
        flash("Company information updated successfully.")
    else:
        flash("Company not found.")
    return redirect(url_for("company_settings"))

@app.route('/manifest/print/selected/generated')
@login_required
@require_permission("manifest", "view")
def manifest_print_selected_generated():
    """
    Print selected generated entries from the manifest list.
    Takes ?entry_ids=1,2,3,... and prints just those entries.
    """
    company_id = get_current_company()
    if not company_id:
        return redirect(url_for('login'))
    cdb = get_customer_session(company_id)

    entry_ids_param = request.args.get('entry_ids', '').strip()
    try:
        entry_ids = [int(x) for x in entry_ids_param.split(',') if x.strip()]
    except ValueError:
        entry_ids = []
    
    if not entry_ids:
        flash('No entries selected to print.', 'danger')
        return redirect(url_for('manifest_list'))

    # Fetch only the selected entries that belong to this company
    entries = cdb.query(ManifestEntry).join(
        CompanyManifest, ManifestEntry.manifest_id == CompanyManifest.id
    ).filter(
        ManifestEntry.id.in_(entry_ids),
        CompanyManifest.company_id == company_id,
        ManifestEntry.status == 'Generated'  # Only allow printing generated entries
    ).all()

    if not entries:
        flash('No generated entries found for the selected IDs.', 'danger')
        return redirect(url_for('manifest_list'))

    # Group entries by manifest for display
    entries_by_manifest = {}
    manifests = set()
    for entry in entries:
        manifest = entry.manifest
        if manifest.id not in entries_by_manifest:
            entries_by_manifest[manifest.id] = []
            manifests.add(manifest)
        entries_by_manifest[manifest.id].append(entry)

    # Same supplier lookup used everywhere else
    suppliers = cdb.query(Supplier).filter_by(company_id=company_id).all()
    brand_to_supplier = {}
    for sup in suppliers:
        brand_to_supplier[sup.name.strip().lower()] = sup.name
        for b in sup.brands:
            brand_to_supplier[b.brand_name.strip().lower()] = sup.name

    # Same object-keyed lookup and single-supplier-or-blank rule as
    # manifest_print_day/manifest_print_selected: only show "TO," when every
    # selected entry maps to the same courier's supplier record; blank it
    # out if multiple companies are represented in this selection.
    supplier_by_brand = {}
    for sup in suppliers:
        supplier_by_brand[sup.name.strip().lower()] = sup
        for b in sup.brands:
            supplier_by_brand[b.brand_name.strip().lower()] = sup

    to_supplier = None
    resolved = None
    ambiguous = False
    for m_id, mfst_entries in entries_by_manifest.items():
        for entry in mfst_entries:
            key = (entry.courier_name or '').strip().lower()
            sup_obj = supplier_by_brand.get(key)
            if sup_obj is None:
                continue
            if resolved is None:
                resolved = sup_obj
            elif resolved.id != sup_obj.id:
                ambiguous = True
                break
        if ambiguous:
            break
    to_supplier = None if ambiguous else resolved

    reg_company = Company.query.filter_by(company_id=company_id).first()
    from_company_name = reg_company.company_name if reg_company else ''

    total_boxes = sum(sum(e.boxes for e in entries_by_manifest[m_id]) for m_id in entries_by_manifest)

    shipment_data = {}
    for m_id, mfst_entries in entries_by_manifest.items():
        for entry in mfst_entries:
            shipment_data[entry.id] = _manifest_entry_shipment_data(cdb, company_id, entry.docket_no)

    total_weight = sum(
        (ship.get('charge_weight') or ship.get('actual_weight') or 0)
        for ship in shipment_data.values() if ship
    )

    return render_template(
        'manifest_print_day.html',
        manifests=list(manifests),
        entries_by_manifest=entries_by_manifest,
        target_date=None,
        total_boxes=total_boxes,
        from_company_name=from_company_name,
        to_supplier=to_supplier,
        total_weight=total_weight,
        shipment_data=shipment_data,
        brand_to_supplier=brand_to_supplier,
        is_selected_print=True,
    )

@app.route("/settings/whatsapp/templates", methods=["GET", "POST"])
@login_required
@owner_required
def whatsapp_template_config():
    from platform_models import WhatsAppTemplate
    from whatsapp_templates import EVENT_DEFS, field_options_for_event, placeholder_for_field
 
    company_id = get_current_company()
    company = get_company_by_id(company_id)
    if not company:
        flash("Company not found")
        return redirect(url_for("dashboard"))
 
    if request.method == "POST":
        for event_key, _label, _desc in EVENT_DEFS:
            prefix = f"tpl_{event_key}_"
            template_name = request.form.get(prefix + "name", "").strip()
            language_code = request.form.get(prefix + "lang", "en").strip() or "en"
            header_type = request.form.get(prefix + "header", "none").strip() or "none"
            field_keys = [f for f in request.form.getlist(prefix + "var") if f]
 
            row = WhatsAppTemplate.query.filter_by(
                company_id=company_id, template_key=event_key
            ).first()
 
            if not template_name:
                # Blank name = "not configuring this event right now."
                # Leave any existing row untouched rather than deleting it.
                continue
 
            # Translate the selected field keys into {{ placeholder }} strings
            # via the SAME catalogue the resolver uses at send time — this is
            # what keeps the dropdown and the actual send logic in sync.
            try:
                placeholders = [placeholder_for_field(event_key, fk) for fk in field_keys]
            except Exception as e:
                flash(f"'{event_key}': {e}")
                continue
 
            if not row:
                row = WhatsAppTemplate(company_id=company_id, template_key=event_key)
                db.session.add(row)
 
            row.template_name = template_name
            row.language_code = language_code
            row.header_type = header_type
            row.param_count = len(placeholders)
            row.variables_json = json.dumps(placeholders) if placeholders else None
            row.is_active = True
 
        db.session.commit()
        flash("WhatsApp template mapping saved.")
        return redirect(url_for("whatsapp_template_config"))
 
    # ── GET: build view data ──────────────────────────────────────────────
    existing = {
        row.template_key: row
        for row in WhatsAppTemplate.query.filter_by(company_id=company_id).all()
    }
 
    events = []
    for event_key, label, desc in EVENT_DEFS:
        row = existing.get(event_key)
        selected_fields = []
        if row and row.variables_json:
            try:
                placeholders = json.loads(row.variables_json)
                # Reverse-lookup each stored placeholder back to its field
                # key so the form can pre-select the right dropdown option.
                opts = field_options_for_event(event_key)
                lookup = {v[1]: k for k, v in opts.items()}
                selected_fields = [lookup.get(p, "") for p in placeholders]
            except (ValueError, TypeError):
                selected_fields = []
 
        events.append({
            "event_key": event_key,
            "label": label,
            "desc": desc,
            "row": row,
            "selected_fields": selected_fields,
            "field_options": field_options_for_event(event_key),
        })
 
    return render_template(
        "admin_whatsapp_templates.html",
        company=company,
        events=events,
        active="whatsapp_settings",
    )

@app.route("/company/change-password", methods=["POST"])
@login_required
@owner_required
def change_company_password():
    current_password = request.form.get("current_password", "")
    new_password      = request.form.get("new_password", "")
    confirm_password  = request.form.get("confirm_password", "")

    if new_password != confirm_password:
        flash("New password and confirmation do not match.", "error")
        return redirect(url_for("company_settings"))
    if len(new_password) < 8:
        flash("New password must be at least 8 characters.", "error")
        return redirect(url_for("company_settings"))

    email = get_current_user().get("email", "").strip().lower()
    reg_user = RegisteredUser.query.filter_by(email=email, is_active=True).first()

    if reg_user:
        if not verify_password(current_password, reg_user.password_hash):
            flash("Current password is incorrect.", "error")
            return redirect(url_for("company_settings"))
        reg_user.password_hash = hash_password(new_password)
        db.session.commit()

        # Keep the matching CompanyUser record(s) in sync across all of this
        # owner's companies, since employee login checks CompanyUser.password_hash.
        for comp in get_owner_companies(email):
            try:
                _cdb = get_customer_session(comp.company_id)
                emp = _cdb.query(CompanyUser).filter_by(email=email).first()
                if emp:
                    emp.password_hash = hash_password(new_password)
                    _cdb.commit()
            except Exception as e:
                # This loop touches OTHER companies' sessions, not the
                # logged-in owner's own company — teardown_request only
                # rolls back get_current_company()'s session, so a failure
                # here on comp.company_id would otherwise poison that
                # company's session for every future request, with nothing
                # to ever clean it up. Must roll back explicitly.
                try:
                    _cdb.rollback()
                except Exception:
                    pass
                print(f"⚠  Could not sync password for {comp.company_id}: {e}")
    else:
        # Fallback: user only exists as a CompanyUser (shouldn't normally reach
        # this owner-only page, but handled defensively).
        cdb = get_cdb()
        company_id = get_current_company()
        emp = cdb.query(CompanyUser).filter_by(company_id=company_id, email=email).first()
        if not emp or not verify_password(current_password, emp.password_hash):
            flash("Current password is incorrect.", "error")
            return redirect(url_for("company_settings"))
        emp.password_hash = hash_password(new_password)
        cdb.commit()

    flash("Password changed successfully.")
    return redirect(url_for("company_settings"))


ALLOWED_COMPANY_ROLES = ["manager", "employee", "accountant"]  # "owner" is reserved, assigned only at company creation

@app.route("/company/add-user", methods=["POST"])
@login_required
@owner_required
def add_company_user():
    owner_email = get_current_user().get("email", "").strip().lower()

    email      = request.form.get("email",     "").strip().lower()
    password   = request.form.get("password",  "")
    full_name  = request.form.get("full_name", "").strip()
    department = request.form.get("department","")
    phone      = request.form.get("phone",     "")

    if not email or not full_name:
        flash("Email and full name are required.")
        return redirect(url_for("company_settings"))

    # Which of the owner's companies should this user get access to?
    # Form sends company_ids[] (checkboxes) and role_<company_id> (per-row select)
    owner_companies = {c.company_id: c for c in get_owner_companies(owner_email)}
    selected_company_ids = [cid for cid in request.form.getlist("company_ids") if cid in owner_companies]

    if not selected_company_ids:
        flash("Select at least one company to grant access to.")
        return redirect(url_for("company_settings"))

    # Seat cap is owner-wide: only a brand-new email consumes a seat.
    # Granting an EXISTING user access to another company doesn't cost a seat.
    current_count, max_u, existing_emails = get_owner_user_stats(owner_email)
    is_new_user = email not in existing_emails
    if is_new_user and max_u != "Unlimited":
        try:
            max_u_int = int(max_u)
            if current_count >= max_u_int:
                flash(f"Maximum {max_u_int} users allowed across all your companies under your plan. Please upgrade.")
                return redirect(url_for("company_settings"))
        except (ValueError, TypeError):
            pass  # "Unlimited"

    if is_new_user and not password:
        flash("Password is required to create a new user.")
        return redirect(url_for("company_settings"))

    pw_hash = hash_password(password) if password else None
    granted_to = []

    for cid in selected_company_ids:
        role = request.form.get(f"role_{cid}", "employee")
        if role not in ALLOWED_COMPANY_ROLES:
            role = "employee"

        _cdb = get_customer_session(cid)
        existing = _cdb.query(CompanyUser).filter_by(company_id=cid, email=email).first()

        if existing:
            # Already has access to this company — update role / reactivate / reset password if given
            existing.role = role
            existing.is_active = True
            existing.full_name = full_name or existing.full_name
            if pw_hash:
                existing.password_hash = pw_hash
            _cdb.commit()
        else:
            emp_id    = _next_numbered_id(_cdb, CompanyUser.user_id, "EMP")
            new_user = CompanyUser(
                user_id=emp_id, company_id=cid,
                email=email, password_hash=pw_hash,
                full_name=full_name, role=role,
                department=department, phone=phone,
                is_active=True, created_at=today_ist()
            )
            _cdb.add(new_user)
            _cdb.commit()

        granted_to.append(owner_companies[cid].company_name)

    flash(f"'{full_name}' now has access to: {', '.join(granted_to)}.")
    return redirect(url_for("company_settings"))


@app.route("/company/revoke-user/<email>/<company_id>")
@login_required
@owner_required
def revoke_company_user(email, company_id):
    """Remove a user's access to ONE specific company (not all of their companies)."""
    owner_email = get_current_user().get("email", "").strip().lower()
    owner_companies = {c.company_id: c for c in get_owner_companies(owner_email)}
    if company_id not in owner_companies:
        flash("Invalid company.")
        return redirect(url_for("company_settings"))

    _cdb = get_customer_session(company_id)
    user = _cdb.query(CompanyUser).filter_by(company_id=company_id, email=email.strip().lower()).first()
    if user and user.role != "owner":
        user.is_active = False
        _cdb.commit()
        flash(f"Access to {owner_companies[company_id].company_name} revoked.")
    else:
        flash("Cannot revoke this user.")
    return redirect(url_for("company_settings"))


@app.route("/company/remove-user/<user_id>")
@login_required
@owner_required
def remove_company_user(user_id):
    cdb = get_cdb()
    company_id = get_current_company()
    user = cdb.query(CompanyUser).filter_by(user_id=user_id, company_id=company_id).first()
    if user and user.role != "owner":
        user.is_active = False
        cdb.commit()
        flash("User removed successfully.")
    else:
        flash("Cannot remove this user.")
    return redirect(url_for("company_settings"))


@app.route("/company/delete-user/<email>")
@login_required
@owner_required
def delete_company_user(email):
    """Remove a person's access to ALL of the owner's companies in one go
    (a full 'delete this person' action). Soft-delete only — sets
    is_active=False everywhere, same as revoke_company_user, so we never
    orphan the invoices/orders/etc. that still reference their user_id.
    """
    owner_email = get_current_user().get("email", "").strip().lower()
    owner_companies = get_owner_companies(owner_email)
    email = email.strip().lower()

    removed_from = []
    for c in owner_companies:
        _cdb = get_customer_session(c.company_id)
        user = _cdb.query(CompanyUser).filter_by(company_id=c.company_id, email=email).first()
        if user and user.role != "owner" and user.is_active:
            user.is_active = False
            _cdb.commit()
            removed_from.append(c.company_name)

    if removed_from:
        flash(f"User removed from: {', '.join(removed_from)}.")
    else:
        flash("Cannot remove this user.")
    return redirect(url_for("company_settings"))


@app.route("/company/edit-user-access/<email>", methods=["POST"])
@login_required
@owner_required
def edit_user_access(email):
    """Update a person's company access + role in one submit: check a
    company to grant/keep access (with the chosen role), uncheck one to
    revoke it. Mirrors add_company_user's creation logic for any newly
    checked company, and revoke_company_user's is_active=False for any
    unchecked one that currently has access.
    """
    owner_email = get_current_user().get("email", "").strip().lower()
    owner_companies = {c.company_id: c for c in get_owner_companies(owner_email)}
    email = email.strip().lower()

    selected_company_ids = set(cid for cid in request.form.getlist("company_ids") if cid in owner_companies)

    # Find an existing row for this email to copy password_hash/full_name/etc.
    # onto any brand-new company rows we create below.
    template_user = None
    for cid in owner_companies:
        _cdb = get_customer_session(cid)
        u = _cdb.query(CompanyUser).filter_by(company_id=cid, email=email).first()
        if u:
            template_user = u
            break
    if not template_user:
        flash("User not found.")
        return redirect(url_for("company_settings"))
    if template_user.role == "owner":
        flash("Owner access can't be edited this way.")
        return redirect(url_for("company_settings"))

    granted, revoked = [], []
    for cid, c in owner_companies.items():
        _cdb = get_customer_session(cid)
        existing = _cdb.query(CompanyUser).filter_by(company_id=cid, email=email).first()

        if cid in selected_company_ids:
            role = request.form.get(f"role_{cid}", "employee")
            if role not in ALLOWED_COMPANY_ROLES:
                role = "employee"
            if existing:
                if existing.role != "owner":
                    existing.role = role
                    existing.is_active = True
                    _cdb.commit()
            else:
                emp_id    = _next_numbered_id(_cdb, CompanyUser.user_id, "EMP")
                new_user = CompanyUser(
                    user_id=emp_id, company_id=cid,
                    email=email, password_hash=template_user.password_hash,
                    full_name=template_user.full_name, role=role,
                    department=template_user.department, phone=template_user.phone,
                    is_active=True, created_at=today_ist()
                )
                _cdb.add(new_user)
                _cdb.commit()
            granted.append(c.company_name)
        else:
            if existing and existing.role != "owner" and existing.is_active:
                existing.is_active = False
                _cdb.commit()
                revoked.append(c.company_name)

    parts = []
    if granted:
        parts.append(f"access: {', '.join(granted)}")
    if revoked:
        parts.append(f"revoked: {', '.join(revoked)}")
    flash(f"Updated {template_user.full_name} — " + "; ".join(parts) if parts else "No changes made.")
    return redirect(url_for("company_settings"))


@app.route("/company/upgrade-plan", methods=["POST"])
@login_required
@owner_required
def upgrade_plan():
    company_id = get_current_company()
    company    = get_company_by_id(company_id)
    new_plan   = request.form.get("plan")
    plan       = SubscriptionPlan.query.get(new_plan)
    if company and plan:
        company.subscription_plan     = new_plan
        company.max_users_per_company = plan.max_users
        company.max_companies_allowed = plan.max_companies
        db.session.commit()
        flash(f"Plan upgraded to {plan.name} successfully!")
    else:
        flash("Invalid plan selected.")
    return redirect(url_for("company_settings"))

# ─────────────────────────────────────────────────────────────────────────────
# ── DEBTORS & CREDITORS ───────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────
def _debtor_summary(company_id):
    cdb = get_cdb()
    
    all_clients = cdb.query(Client).filter_by(company_id=company_id).order_by(Client.name).all()
    today = today_ist()
    rows = []

    for c in all_clients:
        cutoff_date = c.statement_cutoff.date() if c.statement_cutoff else None

        # 🔴 FIX: Filter out VOID invoices
        inv_q = cdb.query(Invoice).filter_by(company_id=company_id, client_id=c.id)
        inv_q = inv_q.filter(Invoice.status.notin_(['Cancelled', 'Void']))  # ← EXCLUDE VOID
        if cutoff_date:
            inv_q = inv_q.filter(Invoice.date >= cutoff_date)
        invoices = inv_q.order_by(Invoice.date.desc()).all()

        cash_q = (cdb.query(CashTransaction)
                  .filter(CashTransaction.company_id == company_id,
                          func.lower(CashTransaction.party_name) == func.lower(c.name))
                  .filter(CashTransaction.category.in_(["Receipt", "Adjustment"]))
                  .filter(CashTransaction.reference != "WRITE-OFF"))
        if cutoff_date:
            cash_q = cash_q.filter(CashTransaction.date >= cutoff_date)
        cash_received = float(sum(t.amount or 0 for t in cash_q.all()))

        bank_q = (cdb.query(BankTransaction)
                  .filter(BankTransaction.company_id == company_id,
                          func.lower(BankTransaction.party_name) == func.lower(c.name))
                  .filter(BankTransaction.type == "credit"))
        if cutoff_date:
            bank_q = bank_q.filter(BankTransaction.date >= cutoff_date)
        bank_received = float(sum(t.amount or 0 for t in bank_q.all()))

        if not invoices:
            total_pending = (c.opening_balance or 0) - cash_received - bank_received
            # ... rest of the code
            continue

        total_invoiced = sum(float(i.grand_total or 0) for i in invoices)
        
        # 🔴 FIX: Opening balance should ONLY be used if there's no cutoff
        # or if we're calculating after the cutoff date
        if cutoff_date:
            # Opening balance already represents everything before cutoff
            total_pending = (c.opening_balance or 0) + total_invoiced - cash_received - bank_received
        else:
            # No cutoff - use the actual balance from invoices
            total_pending = sum(float(getattr(i, 'balance', 0) or 0) for i in invoices)

        # If no invoices, still show client — dues can come from opening
        # balance alone.
        if not invoices:
            total_pending = (c.opening_balance or 0) - cash_received - bank_received
            rows.append({
                "id":                c.id,
                "name":              c.name,
                "phone":             c.phone or "",
                "city":              c.city or "",
                "total_invoiced":    0,
                "total_paid":        cash_received + bank_received,
                "total_pending":     total_pending,
                "last_invoice_id":   None,
                "last_awb":          None,
                "last_invoice_date": None,
                "nearest_due_date":  None,
                "nearest_due_amt":   None,
                "last_payment_date": None,
                "last_payment_amt":  None,
                "invoice_count":     0,
                "overdue":           False,
                "status":            "Fully Paid" if total_pending <= 0 else "Has Dues",
            })
            continue

        # Calculate totals — opening_balance + invoices-since-cutoff −
        # receipts-since-cutoff, matching the statement's closing balance.
        total_invoiced = sum(float(i.grand_total or 0) for i in invoices)
        total_pending = (c.opening_balance or 0) + total_invoiced - cash_received - bank_received
        total_paid = total_invoiced - total_pending
        last_invoice = invoices[0]
        last_invoice_date = last_invoice.date
        last_invoice_id = last_invoice.invoice_id
        last_awb = _get_awb(last_invoice) or None

        # Calculate overdue
        unpaid = [i for i in invoices if (float(getattr(i, "balance", 0) or 0)) > 0]
        due_invoices = [i for i in unpaid if getattr(i, "due_date", None)]
        if due_invoices:
            future = [i for i in due_invoices if i.due_date >= today]
            nearest = min(future, key=lambda i: i.due_date) if future else \
                      max(due_invoices, key=lambda i: i.due_date)
            nearest_due_date = nearest.due_date
            nearest_due_amt = float(getattr(nearest, "balance", 0) or 0)
            overdue = nearest_due_date < today if nearest_due_date else False
        else:
            nearest_due_date = None
            nearest_due_amt = None
            overdue = False

        # Last payment — the most recent actual Cash/Bank transaction for this
        # client, not an invoice-date-based guess. (Same class of bug as the
        # statement pages: picking "the invoice with the latest date among
        # paid ones" is neither the date nor necessarily the invoice a
        # payment was last applied to — a receipt can land against an older
        # invoice, or not be tied to one at all.)
        last_cash = (cdb.query(CashTransaction)
                     .filter_by(company_id=company_id, party_name=c.name, category="Receipt")
                     .order_by(CashTransaction.date.desc()).first())
        last_bank = (cdb.query(BankTransaction)
                     .filter_by(company_id=company_id, party_name=c.name)
                     .filter(BankTransaction.type == "credit")
                     .order_by(BankTransaction.date.desc()).first())
        candidates = [t for t in (last_cash, last_bank) if t is not None]
        if candidates:
            last_txn = max(candidates, key=lambda t: t.date)
            last_payment_date = last_txn.date
            last_payment_amt = last_txn.amount or 0
        else:
            last_payment_date = None
            last_payment_amt = None

        rows.append({
            "id":                c.id,
            "name":              c.name,
            "phone":             c.phone or "",
            "city":              c.city or "",
            "total_invoiced":    total_invoiced,
            "total_paid":        total_paid,
            "total_pending":     total_pending,
            "last_invoice_id":   last_invoice_id,
            "last_awb":          last_awb,
            "last_invoice_date": last_invoice_date,
            "nearest_due_date":  nearest_due_date,
            "nearest_due_amt":   nearest_due_amt,
            "last_payment_date": last_payment_date,
            "last_payment_amt":  last_payment_amt,
            "invoice_count":     len(invoices),
            "overdue":           overdue,
            "status":            "Fully Paid" if total_pending <= 0 else "Has Dues",
        })

    rows.sort(key=lambda r: r["name"].lower())
    return rows

def _creditor_summary(company_id):
    cdb = get_cdb()
    suppliers = cdb.query(Supplier).filter(
        Supplier.company_id == company_id
    ).order_by(Supplier.name).all()

    today = today_ist()
    rows = []

    for s in suppliers:
        cutoff_date = s.statement_cutoff.date() if s.statement_cutoff else None

        # 🔴 FIX: Filter out VOID purchase invoices
        inv_q = cdb.query(PurchaseInvoice).filter_by(company_id=company_id, supplier_id=s.id)
        inv_q = inv_q.filter(PurchaseInvoice.status.notin_(['Cancelled', 'Void']))  # ← EXCLUDE VOID
        if cutoff_date:
            inv_q = inv_q.filter(PurchaseInvoice.date >= cutoff_date)
        invoices = inv_q.order_by(PurchaseInvoice.date.desc()).all()

        # BUG FIX: total_pending must be driven off the ACTUAL cash/bank
        # Payment transactions recorded for this supplier — not off each
        # invoice's own paid_amount. A payment that isn't matched exactly to
        # an invoice's balance (an advance, an overpayment, or a payment
        # recorded with no invoice selected) never touches any invoice's
        # paid_amount/balance, so a paid_amount-based sum silently ignores
        # that money entirely and the "outstanding" figure never moves.
        # (Same fix already applied to _debtor_summary above — this brings
        # creditors in line with it.)
        cash_q = (cdb.query(CashTransaction)
                  .filter(CashTransaction.company_id == company_id,
                          func.lower(CashTransaction.party_name) == func.lower(s.name))
                  .filter(CashTransaction.category == "Payment"))
        if cutoff_date:
            cash_q = cash_q.filter(CashTransaction.date >= cutoff_date)
        cash_paid = float(sum(t.amount or 0 for t in cash_q.all()))

        bank_q = (cdb.query(BankTransaction)
                  .filter(BankTransaction.company_id == company_id,
                          func.lower(BankTransaction.party_name) == func.lower(s.name))
                  .filter(BankTransaction.type == "debit"))
        if cutoff_date:
            bank_q = bank_q.filter(BankTransaction.date >= cutoff_date)
        bank_paid = float(sum(t.amount or 0 for t in bank_q.all()))

        if not invoices:
            total_pending = (s.opening_balance or 0) - cash_paid - bank_paid
            rows.append({
                "id":                s.id,
                "name":              s.name,
                "phone":             s.phone or "",
                "city":              s.city or "",
                "total_pending":     total_pending,
                "last_bill_date":    None,
                "nearest_due_date":  None,
                "nearest_due_amt":   None,
                "last_payment_date": None,
                "last_payment_amt":  None,
                "invoice_count":     0,
                "overdue":           False,
                "status":            "Fully Paid" if total_pending <= 0 else "Has Dues",
            })
            continue

        # total_pending — opening_balance + Σ(grand_total) since cutoff −
        # actual payments made since cutoff, matching the creditor
        # statement's closing balance and correctly reflecting advance /
        # unmatched payments (which don't reduce any single invoice's
        # balance).
        total_invoiced = sum(float(i.grand_total or 0) for i in invoices)
        total_pending = (s.opening_balance or 0) + total_invoiced - cash_paid - bank_paid
        last_bill_date = invoices[0].date

        # Calculate overdue
        unpaid = [i for i in invoices if (i.balance or 0) > 0]
        due_invoices = [i for i in unpaid if i.due_date]
        if due_invoices:
            future = [i for i in due_invoices if i.due_date >= today]
            nearest = min(future, key=lambda i: i.due_date) if future else \
                      max(due_invoices, key=lambda i: i.due_date)
            nearest_due_date = nearest.due_date
            nearest_due_amt = nearest.balance or 0
            overdue = nearest_due_date < today if nearest_due_date else False
        else:
            nearest_due_date = None
            nearest_due_amt = None
            overdue = False

        # Last payment — real most-recent Cash/Bank transaction, not an
        # invoice-date guess (same fix as _debtor_summary above).
        last_cash = (cdb.query(CashTransaction)
                     .filter(CashTransaction.company_id == company_id,
                             func.lower(CashTransaction.party_name) == func.lower(s.name),
                             CashTransaction.category == "Payment")
                     .order_by(CashTransaction.date.desc()).first())
        last_bank = (cdb.query(BankTransaction)
                     .filter(BankTransaction.company_id == company_id,
                             func.lower(BankTransaction.party_name) == func.lower(s.name),
                             BankTransaction.type == "debit")
                     .order_by(BankTransaction.date.desc()).first())
        candidates = [t for t in (last_cash, last_bank) if t is not None]
        if candidates:
            last_txn = max(candidates, key=lambda t: t.date)
            last_payment_date = last_txn.date
            last_payment_amt = last_txn.amount or 0
        else:
            last_payment_date = None
            last_payment_amt = None

        rows.append({
            "id":                s.id,
            "name":              s.name,
            "phone":             s.phone or "",
            "city":              s.city or "",
            "total_pending":     total_pending,
            "last_bill_date":    last_bill_date,
            "nearest_due_date":  nearest_due_date,
            "nearest_due_amt":   nearest_due_amt,
            "last_payment_date": last_payment_date,
            "last_payment_amt":  last_payment_amt,
            "invoice_count":     len(invoices),
            "overdue":           overdue,
            "status":            "Fully Paid" if total_pending <= 0 else "Has Dues",
        })

    rows.sort(key=lambda r: r["name"].lower())
    return rows

@app.route("/debtors")
@login_required
@require_permission("debtors", "view")
def debtors_list():
    company_id        = get_current_company()
    debtors           = _debtor_summary(company_id)
    total_outstanding = sum(d["total_pending"] for d in debtors)
    overdue_count     = sum(1 for d in debtors if d["overdue"])
    return render_template("debtors.html",
                           debtors=debtors,
                           total_outstanding=total_outstanding,
                           overdue_count=overdue_count)


@app.route("/creditors")
@login_required
@require_permission("creditors", "view")
def creditors_list():
    company_id    = get_current_company()
    creditors     = _creditor_summary(company_id)
    total_payable = sum(c["total_pending"] for c in creditors)
    overdue_count = sum(1 for c in creditors if c["overdue"])
    return render_template("creditors.html",
                           creditors=creditors,
                           total_payable=total_payable,
                           overdue_count=overdue_count)


@app.route("/debtors/<int:client_pk>/statement")
@login_required
@require_permission("debtors", "view")
def debtor_statement(client_pk):
    """Short/Standard statement for Debtors (simple format)"""
    cdb = get_cdb()
    company_id = get_current_company()
    c = _first_or_404(cdb.query(Client).filter_by(id=client_pk, company_id=company_id).first())

    # Get date filters from query params
    from_date_str = request.args.get('from_date', '')
    to_date_str = request.args.get('to_date', '')

    from_date = date.fromisoformat(from_date_str) if from_date_str else None
    to_date = date.fromisoformat(to_date_str) if to_date_str else None

    # A statement cutoff (set when outstanding was cleared/shifted from the
    # Clients page) acts as a hard, always-on floor — old, archived
    # transactions must never leak back into this "new" statement, even if
    # someone picks a from_date earlier than the cutoff. It's applied
    # independently of (and in addition to) the from_date/to_date filters.
    cutoff_date = c.statement_cutoff.date() if c.statement_cutoff else None

    # Statement shows customer invoices only — raw bookings never appear
    # here on their own, whether or not they've been rolled into a
    # customer invoice yet.

    # ── Customer invoices (grouped bookings) as their own ledger lines ────
    ci_query = (cdb.query(CustomerInvoice)
                .filter_by(company_id=company_id, client_id=c.id, invoice_type="credit")
                .filter(CustomerInvoice.status != "Void"))
    if cutoff_date:
        ci_query = ci_query.filter(CustomerInvoice.invoice_date >= cutoff_date)
    if from_date:
        ci_query = ci_query.filter(CustomerInvoice.invoice_date >= from_date)
    if to_date:
        ci_query = ci_query.filter(CustomerInvoice.invoice_date <= to_date)
    customer_invoices = ci_query.order_by(CustomerInvoice.invoice_date.asc()).all()

    # Every cash/bank transaction for this client is its own ledger event —
    # with its OWN real date and amount — rather than being inferred from
    # an invoice's grand_total-minus-balance delta (which stamped every
    # payment with the invoice's date and collapsed multiple/advance
    # payments into invisibility). Same date-range filter as invoices above,
    # so a payment outside the selected period doesn't leak in (or a real
    # one inside it doesn't get missed).
    cash_q = cdb.query(CashTransaction).filter(
        CashTransaction.company_id == company_id,
        func.lower(CashTransaction.party_name) == func.lower(c.name)
    ).filter(CashTransaction.category.in_(["Receipt", "Adjustment"]))
    # The write-off/carry-forward adjustment row itself is the mechanism
    # that produces the cutoff — it must never appear as a ledger line.
    cash_q = cash_q.filter(CashTransaction.reference != "WRITE-OFF")
    if cutoff_date:
        cash_q = cash_q.filter(CashTransaction.date >= cutoff_date)
    if from_date:
        cash_q = cash_q.filter(CashTransaction.date >= from_date)
    if to_date:
        cash_q = cash_q.filter(CashTransaction.date <= to_date)
    cash_txns = cash_q.all()

    bank_q = cdb.query(BankTransaction).filter(
        BankTransaction.company_id == company_id,
        func.lower(BankTransaction.party_name) == func.lower(c.name)
    ).filter(BankTransaction.type == "credit")
    if cutoff_date:
        bank_q = bank_q.filter(BankTransaction.date >= cutoff_date)
    if from_date:
        bank_q = bank_q.filter(BankTransaction.date >= from_date)
    if to_date:
        bank_q = bank_q.filter(BankTransaction.date <= to_date)
    bank_txns = bank_q.all()

    events = []

    # One line per customer invoice, standing in for the bookings it groups
    for ci in customer_invoices:
        try:
            booking_count = len(json.loads(ci.booking_ids_json)) if ci.booking_ids_json else 0
        except (ValueError, TypeError):
            booking_count = 0
        events.append({
            "date": ci.invoice_date,
            "type": "Invoice",
            "ref": ci.invoice_number,
            "payment_mode": "",
            "debit": ci.grand_total or 0,
            "credit": 0,
            "status": ci.status,
            "awb": "",
            "consignee": f"{booking_count} booking" + ("" if booking_count == 1 else "s"),
            "destination": "", "carrier_ref": "", "carrier": "",
            "chrg_wt": 0, "act_wt": 0, "vol_wt": 0,
            "grand_total": ci.grand_total or 0,
            "other_charges": 0,
            "billing_amount": ci.grand_total or 0,
            "per_kg": 0,
            "_sort": 0,
        })

    for ct in cash_txns:
        ref = ct.reference or ""
        events.append({
            "date": ct.date,
            "type": "Payment Received",
            "ref": "—" if ref == "ADVANCE" else ref,
            "payment_mode": "Cash",
            "debit": 0,
            "credit": ct.amount or 0,
            "status": "",
            "awb": "", "consignee": "", "destination": "", "carrier_ref": "", "carrier": "",
            "chrg_wt": 0, "act_wt": 0, "vol_wt": 0,
            "grand_total": 0, "other_charges": 0, "billing_amount": 0,
            "per_kg": 0,
            "_sort": 1,
        })

    for bt in bank_txns:
        ref = bt.reference or ""
        events.append({
            "date": bt.date,
            "type": "Payment Received",
            "ref": "—" if ref == "ADVANCE" else ref,
            "payment_mode": bt.transaction_mode or "Bank Transfer",
            "debit": 0,
            "credit": bt.amount or 0,
            "status": "",
            "awb": "", "consignee": "", "destination": "", "carrier_ref": "", "carrier": "",
            "chrg_wt": 0, "act_wt": 0, "vol_wt": 0,
            "grand_total": 0, "other_charges": 0, "billing_amount": 0,
            "per_kg": 0,
            "_sort": 1,
        })

    events.sort(key=lambda e: (e["date"] or date.min, e["_sort"]))

    ledger = []
    running_balance = c.opening_balance or 0.0

    if running_balance:
        ledger.append({
            "date": c.statement_cutoff.date() if c.statement_cutoff else (c.created_at or today_ist()),
            "type": "Balance Carried Forward" if c.statement_cutoff else "Opening Balance",
            "ref": "—",
            "payment_mode": "",
            "debit": running_balance,
            "credit": 0,
            "balance": running_balance,
            "status": "",
            "awb": "", "consignee": "", "destination": "", "carrier_ref": "", "carrier": "",
            "chrg_wt": 0, "act_wt": 0, "vol_wt": 0,
            "grand_total": 0, "other_charges": 0, "billing_amount": 0,
            "per_kg": 0,
        })

    for e in events:
        running_balance += (e["debit"] or 0) - (e["credit"] or 0)
        e["balance"] = running_balance
        del e["_sort"]
        ledger.append(e)

    total_debit = sum(r["debit"] for r in ledger)
    total_credit = sum(r["credit"] for r in ledger)

    archives = (cdb.query(StatementClosing)
                .filter_by(company_id=company_id, entity_type="client", entity_id=c.id)
                .order_by(StatementClosing.closed_at.desc())
                .all())

    return render_template("debtor_creditor_statement.html",
                           entity=_normalize_client(c),
                           company=get_company_by_id(company_id),
                           ledger=ledger,
                           total_debit=total_debit,
                           total_credit=total_credit,
                           closing_balance=running_balance,
                           mode="debtor",
                           back_url="/debtors",
                           archive_base_url=f"/debtors/{client_pk}",
                           archives=archives,
                           archived=False,
                           today=today_ist().strftime("%d %b %Y"),
                           from_date=from_date_str,
                           to_date=to_date_str)


@app.route("/debtors/<int:client_pk>/statement/archive/<int:archive_id>")
@login_required
@require_permission("debtors", "view")
def debtor_statement_archive(client_pk, archive_id):
    """Frozen old debtor statement, same snapshot the Clients-page statement
    archive reads from — both point at the same StatementClosing rows since
    they're the same underlying client/closing action."""
    cdb = get_cdb()
    company_id = get_current_company()
    c = _first_or_404(cdb.query(Client).filter_by(id=client_pk, company_id=company_id).first())
    archive = _first_or_404(cdb.query(StatementClosing).filter_by(
        id=archive_id, company_id=company_id, entity_type="client", entity_id=client_pk).first())

    return render_template("debtor_creditor_statement.html",
                           entity=_normalize_client(c),
                           company=get_company_by_id(company_id),
                           ledger=json.loads(archive.ledger_snapshot or "[]"),
                           total_debit=archive.total_debit,
                           total_credit=archive.total_credit,
                           closing_balance=archive.closing_balance,
                           mode="debtor",
                           back_url=f"/debtors/{client_pk}/statement",
                           archived=True,
                           archived_at=archive.closed_at,
                           today=today_ist().strftime("%d %b %Y"))


@app.route("/creditors/<int:supplier_pk>/statement")
@login_required
@require_permission("creditors", "view")
def creditor_statement(supplier_pk):
    """Short/Standard statement for Creditors (simple format)"""
    cdb = get_cdb()
    company_id = get_current_company()
    s = _first_or_404(cdb.query(Supplier).filter_by(id=supplier_pk, company_id=company_id).first())

    cutoff_date = s.statement_cutoff.date() if s.statement_cutoff else None

    invoices_q = cdb.query(PurchaseInvoice).filter_by(company_id=company_id, supplier_id=s.id)
    invoices_q = invoices_q.filter(PurchaseInvoice.status.notin_(['Cancelled', 'Void']))
    if cutoff_date:
        invoices_q = invoices_q.filter(PurchaseInvoice.date >= cutoff_date)
    invoices = invoices_q.order_by(PurchaseInvoice.date.asc()).all()

    # BUG FIX: every cash/bank Payment transaction is its own ledger event —
    # with its OWN real date, amount and mode — instead of a "Payment Made"
    # line being *derived* from each invoice's paid_amount. Deriving it from
    # paid_amount meant: (1) a payment recorded as an advance / not matched
    # to a specific invoice's exact remaining balance (overpayment, no
    # invoice selected, amount split across invoices) never showed up on the
    # statement at all — the loop below never saw it because it only walked
    # invoices, and (2) multiple partial payments against the same invoice
    # collapsed into a single line stamped with the invoice's date instead
    # of each payment's real date. Same class of fix already applied to
    # debtor_statement above — this brings creditors in line with it.
    cash_q = cdb.query(CashTransaction).filter(
        CashTransaction.company_id == company_id,
        func.lower(CashTransaction.party_name) == func.lower(s.name)
    ).filter(CashTransaction.category == "Payment")
    if cutoff_date:
        cash_q = cash_q.filter(CashTransaction.date >= cutoff_date)
    cash_txns = cash_q.all()

    bank_q = cdb.query(BankTransaction).filter(
        BankTransaction.company_id == company_id,
        func.lower(BankTransaction.party_name) == func.lower(s.name)
    ).filter(BankTransaction.type == "debit")
    if cutoff_date:
        bank_q = bank_q.filter(BankTransaction.date >= cutoff_date)
    bank_txns = bank_q.all()

    events = []

    for inv in invoices:
        events.append({
            "date": inv.date,
            "type": "Purchase Invoice",
            "ref": inv.invoice_number or inv.invoice_id,
            "payment_mode": "",
            "debit": 0,
            "credit": inv.grand_total or 0,
            "status": inv.status,
            "_sort": 0,
        })

    for ct in cash_txns:
        ref = ct.reference or ""
        events.append({
            "date": ct.date,
            "type": "Payment Made",
            "ref": "—" if ref == "ADVANCE" else ref,
            "payment_mode": "Cash",
            "debit": ct.amount or 0,
            "credit": 0,
            "status": "",
            "_sort": 1,
        })

    for bt in bank_txns:
        ref = bt.reference or ""
        events.append({
            "date": bt.date,
            "type": "Payment Made",
            "ref": "—" if ref == "ADVANCE" else ref,
            "payment_mode": bt.transaction_mode or "Bank Transfer",
            "debit": bt.amount or 0,
            "credit": 0,
            "status": "",
            "_sort": 1,
        })

    events.sort(key=lambda e: (e["date"] or date.min, e["_sort"]))

    ledger = []
    running_balance = s.opening_balance or 0.0

    if running_balance:
        ledger.append({
            "date": s.statement_cutoff.date() if s.statement_cutoff else (s.created_at or today_ist()),
            "type": "Balance Carried Forward" if s.statement_cutoff else "Opening Balance",
            "ref": "—",
            "payment_mode": "",
            "debit": 0,
            "credit": running_balance,
            "balance": running_balance,
            "status": "",
        })

    for e in events:
        running_balance += (e["credit"] or 0) - (e["debit"] or 0)
        e["balance"] = running_balance
        del e["_sort"]
        ledger.append(e)

    total_debit = sum(r["debit"] for r in ledger)
    total_credit = sum(r["credit"] for r in ledger)

    archives = (cdb.query(StatementClosing)
                .filter_by(company_id=company_id, entity_type="supplier", entity_id=s.id)
                .order_by(StatementClosing.closed_at.desc())
                .all())

    # Use the SIMPLE template for creditors
    return render_template("debtor_creditor_statement.html",
                           entity=_normalize_supplier(s),
                           company=get_company_by_id(company_id),
                           ledger=ledger,
                           total_debit=total_debit,
                           total_credit=total_credit,
                           closing_balance=running_balance,
                           mode="supplier",
                           back_url="/creditors",
                           archive_base_url=f"/creditors/{supplier_pk}",
                           archives=archives,
                           archived=False,
                           today=today_ist().strftime("%d %b %Y"))


@app.route("/creditors/<int:supplier_pk>/statement/archive/<int:archive_id>")
@login_required
@require_permission("creditors", "view")
def creditor_statement_archive(supplier_pk, archive_id):
    """Frozen old creditor statement — reads the same StatementClosing rows
    the Suppliers-page statement archive uses."""
    cdb = get_cdb()
    company_id = get_current_company()
    s = _first_or_404(cdb.query(Supplier).filter_by(id=supplier_pk, company_id=company_id).first())
    archive = _first_or_404(cdb.query(StatementClosing).filter_by(
        id=archive_id, company_id=company_id, entity_type="supplier", entity_id=supplier_pk).first())

    return render_template("debtor_creditor_statement.html",
                           entity=_normalize_supplier(s),
                           company=get_company_by_id(company_id),
                           ledger=json.loads(archive.ledger_snapshot or "[]"),
                           total_debit=archive.total_debit,
                           total_credit=archive.total_credit,
                           closing_balance=archive.closing_balance,
                           mode="supplier",
                           back_url=f"/creditors/{supplier_pk}/statement",
                           archived=True,
                           archived_at=archive.closed_at,
                           today=today_ist().strftime("%d %b %Y"))


# ─────────────────────────────────────────────────────────────────────────────
# ── Receipts & Payments ───────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────

def _outstanding_invoices_for_client(company_id, client_id):
    """Return list of dicts for invoices with a remaining balance for a client.

    BUG FIX: this used to filter `Invoice.status.in_(["Pending", "Partial"])`.
    New invoices default to status "Draft" (see the invoice-creation routes'
    status logic) whenever nothing has been paid yet — which is the normal
    case for a brand-new invoice — so a finalized, fully unpaid invoice sat in
    "Draft" forever and was invisible here, even though the client's total
    pending balance (computed elsewhere from `balance` directly, not status)
    correctly showed the money owed. That's why the receipts page could show
    "₹18,743 pending" for a client and "No outstanding invoices" in the same
    breath. What actually determines whether an invoice needs settling is its
    balance, not its status label, so filter on that instead.
    """
    cdb = get_cdb()
    invs = (cdb.query(Invoice)
            .filter_by(company_id=company_id, client_id=client_id)
            .filter(Invoice.status != "Paid")
            .order_by(Invoice.date.asc())
            .all())
    result = []
    for inv in invs:
        total   = inv.grand_total or 0
        balance = getattr(inv, "balance", None)
        if balance is None:
            balance = total if inv.status != "Paid" else 0
        if balance > 0:
            result.append({
                "id":      inv.id,
                "ref":     inv.invoice_id,
                "date":    inv.date.strftime("%d %b %Y") if inv.date else "",
                "total":   total,
                "balance": balance,
            })
    return result


def _outstanding_invoices_for_supplier(company_id, supplier_id):
    """Return list of dicts for purchase invoices with a remaining balance.

    Same fix as _outstanding_invoices_for_client above: filter on balance,
    not on a status whitelist that excludes "Draft" purchase invoices.
    """
    cdb = get_cdb()
    invs = (cdb.query(PurchaseInvoice)
            .filter_by(company_id=company_id, supplier_id=supplier_id)
            .filter(PurchaseInvoice.status != "Paid")
            .order_by(PurchaseInvoice.date.asc())
            .all())
    result = []
    for inv in invs:
        total   = inv.grand_total or 0
        balance = inv.balance or total
        if balance > 0:
            result.append({
                "id":      inv.id,
                "ref":     inv.invoice_number or inv.invoice_id,
                "date":    inv.date.strftime("%d %b %Y") if inv.date else "",
                "total":   total,
                "balance": balance,
            })
    return result


def _build_invoices_json(company_id, entities, fetch_fn):
    """Build {entity_id: [invoice list]} dict for JS."""
    data = {}
    for e in entities:
        data[str(e.id)] = fetch_fn(company_id, e.id)
    return json.dumps(data)


# ── Customer-invoice aware receivables (used by the debtor Receipts screen) ──
# Bookings can now be rolled up into a CustomerInvoice (see
# customer_invoice_create). Once that happens the booking must stop being
# separately payable through the raw per-booking picker — otherwise the same
# money could be applied twice, once against the booking and once against
# the customer invoice it belongs to. `_outstanding_invoices_for_client` /
# `_outstanding_invoices_for_supplier` above are left untouched (the Cheques
# page still uses them for plain booking-level cheque linking); everything
# below is additive and only wired into the Receipts screen.

def _used_booking_ids_in_customer_invoices(cdb, company_id):
    """Booking (Invoice) ids already rolled into a live (non-Void) customer
    invoice. Same scan customer_invoice_new/customer_invoice_create already
    do to stop a booking being added to two customer invoices — reused here
    to keep the receipt picker in sync with that rule."""
    ids = set()
    cis = cdb.query(CustomerInvoice).filter(
        CustomerInvoice.company_id == company_id,
        CustomerInvoice.status != "Void",
        CustomerInvoice.booking_ids_json.isnot(None),
    ).all()
    for ci in cis:
        try:
            ids.update(json.loads(ci.booking_ids_json))
        except (ValueError, TypeError):
            continue
    return ids


def _receivables_for_client(company_id, client_id):
    """Everything a client can settle a receipt against: outstanding credit
    customer invoices only. Raw bookings (Invoice rows) never appear here
    on their own — a booking becomes payable through the Receipts screen
    only once it's grouped into a CustomerInvoice."""
    cdb = get_cdb()

    rows = []

    cis = (cdb.query(CustomerInvoice)
           .filter_by(company_id=company_id, client_id=client_id, invoice_type="credit")
           .filter(CustomerInvoice.status != "Void")
           .order_by(CustomerInvoice.invoice_date.asc())
           .all())
    for ci in cis:
        total = ci.grand_total or 0
        balance = ci.balance if ci.balance is not None else max(0, total - (ci.paid_amount or 0))
        if balance > 0:
            try:
                booking_count = len(json.loads(ci.booking_ids_json)) if ci.booking_ids_json else 0
            except (ValueError, TypeError):
                booking_count = 0
            rows.append({
                "kind":     "customer_invoice",
                "id":       ci.id,
                "ref":      ci.invoice_number,
                "date":     ci.invoice_date.strftime("%d %b %Y") if ci.invoice_date else "",
                "sort_dt":  ci.invoice_date or date.min,
                "total":    total,
                "balance":  balance,
                "bookings": booking_count,
            })

    rows.sort(key=lambda r: r["sort_dt"])
    for r in rows:
        r.pop("sort_dt", None)
    return rows


def _expand_ci_token(cdb, company_id, entity_id, ci_id, ci_for_booking, touched_ci_ids):
    """Resolve a 'ci:<id>' receipt token into its constituent booking ids,
    oldest booking first (same order the raw per-booking loop already
    applies money in). Records which CustomerInvoice each booking belongs
    to so receipt_save can enrich the ledger narration and, once the main
    loop is done, re-sync the customer invoice's own paid_amount/balance/
    status from what actually landed on its bookings."""
    ci = cdb.query(CustomerInvoice).filter_by(
        id=ci_id, company_id=company_id, client_id=entity_id
    ).first()
    if not ci:
        return []
    touched_ci_ids.add(ci.id)
    try:
        booking_ids = json.loads(ci.booking_ids_json) if ci.booking_ids_json else []
    except (ValueError, TypeError):
        booking_ids = []
    if not booking_ids:
        return []
    bookings = (cdb.query(Invoice)
                .filter(Invoice.id.in_(booking_ids), Invoice.company_id == company_id)
                .order_by(Invoice.date.asc())
                .all())
    expanded = []
    for b in bookings:
        expanded.append(b.id)
        ci_for_booking[b.id] = ci
    return expanded


def _sync_customer_invoice_payment(cdb, company_id, ci):
    """Re-derive a customer invoice's paid_amount/balance/status from the
    current paid_amount of its own constituent bookings. The customer
    invoice is a view over its bookings' payment state, not a separate
    ledger — it must never drift from them."""
    try:
        booking_ids = json.loads(ci.booking_ids_json) if ci.booking_ids_json else []
    except (ValueError, TypeError):
        booking_ids = []
    bookings = (cdb.query(Invoice)
                .filter(Invoice.id.in_(booking_ids), Invoice.company_id == company_id)
                .all()) if booking_ids else []
    ci.paid_amount = sum((b.paid_amount or 0) for b in bookings)
    ci.balance = max(0, (ci.grand_total or 0) - ci.paid_amount)
    if ci.balance <= 0:
        ci.status = "Paid"
    elif ci.paid_amount > 0:
        ci.status = "Partial"
    else:
        ci.status = "Pending"


@app.route("/receipts/new")
@login_required
@require_permission("receipts_payments", "view")
def receipt_new():
    cdb        = get_cdb()
    company_id = get_current_company()
    all_clients    = cdb.query(Client).filter_by(company_id=company_id).order_by(Client.name).all()
    bank_accounts  = cdb.query(BankAccount).filter_by(company_id=company_id, status='Active').all()
    selected_id    = request.args.get("client_id", type=int)
    invoices_json  = _build_invoices_json(company_id, all_clients, _receivables_for_client)

    # Date-wise filter for history
    date_from_str = request.args.get("date_from", "")
    date_to_str   = request.args.get("date_to", "")
    date_from = date.fromisoformat(date_from_str) if date_from_str else None
    date_to   = date.fromisoformat(date_to_str) if date_to_str else None
    has_date_filter = bool(date_from or date_to)

    # Search filter for history — matches reference, description, notes, or exact amount
    search_q = request.args.get("q", "").strip()
    hist_limit = 1000 if (has_date_filter or search_q) else 100

    def _apply_search(query, model):
        if not search_q:
            return query
        like_q = f"%{search_q}%"
        conditions = [
            model.reference.ilike(like_q),
            model.description.ilike(like_q),
            model.notes.ilike(like_q),
            model.party_name.ilike(like_q),
        ]
        try:
            conditions.append(model.amount == float(search_q))
        except ValueError:
            pass
        return query.filter(or_(*conditions))

    # Build receipt history — cash + bank transactions
    history = []
    cash_q = cdb.query(CashTransaction).filter_by(company_id=company_id, category="Receipt")
    if date_from:
        cash_q = cash_q.filter(CashTransaction.date >= date_from)
    if date_to:
        cash_q = cash_q.filter(CashTransaction.date <= date_to)
    cash_q = _apply_search(cash_q, CashTransaction)
    cash_receipts = cash_q.order_by(CashTransaction.date.desc()).limit(hist_limit).all()
    for t in cash_receipts:
        history.append({
            "id":          t.id,
            "txn_type":    "cash",
            "date":        t.date.strftime("%d %b %Y") if t.date else "",
            "sort_date":   t.date or date.min,
            "reference":   t.reference or "—",
            "client":      t.party_name or "—",
            "description": t.description,
            "amount":      t.amount,
            "mode":        "Cash",
            "bank_name":   "Cash in Hand",
            "notes":       t.notes or "",
        })

    bank_q = (cdb.query(BankTransaction)
              .filter_by(company_id=company_id, type="credit"))
              
    if date_from:
        bank_q = bank_q.filter(BankTransaction.date >= date_from)
    if date_to:
        bank_q = bank_q.filter(BankTransaction.date <= date_to)
    bank_q = _apply_search(bank_q, BankTransaction)
    bank_receipts = bank_q.order_by(BankTransaction.date.desc()).limit(hist_limit).all()
    for t in bank_receipts:
        bank_name = ""
        if t.bank_account:
            bank_name = f"{t.bank_account.bank_name} – {t.bank_account.account_name}"
        history.append({
            "id":          t.id,
            "txn_type":    "bank",
            "date":        t.date.strftime("%d %b %Y") if t.date else "",
            "sort_date":   t.date or date.min,
            "reference":   t.reference or "—",
            "client":      t.party_name or "—",
            "description": t.description,
            "amount":      t.amount,
            "mode":        t.transaction_mode or "Bank",
            "bank_name":   bank_name,
            "notes":       t.notes or "",
        })
    history.sort(key=lambda x: x["sort_date"], reverse=True)

    return render_template(
        "record_receipt.html",
        entities=all_clients,
        bank_accounts=bank_accounts,
        invoices_json=invoices_json,
        selected_id=selected_id,
        today=str(today_ist()),
        history=history,
        date_from=date_from_str,
        date_to=date_to_str,
        search_q=search_q,
    )


@app.route("/receipts/save", methods=["POST"])
@login_required
@require_permission("receipts_payments", "create")
def receipt_save():
    cdb         = get_cdb()
    company_id  = get_current_company()
    entity_id   = request.form.get("entity_id", type=int)
    amount      = request.form.get("amount", type=float, default=0)
    narration   = request.form.get("narration", "")
    pay_mode    = request.form.get("pay_mode", "Cash")
    bank_account_id = request.form.get("bank_account_id", type=int)
    txn_date_str = request.form.get("txn_date")
    txn_date    = date.fromisoformat(txn_date_str) if txn_date_str else today_ist()

    if not entity_id or amount <= 0:
        flash("Please select a client and enter a valid amount.", "error")
        return redirect(url_for("receipt_new"))

    client = cdb.query(Client).filter_by(id=entity_id, company_id=company_id).first()
    client_name = get_party_name(client_id=entity_id)

    # Validate bank account required for non-cash
    bank_account = None
    if pay_mode.lower() != "cash":
        if not bank_account_id:
            flash("Please select a bank account for non-cash payments.", "error")
            return redirect(url_for("receipt_new"))
        bank_account = cdb.query(BankAccount).filter_by(
            id=bank_account_id, company_id=company_id, status='Active'
        ).first()
        if not bank_account:
            flash("Selected bank account not found or inactive.", "error")
            return redirect(url_for("receipt_new"))

    # Selected rows arrive as tokens: "ci:<id>" for a customer invoice,
    # "b:<id>" for a raw booking (a bare numeric id is treated as a raw
    # booking too, for backward compatibility with any already-rendered
    # form). A "ci:" token is expanded into its constituent bookings here
    # so the rest of this function — the apply/settle loop, cash/bank
    # ledger writes, client.pending update — runs exactly as it always has,
    # unchanged, regardless of which picker row the money came in against.
    ci_for_booking = {}   # booking Invoice.id -> its CustomerInvoice
    touched_ci_ids = set()
    invoice_ids = []
    for tok in request.form.get("invoice_ids", "").split(","):
        tok = tok.strip()
        if not tok:
            continue
        if tok.startswith("ci:"):
            try:
                ci_id = int(tok.split(":", 1)[1])
            except ValueError:
                continue
            invoice_ids.extend(
                _expand_ci_token(cdb, company_id, entity_id, ci_id, ci_for_booking, touched_ci_ids)
            )
        elif tok.startswith("b:"):
            try:
                invoice_ids.append(int(tok.split(":", 1)[1]))
            except ValueError:
                continue
        else:
            try:
                invoice_ids.append(int(tok))
            except ValueError:
                continue

    if not invoice_ids:
        for r in _receivables_for_client(company_id, entity_id):
            if r["kind"] == "customer_invoice":
                invoice_ids.extend(
                    _expand_ci_token(cdb, company_id, entity_id, r["id"], ci_for_booking, touched_ci_ids)
                )
            else:
                invoice_ids.append(r["id"])

    remaining = amount
    settled   = 0

    for inv_id in invoice_ids:
        if remaining <= 0:
            break
        inv = cdb.query(Invoice).filter_by(id=inv_id, company_id=company_id).first()
        if not inv:
            continue

        inv_balance = getattr(inv, "balance", None)
        if inv_balance is None:
            inv_balance = inv.grand_total or 0

        apply        = min(remaining, inv_balance)
        remaining   -= apply
        inv_balance -= apply
        settled     += apply

        if hasattr(inv, "balance"):
            inv.balance = inv_balance
        if hasattr(inv, "paid_amount"):
            inv.paid_amount = (inv.paid_amount or 0) + apply

        if inv_balance <= 0:
            inv.status = "Paid"
        elif apply > 0:
            inv.status = "Partial"

        if apply > 0:
            ci = ci_for_booking.get(inv.id)
            ci_suffix = f" (Customer Invoice {ci.invoice_number})" if ci else ""
            # If this booking is inside a customer invoice, the statement
            # shows one debit line for the customer invoice, not the raw
            # booking — so the credit line's reference must match the
            # customer invoice number too, or the two rows won't visually
            # pair up on the ledger.
            txn_reference = ci.invoice_number if ci else inv.invoice_id
            if pay_mode.lower() == "cash":
                # Record in Cash in Hand
                cash_txn = CashTransaction(
                    company_id=company_id,
                    type="income",
                    date=txn_date,
                    category="Receipt",
                    description=f"Payment received for invoice {inv.invoice_id}{ci_suffix} - {narration}",
                    amount=apply,
                    reference=txn_reference,
                    notes=f"Payment from client via Cash",
                    party_name=client_name,
                    created_by=get_current_user().get('email'),
                    applied_ref_type="invoice",
                    applied_ref_id=inv.id,
                    applied_ci_id=ci.id if ci else None,
                )
                cdb.add(cash_txn)
            else:
                # Record in the chosen bank account
                bank_txn = BankTransaction(
                    bank_account_id=bank_account.id,
                    company_id=company_id,
                    type="credit",
                    date=txn_date,
                    description=f"Payment received for invoice {inv.invoice_id}{ci_suffix}",
                    amount=apply,
                    reference=txn_reference,
                    transaction_mode=pay_mode.title(),
                    notes=narration,
                    party_name=client_name,
                    created_by=get_current_user().get('email'),
                    applied_ref_type="invoice",
                    applied_ref_id=inv.id,
                    applied_ci_id=ci.id if ci else None,
                )
                cdb.add(bank_txn)
                bank_account.balance += apply

    # BUG FIX: previously, any amount left over after settling the selected
    # invoices' balances was never recorded anywhere. If a client had no
    # outstanding invoices (or the amount received was more than their total
    # balance due), `remaining` stayed > 0, no CashTransaction/BankTransaction
    # was ever created for it, and the flash message reported `settled`
    # (money actually applied to an invoice) instead of `amount` (money the
    # user said they received) — so it could show "Receipt of ₹0.00 recorded"
    # while the real amount the person typed in just vanished with no ledger
    # entry and no error. Record any unapplied leftover as its own advance
    # receipt so the cash/bank ledger always reflects the full amount received.
    if remaining > 0:
        advance_desc = f"Advance receipt from client (not applied to a specific invoice) - {narration}".strip(" -")
        if pay_mode.lower() == "cash":
            cash_txn = CashTransaction(
                company_id=company_id,
                type="income",
                date=txn_date,
                category="Receipt",
                description=advance_desc,
                amount=remaining,
                reference="ADVANCE",
                notes="Unapplied portion of receipt via Cash",
                party_name=client_name,
                created_by=get_current_user().get('email')
            )
            cdb.add(cash_txn)
        else:
            bank_txn = BankTransaction(
                bank_account_id=bank_account.id,
                company_id=company_id,
                type="credit",
                date=txn_date,
                description=advance_desc,
                amount=remaining,
                reference="ADVANCE",
                transaction_mode=pay_mode.title(),
                notes=narration,
                party_name=client_name,
                created_by=get_current_user().get('email')
            )
            cdb.add(bank_txn)
            bank_account.balance += remaining

    if client and hasattr(client, "pending") and client.pending:
        client.pending = max(0, (client.pending or 0) - settled)

    # ── UPDATE CLIENT LAST PAYMENT DATE ──────────────────────────────
    if client:
        # Find the most recent payment date from CashTransaction or BankTransaction
        last_cash = cdb.query(CashTransaction).filter_by(
            company_id=company_id,
            party_name=client_name,
            category="Receipt"
        ).order_by(CashTransaction.date.desc()).first()
        
        last_bank = cdb.query(BankTransaction).filter_by(
            company_id=company_id,
            party_name=client_name,
            type="credit"
        ).order_by(BankTransaction.date.desc()).first()
        
        if last_cash and last_bank:
            client.last_payment = last_cash.date if last_cash.date > last_bank.date else last_bank.date
        elif last_cash:
            client.last_payment = last_cash.date
        elif last_bank:
            client.last_payment = last_bank.date    

    # ── Re-sync any customer invoices this receipt touched ───────────────
    # The loop above already updated each underlying booking's own
    # paid_amount/balance/status; pull the customer invoice's totals back
    # in line with that so it never shows a balance its bookings disagree
    # with.
    for ci_id in touched_ci_ids:
        ci = cdb.query(CustomerInvoice).filter_by(id=ci_id, company_id=company_id).first()
        if ci:
            _sync_customer_invoice_payment(cdb, company_id, ci)

    cdb.commit()

    dest = bank_account.bank_name if bank_account else "Cash in Hand"
    flash(f"Receipt of ₹{amount:,.2f} recorded via {pay_mode} → {dest}. {narration}", "success")
    return redirect(url_for("debtors_list"))


@app.route("/payments/new")
@login_required
@require_permission("receipts_payments", "view")
def payment_new():
    cdb        = get_cdb()
    company_id = get_current_company()
    all_suppliers  = cdb.query(Supplier).filter_by(company_id=company_id).order_by(Supplier.name).all()
    bank_accounts  = cdb.query(BankAccount).filter_by(company_id=company_id, status='Active').all()
    selected_id    = request.args.get("supplier_id", type=int)
    invoices_json  = _build_invoices_json(company_id, all_suppliers, _outstanding_invoices_for_supplier)

    # Date-wise filter for history
    date_from_str = request.args.get("date_from", "")
    date_to_str   = request.args.get("date_to", "")
    date_from = date.fromisoformat(date_from_str) if date_from_str else None
    date_to   = date.fromisoformat(date_to_str) if date_to_str else None
    has_date_filter = bool(date_from or date_to)

    # Search filter for history — matches reference, description, notes, or exact amount
    search_q = request.args.get("q", "").strip()
    hist_limit = 1000 if (has_date_filter or search_q) else 100

    def _apply_search(query, model):
        if not search_q:
            return query
        like_q = f"%{search_q}%"
        conditions = [
            model.reference.ilike(like_q),
            model.description.ilike(like_q),
            model.notes.ilike(like_q),
            model.party_name.ilike(like_q),
        ]
        try:
            conditions.append(model.amount == float(search_q))
        except ValueError:
            pass
        return query.filter(or_(*conditions))

    # Build payment history — cash + bank transactions
    history = []
    cash_q = cdb.query(CashTransaction).filter_by(company_id=company_id, category="Payment")
    if date_from:
        cash_q = cash_q.filter(CashTransaction.date >= date_from)
    if date_to:
        cash_q = cash_q.filter(CashTransaction.date <= date_to)
    cash_q = _apply_search(cash_q, CashTransaction)
    cash_payments = cash_q.order_by(CashTransaction.date.desc()).limit(hist_limit).all()
    for t in cash_payments:
        history.append({
            "id":          t.id,
            "txn_type":    "cash",
            "date":        t.date.strftime("%d %b %Y") if t.date else "",
            "sort_date":   t.date or date.min,
            "reference":   t.reference or "—",
            "supplier":    t.party_name or "—",
            "description": t.description,
            "amount":      t.amount,
            "mode":        "Cash",
            "bank_name":   "Cash in Hand",
            "notes":       t.notes or "",
        })

    bank_q = (cdb.query(BankTransaction)
              .filter_by(company_id=company_id, type="debit"))
    if date_from:
        bank_q = bank_q.filter(BankTransaction.date >= date_from)
    if date_to:
        bank_q = bank_q.filter(BankTransaction.date <= date_to)
    bank_q = _apply_search(bank_q, BankTransaction)
    bank_payments = bank_q.order_by(BankTransaction.date.desc()).limit(hist_limit).all()
    for t in bank_payments:
        bank_name = ""
        if t.bank_account:
            bank_name = f"{t.bank_account.bank_name} – {t.bank_account.account_name}"
        history.append({
            "id":          t.id,
            "txn_type":    "bank",
            "date":        t.date.strftime("%d %b %Y") if t.date else "",
            "sort_date":   t.date or date.min,
            "reference":   t.reference or "—",
            "supplier":    t.party_name or "—",
            "description": t.description,
            "amount":      t.amount,
            "mode":        t.transaction_mode or "Bank",
            "bank_name":   bank_name,
            "notes":       t.notes or "",
        })
    history.sort(key=lambda x: x["sort_date"], reverse=True)

    return render_template(
        "record_payment.html",
        entities=all_suppliers,
        bank_accounts=bank_accounts,
        invoices_json=invoices_json,
        selected_id=selected_id,
        today=str(today_ist()),
        history=history,
        date_from=date_from_str,
        date_to=date_to_str,
        search_q=search_q,
    )


@app.route("/payments/save", methods=["POST"])
@login_required
@require_permission("receipts_payments", "create")
def payment_save():
    cdb = get_cdb()
    company_id = get_current_company()
    entity_id = request.form.get("entity_id", type=int)
    amount = request.form.get("amount", type=float, default=0)
    invoice_ids = [int(x) for x in request.form.get("invoice_ids", "").split(",") if x.strip()]
    narration = request.form.get("narration", "")
    pay_mode = request.form.get("pay_mode", "Cash")
    bank_account_id = request.form.get("bank_account_id", type=int)
    txn_date_str = request.form.get("txn_date")
    txn_date = date.fromisoformat(txn_date_str) if txn_date_str else today_ist()

    if not entity_id or amount <= 0:
        flash("Please select a supplier and enter a valid amount.", "error")
        return redirect(url_for("payment_new"))

    supplier_entity = cdb.query(Supplier).filter_by(id=entity_id, company_id=company_id).first()
    if not supplier_entity:
        flash("Supplier not found.", "error")
        return redirect(url_for("payment_new"))

    supplier_name = get_party_name(supplier_id=entity_id)

    # Validate bank account required for non-cash
    bank_account = None
    if pay_mode.lower() != "cash":
        if not bank_account_id:
            flash("Please select a bank account for non-cash payments.", "error")
            return redirect(url_for("payment_new"))
        bank_account = cdb.query(BankAccount).filter_by(
            id=bank_account_id, company_id=company_id, status='Active'
        ).first()
        if not bank_account:
            flash("Selected bank account not found or inactive.", "error")
            return redirect(url_for("payment_new"))

    # If no invoices selected, get all outstanding invoices for this supplier
    if not invoice_ids:
        rows = _outstanding_invoices_for_supplier(company_id, entity_id)
        invoice_ids = [r["id"] for r in rows]

    remaining = amount
    settled = 0
    applied_invoice_ids = []

    # ── APPLY PAYMENT TO INVOICES, ONE LEDGER ENTRY PER INVOICE ──
    # Previously this loop only updated invoice fields, then a SINGLE
    # lumped CashTransaction/BankTransaction was written afterwards
    # referencing just the first invoice touched. That made a payment
    # spread across several invoices impossible to reverse correctly —
    # deleting that one row could only ever undo invoice #1, leaving the
    # others permanently marked paid with no transaction behind them.
    # Writing one transaction per invoice (same pattern receipt_save
    # already uses on the receipts side) makes every rupee traceable to
    # exactly one invoice, and therefore correctly reversible.
    for inv_id in invoice_ids:
        if remaining <= 0:
            break
        inv = cdb.query(PurchaseInvoice).filter_by(id=inv_id, company_id=company_id).first()
        if not inv:
            continue

        inv_balance = inv.balance or (inv.grand_total or 0)
        if inv_balance <= 0:
            continue

        apply_amount = min(remaining, inv_balance)
        remaining -= apply_amount
        settled += apply_amount
        applied_invoice_ids.append(inv_id)

        # Update invoice
        inv.balance = inv_balance - apply_amount
        inv.paid_amount = (inv.paid_amount or 0) + apply_amount

        if inv.balance <= 0:
            inv.status = "Paid"
        elif inv.paid_amount > 0:
            inv.status = "Partial"
        else:
            inv.status = "Pending"

        # Update supplier payable
        if inv.supplier:
            inv.supplier.payable = max(0, (inv.supplier.payable or 0) - apply_amount)

        txn_reference = inv.invoice_number or inv.invoice_id
        desc = f"Payment made for purchase invoice {txn_reference}"
        if narration:
            desc += f" - {narration}"

        if pay_mode.lower() == "cash":
            cash_txn = CashTransaction(
                company_id=company_id,
                type="expense",
                date=txn_date,
                category="Payment",
                description=desc,
                amount=apply_amount,
                reference=txn_reference,
                notes=f"Payment of ₹{apply_amount:,.2f} to supplier via Cash",
                party_name=supplier_name,
                created_by=get_current_user().get('email'),
                applied_ref_type="purchase_invoice",
                applied_ref_id=inv.id,
            )
            cdb.add(cash_txn)
        else:
            bank_txn = BankTransaction(
                bank_account_id=bank_account.id,
                company_id=company_id,
                type="debit",
                date=txn_date,
                description=desc,
                amount=apply_amount,
                reference=txn_reference,
                transaction_mode=pay_mode.title(),
                notes=narration,
                party_name=supplier_name,
                created_by=get_current_user().get('email'),
                applied_ref_type="purchase_invoice",
                applied_ref_id=inv.id,
            )
            cdb.add(bank_txn)

    # ── ANY LEFTOVER AFTER SETTLING SELECTED INVOICES → ADVANCE ──
    # Recorded as its own transaction, unattached to any invoice
    # (applied_ref_id stays NULL), so it deletes cleanly with nothing to
    # reverse — matching the "ADVANCE" convention receipt_save uses.
    if remaining > 0:
        advance_desc = f"Advance payment to supplier (not applied to a specific invoice)"
        if narration:
            advance_desc += f" - {narration}"
        if pay_mode.lower() == "cash":
            cash_txn = CashTransaction(
                company_id=company_id,
                type="expense",
                date=txn_date,
                category="Payment",
                description=advance_desc,
                amount=remaining,
                reference="ADVANCE",
                notes=f"Unapplied portion of payment via Cash",
                party_name=supplier_name,
                created_by=get_current_user().get('email'),
            )
            cdb.add(cash_txn)
        else:
            bank_txn = BankTransaction(
                bank_account_id=bank_account.id,
                company_id=company_id,
                type="debit",
                date=txn_date,
                description=advance_desc,
                amount=remaining,
                reference="ADVANCE",
                transaction_mode=pay_mode.title(),
                notes=narration,
                party_name=supplier_name,
                created_by=get_current_user().get('email'),
            )
            cdb.add(bank_txn)

    # A non-cash payment debits the bank account by the FULL amount
    # entered, regardless of how it was split across invoices/advance.
    if pay_mode.lower() != "cash" and amount > 0:
        bank_account.balance -= amount

    cdb.commit()

    # ── FLASH MESSAGE ──
    dest = bank_account.bank_name if bank_account else "Cash in Hand"
    if settled > 0 and amount - settled > 0:
        flash(f"✅ Payment of ₹{amount:,.2f} recorded via {pay_mode} → {dest}. ₹{settled:,.2f} applied to invoices, ₹{amount - settled:,.2f} recorded as advance/unapplied. {narration}", "success")
    elif settled > 0:
        flash(f"✅ Payment of ₹{settled:,.2f} applied to invoices via {pay_mode} → {dest}. {narration}", "success")
    else:
        flash(f"✅ Advance payment of ₹{amount:,.2f} recorded via {pay_mode} → {dest}. {narration}", "success")

    return redirect(url_for("creditors_list"))

# ─────────────────────────────────────────────────────────────────────────────
# ── Backup & Restore Routes ───────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────

@app.route("/backup")
@login_required
@require_permission("backup", "view")
def backup():
    """Backup management page"""
    company_id = get_current_company()
    
    # Define backup destinations (fallback)
    backup_destinations = {
        "local": "Local Storage",
        "s3": "Amazon S3",
        "gcs": "Google Cloud Storage",
        "ftp": "FTP/SFTP Server",
    }
    
    backups = []
    
    try:
        from backup_utils import list_backups, BACKUP_DESTINATIONS
        backups = list_backups(company_id)
        backup_destinations = BACKUP_DESTINATIONS
    except ImportError as e:
        print(f"Could not import backup_utils: {e}")
        flash("Backup utilities not fully configured. Some features may be limited.", "warning")
    except Exception as e:
        print(f"Error loading backups: {e}")
        flash(f"Error loading backups: {str(e)}", "error")
    
    return render_template("backup.html", 
                         active='backup',
                         backups=backups,
                         backup_destinations=backup_destinations)
                         

@app.route("/backup/create", methods=["POST"])
@login_required
@require_permission("backup", "create")
def create_backup():
    """Create a new backup with optional date range"""
    company_id = get_current_company()
    include_attachments = request.form.get("include_attachments", "true") == "true"
    
    # Get date range from form
    from_date_str = request.form.get("from_date", "").strip()
    to_date_str = request.form.get("to_date", "").strip()
    
    from_date = None
    to_date = None
    
    if from_date_str:
        try:
            from_date = date.fromisoformat(from_date_str)
        except ValueError:
            flash("Invalid from date format. Please use YYYY-MM-DD.", "error")
            return redirect(url_for("backup"))
    
    if to_date_str:
        try:
            to_date = date.fromisoformat(to_date_str)
        except ValueError:
            flash("Invalid to date format. Please use YYYY-MM-DD.", "error")
            return redirect(url_for("backup"))
    
    # Validate date range
    if from_date and to_date and from_date > to_date:
        flash("From date cannot be after to date.", "error")
        return redirect(url_for("backup"))
    
    try:
        from backup_utils import create_company_backup, BACKUP_DESTINATIONS
        
        backup_info = create_company_backup(
            company_id, 
            include_attachments,
            from_date=from_date,
            to_date=to_date
        )
        
        # Build date range message
        date_msg = ""
        if from_date and to_date:
            date_msg = f" (from {from_date.strftime('%d %b %Y')} to {to_date.strftime('%d %b %Y')})"
        elif from_date:
            date_msg = f" (from {from_date.strftime('%d %b %Y')})"
        elif to_date:
            date_msg = f" (up to {to_date.strftime('%d %b %Y')})"
        
        flash(f"Backup created successfully! File size: {backup_info['size_mb']} MB{date_msg}", "success")
        
        # Optionally upload to cloud
        if request.form.get("upload_to_cloud"):
            destination = request.form.get("cloud_destination")
            config = {
                'access_key': request.form.get('access_key'),
                'secret_key': request.form.get('secret_key'),
                'bucket': request.form.get('bucket'),
                'region': request.form.get('region', 'us-east-1')
            }
            from backup_utils import upload_backup_to_cloud
            upload_backup_to_cloud(backup_info['backup_id'], destination, config)
            flash("Backup also uploaded to cloud storage!", "success")
            
    except Exception as e:
        flash(f"Backup failed: {str(e)}", "error")
    
    return redirect(url_for("backup"))

@app.route("/backup/restore/<backup_id>", methods=["POST"])
@login_required
@require_permission("backup", "edit")
def restore_backup(backup_id):
    """Restore from a backup"""
    company_id = get_current_company()
    user = get_current_user()
    
    try:
        from backup_utils import restore_from_backup
        result = restore_from_backup(backup_id, user.get('email'))
        
        flash(f"Restore completed successfully! Company data restored from backup {backup_id}", "success")
        
    except Exception as e:
        flash(f"Restore failed: {str(e)}", "error")
    
    return redirect(url_for("backup"))

@app.route("/backup/download/<backup_id>")
@login_required
@require_permission("backup", "view")
def download_backup(backup_id):
    """Download backup file"""
    company_id = get_current_company()
    
    from platform_models import BackupRecord
    backup = BackupRecord.query.filter_by(backup_id=backup_id, company_id=company_id).first()
    
    if not backup or not os.path.exists(backup.backup_file_path):
        flash("Backup file not found", "error")
        return redirect(url_for("backup"))
    
    return send_file(
        backup.backup_file_path,
        as_attachment=True,
        download_name=f"{backup_id}.zip"
    )

@app.route("/backup/delete/<backup_id>", methods=["POST"])
@login_required
@owner_required
def delete_backup_record(backup_id):
    """Delete a backup"""
    company_id = get_current_company()
    
    try:
        from backup_utils import delete_backup
        if delete_backup(backup_id):
            flash("Backup deleted successfully", "success")
        else:
            flash("Backup not found", "error")
    except Exception as e:
        flash(f"Error deleting backup: {str(e)}", "error")
    
    return redirect(url_for("backup"))

@app.route("/backup/schedule", methods=["POST"])
@login_required
@require_permission("backup", "edit")
def schedule_backup():
    """Schedule automatic backups"""
    company_id = get_current_company()
    
    frequency = request.form.get("frequency")
    time_of_day = request.form.get("time_of_day")
    retention_days = request.form.get("retention_days", 30)
    upload_to_cloud = request.form.get("upload_to_cloud") == "true"
    
    from platform_models import BackupSchedule
    
    # Save schedule to database
    schedule = BackupSchedule.query.filter_by(company_id=company_id).first()
    
    if not schedule:
        schedule = BackupSchedule(company_id=company_id)
        db.session.add(schedule)
    
    schedule.frequency = frequency
    schedule.time_of_day = time_of_day
    schedule.retention_days = int(retention_days)
    schedule.upload_to_cloud = upload_to_cloud
    schedule.last_backup = None
    
    # Calculate next backup
    from datetime import datetime, timedelta
    now = datetime.now()
    hour, minute = map(int, time_of_day.split(':'))
    
    if frequency == "daily":
        next_date = now.replace(hour=hour, minute=minute, second=0, microsecond=0)
        if next_date <= now:
            next_date += timedelta(days=1)
    elif frequency == "weekly":
        days_ahead = 6 - now.weekday()
        next_date = (now + timedelta(days=days_ahead)).replace(hour=hour, minute=minute, second=0, microsecond=0)
    elif frequency == "monthly":
        next_date = now.replace(day=1, hour=hour, minute=minute, second=0, microsecond=0)
        if next_date <= now:
            if next_date.month == 12:
                next_date = next_date.replace(year=next_date.year + 1, month=1)
            else:
                next_date = next_date.replace(month=next_date.month + 1)
    else:
        next_date = now + timedelta(days=1)
    
    schedule.next_backup = next_date
    schedule.is_active = True
    
    db.session.commit()
    flash(f"Automatic backup scheduled {frequency} at {time_of_day}", "success")
    
    return redirect(url_for("backup"))

@app.route("/backup/upload-to-cloud/<backup_id>", methods=["POST"])
@login_required
@require_permission("backup", "edit")
def upload_backup_to_cloud_route(backup_id):
    """Upload existing backup to cloud"""
    company_id = get_current_company()
    
    destination = request.form.get("destination")
    config = {
        'access_key': request.form.get('access_key'),
        'secret_key': request.form.get('secret_key'),
        'bucket': request.form.get('bucket'),
        'region': request.form.get('region', 'us-east-1'),
        'host': request.form.get('host'),
        'port': request.form.get('port', 22),
        'username': request.form.get('username'),
        'password': request.form.get('password'),
        'path': request.form.get('path', '/'),
        'credentials_file': request.form.get('credentials_file'),
    }
    
    try:
        from backup_utils import upload_backup_to_cloud
        upload_backup_to_cloud(backup_id, destination, config)
        flash("Backup uploaded to cloud successfully!", "success")
    except Exception as e:
        flash(f"Cloud upload failed: {str(e)}", "error")
    
    return redirect(url_for("backup"))

# Start backup scheduler
try:
    from backup_scheduler import start_backup_scheduler
    start_backup_scheduler()
except Exception as e:
    print(f"Could not start backup scheduler: {e}")


@app.route("/backup/upload", methods=["POST"])
@login_required
@require_permission("backup", "edit")
def upload_backup():
    """
    Upload a backup .zip file and restore it.
    This allows restoring from a downloaded backup file.
    """
    company_id = get_current_company()
    user = get_current_user()
    
    if 'backup_file' not in request.files:
        flash("No file selected", "error")
        return redirect(url_for("backup"))
    
    file = request.files['backup_file']
    if file.filename == '':
        flash("No file selected", "error")
        return redirect(url_for("backup"))
    
    if not file.filename.endswith('.zip'):
        flash("Please upload a .zip backup file", "error")
        return redirect(url_for("backup"))
    
    try:
        from backup_utils import restore_from_uploaded_backup
        
        # Save uploaded file temporarily
        import tempfile
        import os
        
        temp_dir = tempfile.mkdtemp()
        temp_path = os.path.join(temp_dir, file.filename)
        file.save(temp_path)
        
        # Restore from the uploaded file
        result = restore_from_uploaded_backup(
            company_id=company_id, 
            file_path=temp_path,
            user_email=user.get('email', 'unknown')
        )
        
        # Clean up
        import shutil
        shutil.rmtree(temp_dir, ignore_errors=True)
        
        if result and result.get('success'):
            flash(f"✅ Restore completed successfully from {file.filename}", "success")
        else:
            error_msg = result.get('message', 'Unknown error') if result else 'Unknown error'
            flash(f"❌ Restore failed: {error_msg}", "error")
            
    except Exception as e:
        flash(f"❌ Restore failed: {str(e)}", "error")
        print(f"Upload restore error: {e}")
        import traceback
        traceback.print_exc()
    
    return redirect(url_for("backup"))

# ─────────────────────────────────────────────────────────────────────────────
# ── App entry point ───────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────
from platform_models import CompanyApiKey, generate_api_key
 
 
@app.route("/company/api-keys/generate", methods=["POST"])
@login_required
@owner_required   # matches the guard already used on update_company_info
def generate_company_api_key():
    company_id = get_current_company()
    label = request.form.get("label", "").strip() or None
    new_key, _row = generate_api_key(company_id, label=label)
    # Flash category "new_api_key" is what the template's one-time reveal
    # box looks for — it's consumed on the very next render, same as any
    # other flash message, so it can't be re-displayed by refreshing.
    flash(new_key, "new_api_key")
    return redirect(url_for("company_settings", tab="api"))
 
 
@app.route("/company/api-keys/<int:key_id>/revoke", methods=["POST"])
@login_required
@owner_required
def revoke_company_api_key(key_id):
    company_id = get_current_company()
    row = CompanyApiKey.query.filter_by(id=key_id, company_id=company_id).first()
    if not row:
        abort(404)
    row.is_active = False
    db.session.commit()
    flash(f"Revoked key {row.key_prefix}...")
    return redirect(url_for("company_settings", tab="api"))
# ═════════════════════════════════════════════════════════════════════════
# PUBLIC SHIPMENT TRACKING PAGE
# ═════════════════════════════════════════════════════════════════════════
from platform_models import TrackingIndex, CarrierTrackingConfig

try:
    from flask_limiter import Limiter
    from flask_limiter.util import get_remote_address
    limiter = Limiter(get_remote_address, app=app, default_limits=[])
    _limiter_available = True
except ImportError:
    print("[tracking] flask-limiter not installed — public tracking routes "
          "are running WITHOUT rate limiting. Run: pip install flask-limiter")
    _limiter_available = False

    class _NoOpLimiter:
        def limit(self, *a, **kw):
            def decorator(f):
                return f
            return decorator
    limiter = _NoOpLimiter()


def normalize_carrier(raw):
    return re.sub(r"[^A-Z0-9]", "", (raw or "").upper())


def sync_tracking_index(company_id, docket_no, carrier):
    """Fire-and-forget: never blocks or fails the caller."""
    docket_no = (docket_no or "").strip()
    if not docket_no:
        return
    try:
        row = TrackingIndex.query.filter_by(docket_no=docket_no).first()
        if row and row.company_id != company_id:
            print(f"[tracking-index] CONFLICT: docket {docket_no} already "
                  f"belongs to company {row.company_id}, ignoring write from {company_id}")
            return
        if row:
            row.carrier = carrier or row.carrier
        else:
            row = TrackingIndex(company_id=company_id, docket_no=docket_no, carrier=carrier)
            db.session.add(row)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f"[tracking-index] could not sync {docket_no} for {company_id}: {e}")


def get_shipment_status(cdb, docket_no):
    # Invoice.docket_no (the plain column) is NOT what's populated on booking —
    # your booking screen stores it inside the JSON terms blob instead, same as
    # every other docket lookup elsewhere in this codebase (see line ~790, 5777).
    invoice = cdb.query(Invoice).filter(
        Invoice.terms.like(f'%"docket_no": "{docket_no}"%')
    ).first()
    if not invoice:
        return None

    entry = cdb.query(ManifestEntry).filter_by(docket_no=docket_no).first()

    stages = [
        {"label": "Booked", "done": True, "at": invoice.created_at},
        {"label": "Ready for Dispatch", "done": bool(entry and entry.generated_at),
         "at": entry.generated_at if entry else None},
        {"label": f"In Transit to {entry.courier_name}" if entry and entry.courier_name else "In Transit",
         "done": bool(entry and entry.dispatched_at),
         "at": entry.dispatched_at if entry else None},
    ]

    carrier_redirect_url = None
    if entry and entry.dispatched_at and entry.courier_name:
        carrier_key = normalize_carrier(entry.courier_name)
        cfg = CarrierTrackingConfig.query.filter_by(carrier_key=carrier_key, is_active=True).first()
        if cfg:
            carrier_redirect_url = cfg.tracking_url_template.replace("{tracking_number}", docket_no)

    return {
        "docket_no": docket_no,
        "courier_name": entry.courier_name if entry else None,
        "stages": stages,
        "carrier_redirect_url": carrier_redirect_url,
    }


def _lookup_status(company_id, docket_no):
    cdb = get_customer_session(company_id)
    return get_shipment_status(cdb, docket_no)


@app.route("/manifest/entry/<int:entry_id>/dispatch", methods=["POST"])
@login_required
def mark_entry_dispatched(entry_id):
    company_id = session.get("company_id")
    cdb = get_customer_session(company_id)

    entry = cdb.query(ManifestEntry).filter_by(id=entry_id).first()
    if not entry:
        abort(404)
    if not entry.docket_no:
        flash("This manifest entry has no docket number — can't mark dispatched.")
        return redirect(request.referrer or url_for("manifest_list"))

    entry.dispatched_at = datetime.utcnow()
    entry.dispatched_by = session.get("username", "unknown")
    cdb.commit()

    sync_tracking_index(company_id, entry.docket_no, entry.courier_name)

    flash(f"Marked {entry.docket_no} as dispatched to {entry.courier_name}.")
    return redirect(request.referrer or url_for("manifest_list"))


@app.route("/track/<company_slug>", methods=["GET", "POST"])
@limiter.limit("10 per minute")
def public_tracking(company_slug):
    company = Company.query.filter_by(public_slug=company_slug).first_or_404()
    status = None
    error = None
    if request.method == "POST":
        docket_no = request.form.get("docket_no", "").strip()
        status = _lookup_status(company.company_id, docket_no)
        if not status:
            error = "No shipment found for that tracking number."
    return render_template("tracking_status.html", company=company, status=status, error=error)


@app.route("/t/<company_id>/<docket_no>")
@limiter.limit("30 per minute")
def track_magic_link(company_id, docket_no):
    status = _lookup_status(company_id, docket_no)
    if not status:
        abort(404)
    company = Company.query.filter_by(company_id=company_id).first()
    return render_template("tracking_status.html", company=company, status=status, error=None)


@app.route("/track", methods=["GET", "POST"])
@limiter.limit("10 per minute")
def public_tracking_generic():
    status = None
    error = None
    if request.method == "POST":
        docket_no = request.form.get("docket_no", "").strip()
        idx = TrackingIndex.query.filter_by(docket_no=docket_no).first()
        if not idx:
            error = "No shipment found for that tracking number."
        else:
            status = _lookup_status(idx.company_id, docket_no)
    return render_template("tracking_status.html", company=None, status=status, error=error)


def _find_purchase_invoice_for_reversal(cdb, company_id, txn):
    """Resolve the exact PurchaseInvoice a payment transaction should be
    reversed against. Prefers the structural applied_ref_id column
    (unambiguous, set by every payment recorded after this fix). Falls
    back to matching the display `reference` string against
    invoice_number/invoice_id for transactions recorded before the
    applied_ref_id column existed — never falls back to int(reference),
    since invoice_id/invoice_number are non-numeric strings (e.g.
    "PINV-0001") and that cast just silently failed, which is the bug
    this replaces.
    """
    if txn.applied_ref_type == "purchase_invoice" and txn.applied_ref_id:
        return cdb.query(PurchaseInvoice).filter_by(
            id=txn.applied_ref_id, company_id=company_id
        ).first()
    ref = txn.reference
    if ref and ref != "ADVANCE":
        return cdb.query(PurchaseInvoice).filter(
            PurchaseInvoice.company_id == company_id,
            or_(PurchaseInvoice.invoice_number == ref, PurchaseInvoice.invoice_id == ref)
        ).first()
    return None


@app.route("/payment/delete", methods=["GET"])
@login_required
@owner_required
def delete_payment():
    txn_id = request.args.get("id", type=int)
    txn_type = request.args.get("type", "cash")  # "cash" or "bank"
    company_id = get_current_company()
    cdb = get_cdb()

    if not company_id or not cdb:
        flash("Could not connect to database.", "error")
        return redirect(url_for("payment_new"))

    try:
        # ── 1. Find the transaction (cash or bank) ──
        if txn_type == "bank":
            txn = cdb.query(BankTransaction).filter_by(
                id=txn_id,
                company_id=company_id,
                type='debit'
            ).first()
        else:
            txn = cdb.query(CashTransaction).filter_by(
                id=txn_id,
                company_id=company_id,
                type='expense',
                category='Payment'
            ).first()

        if not txn:
            flash("Payment transaction not found.", "error")
            return redirect(url_for("payment_new"))

        amount = txn.amount
        supplier_name = txn.party_name

        print(f"[ADMIN] Deleting {txn_type} payment: ₹{amount} to {supplier_name}")

        # ── 2. Find and reset the invoice this payment was applied to ──
        inv = _find_purchase_invoice_for_reversal(cdb, company_id, txn)
        if inv:
            inv.paid_amount = max(0, (inv.paid_amount or 0) - amount)
            inv.balance = (inv.grand_total or 0) - (inv.paid_amount or 0)
            if inv.balance <= 0:
                inv.status = "Paid"
            elif inv.paid_amount > 0:
                inv.status = "Partial"
            else:
                inv.status = "Pending"

            # Reset supplier payable
            if inv.supplier_id:
                supplier = cdb.query(Supplier).filter_by(id=inv.supplier_id, company_id=company_id).first()
                if supplier:
                    supplier.payable = (supplier.payable or 0) + amount
            print(f"[ADMIN] Reset invoice {inv.invoice_id}")

        # ── 3. If this was a non-cash payment, credit the bank account back ──
        if txn_type == "bank" and getattr(txn, "bank_account", None):
            txn.bank_account.balance += amount

        # ── 4. Delete the transaction ──
        cdb.delete(txn)
        cdb.commit()

        flash(f"✅ Payment of ₹{amount:,.2f} to {supplier_name} has been deleted and reversed.", "success")

    except Exception as e:
        cdb.rollback()
        print(f"[ADMIN] Error deleting payment: {e}")
        import traceback
        traceback.print_exc()
        flash(f"Error deleting payment: {str(e)}", "error")

    return redirect(url_for("payment_new"))


def _find_receivable_for_reversal(cdb, company_id, txn):
    """Resolve the exact (booking) Invoice a receipt transaction should be
    reversed against, plus its parent CustomerInvoice if any (so that can
    be re-synced too). Same applied_ref_id-first, string-fallback approach
    as _find_purchase_invoice_for_reversal — see that function's docstring."""
    if txn.applied_ref_type == "invoice" and txn.applied_ref_id:
        inv = cdb.query(Invoice).filter_by(id=txn.applied_ref_id, company_id=company_id).first()
        ci = None
        if txn.applied_ci_id:
            ci = cdb.query(CustomerInvoice).filter_by(id=txn.applied_ci_id, company_id=company_id).first()
        return inv, ci

    ref = txn.reference
    if not ref or ref == "ADVANCE":
        return None, None

    # Old rows before applied_ref_id existed: reference is either a raw
    # booking invoice_id or a CustomerInvoice invoice_number.
    inv = cdb.query(Invoice).filter_by(company_id=company_id, invoice_id=ref).first()
    if inv:
        return inv, None
    ci = cdb.query(CustomerInvoice).filter_by(company_id=company_id, invoice_number=ref).first()
    if ci:
        # Can't tell which specific booking under this CI the old row
        # belonged to — nothing safe to reverse at the booking level, but
        # the CI itself can still be re-synced after the transaction is
        # gone using whatever the booking totals now say.
        return None, ci
    return None, None


@app.route("/receipt/delete", methods=["GET"])
@login_required
@owner_required
def delete_receipt():
    txn_id = request.args.get("id", type=int)
    txn_type = request.args.get("type", "cash")  # "cash" or "bank"
    company_id = get_current_company()
    cdb = get_cdb()

    if not company_id or not cdb:
        flash("Could not connect to database.", "error")
        return redirect(url_for("receipt_new"))

    try:
        # ── 1. Find the transaction (cash or bank) ──
        if txn_type == "bank":
            txn = cdb.query(BankTransaction).filter_by(
                id=txn_id,
                company_id=company_id,
                type='credit'
            ).first()
        else:
            txn = cdb.query(CashTransaction).filter_by(
                id=txn_id,
                company_id=company_id,
                type='income',
                category='Receipt'
            ).first()

        if not txn:
            flash("Receipt transaction not found.", "error")
            return redirect(url_for("receipt_new"))

        amount = txn.amount
        client_name = txn.party_name

        print(f"[ADMIN] Deleting {txn_type} receipt: ₹{amount} from {client_name}")

        # ── 2. Find and reset the booking this receipt was applied to ──
        inv, ci = _find_receivable_for_reversal(cdb, company_id, txn)
        if inv:
            inv.paid_amount = max(0, (inv.paid_amount or 0) - amount)
            inv.balance = (inv.grand_total or 0) - (inv.paid_amount or 0)
            if inv.balance <= 0:
                inv.status = "Paid"
            elif inv.paid_amount > 0:
                inv.status = "Partial"
            else:
                inv.status = "Pending"

            client = cdb.query(Client).filter_by(id=inv.client_id, company_id=company_id).first() if inv.client_id else None
            if client and hasattr(client, "pending"):
                client.pending = (client.pending or 0) + amount

        # ── 3. Re-sync the parent customer invoice, if any ──
        if ci:
            _sync_customer_invoice_payment(cdb, company_id, ci)

        # ── 4. If this was a non-cash receipt, debit the bank account back ──
        if txn_type == "bank" and getattr(txn, "bank_account", None):
            txn.bank_account.balance -= amount

        # ── 5. Delete the transaction ──
        cdb.delete(txn)
        cdb.commit()

        flash(f"✅ Receipt of ₹{amount:,.2f} from {client_name} has been deleted and reversed.", "success")

    except Exception as e:
        cdb.rollback()
        print(f"[ADMIN] Error deleting receipt: {e}")
        import traceback
        traceback.print_exc()
        flash(f"Error deleting receipt: {str(e)}", "error")

    return redirect(url_for("receipt_new"))
# ═════════════════════════════════════════════════════════════════════════


if __name__ == "__main__":
    with app.app_context():
        db.create_all()
        seed_database()  # Only platform data
        
        # Seed customer databases for existing companies
        companies = Company.query.all()
        for company in companies:
            try:
                seed_customer_database(company.company_id)
            except Exception as e:
                print(f"Could not seed customer DB for {company.company_id}: {e}")
    app.run(debug=True, port=5010)
else:
    # When run by Gunicorn / Render, seed after the app is fully loaded
    with app.app_context():
        seed_database()  # Only platform data
        
        # Seed customer databases for existing companies
        companies = Company.query.all()
        for company in companies:
            try:
                seed_customer_database(company.company_id)
            except Exception as e:
                print(f"Could not seed customer DB for {company.company_id}: {e}")
